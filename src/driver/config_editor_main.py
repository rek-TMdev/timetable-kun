import os
import sys
import re
import darkdetect
from pathlib import Path
import json
import unicodedata # Added for NFKC normalization
import pykakasi
from deep_translator import GoogleTranslator
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import hashlib
import copy

# Logging infrastructure
try:
    from logging_utils import setup_logger, log_exception
    logger = setup_logger("config_editor")
except ImportError:
    import logging
    logger = logging.getLogger("config_editor")
    logger.addHandler(logging.NullHandler())

# 新しいモジュールのインポート
from config_model import ConfigModel
from config_editor.tab_manager import TabManager
from config_editor.tabs import (
    MasterSubjectTab, AliasTab, PrerequisiteTab, NoTogetherTab,
    GeneralSettingsTab, SavePositionTab, YearSettingsTab,
    RequiredSubjectsTab, SubjectDetailsTab, LayoutTab
)

from PySide6.QtWidgets import(
    QApplication, QMainWindow, QWidget, QVBoxLayout, QStackedWidget,
    QMessageBox, QPushButton, QListWidget, QHBoxLayout, QLabel, QInputDialog, QComboBox,
    QFormLayout, QLineEdit, QCheckBox, QScrollArea, QTableWidget, QTableWidgetItem,
    QSpinBox, QMenu, QDialog, QListWidgetItem, QGroupBox, QColorDialog, QFrame, QSlider, QFileDialog, QTreeWidget, QTreeWidgetItem,
    QAbstractItemView, QDialogButtonBox, QRadioButton, QButtonGroup, QGridLayout
)
from PySide6.QtGui import QAction, QIcon, QColor, QPalette, QPixmap, QPainter
from PySide6.QtCore import Qt, Signal, QTimer, QEvent


def set_button_styles(app, is_dark=None):
    """Applies a theme-aware stylesheet for QPushButton and QComboBox."""
    if is_dark is None:
        is_dark = darkdetect.theme() == "Dark"

    light_style = """
        QPushButton, QComboBox {
            background-color: #3478f6FF;
            color: black;
            border: 1px solid #747474;
            padding: 4px;
            border-radius: 4px;
        }
        QPushButton:hover, QComboBox:hover {
            background-color: #5F92F1;
            border-color: #285ec6;
        }
        QPushButton:pressed {
            background-color: #1a479b;
            border-color: #1a479b;
        }
        QPushButton:checked {
            background-color: #007AFF;
        }
        QPushButton:disabled, QComboBox:disabled {
            background-color: #E4E4E4FF;
            color: #999999;
            border-color: #999999;
        }
        QComboBox QAbstractItemView {
            background-color: white;
            color: black;
            border: 1px solid #3478f6;
            selection-background-color: #285ec6;
            selection-color: white;
        }
    """

    dark_style = """
        QPushButton, QComboBox {
            background-color: #0A4C8DFF;
            color: white;
            border: 1px solid #EBEBEB;
            padding: 4px;
            border-radius: 4px;
        }
        QPushButton:hover, QComboBox:hover {
            background-color: #0166C5;
            border-color: #0077E8;
        }
        QPushButton:pressed {
            background-color: #0062C2;
            border-color: #0062C2;
        }
        QPushButton:checked {
            background-color: #0A84FF;
        }
        QPushButton:disabled, QComboBox:disabled {
            background-color: #1e1e1e;
            color: #999999;
            border-color: #999999;
        }
        QComboBox QAbstractItemView {
            background-color: #2D2D2D;
            color: white;
            border: 1px solid #0A84FF;
            selection-background-color: #0077E8;
            selection-color: white;
        }
    """

    app.setStyleSheet(dark_style if is_dark else light_style)

# Function to get the base path of the application (handles PyInstaller)
def get_base_path():
    # アプリケーションのベースパスを決定
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    else:
        base_path = os.path.abspath(os.path.dirname(sys.argv[0]))
        return Path(base_path)

def get_exe_directory():
    """
    ファイルダイアログ用：exeファイルがある場所（_internalの親）を取得する。
    開発時は通常のベースパスを返す。
    """
    if getattr(sys, 'frozen', False):
        # PyInstallerでexe化されている場合、exeのあるディレクトリを返す
        return Path(sys.executable).parent
    else:
        # 開発時は通常のベースパス
        return Path(os.path.abspath(os.path.dirname(sys.argv[0])))

# --- Start: New helper functions and classes for merge functionality ---

def calculate_sha256(file_path):
    """Calculate the SHA256 hash of a file."""
    if not os.path.exists(file_path):
        return None
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()

# --- New Diffing and Merge Logic ---

def diff_and_merge(base, mine, theirs):
    auto_merged = {}
    conflicts = []

    all_keys = set((base or {}).keys()) | set((mine or {}).keys()) | set((theirs or {}).keys())

    for key in sorted(list(all_keys)):
        val_base = (base or {}).get(key)
        val_mine = (mine or {}).get(key)
        val_theirs = (theirs or {}).get(key)

        in_base = key in (base or {})
        in_mine = key in (mine or {})
        in_theirs = key in (theirs or {})

        if val_mine == val_theirs and in_mine == in_theirs:
            if in_mine:
                auto_merged[key] = val_mine
            continue

        if val_mine == val_base and in_mine == in_base:
            if in_theirs:
                auto_merged[key] = val_theirs
            continue

        if val_theirs == val_base and in_theirs == in_base:
            if in_mine:
                auto_merged[key] = val_mine
            continue

        # Conflict found
        is_mine_dict = isinstance(val_mine, dict) and in_mine
        is_theirs_dict = isinstance(val_theirs, dict) and in_theirs
        is_list_merge_key = key in ["MASTER_SUBJECTS", "YEARS_MESSAGE", "SELECTED_ART_SUBJECT"]
        
        # Check if values are list-like (a list, or None which implies an empty list)
        val_is_list_or_none = lambda v: isinstance(v, list) or v is None

        if is_list_merge_key and val_is_list_or_none(val_mine) and val_is_list_or_none(val_theirs):
            base_set = set(val_base or [])

            # A missing key implies no change from base, not an empty list.
            mine_set = set(val_mine or []) if key in (mine or {}) else base_set
            theirs_set = set(val_theirs or []) if key in (theirs or {}) else base_set

            additions = (mine_set - base_set) | (theirs_set - base_set)
            removals = (base_set - mine_set) | (base_set - theirs_set)
            
            final_set = (base_set | additions) - removals
            auto_merged[key] = sorted(list(final_set))
        elif key.startswith("table_layout"):
            conflicts.append({
                "path": [key], "type": "table_layout", "base": val_base,
                "mine": val_mine, "theirs": val_theirs
            })
        elif is_mine_dict and is_theirs_dict:
            # Standard dictionary recursion
            sub_merged, sub_conflicts = diff_and_merge(val_base if isinstance(val_base, dict) else {}, val_mine, val_theirs)
            if sub_merged:
                auto_merged[key] = sub_merged
            for conflict in sub_conflicts:
                conflict['path'] = [key] + conflict['path']
            conflicts.extend(sub_conflicts)
        else: # Scalar or other type mismatch
            conflicts.append({
                "path": [key], "type": "scalar", "base": val_base,
                "mine": val_mine, "theirs": val_theirs
            })
            
    return auto_merged, conflicts

class ScalarConflictWidget(QWidget):
    def __init__(self, conflict, parent=None):
        super().__init__(parent)
        self.conflict = conflict
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        self.button_group = QButtonGroup(self)

        rb_mine = QRadioButton(f"あなたの変更: {json.dumps(conflict['mine'], ensure_ascii=False)}")
        rb_theirs = QRadioButton(f"外部ファイルの変更: {json.dumps(conflict['theirs'], ensure_ascii=False)}")
        
        self.button_group.addButton(rb_mine, 0)
        self.button_group.addButton(rb_theirs, 1)
        rb_mine.setChecked(True)
        
        form_layout.addRow(rb_mine)
        form_layout.addRow(rb_theirs)
        layout.addLayout(form_layout)
        
    def get_resolved_value(self):
        if self.button_group.checkedId() == 0:
            return self.conflict['mine']
        else:
            return self.conflict['theirs']

class TableLayoutConflictWidget(QWidget):
    def __init__(self, conflict, parent=None):
        super().__init__(parent)
        self.conflict = conflict
        layout = QVBoxLayout(self)
        
        # 選択用のラジオボタン
        self.button_group = QButtonGroup(self)
        rb_mine = QRadioButton("あなたの変更を採用")
        rb_theirs = QRadioButton("外部ファイルの変更を採用")
        self.button_group.addButton(rb_mine, 0)
        self.button_group.addButton(rb_theirs, 1)
        rb_mine.setChecked(True)

        radio_layout = QHBoxLayout()
        radio_layout.addWidget(rb_mine)
        radio_layout.addWidget(rb_theirs)
        layout.addLayout(radio_layout)

        # テーブル表示用のレイアウト
        tables_layout = QHBoxLayout()
        layout.addLayout(tables_layout)

        # 'mine' のテーブル
        mine_group = QGroupBox("あなたのレイアウト")
        mine_layout = QVBoxLayout(mine_group)
        self.mine_table = self.create_table_from_data(conflict.get('mine'))
        mine_layout.addWidget(self.mine_table)
        tables_layout.addWidget(mine_group)

        # 'theirs' のテーブル
        theirs_group = QGroupBox("外部ファイルのレイアウト")
        theirs_layout = QVBoxLayout(theirs_group)
        self.theirs_table = self.create_table_from_data(conflict.get('theirs'))
        theirs_layout.addWidget(self.theirs_table)
        tables_layout.addWidget(theirs_group)

    def create_table_from_data(self, data):
        table = QTableWidget()
        if not data:
            return table
        
        num_rows = len(data)
        num_cols = max(len(row) for row in data) if num_rows > 0 else 0
        table.setRowCount(num_rows)
        table.setColumnCount(num_cols)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers) # 読み取り専用

        for i, row_data in enumerate(data):
            for j, cell_text in enumerate(row_data):
                item = QTableWidgetItem(cell_text)
                table.setItem(i, j, item)
        
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        return table

    def get_resolved_value(self):
        if self.button_group.checkedId() == 0:
            return self.conflict['mine']
        else:
            return self.conflict['theirs']

class ListConflictWidget(QWidget):
    def __init__(self, conflict, parent=None):
        super().__init__(parent)
        self.conflict = conflict
        layout = QGridLayout(self)

        base_set = set(conflict.get('base', []))
        mine_set = set(conflict.get('mine', []))
        theirs_set = set(conflict.get('theirs', []))

        # Calculate diffs
        mine_added = sorted(list(mine_set - base_set))
        theirs_added = sorted(list(theirs_set - base_set))
        
        # Initial merged state: base + all additions from both sides
        initial_merged_set = base_set.union(mine_added).union(theirs_added)

        # UI components
        self.to_add_list = QListWidget()
        self.to_add_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.to_add_list.addItems(sorted(list(set(mine_added + theirs_added))))

        self.merged_list = QListWidget()
        self.merged_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.merged_list.addItems(sorted(list(initial_merged_set)))

        # Buttons
        add_button = QPushButton("追加 →")
        remove_button = QPushButton("← 削除")

        # Layout
        layout.addWidget(QLabel("<b>追加可能な項目</b><br>(両者の変更で追加された項目)"), 0, 0)
        layout.addWidget(self.to_add_list, 1, 0)
        
        button_layout = QVBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(add_button)
        button_layout.addWidget(remove_button)
        button_layout.addStretch()
        layout.addLayout(button_layout, 1, 1)

        layout.addWidget(QLabel("<b>マージ後の最終リスト</b>"), 0, 2)
        layout.addWidget(self.merged_list, 1, 2)

        # Connect signals
        add_button.clicked.connect(self.add_items)
        remove_button.clicked.connect(self.remove_items)

    def add_items(self):
        for item in self.to_add_list.selectedItems():
            if not self.merged_list.findItems(item.text(), Qt.MatchExactly):
                self.merged_list.addItem(item.text())
        self.sort_merged_list()

    def remove_items(self):
        for item in self.merged_list.selectedItems():
            self.merged_list.takeItem(self.merged_list.row(item))
            
    def sort_merged_list(self):
        items = [self.merged_list.item(i).text() for i in range(self.merged_list.count())]
        self.merged_list.clear()
        self.merged_list.addItems(sorted(items))

    def get_resolved_value(self):
        return [self.merged_list.item(i).text() for i in range(self.merged_list.count())]

class ConflictResolutionWindow(QDialog):
    def __init__(self, conflicts, parent=None):
        super().__init__(parent)
        self.setWindowTitle("競合の解決")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.resize(800, 600)
        
        self.conflicts = conflicts
        self.resolver_widgets = []

        # Main layout for the whole dialog
        dialog_layout = QVBoxLayout(self)

        # Top part with the two panels
        main_content_layout = QHBoxLayout()
        
        # Left: Conflict list
        self.conflict_list_widget = QListWidget()
        self.conflict_list_widget.setFixedWidth(250)
        main_content_layout.addWidget(self.conflict_list_widget)

        # Right: Stacked widget for resolver UIs
        self.resolver_stack = QStackedWidget()
        main_content_layout.addWidget(self.resolver_stack)

        for conflict in self.conflicts:
            path_str = " -> ".join(conflict['path'])
            self.conflict_list_widget.addItem(path_str)
            
            widget = None
            if conflict['type'] == 'scalar':
                widget = ScalarConflictWidget(conflict)
            elif conflict['type'] == 'list_simple':
                widget = ListConflictWidget(conflict)
            elif conflict['type'] == 'table_layout':
                widget = TableLayoutConflictWidget(conflict)
            else: # Fallback for unhandled types
                widget = QLabel(f"未対応の競合タイプ: {conflict['type']}\n{json.dumps(conflict, ensure_ascii=False, indent=2)}")
            
            if widget:
                self.resolver_stack.addWidget(widget)
                self.resolver_widgets.append(widget)

        self.conflict_list_widget.currentRowChanged.connect(self.resolver_stack.setCurrentIndex)
        if self.conflict_list_widget.count() > 0:
            self.conflict_list_widget.setCurrentRow(0)

        # Add the top part to the main dialog layout
        dialog_layout.addLayout(main_content_layout)

        # Bottom: OK/Cancel buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.button(QDialogButtonBox.Ok).setText("マージを確定")
        button_box.button(QDialogButtonBox.Cancel).setText("キャンセル")
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        
        # Add the button box to the main dialog layout
        dialog_layout.addWidget(button_box)

    def get_all_resolved_data(self):
        resolved_data = {}
        for i, conflict in enumerate(self.conflicts):
            widget = self.resolver_widgets[i]
            
            value = None
            if isinstance(widget, (ScalarConflictWidget, ListConflictWidget, TableLayoutConflictWidget)):
                value = widget.get_resolved_value()

            if value is not None:
                path = conflict['path']
                d = resolved_data
                for key in path[:-1]:
                    d = d.setdefault(key, {})
                d[path[-1]] = value
        return resolved_data

class SlotSelectionDialog(QDialog):
    """スロットを複数選択するためのカスタムダイアログ"""
    def __init__(self, available_slots, parent=None):
        super().__init__(parent)
        self.setWindowTitle("スロットの選択")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.layout = QVBoxLayout(self)

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QListWidget.MultiSelection)
        for slot in available_slots:
            self.list_widget.addItem(QListWidgetItem(slot))
        self.layout.addWidget(self.list_widget)

        button_box = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("キャンセル")
        button_box.addWidget(self.ok_button)
        button_box.addWidget(self.cancel_button)
        self.layout.addLayout(button_box)

        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

    def get_selected_slots(self):
        return [item.text() for item in self.list_widget.selectedItems()]

class RequiredSubjectGroupWidget(QGroupBox):
    """A widget representing a single 'OR' condition group."""
    changed = Signal()

    def __init__(self, group_number, master_subjects, parent=None):
        super().__init__(f"条件グループ {group_number}", parent)
        self.master_subjects = master_subjects
        
        self.setCheckable(True)
        self.setChecked(False)
        self.setStyleSheet("QGroupBox { border: 1px solid gray; margin-top: 1ex; } QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 0 3px; }")

        layout = QVBoxLayout(self)
        
        req_layout = QHBoxLayout()
        req_layout.addWidget(QLabel("この中から"))
        self.required_spinbox = QSpinBox()
        self.required_spinbox.setMinimum(0)
        req_layout.addWidget(self.required_spinbox)
        req_layout.addWidget(QLabel("教科を必ず選択"))
        req_layout.addStretch()
        layout.addLayout(req_layout)

        subjects_layout = QHBoxLayout()
        
        rule_subjects_v_layout = QVBoxLayout()
        rule_subjects_v_layout.addWidget(QLabel("含まれる教科") )
        self.rule_subjects_list = QListWidget()
        self.rule_subjects_list.setSelectionMode(QListWidget.ExtendedSelection)
        rule_subjects_v_layout.addWidget(self.rule_subjects_list)
        subjects_layout.addLayout(rule_subjects_v_layout)

        move_buttons_layout = QVBoxLayout()
        move_buttons_layout.addStretch()
        self.add_button = QPushButton("←")
        self.add_button.setToolTip("選択可能な教科から追加")
        self.remove_button = QPushButton("→")
        self.remove_button.setToolTip("グループから削除")
        move_buttons_layout.addWidget(self.add_button)
        move_buttons_layout.addWidget(self.remove_button)
        move_buttons_layout.addStretch()
        subjects_layout.addLayout(move_buttons_layout)

        available_subjects_v_layout = QVBoxLayout()
        available_subjects_v_layout.addWidget(QLabel("選択可能な教科") )
        self.available_subjects_list = QListWidget()
        self.available_subjects_list.setSelectionMode(QListWidget.ExtendedSelection)
        available_subjects_v_layout.addWidget(self.available_subjects_list)
        subjects_layout.addLayout(available_subjects_v_layout)
        
        layout.addLayout(subjects_layout)

        self.add_button.clicked.connect(self.add_subjects)
        self.remove_button.clicked.connect(self.remove_subjects)

        self.required_spinbox.valueChanged.connect(lambda: self.changed.emit())
        self.rule_subjects_list.model().rowsInserted.connect(lambda: self.changed.emit())
        self.rule_subjects_list.model().rowsRemoved.connect(lambda: self.changed.emit())
        
        self.toggled.connect(self._update_contents_enabled)
        
        self.update_available_list()
        self._update_contents_enabled(self.isChecked())

    def _update_contents_enabled(self, checked):
        for i in range(self.layout().count()):
            item = self.layout().itemAt(i)
            if item.widget():
                item.widget().setEnabled(checked)
            elif item.layout():
                for j in range(item.layout().count()):
                    if item.layout().itemAt(j).widget():
                        item.layout().itemAt(j).widget().setEnabled(checked)

    def add_subjects(self):
        for item in self.available_subjects_list.selectedItems():
            self.rule_subjects_list.addItem(item.text())
        self.update_available_list()
        self.changed.emit()

    def remove_subjects(self):
        for item in self.rule_subjects_list.selectedItems():
            self.rule_subjects_list.takeItem(self.rule_subjects_list.row(item))
        self.update_available_list()
        self.changed.emit()

    def update_available_list(self):
        self.available_subjects_list.clear()
        rule_subjects = {self.rule_subjects_list.item(i).text() for i in range(self.rule_subjects_list.count())}
        available = sorted([s for s in self.master_subjects if s not in rule_subjects])
        self.available_subjects_list.addItems(available)

    def set_data(self, data):
        self.required_spinbox.blockSignals(True)
        self.rule_subjects_list.blockSignals(True)
        
        self.required_spinbox.setValue(data.get("required", 0))
        self.rule_subjects_list.addItems(data.get("subjects", []))
        self.update_available_list()
        self.setChecked(True)
        
        self.required_spinbox.blockSignals(False)
        self.rule_subjects_list.blockSignals(False)

    def get_data(self):
        subjects = sorted([self.rule_subjects_list.item(i).text() for i in range(self.rule_subjects_list.count())])
        return {"required": self.required_spinbox.value(), "subjects": subjects}

class ArtSubjectDialog(QDialog):
    """特別選択教科を設定するためのダイアログ"""
    def __init__(self, config_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("特別選択教科設定")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.config_data = config_data
        self.master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        
        self.resize(600, 500)

        main_layout = QVBoxLayout(self)

        # Header for ART_SUBJECT
        art_subject_group = QGroupBox("特別枠の数（1年芸術科目など）")
        art_subject_layout = QFormLayout(art_subject_group)
        self.art_subject_spinbox = QSpinBox()
        self.art_subject_spinbox.setRange(0, 10)
        art_subject_layout.addRow("枠数:", self.art_subject_spinbox)
        main_layout.addWidget(art_subject_group)

        # Subject selection part
        selection_group = QGroupBox("対象教科の選択")
        selection_layout = QHBoxLayout(selection_group)
        
        available_layout = QVBoxLayout()
        available_layout.addWidget(QLabel("選択可能な教科") )
        self.available_list = QListWidget()
        self.available_list.setSelectionMode(QListWidget.ExtendedSelection)
        available_layout.addWidget(self.available_list)
        selection_layout.addLayout(available_layout)

        move_buttons_layout = QVBoxLayout()
        move_buttons_layout.addStretch()
        add_button = QPushButton("→")
        add_button.setToolTip("特別教科に追加")
        remove_button = QPushButton("←")
        remove_button.setToolTip("特別教科から削除")
        move_buttons_layout.addWidget(add_button)
        move_buttons_layout.addWidget(remove_button)
        move_buttons_layout.addStretch()
        selection_layout.addLayout(move_buttons_layout)

        selected_layout = QVBoxLayout()
        selected_layout.addWidget(QLabel("選択済みの特別教科") )
        self.selected_list = QListWidget()
        self.selected_list.setSelectionMode(QListWidget.ExtendedSelection)
        selected_layout.addWidget(self.selected_list)
        selection_layout.addLayout(selected_layout)
        
        main_layout.addWidget(selection_group)

        # OK/Cancel buttons
        button_box = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("キャンセル")
        button_box.addStretch()
        button_box.addWidget(self.ok_button)
        button_box.addWidget(self.cancel_button)
        main_layout.addLayout(button_box)

        # Connections
        add_button.clicked.connect(self.add_subjects)
        remove_button.clicked.connect(self.remove_subjects)
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        self.load_settings()

    def load_settings(self):
        # Load ART_SUBJECT value
        art_subject_value = self.config_data.get("ART_SUBJECT", 0)
        try:
            self.art_subject_spinbox.setValue(int(art_subject_value))
        except (ValueError, TypeError):
            self.art_subject_spinbox.setValue(0)

        # Load SELECTED_ART_SUBJECT
        selected_subjects = self.config_data.get("SELECTED_ART_SUBJECT", [])
        self.selected_list.addItems(sorted(selected_subjects))
        
        self.update_available_list()

    def update_available_list(self):
        self.available_list.clear()
        selected_subjects = {self.selected_list.item(i).text() for i in range(self.selected_list.count())}
        available = sorted([s for s in self.master_subjects if s not in selected_subjects])
        self.available_list.addItems(available)

    def add_subjects(self):
        for item in self.available_list.selectedItems():
            self.selected_list.addItem(item.text())
        self.update_available_list()

    def remove_subjects(self):
        for item in self.selected_list.selectedItems():
            self.selected_list.takeItem(self.selected_list.row(item))
        self.update_available_list()

    def save_settings(self):
        self.config_data["ART_SUBJECT"] = self.art_subject_spinbox.value()
        selected_subjects = sorted([self.selected_list.item(i).text() for i in range(self.selected_list.count())])
        if selected_subjects:
            self.config_data["SELECTED_ART_SUBJECT"] = selected_subjects
        elif "SELECTED_ART_SUBJECT" in self.config_data:
            del self.config_data["SELECTED_ART_SUBJECT"]

    def accept(self):
        self.save_settings()
        super().accept()

class AliasGeneratorDialog(QDialog):
    """エイリアスを自動生成する対象の教科を選択するためのダイアログ"""
    def __init__(self, master_subjects, pre_selected_subjects, parent=None):
        super().__init__(parent)
        self.setWindowTitle("エイリアス自動生成")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.resize(400, 500)

        layout = QVBoxLayout(self)

        layout.addWidget(QLabel("エイリアスを生成する教科を選択してください:"))

        # Add select/deselect all buttons
        select_buttons_layout = QHBoxLayout()
        select_all_button = QPushButton("すべて選択")
        deselect_all_button = QPushButton("すべて選択解除")
        select_buttons_layout.addWidget(select_all_button)
        select_buttons_layout.addWidget(deselect_all_button)
        select_buttons_layout.addStretch()
        layout.addLayout(select_buttons_layout)

        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)

        for subject in master_subjects:
            item = QListWidgetItem(subject, self.list_widget)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            if subject in pre_selected_subjects:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)

        button_box = QHBoxLayout()
        self.ok_button = QPushButton("生成実行")
        self.cancel_button = QPushButton("キャンセル")
        button_box.addStretch()
        button_box.addWidget(self.ok_button)
        button_box.addWidget(self.cancel_button)
        layout.addLayout(button_box)

        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        select_all_button.clicked.connect(self.select_all)
        deselect_all_button.clicked.connect(self.deselect_all)

    def get_selected_subjects(self):
        selected_subjects = []
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.Checked:
                selected_subjects.append(item.text())
        return selected_subjects

    def select_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.Checked)

    def deselect_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.Unchecked)

def _find_nodes_at_depth(node, target_depth, current_depth=1, path_prefix=[]):
    if not isinstance(node, dict):
        return []
    if current_depth == target_depth:
        return [path_prefix + [key] for key in node.keys()]
    
    all_paths = []
    for key, children in node.items():
        new_prefix = path_prefix + [key]
        all_paths.extend(_find_nodes_at_depth(children, target_depth, current_depth + 1, new_prefix))
    return all_paths

class InitialUnitsDialog(QDialog):
    def __init__(self, config_data, config_model=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("取得単位数の初期値加算 設定")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.config_data = config_data
        self.config_model = config_model
        self.widgets = {}
        layout = QVBoxLayout(self)

        depth = len(self.config_data.get("YEARS_MESSAGE", []))
        if depth == 0:
            layout.addWidget(QLabel("先に「学部学年選択ダイアログメッセージ設定」でメッセージを設定してください。"))
        else:
            hierarchy = self.config_data.get("YEARS_HIERARCHY", {})
            self.target_paths = _find_nodes_at_depth(hierarchy, depth)

            form_layout = QFormLayout()
            if not self.target_paths:
                layout.addWidget(QLabel(f"{depth}階層目に対応する項目が「学年設定」にありません。"))
            else:
                old_1year_units = self.config_data.get("1YEARS_SUBJECTS_UNITS")

                for path in sorted(self.target_paths):
                    path_key = "_".join(path)
                    config_key = f"YEARS_SUBJECTS_UNITS_{path_key}"
                    
                    spinbox = QSpinBox()
                    spinbox.setRange(0, 200)
                    
                    initial_value = self.config_data.get(config_key)
                    
                    if initial_value is None and depth == 1: 
                        intermediate_key = f"YEARS_SUBJECTS_UNITS_{path[0]}" 
                        initial_value = self.config_data.get(intermediate_key)

                    if initial_value is None and old_1year_units is not None:
                        initial_value = old_1year_units
                    
                    if initial_value is None:
                        initial_value = 0
                    
                    spinbox.setValue(initial_value)
                    
                    display_label = " → ".join(path)
                    form_layout.addRow(f"{display_label}:", spinbox)
                    self.widgets[path_key] = spinbox
                layout.addLayout(form_layout)

        button_box = QHBoxLayout()
        ok_button = QPushButton("OK")
        cancel_button = QPushButton("キャンセル")
        button_box.addStretch()
        button_box.addWidget(ok_button)
        button_box.addWidget(cancel_button)
        layout.addLayout(button_box)

        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)

    def accept(self):
        keys_to_delete = [k for k in self.config_data if k.startswith("YEARS_SUBJECTS_UNITS_")]
        for key in keys_to_delete:
            if key in self.config_data:
                del self.config_data[key]
            # config_modelからも削除
            if self.config_model and key in self.config_model._data:
                del self.config_model._data[key]

        for path_key, spinbox in self.widgets.items():
            config_key = f"YEARS_SUBJECTS_UNITS_{path_key}"
            value = spinbox.value()
            if value > 0:
                self.config_data[config_key] = value
                # config_modelにも保存
                if self.config_model:
                    self.config_model._data[config_key] = value
                    self.config_model._is_modified = True
                    print(f"DEBUG: Set {config_key} = {value} in config_model._data")

        super().accept()

class CustomTreeWidget(QTreeWidget):
    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        if not self.itemAt(event.pos()):
            self.clearSelection()

class ExcelImportDialog(QDialog):
    def __init__(self, years, config_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Excelからインポート")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.years = years
        self.config_data = config_data
        self.workbook = None
        self.sheet_widgets = {}
        self.generated_data = None

        self.layout = QVBoxLayout(self)

        icon_path = get_base_path() / "時間割くんドライバアイコン.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))

        file_layout = QHBoxLayout()
        self.file_path_label = QLabel("ファイルが選択されていません")
        load_file_button = QPushButton("Excelファイルを開く")
        load_file_button.clicked.connect(self.load_excel_file)
        file_layout.addWidget(QLabel("ファイル:"))
        file_layout.addWidget(self.file_path_label)
        file_layout.addWidget(load_file_button)
        self.layout.addLayout(file_layout)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        self.layout.addWidget(scroll_area)

        container = QWidget()
        self.main_settings_layout = QVBoxLayout(container)
        scroll_area.setWidget(container)

        button_box = QHBoxLayout()
        self.import_button = QPushButton("インポート実行")
        self.import_button.setEnabled(False)
        self.import_button.clicked.connect(self.run_import)
        cancel_button = QPushButton("キャンセル")
        cancel_button.clicked.connect(self.reject)
        button_box.addStretch()
        button_box.addWidget(self.import_button)
        button_box.addWidget(cancel_button)
        self.layout.addLayout(button_box)

    def load_excel_file(self):
        dialog = QFileDialog(self, "Excelファイルを選択", str(get_base_path()), "Excel Files (*.xlsx *.xlsm)")
        dialog.setWindowFlags(dialog.windowFlags() | Qt.WindowStaysOnTopHint)
        dialog.raise_()
        dialog.activateWindow()

        file_path = None
        if dialog.exec():
            selected_files = dialog.selectedFiles()
            if selected_files:
                file_path = selected_files[0]
        
        if not file_path:
            return

        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.file_path_label.setText(Path(file_path).name)
            self.populate_settings()
            self.import_button.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"Excelファイルの読み込みに失敗しました: {e}")
            self.workbook = None
            self.file_path_label.setText("ファイルが選択されていません")
            self.import_button.setEnabled(False)

    def populate_settings(self):
        for i in reversed(range(self.main_settings_layout.count())):
            widget = self.main_settings_layout.itemAt(i).widget()
            if widget is not None:
                widget.setParent(None)
        self.sheet_widgets.clear()

        sheet_names = self.workbook.sheetnames

        for year in self.years:
            group_box = QGroupBox(year)
            form_layout = QFormLayout(group_box)

            sheet_combo = QComboBox()
            sheet_combo.addItem("(未選択)")
            sheet_combo.addItems(sheet_names)
            
            start_cell_edit = QLineEdit("B8")
            end_cell_edit = QLineEdit("Y102")
            range_layout = QHBoxLayout()
            range_layout.addWidget(start_cell_edit)
            range_layout.addWidget(QLabel("から"))
            range_layout.addWidget(end_cell_edit)

            form_layout.addRow("シート名:", sheet_combo)
            form_layout.addRow("データ範囲 (ヘッダーを含まない):", range_layout)
            
            self.sheet_widgets[year] = {
                "sheet_combo": sheet_combo,
                "start_cell_edit": start_cell_edit,
                "end_cell_edit": end_cell_edit
            }
            self.main_settings_layout.addWidget(group_box)
        self.main_settings_layout.addStretch()

    def run_import(self):
        import logging
        generated_data = {}
        error_occurred = False
        try:
            for year, widgets in self.sheet_widgets.items():
                sheet_name = widgets["sheet_combo"].currentText()
                start_cell = widgets["start_cell_edit"].text().strip()
                end_cell = widgets["end_cell_edit"].text().strip()
                
                if not sheet_name or sheet_name == "(未選択)" or not start_cell or not end_cell:
                    continue
                
                range_str = f"{start_cell}:{end_cell}"
                ws = self.workbook[sheet_name]
                
                try:
                    min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(range_str)
                except Exception as e:
                    QMessageBox.warning(self, "範囲エラー", f"""無効なデータ範囲です: {range_str}\n{e}""" )
                    error_occurred = True
                    continue

                if min_row < 2:
                    QMessageBox.warning(self, "範囲エラー", f"データ範囲 ({range_str}) は、スロット名ヘッダーのために2行目以降から開始する必要があります。" )
                    error_occurred = True
                    continue

                slot_headers = {}
                header_row = min_row - 1
                for col_idx in range(min_col, max_col + 1):
                    slot_name_raw = ws.cell(row=header_row, column=col_idx).value
                    if slot_name_raw and str(slot_name_raw).strip():
                        full_name = str(slot_name_raw).strip()
                        match = re.match(r'^[A-Z]+\d+', full_name)
                        slot_headers[col_idx] = match.group(0) if match else full_name

                if not slot_headers:
                    QMessageBox.warning(self, "ヘッダーエラー", f"""スロット名ヘッダーが見つかりませんでした。\nデータ範囲 ({range_str}) の【一行上】にヘッダー行が設定されているか確認してください。""" )
                    error_occurred = True
                    continue

                config_slots_key = f"ALL_SLOTS_{year}"
                config_slots = self.config_data.get(config_slots_key)

                if not config_slots:
                    reply = QMessageBox.warning(self, "スロットレイアウト未定義",
                        f"「{year}」のスロットレイアウトが「スロットレイアウト設定」で定義されていません。\nExcelファイルからスロットを読み込んで設定を続行しますか？",
                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                    if reply == QMessageBox.No:
                        error_occurred = True
                        continue
                else:
                    excel_slots = set(slot_headers.values())
                    config_slots_set = set(config_slots)
                    
                    missing_slots = config_slots_set - excel_slots
                    if missing_slots:
                        reply = QMessageBox.warning(self, "スロット不足",
                            f"「{year}」のレイアウトで定義されているスロットのうち、以下のものがExcelに見つかりません:"
                            f"{', '.join(sorted(list(missing_slots)))}"
                            "このままインポートを続行しますか？"
                            "（不足しているスロットのデータはインポートされません）",
                            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                        if reply == QMessageBox.No:
                            error_occurred = True
                            continue

                temp_subject_data = {}
                for row_idx in range(min_row, max_row + 1, 2):
                    if row_idx + 1 > max_row: continue

                    for col_idx in range(min_col, max_col + 1):
                        subject_cell = ws.cell(row=row_idx, column=col_idx)
                        subject_name_raw = subject_cell.value
                        
                        if subject_name_raw and str(subject_name_raw).strip() != "":
                            full_subject_name = unicodedata.normalize('NFKC', str(subject_name_raw).strip())
                            
                            match = re.match(r'^([A-Z]+)(.*)$', full_subject_name)
                            if match:
                                group_prefix = match.group(1)
                                subject_name = match.group(2).strip()
                            else:
                                group_prefix = "DEFAULT"
                                subject_name = full_subject_name

                            number_cell = ws.cell(row=row_idx + 1, column=col_idx)
                            number_value_raw = number_cell.value

                            if number_value_raw is not None and str(number_value_raw).strip() != "":
                                slot_name = slot_headers.get(col_idx)
                                if slot_name:
                                    if subject_name not in temp_subject_data:
                                        temp_subject_data[subject_name] = {}
                                    if group_prefix not in temp_subject_data[subject_name]:
                                        temp_subject_data[subject_name][group_prefix] = {}
                                    
                                    number_value = str(number_value_raw).strip()
                                    temp_subject_data[subject_name][group_prefix][slot_name] = int(number_value) if number_value.isdigit() else number_value

                subject_number_list = []
                subject_slots_base_list = []

                for subject_name, groups in temp_subject_data.items():
                    all_slot_data_for_subject = {}
                    all_slot_groups_for_subject = []

                    for group_prefix, slot_data in groups.items():
                        all_slot_data_for_subject.update(slot_data)
                        all_slot_groups_for_subject.append(sorted(list(slot_data.keys())))

                    if all_slot_data_for_subject:
                        subject_number_list.append({"name": subject_name, "data": all_slot_data_for_subject})
                    if all_slot_groups_for_subject:
                        subject_slots_base_list.append({"name": subject_name, "data": sorted(all_slot_groups_for_subject)})

                generated_data[f"subject_number{year}"] = subject_number_list
                generated_data[f"subject_slots_base{year}"] = subject_slots_base_list

            if not error_occurred:
                self.generated_data = generated_data
                self.accept()

        except Exception as e:
            logging.error(f"An unexpected error occurred during import: {e}", exc_info=True)
            QMessageBox.critical(self, "インポートエラー", f"データの処理中に予期せぬエラーが発生しました: {e}")

class ExcelExportDialog(QDialog):
    def __init__(self, years, config_data, parent=None):
          super().__init__(parent)
          self.setWindowTitle("Excelへエクスポート")
          self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
          self.years = years
          self.config_data = config_data
          self.workbook = None
          self.file_path = None
          self.sheet_widgets = {}

          self.layout = QVBoxLayout(self)

          file_layout = QHBoxLayout()
          self.file_path_label = QLabel("ファイルが選択されていません")
          select_file_button = QPushButton("Excelファイルを開く")
          select_file_button.clicked.connect(self.select_file)
          file_layout.addWidget(QLabel("ファイル:"))
          file_layout.addWidget(self.file_path_label)
          file_layout.addWidget(select_file_button)
          self.layout.addLayout(file_layout)

          scroll_area = QScrollArea()
          scroll_area.setWidgetResizable(True)
          self.layout.addWidget(scroll_area)
          container = QWidget()
          self.main_settings_layout = QVBoxLayout(container)
          scroll_area.setWidget(container)

          button_box = QHBoxLayout()
          self.export_button = QPushButton("エクスポート実行")
          self.export_button.setEnabled(False)
          self.export_button.clicked.connect(self.run_export)
          cancel_button = QPushButton("キャンセル")
          cancel_button.clicked.connect(self.reject)
          button_box.addStretch()
          button_box.addWidget(self.export_button)
          button_box.addWidget(cancel_button)
          self.layout.addLayout(button_box)

    def select_file(self):
          dialog = QFileDialog(self, "保存先Excelファイルを選択", str(get_base_path()), "Excel Files(*.xlsx)")
          dialog.setFileMode(QFileDialog.AnyFile)
          dialog.setAcceptMode(QFileDialog.AcceptSave)
          dialog.setWindowFlags(dialog.windowFlags() | Qt.WindowStaysOnTopHint)
          dialog.raise_()
          dialog.activateWindow()

          file_path = None
          if dialog.exec():
              selected_files = dialog.selectedFiles()
              if selected_files:
                  file_path = selected_files[0]
          
          if not file_path: return
          self.file_path = file_path
          self.file_path_label.setText(Path(file_path).name)

          try:
              self.workbook = openpyxl.Workbook() if not Path(self.file_path).exists() else openpyxl.load_workbook(self.file_path)
              self.populate_settings()
              self.export_button.setEnabled(True)
          except Exception as e:
              QMessageBox.critical(self, "エラー", f"Excelファイルの準備中にエラーが発生しました: {e}")
              self.export_button.setEnabled(False)

    def populate_settings(self):
          for i in reversed(range(self.main_settings_layout.count())):
              widget = self.main_settings_layout.itemAt(i).widget()
              if widget: widget.setParent(None)
          self.sheet_widgets.clear()

          sheet_names = self.workbook.sheetnames

          for year in self.years:
              group_box = QGroupBox(year)
              form_layout = QFormLayout(group_box)
              sheet_combo = QComboBox()
              sheet_combo.addItem("(未選択)")
              sheet_combo.addItems(sheet_names)

              start_cell_edit = QLineEdit("B8")
              range_layout = QHBoxLayout()
              range_layout.addWidget(start_cell_edit)
              range_layout.addWidget(QLabel("から"))

              form_layout.addRow("書き込み先シート名:", sheet_combo)
              form_layout.addRow("データ書き込み開始セル:", range_layout)
              self.sheet_widgets[year] = {
                  "sheet_combo": sheet_combo,
                  "start_cell_edit": start_cell_edit,
              }
              self.main_settings_layout.addWidget(group_box)
          self.main_settings_layout.addStretch()

    def run_export(self):
        try:
            for year, widgets in self.sheet_widgets.items():
                sheet_name = widgets["sheet_combo"].currentText()
                start_cell = widgets["start_cell_edit"].text().strip()

                if not sheet_name or sheet_name == "(未選択)" or not start_cell:
                    continue

                ws = self.workbook.create_sheet(title=sheet_name) if sheet_name not in self.workbook.sheetnames else self.workbook[sheet_name]

                min_col, min_row, _, _ = openpyxl.utils.cell.range_boundaries(f"{start_cell}:{start_cell}")

                subject_number_data = self.config_data.get(f"subject_number{year}", [])
                all_slots = sorted(self.config_data.get(f"ALL_SLOTS_{year}", []))
                if not all_slots: continue

                slot_header_row = min_row - 1
                row_num_header_col = min_col - 1
                if slot_header_row < 1 or row_num_header_col < 1:
                    QMessageBox.warning(self, "範囲エラー", "ヘッダーを書き込むため、データ範囲はB2セル以降から開始してください。" )
                    return

                for i, slot_name in enumerate(all_slots):
                    ws.cell(row=slot_header_row, column=min_col + i, value=slot_name).alignment = Alignment(horizontal='center', vertical='center')

                col_to_next_row = {i + min_col: min_row for i, _ in enumerate(all_slots)}
                for subject_info in subject_number_data:
                    subject_name = subject_info.get("name")
                    if not subject_name: continue
                    data = subject_info.get("data", {})
                    for slot, value in data.items():
                        if slot in all_slots:
                           col_idx = all_slots.index(slot) + min_col
                           row_idx = col_to_next_row[col_idx]
                           ws.cell(row=row_idx, column=col_idx, value=subject_name).alignment = Alignment(horizontal='center', vertical='center')
                           ws.cell(row=row_idx + 1, column=col_idx, value=value).alignment = Alignment(horizontal='center', vertical='center')
                           col_to_next_row[col_idx] = row_idx + 2

            self.workbook.save(self.file_path)
            QMessageBox.information(self, "成功", f"設定をExcelファイルに書き込みました:\n{self.file_path}")
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "エクスポートエラー", f"Excelファイルへの書き込み中にエラーが発生しました:\n{e}")

class MainWindow(QMainWindow):
    leaf_selection_changed = Signal()
    subject_details_updated = Signal()
    table_layout_leaf_selection_changed = Signal()
    subject_added_to_master = Signal(str)
    save_pos_leaf_selection_changed = Signal()
    prereq_selection_changed = Signal()
    prerequisite_subject_added = Signal(str, str)
    no_together_group_added = Signal()
    no_together_subject_added = Signal()
    def __init__(self, tutorial_mode=False, initial_data=None, splash=None):
        super().__init__()
        self.splash = splash # スプラッシュを保持
        self.setWindowTitle("時間割くん - ドライバ")
        self.resize(1000, 900)
        self.config_data = {}
        self.tutorial_mode = tutorial_mode
        self.initial_data = initial_data
        self.splash = splash
        self.translator = GoogleTranslator(source='ja', target='en')
        
        # ConfigModelを初期化（新しいアーキテクチャ）
        self.config_model = ConfigModel()
        
        # --- Attributes for conflict detection ---
        self.file_path = None
        self.config_hash = None
        self.original_config_data = {}
        # ---

        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            self.base_path = sys._MEIPASS
        else:
            self.base_path = os.path.abspath(os.path.dirname(sys.argv[0]))

        # Define the single color to use based on the theme.
        self.target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"

        self.years = []
        self.current_selected_subject_for_alias = None
        self.general_settings_widgets = {}
        self.current_selected_table_layout_key = None
        self.current_details_context = {"year": None, "subject": None}
        self.details_number_widgets = {}
        self.current_rule_key = None
        self.current_prereq_parent = None
        self.current_save_pos_year = None
        self.current_editing_item_text = None
        self.previous_use_numbers_state = False

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)

        content_layout = QHBoxLayout()

        # 左側のパネルを作成
        left_panel = QWidget()
        left_panel_layout = QVBoxLayout(left_panel)
        left_panel.setFixedWidth(150)

        # タブリストを左側のパネルに追加
        self.tab_list = QListWidget()
        left_panel_layout.addWidget(self.tab_list)

        # ファイル操作ボタンとメニューを作成
        icon_path_load = os.path.join(self.base_path, "svgs", "load_file.svg")
        load_file_icon = self._create_icon_from_svg_data(icon_path_load, self.target_color)
        self.file_button = QPushButton(load_file_icon, "ファイルを開く")
        file_menu = QMenu(self.file_button)
        
        # 設定ファイルを開くアクション
        open_config_action = file_menu.addAction("設定ファイルを開く")
        open_config_action.triggered.connect(self.open_config_file_dialog)

        integration_action = file_menu.addAction("設定ファイルを統合")
        integration_action.triggered.connect(self.integrate_config_files)
        file_menu.addSeparator() # セパレータを追加

        # Excelファイルから読み込みアクション
        import_excel_action = file_menu.addAction("Excelファイルから読み込み")
        import_excel_action.triggered.connect(self.open_excel_import_dialog)
        
        # Excelファイルへ書き込みアクション
        export_excel_action = file_menu.addAction("Excelファイルへ書き込み")
        export_excel_action.triggered.connect(self.export_to_excel)
        
        self.file_button.setMenu(file_menu)
        left_panel_layout.addWidget(self.file_button)

        # 設定を保存ボタンを左側のパネルに追加
        icon_path_save = os.path.join(self.base_path, "svgs", "save.svg")
        save_file_icon = self._create_icon_from_svg_data(icon_path_save, self.target_color)
        self.save_button = QPushButton(save_file_icon, "設定を保存")
        save_memu = QMenu(self.save_button)

        save_own_action = save_memu.addAction("上書き保存")
        save_own_action.triggered.connect(self.save_config)

        save_as_action = save_memu.addAction("名前を付けて保存")
        save_as_action.triggered.connect(self.save_config_as)

        self.save_button.setMenu(save_memu)
        left_panel_layout.addWidget(self.save_button)

        
        # 左側のパネルをメインのレイアウトに追加
        content_layout.addWidget(left_panel)

        self.stacked_widget = QStackedWidget()
        content_layout.addWidget(self.stacked_widget)
        self.main_layout.addLayout(content_layout)

        self.create_subject_details_tab()
        self.create_master_subject_tab()
        self.create_alias_tab()
        self.create_table_layout_tab()
        self.create_required_subjects_tab()
        self.create_prerequisite_tab()
        self.create_no_together_tab()
        self.create_save_position_tab()
        self.create_year_settings_tab()
        self.create_general_settings_tab()

        self.tab_list.currentRowChanged.connect(self.on_tab_changed)

        bottom_layout = QHBoxLayout()
        bottom_layout.addWidget(QLabel("フォントサイズ:"))
        self.font_size_spinbox = QSpinBox()
        self.font_size_spinbox.setRange(8, 24)
        self.font_size_spinbox.valueChanged.connect(self.change_font_size)
        bottom_layout.addWidget(self.font_size_spinbox)
        bottom_layout.addStretch()
        self.main_layout.addLayout(bottom_layout)

        icon_path = get_base_path() / "時間割くんドライバアイコン.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        else:
            print(f"アイコンファイルが見つかりません:{icon_path}")

        if self.tutorial_mode:
            self.save_button.setEnabled(False)

        self.load_config()

    def is_windows_light_theme():
        """Checks if the user is using a light theme on Windows."""
        if sys.platform != "win32":
            return True # Default to light theme on non-Windows platforms
        try:
            import winreg
            key_path = r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path)
            value, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
            return value == 1
        except (ImportError, FileNotFoundError, OSError):
            # In case of any error, default to a light theme.
            return True

    def is_dark_theme(self):
        """Determines if a dark theme is active, based on config or OS settings."""
        ini_theme= darkdetect.theme()
        if ini_theme == "Dark":
            return True
        if ini_theme == "Light":
            return False
        # If config theme is not set or invalid, fallback to OS detection
        os_is_light = self.is_windows_light_theme()
        return not os_is_light


    def changeEvent(self, event):
        """Handle theme changes dynamically."""
        if event.type() == QEvent.ThemeChange:
             # Update target color for icons
            self.target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"
            
            # Apply button styles
            set_button_styles(QApplication.instance(), is_dark=self.is_dark_theme())
            
            # Refresh all icons
            self.update_icons()
            
        super().changeEvent(event)

    def update_icons(self):
        """Update all SVG icons with the current target color."""
        # Update Main Window buttons
        if hasattr(self, 'file_button'):
            icon_path_load = os.path.join(self.base_path, "svgs", "load_file.svg")
            new_load_icon = self._create_icon_from_svg_data(icon_path_load, self.target_color)
            if new_load_icon:
                self.file_button.setIcon(new_load_icon)

        if hasattr(self, 'save_button'):
            icon_path_save = os.path.join(self.base_path, "svgs", "save.svg")
            new_save_icon = self._create_icon_from_svg_data(icon_path_save, self.target_color)
            if new_save_icon:
                self.save_button.setIcon(new_save_icon)

        # Update icons in all tabs
        for i in range(self.stacked_widget.count()):
            tab = self.stacked_widget.widget(i)
            if hasattr(tab, 'update_icons'):
                tab.update_icons()

    def _create_icon_from_svg_data(self, svg_path, color_hex):
        """ 指定された色のSVGアイコンを作成する """
        try:
            from PySide6.QtSvg import QSvgRenderer
        except ImportError:
            return None

        if not os.path.exists(svg_path):
            print(f"SVG file not found: {svg_path}")
            return None

        with open(svg_path, 'r', encoding='utf-8') as f:
            svg_data = f.read()
        
        colored_svg_data = re.sub('#1f1f1f', color_hex, svg_data, flags=re.IGNORECASE)
        
        renderer = QSvgRenderer(colored_svg_data.encode('utf-8'))
        if not renderer.isValid():
            print(f"Invalid SVG data for: {svg_path}")
            return None
        
        pixmap = QPixmap(32, 32)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        renderer.render(painter)
        painter.end()

        return QIcon(pixmap)

    def change_font_size(self, size):
        """フォントサイズを変更し、すべてのウィジェットに適用"""
        font = self.font()
        font.setPointSize(size)
        self.setFont(font)
        # 全ての子ウィジェットにもフォントを適用
        for widget in self.findChildren(QWidget):
            widget.setFont(font)
        # アプリケーション全体にも設定
        QApplication.setFont(font)

    def on_tab_changed(self, index):
        previous_index = self.stacked_widget.currentIndex()
        if previous_index < 0:
            self.stacked_widget.setCurrentIndex(index)
            return

        previous_tab_text = self.tab_list.item(previous_index).text()

        if hasattr(self, 'table_widget'):
            self.table_widget.clearSelection()

        # All calls to update_..._in_config methods are removed.
        # Data saving/updating should only happen on explicit save actions.

        self.stacked_widget.setCurrentIndex(index)

    def closeEvent(self, event):
        """ウィンドウを閉じる際に、未保存の変更を確認します。"""
        # チュートリアルモードでは確認しない
        if self.tutorial_mode:
            event.accept()
            return

        current_config = self.get_current_ui_config()
        original_config = copy.deepcopy(self.original_config_data)

        # 比較のために、両方の辞書から揮発性のキーを削除し、リストを正規化（ソート）する
        for config in [original_config, current_config]:
            if "LAST_UPDATE_DAY" in config:
                del config["LAST_UPDATE_DAY"]
            if "MASTER_SUBJECTS" in config and config["MASTER_SUBJECTS"] is not None:
                config["MASTER_SUBJECTS"] = sorted(list(set(config["MASTER_SUBJECTS"])))
            if "SUBJECT_ALIASES" in config and config["SUBJECT_ALIASES"] is not None:
                for subject in config["SUBJECT_ALIASES"]:
                    config["SUBJECT_ALIASES"][subject] = sorted(list(set(config["SUBJECT_ALIASES"].get(subject, []))))

        if original_config != current_config:
            reply = QMessageBox.question(self, "未保存の変更",
                                           "保存されていないデータは削除されます。保存せずに終了しますか？",
                                           QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                           QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    def load_config(self):
        if self.tutorial_mode:
            if self.initial_data:
                self.config_data = self.initial_data
                self.original_config_data = copy.deepcopy(self.initial_data)
                self.config_hash = None
                self.refresh_all_ui_from_config_data()
            else:
                QMessageBox.critical(self, "エラー", "チュートリアルデータが見つかりません。" )
                self.config_data = {}
        else:
            if not self.file_path:
                # 1. ダイアログのインスタンス化
                dialog = QFileDialog(self)
                
                # 【最重要】OS標準のダイアログを使わない設定にする（これでアイコンが表示可能になる）
                dialog.setOption(QFileDialog.DontUseNativeDialog)
                
                dialog.setWindowTitle("編集するconfigファイルを選択")
                dialog.setDirectory(str(get_base_path()))
                dialog.setNameFilter("jsonファイル (*.json)")
                dialog.setFileMode(QFileDialog.ExistingFile)

                # 2. アイコンをセット
                icon_path = get_base_path() / "時間割くんドライバアイコン.ico"
                if icon_path.exists():
                    app_icon = QIcon(str(icon_path))
                    dialog.setWindowIcon(app_icon)
                    # 念のためメインウィンドウ（自分自身）にもセットしておく
                    self.setWindowIcon(app_icon)
                
                # 3. 最前面表示とフォーカスの設定
                dialog.setWindowFlags(dialog.windowFlags() | Qt.WindowStaysOnTopHint)

                # 4. スプラッシュの制御
                if self.splash: self.splash.hide()
                
                path_to_set = None
                # ダイアログを前面へ引き出す
                dialog.raise_()
                dialog.activateWindow()
                
                if dialog.exec():
                    selected_files = dialog.selectedFiles()
                    if selected_files:
                        path_to_set = selected_files[0]
                
                if self.splash: 
                    self.splash.show()
                    # スプラッシュ表示後にイベントを処理して画面を更新
                    QApplication.processEvents()

                self.file_path = path_to_set

            if not self.file_path:
                return
            
            self.reload_from_file(self.file_path, show_success_message=True)
    def reload_from_file(self, file_path, show_success_message=False):
        """指定されたファイルパスから設定を再読み込みし、UIを更新します。"""
        
        # Master Config Protection: マスター設定ファイル名を検出
        file_name = Path(file_path).name.lower()
        is_master_config = "master" in file_name or "マスター" in file_name
        
        if is_master_config and not getattr(self, '_master_config_warning_shown', False):
            reply = QMessageBox.warning(
                self,
                "マスター設定ファイルの編集",
                "⚠️ これはマスター設定ファイルです。\n\n"
                "この設定を変更すると、すべてのユーザーに影響を与える可能性があります。\n\n"
                "本当に開きますか？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                logger.info(f"マスター設定ファイル編集をキャンセル: {file_path}")
                return
            self._master_config_warning_shown = True
            logger.warning(f"マスター設定ファイルを編集モードで開きます: {file_path}")
        
        self.file_path = file_path
        try:
            with open(self.file_path, "rb") as f:
                content_bytes = f.read()
            self.config_hash = hashlib.sha256(content_bytes).hexdigest()
            self.config_data = json.loads(content_bytes.decode('utf-8'))
            self.original_config_data = copy.deepcopy(self.config_data)
            
            # ConfigModelにもデータを同期（新しいアーキテクチャ）
            self.config_model.update_data(copy.deepcopy(self.config_data))
            
        except FileNotFoundError:
            QMessageBox.critical(self, "エラー", "config.jsonが見つかりません。" )
            self.config_data, self.original_config_data, self.config_hash = {}, {}, None
        except json.JSONDecodeError:
            QMessageBox.critical(self, "エラー", "config.jsonの解析に失敗しました。" )
            self.config_data, self.original_config_data, self.config_hash = {}, {}, None
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"ファイルの読み込み中に予期せぬエラーが発生しました: {e}")
            return

        self.refresh_all_ui_from_config_data()

    def refresh_all_ui_from_config_data(self):
        """self.config_dataの内容に基づいてUI全体を再描画します。"""
        font_size = self.config_data.get("EDITOR_FONT_SIZE", 10)
        self.font_size_spinbox.blockSignals(True)
        self.font_size_spinbox.setValue(font_size)
        self.font_size_spinbox.blockSignals(False)
        self.change_font_size(font_size)

        if "YEARS_HIERARCHY" not in self.config_data:
            self.config_data["YEARS_HIERARCHY"] = {}

        self.update_years_in_config()
        self._pre_populate_all_slots()
        self.populate_master_subject_list()
        self.reload_all_selectors()
        self.populate_general_settings()
        self.populate_required_subjects_settings()
        self.populate_theme_settings()
        self.populate_no_together_groups_list()
        self.populate_year_list()
        
        current_index = self.tab_list.currentRow()
        if current_index > -1:
            # This logic might need adjustment, but let's keep it for now
            self.on_tab_changed(-1) 
            self.on_tab_changed(current_index)
            tab_text = self.tab_list.item(current_index).text()
            if tab_text == "科目詳細設定":
                self.update_details_subject_lists()
                self.populate_subject_details()
            elif tab_text == "科目エイリアス設定":
                 if self.alias_subject_list.currentItem():
                    self.on_alias_subject_selected(self.alias_subject_list.currentRow())
            elif tab_text == "必須教科設定":
                if self.rules_list_widget.currentItem():
                    self.on_rule_selected()
            elif tab_text == "前提教科設定":
                 if self.prereq_parent_list.currentItem():
                    self.on_prereq_parent_selected(self.prereq_parent_list.currentRow())
            elif tab_text == "保存位置設定":
                if self.save_pos_leaf_list.currentItem():
                    self.on_save_pos_leaf_selected(self.save_pos_leaf_list.currentRow())

    def get_current_ui_config(self):
        """現在のUIの状態をすべて反映したconfigのディクショナリを返します。"""
        # Create a brand new dictionary to hold the UI's current state.
        # Start with a deep copy of self.config_data to retain any keys not managed by the UI
        # and then overwrite/add values based on the current UI state.
        current_ui_state = copy.deepcopy(self.config_data)

        # Call helper methods to populate current_ui_state from UI widgets.
        # These helper methods will take `current_ui_state` as an argument.

        self._get_subject_details_from_config_data(current_ui_state)

        self.update_master_subjects_in_config(current_ui_state)
        self._get_aliases_from_ui(current_ui_state)

        # Update table layouts and fixed slots (assuming self.config_data is kept current by table_layout methods)
        for key in list(self.config_data.keys()):
            if key.startswith("table_layout") or key.startswith("FIXED_SLOTS") or key.startswith("ALL_SLOTS_"):
                current_ui_state[key] = copy.deepcopy(self.config_data[key])
        
        # self.update_table_layout_in_config(current_ui_state) # This call is no longer needed with the copying approach

        self.update_general_settings_in_config(current_ui_state)
        self._get_required_subjects_from_ui(current_ui_state)
        self._get_prerequisites_from_ui(current_ui_state)
        self._get_no_together_from_ui(current_ui_state)
        self._get_save_position_from_config_data(current_ui_state)
        self._rebuild_hierarchy_from_tree_and_update_config(current_ui_state) # New helper for hierarchy
        self._get_theme_settings_from_ui_and_update_config(current_ui_state)

        current_ui_state['EDITOR_FONT_SIZE'] = self.font_size_spinbox.value()

        # Perform cleanup for keys that might exist but should be removed if empty or not applicable
        keys_to_delete_on_save = [k for k in current_ui_state if k.startswith("YEARS_SUBJECTS_UNITS_") or k in ["YEAR", "1YEARS_SUBJECTS_UNITS", "ABNORMAL_SUBJECTS_UNITS", "LAST_SELECTED_HIERARCHY", "LAST_SELECTED_PATH", "USE_ONLY_NAME"]]
        for key in keys_to_delete_on_save:
            if key in current_ui_state: del current_ui_state[key]

        # Consolidated cleanup for empty config sections
        for key in list(current_ui_state.keys()):
            if (key.startswith(("REQUIRED_SUBJECTS_", "FIXED_SLOTS", "table_layout", "ALL_SLOTS", "SAVE_POSITION")) and not current_ui_state[key]) or \
               (key in ["SUBJECT_ALIASES", "PREREQUISITE_SUBJECTS", "NO_TOGETHER_SUBJECTS", "YEARS_MESSAGE", "YEARS_HIERARCHY"] and not current_ui_state[key]):
                del current_ui_state[key]
        
        return current_ui_state




    def populate_required_subjects_settings(self, reselect_key=None):
        self.rules_list_widget.blockSignals(True)
        self.rules_list_widget.clear()
        
        self.details_pane.setVisible(False)
        self.current_rule_key = None # This might become obsolete, we'll use the selected item directly
        
        all_req_keys = [k for k in self.config_data if k.startswith("REQUIRED_SUBJECTS_")]
        
        all_rules_for_list = []
        for key in all_req_keys:
            target_suffix = key.replace("REQUIRED_SUBJECTS_", "")
            rules_in_target = self.config_data.get(key, {})
            for rule_id, rule_data in rules_in_target.items():
                # Augment the rule_data with its id and target for easier handling
                full_rule_data = rule_data.copy()
                full_rule_data['id'] = rule_id
                full_rule_data['target'] = target_suffix
                all_rules_for_list.append(full_rule_data)
        
        all_rules_for_list.sort(key=lambda x: (x.get('target', ''), x.get('name', '')))

        row_to_select = -1
        for i, rule_data in enumerate(all_rules_for_list):
            name = rule_data.get("name", f"ルール {rule_data['id']}")
            target_suffix = rule_data['target']
            display_target = "全体" if target_suffix == "ALL" else target_suffix
            
            item = QListWidgetItem(f"{name} ({display_target})")
            item.setData(Qt.UserRole, rule_data) # Store the whole dictionary
            self.rules_list_widget.addItem(item)
            
            if reselect_key and (rule_data['id'], rule_data['target']) == reselect_key:
                row_to_select = i

        self.rules_list_widget.blockSignals(False)

        if row_to_select != -1:
            self.rules_list_widget.setCurrentRow(row_to_select)

    def reload_all_selectors(self):
        self.update_years_in_config()
        self.populate_alias_subject_list()
        
        self.initialize_year_selectors()
        self.initialize_table_layout_selectors()
        self.initialize_save_position_selectors()
        
        self.rule_target_combo.blockSignals(True)
        self.rule_target_combo.clear()
        all_paths = ["全体"] + sorted(self.years)
        self.rule_target_combo.addItems(all_paths)
        self.rule_target_combo.blockSignals(False)
        
        self.populate_prerequisite_parent_list()

    def check_config_validity(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                json.load(f)
            return True
        except (json.JSONDecodeError, UnicodeDecodeError, FileNotFoundError):
            QMessageBox.warning(self, "無効なファイル", "指定されたファイルは有効なJSONファイルではありません。" )
            return False

    def create_new_config(self):
        """JSON設定ファイルを新規作成します。"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "新しいconfigファイルを作成", str(get_base_path()), "jsonファイル (*.json)")
        if not file_path:
            return

        self.config_data = {
            "LAST_UPDATE_DAY": "",
            "SUBJECTS": {},
            "ALIAS_SUBJECTS": {},
            "PREREQUISITE_SUBJECTS": {},
            "TABLE_LAYOUT": {},
            "SAVE_POSITION": {},
            "GENERAL_SETTINGS": {},
            "REQUIRED_SUBJECTS": {},
            "YEARS_HIERARCHY": {}
        }

        return file_path

    def save_config(self):
        # First, capture the complete current state of the UI.
        self.config_data = self.get_current_ui_config()

        if self.tutorial_mode:
            QMessageBox.warning(self, "チュートリアルモード", "チュートリアル中は設定を保存できません。" )
            return

        if not self.file_path:
            new_path, _ = QFileDialog.getSaveFileName(self, "保存先configファイルを選択", str(get_base_path()), "jsonファイル (*.json)")
            if not new_path: return
            self.file_path = new_path

        latest_config = None
        if os.path.exists(self.file_path):
            try:
                latest_hash = calculate_sha256(self.file_path)
                if latest_hash != self.config_hash:
                    with open(self.file_path, "r", encoding="utf-8") as f:
                        latest_config = json.load(f)
            except Exception as e:
                QMessageBox.critical(self, "読み込みエラー", f"ファイルの読み込み中に予期せぬエラーが発生しました: {e}")
                return

        if latest_config: # A conflict was detected
            # Use self.config_data which was captured at the start
            auto_merged, conflicts = diff_and_merge(self.original_config_data, self.config_data, latest_config)

            if not conflicts:
                # 競合なし
                self.config_data = auto_merged
                self.refresh_all_ui_from_config_data()
                QMessageBox.information(self, "自動マージ完了", "競合はなく、変更は自動的にマージされました。" )
            else:
                # 競合あり
                QMessageBox.information(self, "競合の検出", """現在の変更と外部ファイルの間で競合が検出されました。
競合解決ダイアログを開きます。""")

                # 競合をフィルタリングし、自動解決するキーを処理
                filtered_conflicts = []
                auto_resolved_data_from_mine = {}

                keys_to_auto_resolve_to_mine = [
                    # 更新日数
                    "LAST_UPDATE_DAY",

                    # GENERAL_SETTINGSというキー自体が存在する場合のフォールバック
                    "GENERAL_SETTINGS", 
                    
                    # テーマ関連
                    "THEME_SETTING",
                    "THEME_COLORS",
                    "EDITOR_FONT_SIZE",
                    
                    # 絞り込み設定
                    "ACTIVE_FILTER_SUBJECT",
                    "MIN_SUBJECT_COUNT",
                    "MAX_SUBJECT_COUNT",
                    "ACTIVE_FILTER_SUBJECT_AMOUNT", 
                    "ACTIVE_MIN_SUBJECT",          
                    "ACTIVE_MAX_SUBJECT",          

                    "ACTIVE_FILTER_SUBJECT_UNITS",
                    "MIN_SUBJECT_COUNT_UNITS",
                    "MAX_SUBJECT_COUNT_UNITS",
                    "ACTIVE_MIN_SUBJECT_UNITS",    
                    "ACTIVE_MAX_SUBJECT_UNITS",    
                    
                    # その他チェックボックス設定
                    "CHECK_MISSING_SUBJECTS",
                    "CHECK_CLASSROOM_CONFLICTS",
                    "CHECK_TEACHER_CONFLICTS",
                    "CHECK_SLOT_CONSISTENCY",
                    "ENABLE_COMPATIBILITY_MODE",
                    "AUTO_SAVE_ENABLED",
                    
                    # その他スピンボックス設定
                    "SUBJECT_COMBINATION_COUNT",
                    "MORE_SUBJECT_COMBINATION",
                    "MAX_MEMORY_LIMIT",
                    "ART_SUBJECT", 

                    # その他ラインエディット設定
                    "MASTER_CONFIG_PATH",
                    
                    # その他トグル設定
                    "TIMETABLE_ORDER",
                    "RUN_TUTORIAL_ON_STARTUP",
                ]

                for conflict in conflicts:
                    # conflict['path'] はリスト形式 (例: ['GENERAL_SETTINGS', 'CHECK_MISSING_SUBJECTS'])
                    # conflict['path'][0] がトップレベルキー
                    top_level_key = conflict['path'][0]

                    if top_level_key in keys_to_auto_resolve_to_mine:
                        # mine の値を自動的に採用する
                        # auto_resolved_data_from_mine にこの解決結果を格納
                        # pathを使ってネストされた辞書を構築
                        d = auto_resolved_data_from_mine
                        for key_part in conflict['path'][:-1]:
                            d = d.setdefault(key_part, {})
                        d[conflict['path'][-1]] = conflict['mine'] # mineの値を設定
                    else:
                        # それ以外の競合はユーザーに解決させる
                        filtered_conflicts.append(conflict)

            # ユーザーに解決させる競合があるかチェック
            if filtered_conflicts:
                dialog = ConflictResolutionWindow(filtered_conflicts, self)
                if dialog.exec() == QDialog.Accepted:
                    resolved_data = dialog.get_all_resolved_data() # ユーザーが解決したデータ
                    
                    # recursive_merge は local function に移動
                    def recursive_merge(target, source):
                        for key, value in source.items():
                            if isinstance(value, dict) and key in target and isinstance(target[key], dict):
                                target[key] = recursive_merge(target.get(key, {}), value)
                            else:
                                target[key] = value
                        return target
                    
                    # auto_merged に、自動解決したデータとユーザーが解決したデータをマージ
                    # auto_merged は diff_and_merge の結果で、まだ競合箇所は解決されていない
                    # その上に、auto_resolved_data_from_mine (mine優先解決) と user_resolved_data (ユーザー解決) をマージする
                    temp_merged_config = recursive_merge(auto_merged, auto_resolved_data_from_mine)
                    final_config = recursive_merge(temp_merged_config, resolved_data)
                    
                    self.config_data = final_config
                    self.refresh_all_ui_from_config_data()
                    QMessageBox.information(self, "統合完了", """競合が解決され、設定が統合されました。
内容を確認して保存してください。""" )
                else:
                    # ユーザーが競合解決をキャンセルした
                    QMessageBox.warning(self, "キャンセル", "統合処理はキャンセルされました。" )
                    return
            else:
                # ユーザーに解決させる競合が一つもなかった場合 (すべて自動解決された)
                # auto_merged に auto_resolved_data_from_mine をマージするだけで良い
                def recursive_merge(target, source):
                    for key, value in source.items():
                        if isinstance(value, dict) and key in target and isinstance(target[key], dict):
                            target[key] = recursive_merge(target.get(key, {}), value)
                        else:
                            target[key] = value
                    return target
                
                final_config = recursive_merge(auto_merged, auto_resolved_data_from_mine)
                self.config_data = final_config
                self.refresh_all_ui_from_config_data()
                QMessageBox.information(self, "統合完了", """「その他設定」の競合は自動的に解決され、設定が統合されました。
内容を確認して保存してください。""" )
        # perform_save will now save the final state of self.config_data
        self.perform_save()

    def save_config_as(self):
        """Saves the current configuration to a new file."""
        self.config_data = self.get_current_ui_config()

        if self.tutorial_mode:
            QMessageBox.warning(self, "チュートリアルモード", "チュートリアル中は設定を保存できません。" )
            return

        new_file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "名前を付けて保存", 
            str(get_base_path()), 
            "jsonファイル (*.json)"
        )

        if not new_file_path:
            return

        self.file_path = new_file_path
        self.perform_save()

    def perform_save(self):
        """Writes the current self.config_data to disk after cleanup."""
        # NOTE: UI data should be gathered into self.config_data BEFORE calling this.
        
        # YEARS_SUBJECTS_UNITS_キーをconfig_model._dataからconfig_dataにコピー
        # (InitialUnitsDialogで設定された値はconfig_model._dataにのみ保存されるため)
        if hasattr(self, 'config_model') and self.config_model:
            for key in list(self.config_model._data.keys()):
                if key.startswith("YEARS_SUBJECTS_UNITS_"):
                    self.config_data[key] = self.config_model._data[key]
        
        # DEBUG: Check YEARS_SUBJECTS_UNITS_ keys before cleanup
        years_keys = [k for k in self.config_data if k.startswith("YEARS_SUBJECTS_UNITS_")]
        print(f"DEBUG perform_save: YEARS_SUBJECTS_UNITS_ keys in config_data = {years_keys}")

        # --- Cleanup ---
        # Note: YEARS_SUBJECTS_UNITS_ keys should NOT be deleted - they store initial units settings
        keys_to_delete = [k for k in self.config_data if k in ["YEAR", "1YEARS_SUBJECTS_UNITS", "ABNORMAL_SUBJECTS_UNITS", "LAST_SELECTED_HIERARCHY", "LAST_SELECTED_PATH", "USE_ONLY_NAME"]]
        for key in keys_to_delete:
            if key in self.config_data: del self.config_data[key]

        for key in list(self.config_data.keys()):
            if (key.startswith(("REQUIRED_SUBJECTS_", "FIXED_SLOTS", "table_layout", "ALL_SLOTS", "SAVE_POSITION")) and not self.config_data[key]) or \
               (key in ["SUBJECT_ALIASES", "PREREQUISITE_SUBJECTS", "NO_TOGETHER_SUBJECTS", "YEARS_MESSAGE", "YEARS_HIERARCHY"] and not self.config_data[key]):
                del self.config_data[key]

        self.config_data["LAST_UPDATE_DAY"] = datetime.now().strftime("%Y/%m/%d/%H:%M")

        # 保存前に自動バックアップを作成
        try:
            from config_editor.backup import create_backup
            backup_info = create_backup(self.file_path, max_backups=5)
            if backup_info:
                print(f"バックアップ作成: {backup_info.path}")
        except Exception as e:
            print(f"バックアップ作成に失敗（続行します）: {e}")

        # Write the updated config_data to the file
        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(self.config_data, f, indent=2, ensure_ascii=False)
            
            # After a successful save, update the hash and original data
            self.config_hash = calculate_sha256(self.file_path)
            self.original_config_data = copy.deepcopy(self.config_data)
            QMessageBox.information(self, "成功", "設定を保存しました。" )
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"設定の保存中にエラーが発生しました: {e}")

    def open_config_file_dialog(self):
        """ファイルダイアログを開き、configファイルを読み込みます。"""
        
        # 初期ディレクトリ（前回開いたファイル or カレントディレクトリ）
        initial_dir = str(Path(self.file_path).parent) if self.file_path else str(Path.cwd())

        # QFileDialog をインスタンスとして生成（これが重要）
        dialog = QFileDialog(self)
        dialog.setWindowTitle("設定ファイルを選択")
        dialog.setDirectory(initial_dir)
        dialog.setNameFilter("jsonファイル (*.json)")

        # ウィンドウアイコンを強制適用（アイコンが出ない問題の完全対策）
        icon_path = get_base_path() / "時間割くんドライバアイコン.ico"
        if icon_path.exists():
            dialog.setWindowIcon(QIcon(str(icon_path)))

        # 実際に開く
        if dialog.exec():
            selected_files = dialog.selectedFiles()
            if selected_files:
                self.reload_from_file(selected_files[0], show_success_message=True)

    def integrate_config_files(self):
        """外部の設定ファイルを現在の設定にマージします。"""
        if self.tutorial_mode:
            QMessageBox.warning(self, "チュートリアルモード", "この機能はチュートリアルモードでは使用できません。" )
            return

        # 1. 統合する外部ファイルを選択
        theirs_path, _ = QFileDialog.getOpenFileName(
            self, "統合する設定ファイルを選択", str(get_exe_directory()), "jsonファイル (*.json)")
        
        if not theirs_path:
            return # ユーザーがキャンセルした

        # 2. 外部ファイルの読み込みと検証
        try:
            with open(theirs_path, 'r', encoding='utf-8') as f:
                theirs_config = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError, Exception) as e:
            QMessageBox.critical(self, "読み込みエラー", f"選択されたファイルの読み込みに失敗しました:\n{e}")
            return

        # 3. 2-wayマージのための各バージョンを取得
        # mine: 現在のUIの変更をすべて反映した状態
        mine_config = self.get_current_ui_config()
        # baseは空にして、mineとtheirsの2者間での比較を行う
        
        # 4. diff_and_merge を実行
        auto_merged, conflicts = diff_and_merge(self.original_config_data, mine_config, theirs_config)

        # 5. 結果を処理
        if not conflicts:
            # 競合なし
            self.config_data = auto_merged
            self.refresh_all_ui_from_config_data()
            QMessageBox.information(self, "統合完了", """競合はなく、設定は自動的に統合されました。
内容を確認して保存してください。""")
        else:
            # 競合あり
            QMessageBox.information(self, "競合の検出", """現在の変更と外部ファイルの間で競合が検出されました。
競合解決ダイアログを開きます。""")

            # 競合をフィルタリングし、自動解決するキーを処理
            filtered_conflicts = []
            auto_resolved_data_from_mine = {}

            keys_to_auto_resolve_to_mine = [
                # 更新日数
                "LAST_UPDATE_DAY",

                # GENERAL_SETTINGSというキー自体が存在する場合のフォールバック
                "GENERAL_SETTINGS", 
                
                # テーマ関連
                "THEME_SETTING",
                "THEME_COLORS",
                "EDITOR_FONT_SIZE",
                
                # 絞り込み設定
                "ACTIVE_FILTER_SUBJECT",
                "MIN_SUBJECT_COUNT",
                "MAX_SUBJECT_COUNT",
                "ACTIVE_FILTER_SUBJECT_AMOUNT", 
                "ACTIVE_MIN_SUBJECT",          
                "ACTIVE_MAX_SUBJECT",          

                "ACTIVE_FILTER_SUBJECT_UNITS",
                "MIN_SUBJECT_COUNT_UNITS",
                "MAX_SUBJECT_COUNT_UNITS",
                "ACTIVE_MIN_SUBJECT_UNITS",    
                "ACTIVE_MAX_SUBJECT_UNITS",    
                
                # その他チェックボックス設定
                "CHECK_MISSING_SUBJECTS",
                "CHECK_CLASSROOM_CONFLICTS",
                "CHECK_TEACHER_CONFLICTS",
                "CHECK_SLOT_CONSISTENCY",
                "ENABLE_COMPATIBILITY_MODE",
                "AUTO_SAVE_ENABLED",
                
                # その他スピンボックス設定
                "SUBJECT_COMBINATION_COUNT",
                "MORE_SUBJECT_COMBINATION",
                "MAX_MEMORY_LIMIT",
                "ART_SUBJECT", 

                # その他ラインエディット設定
                "MASTER_CONFIG_PATH",
                
                # その他トグル設定
                "TIMETABLE_ORDER",
                "RUN_TUTORIAL_ON_STARTUP",
            ]

            for conflict in conflicts:
                # conflict['path'] はリスト形式 (例: ['GENERAL_SETTINGS', 'CHECK_MISSING_SUBJECTS'])
                # conflict['path'][0] がトップレベルキー
                top_level_key = conflict['path'][0]

                if top_level_key in keys_to_auto_resolve_to_mine:
                    # mine の値を自動的に採用する
                    # auto_resolved_data_from_mine にこの解決結果を格納
                    # pathを使ってネストされた辞書を構築
                    d = auto_resolved_data_from_mine
                    for key_part in conflict['path'][:-1]:
                        d = d.setdefault(key_part, {})
                    d[conflict['path'][-1]] = conflict['mine'] # mineの値を設定
                else:
                    # それ以外の競合はユーザーに解決させる
                    filtered_conflicts.append(conflict)

            # ユーザーに解決させる競合があるかチェック
            if filtered_conflicts:
                dialog = ConflictResolutionWindow(filtered_conflicts, self)
                if dialog.exec() == QDialog.Accepted:
                    resolved_data = dialog.get_all_resolved_data() # ユーザーが解決したデータ
                    
                    # recursive_merge は local function に移動
                    def recursive_merge(target, source):
                        for key, value in source.items():
                            if isinstance(value, dict) and key in target and isinstance(target[key], dict):
                                target[key] = recursive_merge(target.get(key, {}), value)
                            else:
                                target[key] = value
                        return target
                    
                    # auto_merged に、自動解決したデータとユーザーが解決したデータをマージ
                    # auto_merged は diff_and_merge の結果で、まだ競合箇所は解決されていない
                    # その上に、auto_resolved_data_from_mine (mine優先解決) と user_resolved_data (ユーザー解決) をマージする
                    temp_merged_config = recursive_merge(auto_merged, auto_resolved_data_from_mine)
                    final_config = recursive_merge(temp_merged_config, resolved_data)
                    
                    self.config_data = final_config
                    self.refresh_all_ui_from_config_data()
                    QMessageBox.information(self, "統合完了", """競合が解決され、設定が統合されました。
内容を確認して保存してください。""" )
                else:
                    # ユーザーが競合解決をキャンセルした
                    QMessageBox.warning(self, "キャンセル", "統合処理はキャンセルされました。" )
                    return
            else:
                # ユーザーに解決させる競合が一つもなかった場合 (すべて自動解決された)
                # auto_merged に auto_resolved_data_from_mine をマージするだけで良い
                def recursive_merge(target, source):
                    for key, value in source.items():
                        if isinstance(value, dict) and key in target and isinstance(target[key], dict):
                            target[key] = recursive_merge(target.get(key, {}), value)
                        else:
                            target[key] = value
                    return target
                
                final_config = recursive_merge(auto_merged, auto_resolved_data_from_mine)
                self.config_data = final_config
                self.refresh_all_ui_from_config_data()
                QMessageBox.information(self, "統合完了", """「その他設定」の競合は自動的に解決され、設定が統合されました。
内容を確認して保存してください。""" )

    def _pre_populate_all_slots(self):
        """
        Populates the ALL_SLOTS_{year} config entries based on the table_layout{year} data.
        This should be called at startup to ensure the data is available for all tabs.
        """
        for key in list(self.config_data.keys()):
            if key.startswith("table_layout"):
                year_str = key.replace("table_layout", "")
                layout_data = self.config_data.get(key, [])
                if not year_str or not layout_data:
                    continue
                
                all_slots_from_layout = set()
                for row_data in layout_data:
                    for slot_name in row_data:
                        if slot_name:
                            all_slots_from_layout.add(slot_name)
                
                self.config_data[f"ALL_SLOTS_{year_str}"] = sorted(list(all_slots_from_layout))

    # --- 教科マスタ関連 ---
    def create_master_subject_tab(self):
        """教科マスタタブを作成（MasterSubjectTabクラスを使用）"""
        from config_editor.tabs import MasterSubjectTab
        
        # MasterSubjectTabクラスをインスタンス化
        self.master_subject_tab_widget = MasterSubjectTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # シグナル接続（MainWindowの他のタブとの連携）
        self.master_subject_tab_widget.subjects_reloaded.connect(self.soft_reload_subject_lists)
        self.master_subject_tab_widget.subject_added.connect(self.subject_added_to_master.emit)
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.master_subject_tab_widget)
        self.tab_list.addItem("教科マスタ設定")
        
        # 後方互換性のため古い属性名も維持
        self.master_subject_list = self.master_subject_tab_widget.subject_list

    def populate_master_subject_list(self):
        """教科マスタリストを更新（MasterSubjectTabに委譲）"""
        if hasattr(self, 'master_subject_tab_widget'):
            # ConfigModelにデータを同期してからpopulate
            self.config_model.update_data(self.config_data)
            self.master_subject_tab_widget.populate()
        else:
            # 後方互換性（MasterSubjectTabが初期化される前）
            self.master_subject_list.clear()
            master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
            sorted_subjects = sorted(list(set(master_subjects)))
            self.master_subject_list.addItems(sorted_subjects)

    def soft_reload_subject_lists(self):
        """Reloads UI elements that list subjects, without resetting major selections like the current year."""
        self.populate_alias_subject_list()
        self.populate_prerequisite_parent_list()
        self.update_details_subject_lists()
        self.populate_required_subjects_settings(reselect_key=self.current_rule_key)

    def add_master_subject(self):
        text, ok = QInputDialog.getText(self, "教科の追加", "新しい教科名を入力してください:")
        if ok and text:
            if not self.master_subject_list.findItems(text, Qt.MatchExactly):
                self.master_subject_list.addItem(text)
                self.soft_reload_subject_lists()
                if self.tutorial_mode:
                    self.subject_added_to_master.emit(text)
            else:
                QMessageBox.warning(self, "重複", "その教科名はすでに存在します。" )

    def edit_master_subject(self):
        selected_item = self.master_subject_list.currentItem()
        if selected_item:
            old_text = selected_item.text()
            new_text, ok = QInputDialog.getText(self, "教科の編集", "新しい教科名を入力してください:", text=old_text)
            if ok and new_text and new_text != old_text:
                if not self.master_subject_list.findItems(new_text, Qt.MatchExactly):
                    self.rename_subject_in_config(old_text, new_text)
                    selected_item.setText(new_text)
                    self.soft_reload_subject_lists()
                else:
                    QMessageBox.warning(self, "重複", "その教科名はすでに存在します。" )

    def delete_master_subject(self):
        selected_item = self.master_subject_list.currentItem()
        if selected_item:
            reply = QMessageBox.question(self, "削除の確認", f"「{selected_item.text()}」をすべての設定から削除しますか？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                subject_to_delete = selected_item.text()
                self.remove_subject_from_config(subject_to_delete)
                self.master_subject_list.takeItem(self.master_subject_list.row(selected_item))
                self.soft_reload_subject_lists()

    def update_master_subjects_in_config(self, config_dict):
        subjects = [self.master_subject_list.item(i).text() for i in range(self.master_subject_list.count())]
        config_dict["MASTER_SUBJECTS"] = sorted(list(set(subjects)))

    def rename_subject_in_config(self, old_name, new_name):
        # This is a complex operation that needs to touch many parts of the config
        # For simplicity, we'll just update the master list and rely on other parts
        # of the UI to handle the rename gracefully or require manual correction.
        # For now, we just do this:
        master = self.config_data.get("MASTER_SUBJECTS", [])
        if old_name in master:
            master.remove(old_name)
            master.append(new_name)
            self.config_data["MASTER_SUBJECTS"] = sorted(master)
        
        # Also update aliases
        aliases = self.config_data.get("SUBJECT_ALIASES", {})
        if old_name in aliases:
            aliases[new_name] = aliases.pop(old_name)

    def remove_subject_from_config(self, subject_name):
        """Recursively removes a subject from various parts of the config."""
        # Create a copy of keys to iterate over, as we might modify the dict
        for key in list(self.config_data.keys()):
            # Handle simple dicts like SUBJECT_ALIASES, PREREQUISITE_SUBJECTS (as parent)
            if key in ["SUBJECT_ALIASES", "PREREQUISITE_SUBJECTS"]:
                if subject_name in self.config_data.get(key, {}):
                    del self.config_data[key][subject_name]

            # Handle year-suffixed dicts of dicts (old format)
            if key.startswith(("subject_number", "subject_slots_base", "ABNORMAL_SUBJECTS_UNITS")):
                if isinstance(self.config_data.get(key), dict) and subject_name in self.config_data[key]:
                    del self.config_data[key][subject_name]
            
            # Handle year-suffixed dicts of lists (new format)
            if key.startswith(("subject_number", "subject_slots_base")):
                 if isinstance(self.config_data.get(key), list):
                     self.config_data[key] = [item for item in self.config_data[key] if item.get("name") != subject_name]


        # Handle nested lists in PREREQUISITE_SUBJECTS (as child)
        prereqs = self.config_data.get("PREREQUISITE_SUBJECTS", {})
        for parent in list(prereqs.keys()):
            if subject_name in prereqs.get(parent, []):
                prereqs[parent].remove(subject_name)
                if not prereqs[parent]:
                    del prereqs[parent]

        # Handle REQUIRED_SUBJECTS
        for key in list(self.config_data.keys()):
            if key.startswith("REQUIRED_SUBJECTS_"):
                rules = self.config_data.get(key, {})
                for rule_id in list(rules.keys()):
                    if "conditions" in rules.get(rule_id, {}):
                        for condition in rules[rule_id].get("conditions", []):
                            if "subjects" in condition and subject_name in condition.get("subjects", []):
                                condition["subjects"].remove(subject_name)

        # Handle NO_TOGETHER_SUBJECTS
        if "NO_TOGETHER_SUBJECTS" in self.config_data:
            new_no_together = []
            for group in self.config_data.get("NO_TOGETHER_SUBJECTS", []):
                new_group = [s for s in group if s != subject_name]
                if len(new_group) > 1: # A group with 1 or 0 subjects is meaningless
                    new_no_together.append(new_group)
            self.config_data["NO_TOGETHER_SUBJECTS"] = new_no_together

        # Handle FIXED_SLOTS
        for key in list(self.config_data.keys()):
            if key.startswith("FIXED_SLOTS"):
                slots_to_del = [slot for slot, subj in self.config_data.get(key, {}).items() if subj == subject_name]
                for slot in slots_to_del:
                    del self.config_data[key][slot]

    def sort_master_subjects(self):
        reply = QMessageBox.question(self, "確認", "教科マスタリストを50音順に並び替えますか？\nこの操作は保存され、元に戻せません。", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            items = [self.master_subject_list.item(i).text() for i in range(self.master_subject_list.count())]
            items.sort(key=self.japanese_sort_key) # Use custom sort key
            self.master_subject_list.clear()
            self.master_subject_list.addItems(items)

    # --- 科目エイリアス関連 ---
    def japanese_sort_key(self, text):
        # 正規化（NFKC）してからソート
        normalized = unicodedata.normalize('NFKC', text)
        return normalized

    def create_alias_tab(self):
        """科目エイリアスタブを作成（AliasTabクラスを使用）"""
        from config_editor.tabs import AliasTab
        
        # AliasTabクラスをインスタンス化
        self.alias_tab_widget = AliasTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path,
            alias_generator_dialog_class=AliasGeneratorDialog
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.alias_tab_widget)
        self.tab_list.addItem("科目エイリアス設定")
        
        # 後方互換性のため古い属性名も維持
        self.alias_subject_list = self.alias_tab_widget.subject_list
        self.alias_list_widget = self.alias_tab_widget.alias_list

    def populate_alias_subject_list(self):
        """エイリアス設定リストを更新（AliasTabに委譲）"""
        if hasattr(self, 'alias_tab_widget'):
            # ConfigModelにデータを同期してからpopulate
            self.config_model.update_data(self.config_data)
            self.alias_tab_widget.populate()
        else:
            # 後方互換性（AliasTabが初期化される前）
            current_subject_text = self.alias_subject_list.currentItem().text() if self.alias_subject_list.currentItem() else None
            self.alias_subject_list.blockSignals(True)
            self.alias_subject_list.clear()

            master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
            all_aliases = self.config_data.get("SUBJECT_ALIASES", {})

            for subject in master_subjects:
                item = QListWidgetItem(subject)
                subject_aliases = all_aliases.get(subject, [])
                item.setData(Qt.UserRole, subject_aliases)
                self.alias_subject_list.addItem(item)

            if current_subject_text:
                items = self.alias_subject_list.findItems(current_subject_text, Qt.MatchExactly)
                if items:
                    self.alias_subject_list.setCurrentItem(items[0])

            self.alias_subject_list.blockSignals(False)

    def _update_alias_display(self, aliases):
        self.alias_list_widget.clear()
        self.alias_list_widget.addItems(sorted(aliases))

    def _get_aliases_from_widget_and_update_item_data(self):
        current_master_item = self.alias_subject_list.currentItem()
        if not current_master_item:
            return
        
        new_aliases = [self.alias_list_widget.item(i).text() for i in range(self.alias_list_widget.count())]
        current_master_item.setData(Qt.UserRole, new_aliases)

    def _get_aliases_from_ui(self, config_dict):
        subject_aliases = {}
        for i in range(self.alias_subject_list.count()):
            item = self.alias_subject_list.item(i)
            subject_name = item.text()
            aliases = item.data(Qt.UserRole) or []
            if aliases:
                subject_aliases[subject_name] = sorted(aliases)
        
        if subject_aliases:
            config_dict["SUBJECT_ALIASES"] = subject_aliases
        elif "SUBJECT_ALIASES" in config_dict:
            del config_dict["SUBJECT_ALIASES"]

    def on_alias_subject_selected(self, index):
        # No longer need to save on selection change.
        item = self.alias_subject_list.item(index)
        if not item:
            self.current_selected_subject_for_alias = None
            self.alias_list_widget.clear()
            return

        subject = item.text()
        self.current_selected_subject_for_alias = subject
        
        aliases = item.data(Qt.UserRole) or []
        self._update_alias_display(aliases)

    def add_alias(self):
        if not self.current_selected_subject_for_alias: return
        text, ok = QInputDialog.getText(self, "エイリアスの追加", "新しいエイリアスを入力してください:")
        if ok and text:
            self.alias_list_widget.addItem(text)
            self._get_aliases_from_widget_and_update_item_data()

    def edit_alias(self):
        selected_item = self.alias_list_widget.currentItem()
        if selected_item:
            old_text = selected_item.text()
            new_text, ok = QInputDialog.getText(self, "エイリアスの編集", "新しいエイリアスを入力してください:", text=old_text)
            if ok and new_text and new_text != old_text:
                self.alias_list_widget.currentItem().setText(new_text)
                self._get_aliases_from_widget_and_update_item_data()

    def delete_alias(self):
        selected_item = self.alias_list_widget.currentItem()
        if selected_item:
            self.alias_list_widget.takeItem(self.alias_list_widget.row(selected_item))
            self._get_aliases_from_widget_and_update_item_data()



    def open_alias_generator(self):
        master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        if not master_subjects:
            QMessageBox.warning(self, "エラー", "教科マスタが空です。" )
            return

        pre_selected = []
        if self.alias_subject_list.currentItem():
            pre_selected.append(self.alias_subject_list.currentItem().text())

        dialog = AliasGeneratorDialog(master_subjects, pre_selected, self)
        if dialog.exec():
            subjects_to_process = dialog.get_selected_subjects()
            if subjects_to_process:
                self.generate_aliases(subjects_to_process)

    def generate_aliases(self, subjects):
        kks = pykakasi.kakasi()

        if "SUBJECT_ALIASES" not in self.config_data:
            self.config_data["SUBJECT_ALIASES"] = {}

        processed_count = 0
        error_subjects = []
        
        for subject in subjects:
            try:
                if subject not in self.config_data["SUBJECT_ALIASES"]:
                    self.config_data["SUBJECT_ALIASES"].setdefault(subject, []) # Ensure list exists

                existing_aliases = set(self.config_data["SUBJECT_ALIASES"].get(subject, []))

                # pykakasiによる変換
                converted = kks.convert(subject)
                hira_alias = "".join([item['hira'] for item in converted])
                kana_alias = "".join([item['kana'] for item in converted])
                hepburn_alias = "".join([item['hepburn'] for item in converted])

                new_aliases = {hira_alias, kana_alias, hepburn_alias}

                # 翻訳による変換
                try:
                    # 記号や数字などを取り除いた方が精度が上がる場合がある
                    clean_subject = re.sub(r'[0-9０-９A-ZＡ-ＺⅠ-Ⅲ\s]', '', subject).strip()
                    if clean_subject:
                        translated = self.translator.translate(clean_subject)
                        if translated:
                            # 翻訳結果が元のテキストや他の読み方と異なる場合のみ追加
                            if translated.lower() != clean_subject.lower() and translated.lower() != hepburn_alias.lower():
                                new_aliases.add(translated)
                except Exception as e:
                    print(f"Warning: Could not translate '{subject}': {e}")


                # 元の教科名自身はエイリアスに含めない
                if subject in new_aliases:
                    new_aliases.remove(subject)

                # 既存のエイリアスと重複しないものを追加
                for alias in sorted(list(new_aliases)):
                    if alias and alias not in existing_aliases:
                        self.config_data["SUBJECT_ALIASES"].get(subject, []).append(alias)
                processed_count += 1
            except Exception as e:
                print(f"Error processing alias for '{subject}': {e}")
                error_subjects.append(subject)

        # UIの更新
        current_item = self.alias_subject_list.currentItem()
        if current_item and current_item.text() in subjects:
            self._update_alias_display(current_item.text())

        if not error_subjects:
            QMessageBox.information(self, "成功", f"{processed_count}件の教科について、読み方と英単語のエイリアス生成を試みました。" )
        else:
            QMessageBox.warning(self, "一部エラー", f"{processed_count}件の教科は処理しましたが、以下の教科でエラーが発生しました：\n{', '.join(error_subjects)}")
    # --- 科目詳細設定タブ ---
    def create_subject_details_tab(self):
        """科目詳細設定タブを作成（SubjectDetailsTabクラスを使用）"""
        from config_editor.tabs import SubjectDetailsTab
        
        # SubjectDetailsTabクラスをインスタンス化
        self.subject_details_tab_widget = SubjectDetailsTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.subject_details_tab_widget)
        self.tab_list.addItem("科目詳細設定")
        
        # 後方互換性のため古い属性名も維持
        self.assigned_subjects_list = self.subject_details_tab_widget.assigned_list
        self.unassigned_subjects_list = self.subject_details_tab_widget.unassigned_list
        self.details_year_combo_layout = self.subject_details_tab_widget.combo_layout
        self.details_year_cascading_combos = []
        self.normal_units_label = self.subject_details_tab_widget.normal_units_label
        self.final_units_input = self.subject_details_tab_widget.final_units_input
        self.number_group = self.subject_details_tab_widget.number_group
        self.number_layout = self.subject_details_tab_widget.number_layout
        self.slots_base_group = self.subject_details_tab_widget.slots_group
        self.slots_base_table = self.subject_details_tab_widget.slots_table
        
        # シグナル接続を追加（MainWindowのハンドラを使用）
        # 既存のSubjectDetailsTab内部接続を切断し、MainWindowのメソッドに接続し直す
        try:
            self.assigned_subjects_list.itemSelectionChanged.disconnect()
        except RuntimeError:
            pass
        try:
            self.unassigned_subjects_list.itemSelectionChanged.disconnect()
        except RuntimeError:
            pass
        try:
            self.subject_details_tab_widget.sort_button.clicked.disconnect()
        except RuntimeError:
            pass
        try:
            self.subject_details_tab_widget.add_slot_btn.clicked.disconnect()
        except RuntimeError:
            pass
        try:
            self.subject_details_tab_widget.del_slot_btn.clicked.disconnect()
        except RuntimeError:
            pass
        
        # MainWindowのハンドラに接続
        self.assigned_subjects_list.itemSelectionChanged.connect(self.handle_assigned_selection)
        self.unassigned_subjects_list.itemSelectionChanged.connect(self.handle_unassigned_selection)
        self.assigned_subjects_list.model().rowsMoved.connect(self.on_assigned_subjects_reordered)
        self.subject_details_tab_widget.sort_button.clicked.connect(self.sort_assigned_subjects)
        self.subject_details_tab_widget.add_slot_btn.clicked.connect(self.add_slot_group)
        self.subject_details_tab_widget.del_slot_btn.clicked.connect(self.delete_slot_group)
        
        # スロットグループテーブル変更時にUIを更新
        self.slots_base_table.cellChanged.connect(self._on_slots_base_changed)



    def handle_assigned_selection(self):
        if self.assigned_subjects_list.currentItem() and self.assigned_subjects_list.hasFocus():
            self.unassigned_subjects_list.clearSelection()
            self._on_subject_selection_changed(self.assigned_subjects_list.currentItem())

    def handle_unassigned_selection(self):
        if self.unassigned_subjects_list.currentItem() and self.unassigned_subjects_list.hasFocus():
            self.assigned_subjects_list.clearSelection()
            self._on_subject_selection_changed(self.unassigned_subjects_list.currentItem())

    def update_details_subject_lists(self):
        current_subject = self.current_details_context.get("subject")
        self.assigned_subjects_list.blockSignals(True)
        self.unassigned_subjects_list.blockSignals(True)
        self.assigned_subjects_list.clear()
        self.unassigned_subjects_list.clear()

        year = self.current_details_context["year"]
        if not year:
            self.assigned_subjects_list.blockSignals(False)
            self.unassigned_subjects_list.blockSignals(False)
            return

        master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        
        # Consolidate all subject data for the current year
        all_subject_details = {}
        
        num_key = f"subject_number{year}"
        slots_key = f"subject_slots_base{year}"
        abnormal_key = f"ABNORMAL_SUBJECTS_UNITS{year}"

        num_data_list = self.config_data.get(num_key, [])
        slots_data_list = self.config_data.get(slots_key, [])
        abnormal_data = self.config_data.get(abnormal_key, {})

        if isinstance(num_data_list, dict):
            num_data_list = [{"name": s, "data": d} for s, d in num_data_list.items()]
        if isinstance(slots_data_list, dict):
            slots_data_list = [{"name": s, "data": d} for s, d in slots_data_list.items()]
            
        for subject_name in master_subjects:
            all_subject_details[subject_name] = {
                "number_data": {},
                "slots_base_data": [],
                "abnormal_units_adjustment": abnormal_data.get(subject_name, 0)
            }

        for item in num_data_list:
            if item.get("name") in all_subject_details:
                all_subject_details[item["name"]]["number_data"] = item.get("data", {})
        
        for item in slots_data_list:
            if item.get("name") in all_subject_details:
                all_subject_details[item["name"]]["slots_base_data"] = item.get("data", [])

        # Populate lists and attach data
        assigned_subjects_set = set()
        for subject, details in all_subject_details.items():
            is_assigned = details["number_data"] or details["slots_base_data"]
            if is_assigned:
                assigned_subjects_set.add(subject)

        for subject in master_subjects:
            details = all_subject_details.get(subject)
            item = QListWidgetItem(subject)
            item.setData(Qt.UserRole, details)
            if subject in assigned_subjects_set:
                self.assigned_subjects_list.addItem(item)
            else:
                self.unassigned_subjects_list.addItem(item)
        
        # Restore selection
        self.assigned_subjects_list.blockSignals(False)
        self.unassigned_subjects_list.blockSignals(False)

        if current_subject:
            # This part needs to find the item in either list and select it
            # スクロール位置を維持するためblockSignalsを使用
            for i in range(self.assigned_subjects_list.count()):
                if self.assigned_subjects_list.item(i).text() == current_subject:
                    self.assigned_subjects_list.blockSignals(True)
                    self.assigned_subjects_list.setCurrentRow(i)
                    self.assigned_subjects_list.blockSignals(False)
                    break
            for i in range(self.unassigned_subjects_list.count()):
                if self.unassigned_subjects_list.item(i).text() == current_subject:
                    self.unassigned_subjects_list.blockSignals(True)
                    self.unassigned_subjects_list.setCurrentRow(i)
                    self.unassigned_subjects_list.blockSignals(False)
                    break

    def set_details_panel_enabled(self, enabled):
        self.assigned_subjects_list.setEnabled(enabled)
        self.unassigned_subjects_list.setEnabled(enabled)
        self.number_group.setEnabled(enabled)
        self.slots_base_group.setEnabled(enabled)
        
        if not enabled:
            self.assigned_subjects_list.clear()
            self.unassigned_subjects_list.clear()
            self.final_units_input.setValue(0)
            while self.number_layout.count(): self.number_layout.takeAt(0).widget().deleteLater()
            self.details_number_widgets.clear()
            self.slots_base_table.clearContents()
            self.slots_base_table.setRowCount(0)
            self.number_group.setTitle("スロット番号 (最下層の項目を選択)")
            self.slots_base_group.setTitle("スロットグループ (最下層の項目を選択)")

    def initialize_year_selectors(self):
        for combo in self.details_year_cascading_combos:
            self.details_year_combo_layout.removeWidget(combo)
            combo.deleteLater()
        self.details_year_cascading_combos.clear()

        top_level_items = sorted(list(self.config_data.get("YEARS_HIERARCHY", {}).keys()))
        first_combo = QComboBox()
        first_combo.addItem("(未選択)")
        first_combo.addItems(top_level_items)
        first_combo.currentIndexChanged.connect(
            lambda: self.update_details_year_selectors(level=1)
        )
        self.details_year_cascading_combos.append(first_combo)
        self.details_year_combo_layout.addWidget(first_combo)
        
        self.update_details_year_selectors(level=1)

    def _on_subject_selection_changed(self, current_item):
        self.assigned_subjects_list.blockSignals(True)
        self.unassigned_subjects_list.blockSignals(True)
        try:
            new_subject = current_item.text() if current_item else None
            old_subject = self.current_details_context["subject"]
            if new_subject == old_subject: return
            if old_subject:
                # 科目切り替え前に現在のUIデータをconfig_dataに保存
                self._save_current_subject_details_to_config(old_subject)
            self.current_details_context["subject"] = new_subject
            self.populate_subject_details()
            self.subject_details_updated.emit()
            self.update_details_subject_lists()
        finally:
            self.assigned_subjects_list.blockSignals(False)
            self.unassigned_subjects_list.blockSignals(False)

    def _save_current_subject_details_to_config(self, subject):
        """現在のUIのスロットグループデータをconfig_dataに保存"""
        year = self.current_details_context.get("year")
        if not year or not subject:
            return
        
        # UIからデータを収集
        details = self._get_current_details_from_widgets()
        
        # subject_number を更新
        num_key = f"subject_number{year}"
        if num_key not in self.config_data:
            self.config_data[num_key] = {}
        
        # 既存のサブジェクトデータを探して更新
        if isinstance(self.config_data[num_key], dict):
            self.config_data[num_key][subject] = details["number_data"]
        elif isinstance(self.config_data[num_key], list):
            # リスト形式の場合
            found = False
            for item in self.config_data[num_key]:
                if item.get("name") == subject:
                    item["data"] = details["number_data"]
                    found = True
                    break
            if not found:
                self.config_data[num_key].append({"name": subject, "data": details["number_data"]})
        
        # subject_slots_base を更新
        slots_key = f"subject_slots_base{year}"
        if slots_key not in self.config_data:
            self.config_data[slots_key] = []
        
        if isinstance(self.config_data[slots_key], list):
            found = False
            for item in self.config_data[slots_key]:
                if item.get("name") == subject:
                    item["data"] = details["slots_base_data"]
                    found = True
                    break
            if not found and details["slots_base_data"]:
                self.config_data[slots_key].append({"name": subject, "data": details["slots_base_data"]})

    def _on_slots_base_changed(self, row, col):
        """スロットグループテーブルが変更された時の処理"""
        subject = self.current_details_context.get("subject")
        year = self.current_details_context.get("year")
        if not subject or not year:
            return
        
        # 1. 現在の変更を保存
        self._save_current_subject_details_to_config(subject)
        
        # 2. 未割り当てにいる場合は割り当て済みに移動
        for i in range(self.unassigned_subjects_list.count()):
            item = self.unassigned_subjects_list.item(i)
            if item and item.text() == subject:
                self.unassigned_subjects_list.takeItem(i)
                self.unassigned_subjects_list.clearSelection()  # 選択状態をクリア
                # リストの先頭に追加（スクロール位置を維持）
                self.assigned_subjects_list.insertItem(0, subject)
                self.assigned_subjects_list.setCurrentRow(0)
                break
        
        # 3. スロット番号エリアを更新（新しいスロットを追加）
        self._update_slot_number_from_table()
        
        # 4. 通常単位数ラベルを更新
        first_group_slots = self._get_first_group_slot_count()
        self.normal_units_label.setText(f"{first_group_slots} 単位")
        
        # 最終単位数は現在値が0の時のみ初期値として設定（ユーザー編集を維持）
        if self.final_units_input.value() == 0:
            self.final_units_input.setValue(first_group_slots)

    def _get_first_group_slot_count(self):
        """スロットグループテーブルの最初のグループのスロット数を取得"""
        if self.slots_base_table.rowCount() == 0:
            return 0
        
        count = 0
        for c in range(self.slots_base_table.columnCount()):
            item = self.slots_base_table.item(0, c)
            if item and item.text().strip():
                count += 1
        return count

    def _update_slot_number_from_table(self):
        """スロットグループテーブルからスロット番号エリアを更新"""
        # テーブルから現在のスロット一覧を収集
        current_slots = set()
        for r in range(self.slots_base_table.rowCount()):
            for c in range(self.slots_base_table.columnCount()):
                item = self.slots_base_table.item(r, c)
                if item and item.text().strip():
                    current_slots.add(item.text().strip())
        
        # 既存のスロット番号ウィジェットにないスロットを追加
        existing_slots = set(self.details_number_widgets.keys())
        new_slots = current_slots - existing_slots
        
        for slot in sorted(new_slots):
            line_edit = QLineEdit()
            line_edit.setReadOnly(False)
            self.number_layout.addRow(QLabel(slot), line_edit)
            self.details_number_widgets[slot] = line_edit

    def _get_selected_subject_item(self):
        """Gets the currently selected subject item from either the assigned or unassigned list."""
        assigned_item = self.assigned_subjects_list.currentItem()
        if assigned_item and self.assigned_subjects_list.hasFocus():
            return assigned_item
        
        unassigned_item = self.unassigned_subjects_list.currentItem()
        if unassigned_item and self.unassigned_subjects_list.hasFocus():
            return unassigned_item
            
        # Fallback if focus is not clear, checking assigned first
        if assigned_item:
            return assigned_item
        if unassigned_item:
            return unassigned_item
            
        return None

    def populate_subject_details(self):
        subject = self.current_details_context.get("subject")
        year = self.current_details_context.get("year")

        # --- Clear UI ---
        while self.number_layout.count():
            item = self.number_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.details_number_widgets.clear()
        self.slots_base_table.clearContents()
        self.slots_base_table.setRowCount(0)

        # --- Get the selected item and its data ---
        selected_item = self._get_selected_subject_item()
        if not subject or not year or not selected_item:
            self.number_group.setTitle("スロット番号 (subject_number)")
            self.slots_base_group.setTitle("スロットグループ (subject_slots_base)")
            self.final_units_input.setEnabled(False)
            self.final_units_input.setValue(0)
            self.normal_units_label.setText("- 単位")
            return

        subject_details = selected_item.data(Qt.UserRole)
        if not subject_details:
             subject_details = {} # Should not happen, but a safeguard
        
        # --- Populate UI from item data ---
        self.number_group.setTitle(f"スロット番号 (subject_number_{year}) - {subject}")
        self.slots_base_group.setTitle(f"スロットグループ (subject_slots_base_{year}) - {subject}")
        self.final_units_input.setEnabled(True)
        
        normal_units = self._get_normal_units(subject, year) # This method still reads from config, which is fine for this calculation
        self.normal_units_label.setText(f"{normal_units} 単位")
        
        unit_adjustment = subject_details.get("abnormal_units_adjustment", 0)
        final_units = normal_units + unit_adjustment
        self.final_units_input.blockSignals(True)
        self.final_units_input.setValue(final_units)
        self.final_units_input.blockSignals(False)

        current_subject_num_data = subject_details.get("number_data", {})
        current_subject_slots_data = subject_details.get("slots_base_data", [])

        all_slots_for_subject = set(current_subject_num_data.keys())
        for group in current_subject_slots_data:
            all_slots_for_subject.update(group)

        # --- Update number/name area ---
        for slot in sorted(list(all_slots_for_subject)):
            line_edit = QLineEdit()
            number = current_subject_num_data.get(slot, "")
            line_edit.setText(str(number))
            line_edit.setReadOnly(False)
            self.number_layout.addRow(QLabel(slot), line_edit)
            self.details_number_widgets[slot] = line_edit

        # --- Update slots_base table ---
        if current_subject_slots_data:
            self.slots_base_table.setRowCount(len(current_subject_slots_data))
            for r, group in enumerate(current_subject_slots_data):
                for c, slot in enumerate(group):
                    if c >= self.slots_base_table.columnCount():
                        self.slots_base_table.setColumnCount(c + 1)
                    self.slots_base_table.setItem(r, c, QTableWidgetItem(slot))

    def _get_selected_subject_from_lists(self):
        """Gets the currently selected subject from either the assigned or unassigned list."""
        assigned_item = self.assigned_subjects_list.currentItem()
        if assigned_item and self.assigned_subjects_list.hasFocus():
            return assigned_item.text()
        
        unassigned_item = self.unassigned_subjects_list.currentItem()
        if unassigned_item and self.unassigned_subjects_list.hasFocus():
            return unassigned_item.text()
            
        # Fallback if focus is not clear
        if assigned_item:
            return assigned_item.text()
        if unassigned_item:
            return unassigned_item.text()

        return None

    def _get_current_details_from_widgets(self):
        """
        Gathers details for the currently displayed subject from the UI widgets.
        """
        details = {
            "number_data": {},
            "slots_base_data": [],
            "abnormal_units_adjustment": 0
        }

        # 1. Gather number_data from the QLineEdit widgets
        for slot, widget in self.details_number_widgets.items():
            value = widget.text().strip()
            if value:
                details["number_data"][slot] = value

        # 2. Gather slots_base_data from the QTableWidget
        slots_data = []
        for r in range(self.slots_base_table.rowCount()):
            row_data = []
            for c in range(self.slots_base_table.columnCount()):
                item = self.slots_base_table.item(r, c)
                if item and item.text():
                    row_data.append(item.text().strip())
            if row_data:
                slots_data.append(sorted(row_data))
        details["slots_base_data"] = sorted(slots_data)

        # 3. Calculate abnormal_units_adjustment
        subject = self.current_details_context.get("subject")
        year = self.current_details_context.get("year")
        if subject and year:
            normal_units = self._get_normal_units(subject, year)
            final_units = self.final_units_input.value()
            details["abnormal_units_adjustment"] = final_units - normal_units

        return details

    def _get_subject_details_from_config_data(self, config_dict):
        """
        Gathers subject details for the currently selected year from the UI,
        and populates the config_dict. This is called on save.
        """
        year = self.current_details_context.get("year")
        if not year:
            # If no year is selected in the details tab, we copy existing data to avoid losing it.
            for key in list(self.config_data.keys()):
                if key.startswith(("subject_number", "subject_slots_base", "ABNORMAL_SUBJECTS_UNITS")):
                    config_dict[key] = copy.deepcopy(self.config_data.get(key))
            return

        # 1. Collect details for all subjects for the current year from the UI.
        all_subject_details_for_year = {}
        
        # Helper to process a list widget (assigned or unassigned)
        def collect_from_list(list_widget):
            for i in range(list_widget.count()):
                item = list_widget.item(i)
                subject_name = item.text()
                # If this is the currently edited subject, get fresh data from widgets.
                if subject_name == self.current_details_context.get("subject"):
                    all_subject_details_for_year[subject_name] = self._get_current_details_from_widgets()
                else:
                    # Otherwise, use the data stored on the QListWidgetItem.
                    stored_data = item.data(Qt.UserRole)
                    if stored_data:
                        all_subject_details_for_year[subject_name] = stored_data

        collect_from_list(self.assigned_subjects_list)
        collect_from_list(self.unassigned_subjects_list)

        # 2. Rebuild the config data structures for the current year.
        num_key = f"subject_number{year}"
        slots_key = f"subject_slots_base{year}"
        abnormal_key = f"ABNORMAL_SUBJECTS_UNITS{year}"

        num_data_list = []
        slots_data_list = []
        abnormal_data_dict = {}

        # Use the order from assigned_subjects_list as the definitive order for assigned subjects
        for i in range(self.assigned_subjects_list.count()):
            subject_name = self.assigned_subjects_list.item(i).text()
            details = all_subject_details_for_year.get(subject_name)
            if not details: continue
            
            if details.get("number_data"):
                num_data_list.append({"name": subject_name, "data": details["number_data"]})
            if details.get("slots_base_data"):
                slots_data_list.append({"name": subject_name, "data": details["slots_base_data"]})

        # Process abnormal units for all subjects
        for subject_name, details in all_subject_details_for_year.items():
             if details and details.get("abnormal_units_adjustment") != 0:
                abnormal_data_dict[subject_name] = details["abnormal_units_adjustment"]

        # Update config_dict for the current year
        if num_data_list: config_dict[num_key] = num_data_list
        elif num_key in config_dict: del config_dict[num_key]

        if slots_data_list: config_dict[slots_key] = slots_data_list
        elif slots_key in config_dict: del config_dict[slots_key]
        
        if abnormal_data_dict: config_dict[abnormal_key] = abnormal_data_dict
        elif abnormal_key in config_dict: del config_dict[abnormal_key]

        # 3. Preserve data for other years by copying it from self.config_data.
        for y in self.years:
            if y == year: continue  # Skip the current year as we just processed it
            
            other_num_key = f"subject_number{y}"
            other_slots_key = f"subject_slots_base{y}"
            other_abnormal_key = f"ABNORMAL_SUBJECTS_UNITS{y}"

            if other_num_key in self.config_data:
                config_dict[other_num_key] = copy.deepcopy(self.config_data[other_num_key])
            elif other_num_key in config_dict:
                 del config_dict[other_num_key]

            if other_slots_key in self.config_data:
                config_dict[other_slots_key] = copy.deepcopy(self.config_data[other_slots_key])
            elif other_slots_key in config_dict:
                 del config_dict[other_slots_key]

            if other_abnormal_key in self.config_data:
                config_dict[other_abnormal_key] = copy.deepcopy(self.config_data[other_abnormal_key])
            elif other_abnormal_key in config_dict:
                 del config_dict[other_abnormal_key]

        # YEARS_SUBJECTS_UNITS_キーをコピー（取得単位数の初期値加算）
        for key in list(self.config_data.keys()):
            if key.startswith("YEARS_SUBJECTS_UNITS_"):
                config_dict[key] = self.config_data[key]

    def sort_assigned_subjects(self):
        reply = QMessageBox.question(self, "確認", "割り当て済み教科リストを50音順に並び替えますか？\nこの並び替えはこの画面でのみ有効です。", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            items = [self.assigned_subjects_list.item(i).text() for i in range(self.assigned_subjects_list.count())]
            items.sort(key=self.japanese_sort_key)
            self.assigned_subjects_list.clear()
            self.assigned_subjects_list.addItems(items)

    def on_assigned_subjects_reordered(self, parent, start, end, destination, row):
        pass # No longer update config_data interactively

    def add_slot_group(self):
        year_text = self.current_details_context.get("year")
        if not year_text: return
        available_slots = self.config_data.get(f"ALL_SLOTS_{year_text}", [])
        if not available_slots:
            QMessageBox.warning(self, "エラー", f"{year_text}のスロットが「スロットレイアウト設定」で定義されていません。" )
            return
        dialog = SlotSelectionDialog(available_slots, self)
        if dialog.exec():
            selected_slots = dialog.get_selected_slots()
            if selected_slots:
                row_count = self.slots_base_table.rowCount()
                self.slots_base_table.insertRow(row_count)
                for col, slot in enumerate(selected_slots):
                    if col >= self.slots_base_table.columnCount():
                        self.slots_base_table.setColumnCount(col + 1)
                    self.slots_base_table.setItem(row_count, col, QTableWidgetItem(slot))

    def delete_slot_group(self):
        current_row = self.slots_base_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "行が選択されていません", "削除したいスロットグループの行を選択してください。" )
            return
        deleted_slots = []
        for c in range(self.slots_base_table.columnCount()):
            item = self.slots_base_table.item(current_row, c)
            if item and item.text():
                deleted_slots.append(item.text())
        self.slots_base_table.removeRow(current_row)
        remaining_slots_in_table = set()
        for r in range(self.slots_base_table.rowCount()):
            for c in range(self.slots_base_table.columnCount()):
                item = self.slots_base_table.item(r, c)
                if item and item.text():
                    remaining_slots_in_table.add(item.text())
        orphaned_slots = [slot for slot in deleted_slots if slot not in remaining_slots_in_table]
        if not orphaned_slots: return
        for i in range(self.number_layout.rowCount() - 1, -1, -1):
            label_item = self.number_layout.itemAt(i, QFormLayout.LabelRole)
            if label_item:
                label_widget = label_item.widget()
                if label_widget:
                    slot_name = label_widget.text()
                    if slot_name in orphaned_slots:
                        self.number_layout.removeRow(i)
                        if slot_name in self.details_number_widgets:
                            del self.details_number_widgets[slot_name]

    # --- スロットレイアウト設定タブ ---
    def create_table_layout_tab(self):
        """スロットレイアウト設定タブを作成（LayoutTabクラスを使用）"""
        from config_editor.tabs import LayoutTab
        
        # LayoutTabクラスをインスタンス化
        self.layout_tab_widget = LayoutTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.layout_tab_widget)
        self.tab_list.addItem("スロットレイアウト設定")
        
        # 後方互換性のため古い属性名も維持
        self.table_widget = self.layout_tab_widget.table
        self.table_layout_combo_layout = self.layout_tab_widget.combo_layout
        self.table_layout_cascading_combos = []
        self.fixed_slot_panel = self.layout_tab_widget.fixed_panel
        self.selected_slot_label = self.layout_tab_widget.slot_label
        self.fixed_subject_selector = self.layout_tab_widget.subject_selector
        self.add_row_button = self.layout_tab_widget.add_row_btn
        self.add_col_button = self.layout_tab_widget.add_col_btn
        
        # シグナル接続（MainWindowのハンドラを使用）
        self.table_widget.itemSelectionChanged.connect(self.on_slot_selection_changed)
        self.fixed_subject_selector.currentIndexChanged.connect(self.update_fixed_slot_from_panel)
        self.add_row_button.clicked.connect(self.add_table_row)
        self.add_col_button.clicked.connect(self.add_table_column)

    def on_slot_selection_changed(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            self.fixed_slot_panel.setEnabled(False)
            self.selected_slot_label.setText("(セルを選択してください)")
            self.fixed_subject_selector.clear()
            return

        item = selected_items[0]
        slot_name = item.data(Qt.UserRole)

        if not slot_name:
            self.fixed_slot_panel.setEnabled(False)
            self.selected_slot_label.setText("(名前のないスロット)")
            self.fixed_subject_selector.clear()
            return

        self.fixed_slot_panel.setEnabled(True)
        self.selected_slot_label.setText(slot_name)

        year_str = self.current_selected_table_layout_key
        
        fixed_slots_key_no_underscore = f"FIXED_SLOTS{year_str}"
        fixed_slots_key_with_underscore = f"FIXED_SLOTS_{year_str}"
        fixed_slots = self.config_data.get(fixed_slots_key_no_underscore)
        if fixed_slots is None:
            fixed_slots = self.config_data.get(fixed_slots_key_with_underscore, {})
        current_fixed_subject = fixed_slots.get(slot_name)

        self.fixed_subject_selector.blockSignals(True)
        self.fixed_subject_selector.clear()
        master_subjects = ["(未選択)"] + self.config_data.get("MASTER_SUBJECTS", [])
        self.fixed_subject_selector.addItems(master_subjects)

        if current_fixed_subject:
            self.fixed_subject_selector.setCurrentText(current_fixed_subject)
        else:
            self.fixed_subject_selector.setCurrentIndex(0)
        self.fixed_subject_selector.blockSignals(False)

    def update_fixed_slot_from_panel(self, index):
        selected_items = self.table_widget.selectedItems()
        if index == -1 or not selected_items:
            return

        # ★★★ 1. 現在の選択位置を保存 ★★★
        current_row = self.table_widget.currentRow()
        current_col = self.table_widget.currentColumn()

        slot_name = selected_items[0].data(Qt.UserRole)
        selected_subject = self.fixed_subject_selector.currentText()

        if not slot_name: return

        year_str = self.current_selected_table_layout_key
        
        fixed_slots_key = f"FIXED_SLOTS{year_str}"
        
        if fixed_slots_key not in self.config_data:
            self.config_data[fixed_slots_key] = {}

        if selected_subject == "(未選択)":
            if slot_name in self.config_data[fixed_slots_key]:
                del self.config_data[fixed_slots_key][slot_name]
                if not self.config_data[fixed_slots_key]:
                    del self.config_data[fixed_slots_key]
        elif selected_subject:
            self.config_data[fixed_slots_key][slot_name] = selected_subject
        
        self.populate_table_layout()

        # ★★★ 2. 選択状態を復元 ★★★
        if current_row >= 0 and current_col >= 0:
            self.table_widget.setCurrentCell(current_row, current_col)
        self.on_slot_selection_changed()

    def show_header_context_menu(self, pos):
        header = self.sender()
        menu = QMenu(self)
        orientation = header.orientation()
        index = header.logicalIndexAt(pos.y() if orientation == Qt.Vertical else pos.x())
        if index < 0: return
        item_type = "行" if orientation == Qt.Vertical else "列"
        move_action = QAction(f"{index+1}{item_type}目を移動...", self)
        delete_action = QAction(f"{index+1}{item_type}目を削除", self)
        if orientation == Qt.Vertical:
            move_action.triggered.connect(lambda: self.move_row(index))
            delete_action.triggered.connect(lambda: self.remove_table_row(index))
        else:
            move_action.triggered.connect(lambda: self.move_column(index))
            delete_action.triggered.connect(lambda: self.remove_table_column(index))
        menu.addAction(move_action)
        menu.addAction(delete_action)
        menu.exec(header.mapToGlobal(pos))

    def move_row(self, source_row):
        max_rows = self.table_widget.rowCount()
        dest_row, ok = QInputDialog.getInt(self, "行の移動", f"{source_row + 1}行目をどこに移動しますか？ (1〜{max_rows})", source_row + 1, 1, max_rows)
        if ok and dest_row - 1 != source_row:
            dest_row -= 1
            row_data = [self.table_widget.takeItem(source_row, col) for col in range(self.table_widget.columnCount())]
            self.table_widget.removeRow(source_row)
            self.table_widget.insertRow(dest_row)
            for col, item in enumerate(row_data):
                self.table_widget.setItem(dest_row, col, item)
            self.table_widget.clearSelection()

    def move_column(self, source_col):
        max_cols = self.table_widget.columnCount()
        dest_col, ok = QInputDialog.getInt(self, "列の移動", f"{source_col + 1}列目をどこに移動しますか？ (1〜{max_cols})", source_col + 1, 1, max_cols)
        if ok and dest_col - 1 != source_col:
            dest_col -= 1
            col_data = [self.table_widget.takeItem(row, source_col) for row in range(self.table_widget.rowCount())]
            self.table_widget.removeColumn(source_col)
            self.table_widget.insertColumn(dest_col)
            for row, item in enumerate(col_data):
                self.table_widget.setItem(row, dest_col, item)
            self.table_widget.clearSelection()

    def add_table_row(self): self.table_widget.insertRow(self.table_widget.rowCount())
    def add_table_column(self): self.table_widget.insertColumn(self.table_widget.columnCount())
    def remove_table_row(self, index): 
        if index >= 0: self.table_widget.removeRow(index); self.table_widget.clearSelection()
    def remove_table_column(self, index):
        if index >= 0: self.table_widget.removeColumn(index); self.table_widget.clearSelection()

    def initialize_table_layout_selectors(self):
        for combo in self.table_layout_cascading_combos:
            self.table_layout_combo_layout.removeWidget(combo)
            combo.deleteLater()
        self.table_layout_cascading_combos.clear()

        hierarchy = self.config_data.get("YEARS_HIERARCHY", {})
        
        top_level_items = []
        if hierarchy:
            top_level_items = sorted(list(hierarchy.keys()))
        else:
            # Fallback: Infer years from table_layout keys if hierarchy is missing
            inferred_years = set()
            for key in self.config_data.keys():
                if key.startswith("table_layout"):
                    year_str = key.replace("table_layout", "")
                    if year_str:
                        inferred_years.add(year_str)
            
            if inferred_years:
                # If we inferred years, treat them as a single-level hierarchy
                # and reconstruct a simple hierarchy to fit the existing logic.
                hierarchy = {year: {} for year in inferred_years}
                self.config_data["YEARS_HIERARCHY"] = hierarchy # Store it back for consistency
                top_level_items = sorted(list(hierarchy.keys()))

        first_combo = QComboBox()
        first_combo.addItem("(未選択)")
        first_combo.addItems(top_level_items)
        first_combo.currentIndexChanged.connect(
            lambda: self.update_table_layout_selectors(level=1)
        )
        self.table_layout_cascading_combos.append(first_combo)
        self.table_layout_combo_layout.addWidget(first_combo)
        
        self.update_table_layout_selectors(level=1)

    def update_table_layout_selectors(self, level):
        while len(self.table_layout_cascading_combos) > level:
            combo = self.table_layout_cascading_combos.pop()
            self.table_layout_combo_layout.removeWidget(combo)
            combo.deleteLater()

        path = []
        for i in range(level):
            combo = self.table_layout_cascading_combos[i]
            if combo.currentIndex() > 0:
                 path.append(combo.currentText())
            else:
                self.on_final_table_layout_selection_changed()
                return

        node = self.config_data.get("YEARS_HIERARCHY", {})
        for key in path:
            node = node.get(key, {})
        
        children = node.keys() if node else []

        if children:
            new_combo = QComboBox()
            new_combo.addItem("(未選択)")
            new_combo.addItems(sorted(list(children)))
            new_combo.currentIndexChanged.connect(
                lambda state, l=level + 1: self.update_table_layout_selectors(level=l)
            )
            self.table_layout_cascading_combos.append(new_combo)
            self.table_layout_combo_layout.addWidget(new_combo)

        self.on_final_table_layout_selection_changed()

    def on_final_table_layout_selection_changed(self):
        # ★★★ 1. 現在の選択位置を保存 ★★★
        current_row = self.table_widget.currentRow()
        current_col = self.table_widget.currentColumn()

        year_key = None
        path = []
        for combo in self.table_layout_cascading_combos:
            if combo.currentIndex() > 0:
                path.append(combo.currentText())
            else:
                break
        if path:
            year_key = "_".join(path)

        is_leaf = False
        if year_key:
            node = self.config_data.get("YEARS_HIERARCHY", {})
            for key in path:
                node = node.get(key, {})
            if not node:
                is_leaf = True
        
        self.table_widget.setEnabled(is_leaf)
        self.add_row_button.setEnabled(is_leaf)
        self.add_col_button.setEnabled(is_leaf)
        if not is_leaf:
            self.table_widget.clear()
            self.table_widget.setRowCount(0)
            self.table_widget.setColumnCount(0)
        if is_leaf:
            self.table_layout_leaf_selection_changed.emit()

        if self.current_selected_table_layout_key != year_key:
            self.update_table_layout_in_config()
            self.current_selected_table_layout_key = year_key
            self.populate_table_layout()
            self.on_slot_selection_changed()
        
        # ★★★ 2. 選択状態を復元 ★★★
        if current_row >= 0 and current_col >= 0:
            self.table_widget.setCurrentCell(current_row, current_col)

    def populate_table_layout(self):
        self.table_widget.clear()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)

        year_str = self.current_selected_table_layout_key
        if not year_str: return
        
        # Try fetching data with and without a leading underscore for compatibility
        layout_key_no_underscore = f"table_layout{year_str}"
        layout_key_with_underscore = f"table_layout_{year_str}"
        layout_data = self.config_data.get(layout_key_no_underscore)
        if layout_data is None:
            layout_data = self.config_data.get(layout_key_with_underscore, [])

        fixed_slots_key_no_underscore = f"FIXED_SLOTS{year_str}"
        fixed_slots_key_with_underscore = f"FIXED_SLOTS_{year_str}"
        fixed_slots = self.config_data.get(fixed_slots_key_no_underscore)
        if fixed_slots is None:
            fixed_slots = self.config_data.get(fixed_slots_key_with_underscore, {})

        if not layout_data: return
        rows, cols = len(layout_data), len(layout_data[0]) if layout_data else 0
        self.table_widget.setRowCount(rows)
        self.table_widget.setColumnCount(cols)

        # Theme-aware colors
        color_empty_slot = QApplication.palette().color(QPalette.AlternateBase)
        
        color_fixed_slot = QApplication.palette().color(QPalette.Highlight)
        color_fixed_slot.setAlpha(60) # Use a semi-transparent highlight

        for i, row_data in enumerate(layout_data):
            for j, cell_data in enumerate(row_data):
                slot_name = str(cell_data) if cell_data is not None else ""
                item = QTableWidgetItem()
                item.setData(Qt.UserRole, slot_name)
                
                if not slot_name:
                    item.setText("")
                    item.setBackground(color_empty_slot)
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                elif slot_name in fixed_slots:
                    item.setText(f"{slot_name}\n[{fixed_slots[slot_name]}]")
                    item.setBackground(color_fixed_slot)
                    # Text color for fixed slots should contrast with the highlight
                    item.setForeground(QApplication.palette().color(QPalette.Text))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                else:
                    item.setText(slot_name)
                    # No background color set for normal items, uses theme default
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                
                self.table_widget.setItem(i, j, item)
        self.table_widget.resizeRowsToContents()

    def update_table_layout_in_config(self):
        year_str = self.current_selected_table_layout_key
        if not year_str: return
        
        layout_key = f"table_layout{year_str}"
        fixed_slots_key = f"FIXED_SLOTS{year_str}"

        rows, cols = self.table_widget.rowCount(), self.table_widget.columnCount()
        new_layout_data = []
        all_slot_names_in_table = set()
        for i in range(rows):
            row_data = []
            for j in range(cols):
                item = self.table_widget.item(i, j)
                if not item:
                    row_data.append("")
                    continue

                original_slot_name = item.data(Qt.UserRole)
                is_fixed = original_slot_name in self.config_data.get(fixed_slots_key, {})

                if is_fixed:
                    slot_name = original_slot_name
                else:
                    slot_name = item.text().strip()
                
                row_data.append(slot_name)
                if slot_name:
                    all_slot_names_in_table.add(slot_name)
            new_layout_data.append(row_data)
        
        self.config_data[layout_key] = new_layout_data
        all_slots_key = f"ALL_SLOTS_{year_str}"
        self.config_data[all_slots_key] = sorted(list(all_slot_names_in_table))

        if fixed_slots_key in self.config_data:
            current_fixed_slots = self.config_data[fixed_slots_key]
            orphaned_slots = set(current_fixed_slots.keys()) - all_slot_names_in_table
            for slot in orphaned_slots:
                del current_fixed_slots[slot]

    def open_initial_units_dialog(self):
        dialog = InitialUnitsDialog(self.config_data, self.config_model, self)
        dialog.exec()

    def open_art_subject_dialog(self):
        dialog = ArtSubjectDialog(self.config_data, self)
        dialog.exec()

    ''' --- その他設定タブ --- '''
    def create_general_settings_tab(self):
        """その他設定タブを作成（GeneralSettingsTabクラスを使用）"""
        from config_editor.tabs import GeneralSettingsTab
        
        # GeneralSettingsTabクラスをインスタンス化
        self.general_settings_tab_widget = GeneralSettingsTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.general_settings_tab_widget)
        self.tab_list.addItem("その他設定")
        
        # 後方互換性のため古い属性名も維持
        self.settings_layout = self.general_settings_tab_widget.settings_layout
        self.theme_widgets = self.general_settings_tab_widget._theme_widgets
        if not hasattr(self, 'general_settings_widgets'):
            self.general_settings_widgets = {}
        self.general_settings_widgets.update(self.general_settings_tab_widget._settings_widgets)
        
        # マスターconfig保存位置ボタンのシグナル接続
        self.general_settings_tab_widget.master_path_btn.clicked.connect(self.select_master_config_path)
        self.master_config_path_display = self.general_settings_tab_widget.master_path_edit
        
        # 動的な設定グループ（絞り込み設定）を作成
        self._create_filter_settings_group()
        
        # 静的な設定定義
        self.general_settings_definition = {
            "組み合わせ設定": {
                "SUBJECT_COMBINATION_COUNT": {"label": "1ページ当たりの時間割の最大数", "type": "spinbox", "range": [1, 10000]},
                "MAX_MEMORY_LIMIT": {"label": "最大メモリ使用量 (MB)", "type": "spinbox", "range": [100, 10000]},
            },
            "その他設定": {
                "SCHOOL_NAME": {"label": "学校名", "type": "lineedit"},
                "TIMETABLE_ORDER": {"label": "時間割を単位数が多い順に表示", "type": "checkbox"},
                "RUN_TUTORIAL_ON_STARTUP": {"label": "チュートリアルを有効化する", "type": "checkbox"}
            }
        }
        
        self.theme_color_definitions = {
            "duplicate_subject": {"label": "重複教科の背景色", "default": "#FFD699"},
            "prerequisite_missing": {"label": "前提教科不足の背景色", "default": "#FF9999"},
            "no_together_conflict": {"label": "同時選択不可の背景色", "default": "#7CAAFA"},
            "user_locked_subject": {"label": "手動固定した教科の背景色", "default": "#D8E6FF"},
            "user_locked_conflict": {"label": "手動固定により競合した教科の背景色", "default": "#E0E0E0"}
        }
        
        # GeneralSettingsTabの「その他設定」グループに追加ボタンを挿入
        # (取得単位数の初期値加算、特別選択教科設定)
        setting_icon = self._create_icon_from_svg_data(os.path.join(self.base_path, "svgs", "edit.svg"), self.target_color)
        
        # 既存の「その他設定」グループのレイアウトを取得
        other_group_layout = self.general_settings_tab_widget.settings_layout.itemAt(2).widget().layout()
        
        initial_units_button = QPushButton(setting_icon, "設定を開く")
        initial_units_button.clicked.connect(self.open_initial_units_dialog)
        other_group_layout.addRow("取得単位数の初期値加算", initial_units_button)
        
        art_subject_button = QPushButton(setting_icon, "設定を開く")
        art_subject_button.clicked.connect(self.open_art_subject_dialog)
        other_group_layout.addRow("特別選択教科設定（1年芸術科目など）", art_subject_button)
        
        # 初期状態の更新
        QTimer.singleShot(0, self._update_general_settings_enabled_state)

    def populate_theme_settings(self):
        theme_data = self.config_data.get("THEME_COLORS", {})
        for key, props in self.theme_color_definitions.items():
            default_light = props['default']
            default_dark = props.get('default_dark', props['default'])
            light_color = theme_data.get(f"{key}_light", default_light)
            dark_color = theme_data.get(f"{key}_dark", default_dark)
            btn_light = self.theme_widgets.get(f"{key}_light")
            btn_dark = self.theme_widgets.get(f"{key}_dark")
            if btn_light:
                # Light mode buttons use black text
                btn_light.setStyleSheet(f"background-color: {light_color}; color: #000000;")
                btn_light.setProperty("color_val", light_color)
            if btn_dark:
                # Dark mode buttons use white text
                btn_dark.setStyleSheet(f"background-color: {dark_color}; color: #ffffff;")
                btn_dark.setProperty("color_val", dark_color)

    def select_color_for_theme(self, button, key, mode='light'):
        current_color = button.property("color_val")
        color = QColorDialog.getColor(QColor(current_color or "#ffffff"), self, "カラーを選択")
        if color.isValid():
            hex_color = color.name()
            if mode == 'light':
                button.setStyleSheet(f"background-color: {hex_color}; color: #000000;")
            else:
                button.setStyleSheet(f"background-color: {hex_color}; color: #ffffff;")
            button.setProperty("color_val", hex_color)
            # do not override the other mode, only set this one

    def _get_theme_settings_from_ui_and_update_config(self, config_dict):
        if not hasattr(self, 'theme_widgets'): return

        if "THEME_COLORS" not in config_dict or not isinstance(config_dict.get("THEME_COLORS"), dict):
            config_dict["THEME_COLORS"] = {}
        
        for key, props in self.theme_color_definitions.items():
            light = self.theme_widgets.get(f"{key}_light").property("color_val") if self.theme_widgets.get(f"{key}_light") else None
            dark = self.theme_widgets.get(f"{key}_dark").property("color_val") if self.theme_widgets.get(f"{key}_dark") else None
            if light:
                config_dict["THEME_COLORS"][f"{key}_light"] = light
            if dark:
                config_dict["THEME_COLORS"][f"{key}_dark"] = dark


    def select_master_path(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "マスターconfigファイルを選択", str(get_exe_directory()), "Json Files (*.json)")
        if not file_path:
            return
        elif file_path:
            if "MASTER_CONFIG_PATH" in self.general_settings_widgets:
                self.general_settings_widgets["MASTER_CONFIG_PATH"].setText(file_path)

    def _create_widget_for_setting(self, key, props):
        widget_type = props.get("type")
        if widget_type == "spinbox":
            widget = QSpinBox()
            if "range" in props:
                widget.setRange(props["range"][0], props["range"][1])
        elif widget_type == "checkbox":
            widget = QCheckBox()
        else:
            widget = QLineEdit()
        self.general_settings_widgets[key] = widget
        return widget

    def _create_filter_settings_group(self):
        # 絞り込み設定グループ
        filter_group = QGroupBox("絞り込み設定")
        filter_layout = QVBoxLayout(filter_group)

        # メインの有効化チェックボックスと追加設定
        top_row_layout = QHBoxLayout()
        cb_main_filter = QCheckBox("絞り込み有効化")
        self.general_settings_widgets["ACTIVE_FILTER_SUBJECT"] = cb_main_filter
        top_row_layout.addWidget(cb_main_filter)
        filter_layout.addLayout(top_row_layout)

        # --- 教科数での絞り込み ---
        separator1 = QFrame()
        separator1.setFrameShape(QFrame.HLine)
        separator1.setFrameShadow(QFrame.Sunken)
        filter_layout.addWidget(separator1)

        cb_amount = self._add_setting_checkbox(filter_layout, "ACTIVE_FILTER_SUBJECT_AMOUNT", "教科数での絞り込みを有効にする")
        min_subj_row = self._create_slider_spinbox_row("最小教科数:", "MIN_SUBJECT_COUNT", "ACTIVE_MIN_SUBJECT", [0, 30])
        max_subj_row = self._create_slider_spinbox_row("最大教科数:", "MAX_SUBJECT_COUNT", "ACTIVE_MAX_SUBJECT", [0, 30])
        filter_layout.addWidget(min_subj_row)
        filter_layout.addWidget(max_subj_row)

        # --- 単位数での絞り込み ---
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.HLine)
        separator2.setFrameShadow(QFrame.Sunken)
        filter_layout.addWidget(separator2)

        cb_units = self._add_setting_checkbox(filter_layout, "ACTIVE_FILTER_SUBJECT_UNITS", "単位数での絞り込みを有効にする")
        min_units_row = self._create_slider_spinbox_row("最小単位数:", "MIN_SUBJECT_COUNT_UNITS", "ACTIVE_MIN_SUBJECT_UNITS", [0, 100])
        max_units_row = self._create_slider_spinbox_row("最大単位数:", "MAX_SUBJECT_COUNT_UNITS", "ACTIVE_MAX_SUBJECT_UNITS", [0, 100])
        filter_layout.addWidget(min_units_row)
        filter_layout.addWidget(max_units_row)

        self.settings_layout.addWidget(filter_group)

        # シグナル接続
        cb_main_filter.stateChanged.connect(self._update_general_settings_enabled_state)
        cb_amount.stateChanged.connect(self._update_general_settings_enabled_state)
        cb_units.stateChanged.connect(self._update_general_settings_enabled_state)
        self.general_settings_widgets["ACTIVE_MIN_SUBJECT"].stateChanged.connect(self._update_general_settings_enabled_state)
        self.general_settings_widgets["ACTIVE_MAX_SUBJECT"].stateChanged.connect(self._update_general_settings_enabled_state)
        self.general_settings_widgets["ACTIVE_MIN_SUBJECT_UNITS"].stateChanged.connect(self._update_general_settings_enabled_state)
        self.general_settings_widgets["ACTIVE_MAX_SUBJECT_UNITS"].stateChanged.connect(self._update_general_settings_enabled_state)

    def _add_setting_checkbox(self, layout, key, label):
        cb = QCheckBox(label)
        self.general_settings_widgets[key] = cb
        layout.addWidget(cb)
        return cb

    def _create_slider_spinbox_row(self, label_text, key_spin, key_check, s_range):
        row_frame = QFrame()
        row_layout = QHBoxLayout(row_frame)
        row_layout.setContentsMargins(0, 0, 0, 0)

        row_layout.addWidget(QLabel(label_text))

        spinbox = QSpinBox()
        spinbox.setRange(s_range[0], s_range[1])
        self.general_settings_widgets[key_spin] = spinbox
        row_layout.addWidget(spinbox)

        slider = QSlider(Qt.Horizontal)
        slider.setRange(s_range[0], s_range[1])
        self.general_settings_widgets[key_spin + "_slider"] = slider # Add slider to widgets
        row_layout.addWidget(slider)

        checkbox = QCheckBox("有効化")
        self.general_settings_widgets[key_check] = checkbox
        row_layout.addWidget(checkbox)

        slider.valueChanged.connect(spinbox.setValue)
        spinbox.valueChanged.connect(slider.setValue)

        return row_frame

    def _update_general_settings_enabled_state(self):
        # メインの絞り込みが有効か
        is_main_filter_active = self.general_settings_widgets["ACTIVE_FILTER_SUBJECT"].isChecked()

        # --- 教科数 ---
        self.general_settings_widgets["ACTIVE_FILTER_SUBJECT_AMOUNT"].setEnabled(is_main_filter_active)
        is_amount_active = is_main_filter_active and self.general_settings_widgets["ACTIVE_FILTER_SUBJECT_AMOUNT"].isChecked()
        is_min_subj_active = self.general_settings_widgets["ACTIVE_MIN_SUBJECT"].isChecked()
        is_max_subj_active = self.general_settings_widgets["ACTIVE_MAX_SUBJECT"].isChecked()
        
        self.general_settings_widgets["ACTIVE_MIN_SUBJECT"].setEnabled(is_amount_active)
        self.general_settings_widgets["MIN_SUBJECT_COUNT"].setEnabled(is_amount_active and is_min_subj_active)
        self.general_settings_widgets["MIN_SUBJECT_COUNT_slider"].setEnabled(is_amount_active and is_min_subj_active)
        
        self.general_settings_widgets["ACTIVE_MAX_SUBJECT"].setEnabled(is_amount_active)
        self.general_settings_widgets["MAX_SUBJECT_COUNT"].setEnabled(is_amount_active and is_max_subj_active)
        self.general_settings_widgets["MAX_SUBJECT_COUNT_slider"].setEnabled(is_amount_active and is_max_subj_active)

        # --- 単位数 ---
        self.general_settings_widgets["ACTIVE_FILTER_SUBJECT_UNITS"].setEnabled(is_main_filter_active)
        is_units_active = is_main_filter_active and self.general_settings_widgets["ACTIVE_FILTER_SUBJECT_UNITS"].isChecked()
        is_min_units_active = self.general_settings_widgets["ACTIVE_MIN_SUBJECT_UNITS"].isChecked()
        is_max_units_active = self.general_settings_widgets["ACTIVE_MAX_SUBJECT_UNITS"].isChecked()
        
        self.general_settings_widgets["ACTIVE_MIN_SUBJECT_UNITS"].setEnabled(is_units_active)
        self.general_settings_widgets["MIN_SUBJECT_COUNT_UNITS"].setEnabled(is_units_active and is_min_units_active)
        self.general_settings_widgets["MIN_SUBJECT_COUNT_UNITS_slider"].setEnabled(is_units_active and is_min_units_active)
        
        self.general_settings_widgets["ACTIVE_MAX_SUBJECT_UNITS"].setEnabled(is_units_active)
        self.general_settings_widgets["MAX_SUBJECT_COUNT_UNITS"].setEnabled(is_units_active and is_max_units_active)
        self.general_settings_widgets["MAX_SUBJECT_COUNT_UNITS_slider"].setEnabled(is_units_active and is_max_units_active)

    def select_master_config_path(self):
        """マスターconfig保存位置を選択するダイアログを開く"""
        from PySide6.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "マスターconfigファイルを選択",
            str(get_exe_directory()),
            "JSONファイル (*.json)"
        )
        if file_path:
            self.master_config_path_display.setText(file_path)
            self.config_data["MASTER_CONFIG_PATH"] = file_path

    def populate_general_settings(self):
        general_settings = self.config_data.get("GENERAL_SETTINGS", {})

        for key, widget in self.general_settings_widgets.items():
            if "_slider" in key: continue
            
            value = general_settings.get(key)
            if value is None:
                continue

            widget.blockSignals(True)
            try:
                if isinstance(widget, (QSpinBox, QSlider)):
                    widget.setValue(int(value))
                elif isinstance(widget, QCheckBox):
                    widget.setChecked(bool(value))
                elif isinstance(widget, QLineEdit):
                    widget.setText(str(value))
            except (ValueError, TypeError):
                print(f"Could not set value for {key}: {value}")
            finally:
                widget.blockSignals(False)
        self._update_general_settings_enabled_state()

    def update_general_settings_in_config(self, config_dict):
        if "GENERAL_SETTINGS" not in config_dict:
            config_dict["GENERAL_SETTINGS"] = {}

        for key, widget in self.general_settings_widgets.items():
            if "_slider" in key: continue # Sliders are linked to spinboxes, no need to save separately
            if isinstance(widget, QSpinBox):
                config_dict["GENERAL_SETTINGS"][key] = widget.value()
            elif isinstance(widget, QCheckBox):
                config_dict["GENERAL_SETTINGS"][key] = widget.isChecked()
            elif isinstance(widget, QLineEdit):
                # Attempt to preserve original type if possible, otherwise save as string
                original_value = self.config_data.get("GENERAL_SETTINGS", {}).get(key)
                try:
                    if isinstance(original_value, int):
                        config_dict["GENERAL_SETTINGS"][key] = int(widget.text())
                    elif isinstance(original_value, float):
                        config_dict["GENERAL_SETTINGS"][key] = float(widget.text())
                    else: # Default to string
                        config_dict["GENERAL_SETTINGS"][key] = widget.text()
                except (ValueError, TypeError): # Fallback if conversion fails
                    config_dict["GENERAL_SETTINGS"][key] = widget.text()
            # If a widget was meant to hold a value but now it's empty, remove its entry from config_dict
            if (isinstance(widget, QLineEdit) and not widget.text()) or \
               (isinstance(widget, QSpinBox) and widget.value() == 0 and key not in ["MIN_SUBJECT_COUNT", "MAX_SUBJECT_COUNT", "MIN_SUBJECT_COUNT_UNITS", "MAX_SUBJECT_COUNT_UNITS"]): # Some spinboxes can legitimately be 0
                if key in config_dict["GENERAL_SETTINGS"]:
                    del config_dict["GENERAL_SETTINGS"][key]

    # --- 必須教科設定タブ ---
    def create_required_subjects_tab(self):
        """必須教科設定タブを作成（RequiredSubjectsTabクラスを使用）"""
        from config_editor.tabs import RequiredSubjectsTab
        
        # RequiredSubjectsTabクラスをインスタンス化
        self.required_subjects_tab_widget = RequiredSubjectsTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.required_subjects_tab_widget)
        self.tab_list.addItem("必須教科設定")
        
        # 後方互換性のため古い属性名も維持
        self.rules_list_widget = self.required_subjects_tab_widget.rules_list
        self.details_pane = self.required_subjects_tab_widget.details_pane
        self.rule_name_edit = self.required_subjects_tab_widget.rule_name_edit
        self.rule_target_combo = self.required_subjects_tab_widget.rule_target_combo
        self.rule_color_button = self.required_subjects_tab_widget.rule_color_btn
        self.conditions_widget = self.required_subjects_tab_widget.conditions_widget
        self.conditions_layout = self.required_subjects_tab_widget.conditions_layout
        # シグナル接続（MainWindowのハンドラを使用）
        # 既存の接続を切断してから再接続
        try:
            self.rules_list_widget.currentRowChanged.disconnect()
        except RuntimeError:
            pass
        try:
            self.required_subjects_tab_widget.add_rule_btn.clicked.disconnect()
        except RuntimeError:
            pass
        try:
            self.required_subjects_tab_widget.remove_rule_btn.clicked.disconnect()
        except RuntimeError:
            pass
        
        self.rules_list_widget.currentRowChanged.connect(self.on_rule_selected)
        self.required_subjects_tab_widget.add_rule_btn.clicked.connect(self.add_new_rule)
        self.required_subjects_tab_widget.remove_rule_btn.clicked.connect(self.remove_selected_rule)

    def _create_required_subjects_details_pane(self):
        details_layout = QVBoxLayout(self.details_pane)
        details_layout.setContentsMargins(10, 0, 10, 0)

        details_layout.addWidget(QLabel("<b>ルール詳細設定</b>") )
        
        form_layout = QFormLayout()
        self.rule_name_edit = QLineEdit()
        self.rule_target_combo = QComboBox()
        # Items are now populated by reload_all_selectors
        self.rule_color_button = QPushButton("カラーを選択")
        self.rule_color_button.clicked.connect(self.select_rule_color)
        form_layout.addRow("ルール名:", self.rule_name_edit)
        form_layout.addRow("対象学年:", self.rule_target_combo)
        form_layout.addRow("表示カラー:", self.rule_color_button)
        details_layout.addLayout(form_layout)

        group_area_header = QHBoxLayout()
        group_area_header.addWidget(QLabel("<b>条件グループ (OR条件)</b>") )
        group_area_header.addStretch()
        add_icon = self._create_icon_from_svg_data(os.path.join(self.base_path, "svgs", "plus.svg"), self.target_color)
        add_group_button = QPushButton(add_icon, "グループ追加")
        add_group_button.clicked.connect(lambda: self.add_condition_group_widget())
        delete_icon = self._create_icon_from_svg_data(os.path.join(self.base_path, "svgs", "eraser.svg"), self.target_color)
        remove_group_button = QPushButton(delete_icon, "選択グループ削除")
        remove_group_button.clicked.connect(self.remove_condition_group_widget)
        group_area_header.addWidget(add_group_button)
        group_area_header.addWidget(remove_group_button)
        details_layout.addLayout(group_area_header)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        self.conditions_widget = QWidget()
        self.conditions_layout = QVBoxLayout(self.conditions_widget)
        self.conditions_layout.setAlignment(Qt.AlignTop)
        scroll_area.setWidget(self.conditions_widget)
        details_layout.addWidget(scroll_area)

        self.rule_name_edit.editingFinished.connect(self.update_rule_details)
        self.rule_target_combo.currentIndexChanged.connect(self.update_rule_details)

    def add_condition_group_widget(self, data=None, master_subjects=None):
        if master_subjects is None:
            master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        group_count = self.conditions_layout.count() + 1
        group_widget = RequiredSubjectGroupWidget(group_count, master_subjects)
        if data and isinstance(data, dict):
            group_widget.set_data(data)
        group_widget.changed.connect(self.update_rule_details)
        self.conditions_layout.addWidget(group_widget)
        if self.details_pane.isVisible():
            self.update_rule_details()

    def remove_condition_group_widget(self):
        checked_groups = []
        for i in range(self.conditions_layout.count()):
            widget = self.conditions_layout.itemAt(i).widget()
            if widget and widget.isChecked():
                checked_groups.append(widget)
        
        if not checked_groups:
            QMessageBox.warning(self, "", "削除する条件グループのチェックボックスをONにしてください。" )
            return

        reply = QMessageBox.question(self, "確認", f"{len(checked_groups)}個のグループを削除しますか？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            for widget in checked_groups:
                widget.setParent(None)
                widget.deleteLater()
            QTimer.singleShot(0, self._renumber_condition_groups)
            QTimer.singleShot(0, self.update_rule_details)

    def _renumber_condition_groups(self):
        for i in range(self.conditions_layout.count()):
            widget = self.conditions_layout.itemAt(i).widget()
            if widget:
                widget.setTitle(f"条件グループ {i+1}")

    def _clear_details_pane(self):
        self.rule_name_edit.clear()
        self.rule_color_button.setStyleSheet("")
        while self.conditions_layout.count():
            child = self.conditions_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

    def on_rule_selected(self, index):
        # No longer save on selection change
        if index < 0:
            self.details_pane.setVisible(False)
            self.details_pane.setEnabled(False)
            self.current_rule_key = None # Keep for now, might be useful for re-selection
            return

        item = self.rules_list_widget.item(index)
        if not item:
            return

        rule_data = item.data(Qt.UserRole)
        if not rule_data:
            return

        # Use the stored 'id' and 'target' from the rule_data itself
        self.current_rule_key = (rule_data.get('id'), rule_data.get('target'))

        self._clear_details_pane()
        self.details_pane.setVisible(True)
        self.details_pane.setEnabled(True)

        self.rule_name_edit.setText(rule_data.get("name", f"ルール {rule_data.get('id')}"))
        
        display_target = "全体" if rule_data.get('target') == "ALL" else rule_data.get('target')
        self.rule_target_combo.blockSignals(True)
        self.rule_target_combo.setCurrentText(display_target)
        self.rule_target_combo.blockSignals(False)
        
        color = rule_data.get("color", "#FFFFFF")
        self.rule_color_button.setStyleSheet(f"background-color: {color};")
        self.rule_color_button.setProperty("color", color)

        conditions = rule_data.get("conditions", [])
        master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        for cond_data in conditions:
            self.add_condition_group_widget(cond_data, master_subjects)

    def generate_new_rule_id(self):
        existing_ids = set()
        all_req_keys = [k for k in self.config_data if k.startswith("REQUIRED_SUBJECTS_")]
        for key in all_req_keys:
            if key in self.config_data and isinstance(self.config_data[key], dict):
                existing_ids.update(self.config_data[key].keys())

        highest_id = 0
        for an_id in existing_ids:
            if an_id.isdigit():
                highest_id = max(highest_id, int(an_id))
        
        return str(highest_id + 1)

    def add_new_rule(self):
        rule_name, ok = QInputDialog.getText(self, "新規ルール作成", "新しいルール名を入力してください:")
        if not (ok and rule_name):
            return

        new_id = self.generate_new_rule_id() # Still useful to have unique IDs

        new_rule_data = {
            "id": new_id,
            "name": rule_name,
            "target": "ALL",
            "color": "#FFFFFF",
            "conditions": []
        }
        
        item = QListWidgetItem(f"{rule_name} (全体)")
        item.setData(Qt.UserRole, new_rule_data)
        self.rules_list_widget.addItem(item)
        self.rules_list_widget.setCurrentItem(item)

    def remove_selected_rule(self):
        current_item = self.rules_list_widget.currentItem()
        if not current_item:
            QMessageBox.warning(self, "エラー", "削除するルールを選択してください。" )
            return
            
        rule_data = current_item.data(Qt.UserRole)
        name_to_delete = rule_data.get("name", rule_data.get("id", "不明"))

        reply = QMessageBox.question(self, "確認", f"ルール「{name_to_delete}」を削除しますか？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.rules_list_widget.takeItem(self.rules_list_widget.row(current_item))
            # No need to repopulate, just hide the details pane
            self.details_pane.setVisible(False)
            self.current_rule_key = None

    def select_rule_color(self):
        current_color = self.rule_color_button.property("color") or "#FFFFFF"
        new_color = QColorDialog.getColor(QColor(current_color), self)
        if new_color.isValid():
            color_hex = new_color.name()
            self.rule_color_button.setStyleSheet(f"background-color: {color_hex};")
            self.rule_color_button.setProperty("color", color_hex)
            self.update_rule_details()

    def update_rule_details(self):
        current_item = self.rules_list_widget.currentItem()
        if not current_item or not self.details_pane.isVisible():
            return

        # Get existing data from the item
        rule_data = current_item.data(Qt.UserRole)
        if not rule_data:
             return # Should not happen if item is selected

        # Read all UI elements in the details pane
        new_name = self.rule_name_edit.text()
        if not new_name:
            # Maybe show a warning, but for now just return
            return
            
        current_combo_text = self.rule_target_combo.currentText()
        new_target_key = "ALL" if current_combo_text == "全体" else current_combo_text
        new_color = self.rule_color_button.property("color") or "#FFFFFF"
        
        new_conditions = []
        for i in range(self.conditions_layout.count()):
            widget = self.conditions_layout.itemAt(i).widget()
            if widget:
                new_conditions.append(widget.get_data())

        # Update the dictionary
        rule_data['name'] = new_name
        rule_data['target'] = new_target_key
        rule_data['color'] = new_color
        rule_data['conditions'] = new_conditions

        # Write the updated dictionary back to the item
        current_item.setData(Qt.UserRole, rule_data)
        
        # Update the display text of the item
        current_item.setText(f"{new_name} ({current_combo_text})")

    def _get_required_subjects_from_ui(self, config_dict):
        # Clear all existing required subjects data first
        for key in list(config_dict.keys()):
            if key.startswith("REQUIRED_SUBJECTS_"):
                del config_dict[key]

        # Rebuild from the UI list
        for i in range(self.rules_list_widget.count()):
            item = self.rules_list_widget.item(i)
            rule_data = item.data(Qt.UserRole)
            if not rule_data:
                continue

            target = rule_data.get('target', 'ALL')
            rule_id = rule_data.get('id')
            
            # Data to be saved doesn't need the id or target keys inside it
            saved_data = {
                "name": rule_data.get("name"),
                "color": rule_data.get("color"),
                "conditions": rule_data.get("conditions")
            }

            req_key = f"REQUIRED_SUBJECTS_{target}"
            if req_key not in config_dict:
                config_dict[req_key] = {}
            
            config_dict[req_key][rule_id] = saved_data

    # --- 前提教科設定タブ ---
    def create_prerequisite_tab(self):
        """前提教科設定タブを作成（PrerequisiteTabクラスを使用）"""
        from config_editor.tabs import PrerequisiteTab
        
        # PrerequisiteTabクラスをインスタンス化
        self.prerequisite_tab_widget = PrerequisiteTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.prerequisite_tab_widget)
        self.tab_list.addItem("前提教科設定")
        
        # 後方互換性のため古い属性名も維持
        self.prereq_parent_list = self.prerequisite_tab_widget.parent_list
        self.prereq_assigned_list = self.prerequisite_tab_widget.assigned_list
        self.prereq_available_list = self.prerequisite_tab_widget.available_list
        self.prereq_details_pane = self.prerequisite_tab_widget.details_pane
        self.prereq_parent_label = self.prerequisite_tab_widget.parent_label

    def populate_prerequisite_parent_list(self):
        """前提教科リストを更新（PrerequisiteTabに委譲）"""
        if hasattr(self, 'prerequisite_tab_widget'):
            # ConfigModelにデータを同期してからpopulate
            self.config_model.update_data(self.config_data)
            self.prerequisite_tab_widget.populate()
        else:
            # 後方互換性
            self.prereq_parent_list.blockSignals(True)
            current_selection = self.prereq_parent_list.currentItem().text() if self.prereq_parent_list.currentItem() else None
            self.prereq_parent_list.clear()
            
            prereq_subjects = self.config_data.get("PREREQUISITE_SUBJECTS", {})
            sorted_parents = sorted(prereq_subjects.keys())
            
            for parent in sorted_parents:
                item = QListWidgetItem(parent)
                children = prereq_subjects.get(parent, [])
                item.setData(Qt.UserRole, children)
                self.prereq_parent_list.addItem(item)

            if current_selection in sorted_parents:
                for i in range(len(sorted_parents)):
                    if sorted_parents[i] == current_selection:
                        self.prereq_parent_list.setCurrentRow(i)
                        break
            self.prereq_parent_list.blockSignals(False)

    def on_prereq_parent_selected(self, index):
        if index < 0:
            self.prereq_details_pane.setVisible(False)
            self.prereq_details_pane.setEnabled(False)
            self.current_prereq_parent = None
            return

        parent_item = self.prereq_parent_list.item(index)
        self.current_prereq_parent = parent_item.text()
        self.prereq_details_pane.setVisible(True)
        self.prereq_details_pane.setEnabled(True)
        self.prereq_parent_label.setText(f"親教科: [ {self.current_prereq_parent} ]")

        assigned_subjects = parent_item.data(Qt.UserRole) or []
        self.prereq_assigned_list.clear()
        self.prereq_assigned_list.addItems(sorted(assigned_subjects))

        master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        available_subjects = [s for s in master_subjects if s not in assigned_subjects and s != self.current_prereq_parent]
        self.prereq_available_list.clear()
        self.prereq_available_list.addItems(sorted(available_subjects))
        self.prereq_selection_changed.emit()

    def add_new_prereq_rule(self):
        # Determine which subjects are already in the prereq list
        existing_parents = set()
        for i in range(self.prereq_parent_list.count()):
            existing_parents.add(self.prereq_parent_list.item(i).text())

        master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        available_parents = [s for s in master_subjects if s not in existing_parents]
        
        if not available_parents:
            QMessageBox.information(self, "情報", "すべての教科が既に追加されています。" )
            return

        parent_subject, ok = QInputDialog.getItem(self, "親教科の追加", 
                                                  "前提教科を設定する親教科を選択してください:", 
                                                  sorted(available_parents), 0, False)
        
        if ok and parent_subject:
            item = QListWidgetItem(parent_subject)
            item.setData(Qt.UserRole, []) # Start with an empty list of children
            self.prereq_parent_list.addItem(item)
            self.prereq_parent_list.setCurrentItem(item)

    def remove_selected_prereq_rule(self):
        current_item = self.prereq_parent_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "エラー", "削除する親教科を選択してください。" )
            return

        parent_subject = current_item.text()
        reply = QMessageBox.question(self, "確認", f"親教科「{parent_subject}」とその前提教科設定を削除しますか？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            # No need to touch self.config_data. Just remove the item from the UI list.
            self.prereq_parent_list.takeItem(self.prereq_parent_list.row(current_item))
            self.prereq_details_pane.setVisible(False)
            self.current_prereq_parent = None

    def move_subject_to_prereq(self):
        selected_items = self.prereq_available_list.selectedItems()
        if not selected_items:
            return

        rows_to_remove = sorted([self.prereq_available_list.row(item) for item in selected_items], reverse=True)

        for item in selected_items:
            self.prereq_assigned_list.addItem(item.text())

        for row in rows_to_remove:
            self.prereq_available_list.takeItem(row)

        self._get_prereqs_from_widget_and_update_item_data()
        
        if self.tutorial_mode and selected_items:
            parent_subject = self.current_prereq_parent
            added_subject = selected_items[0].text()
            self.prerequisite_subject_added.emit(parent_subject, added_subject)

    def move_subject_from_prereq(self):
        for item in self.prereq_assigned_list.selectedItems():
            self.prereq_available_list.addItem(item.text())
            self.prereq_assigned_list.takeItem(self.prereq_assigned_list.row(item))
        self._get_prereqs_from_widget_and_update_item_data()



    def _get_prerequisites_from_ui(self, config_dict):
        prereq_subjects = {}
        for i in range(self.prereq_parent_list.count()):
            item = self.prereq_parent_list.item(i)
            parent_name = item.text()
            children = item.data(Qt.UserRole) or []
            if children:
                prereq_subjects[parent_name] = sorted(children)
        
        if prereq_subjects:
            config_dict["PREREQUISITE_SUBJECTS"] = prereq_subjects
        elif "PREREQUISITE_SUBJECTS" in config_dict:
            del config_dict["PREREQUISITE_SUBJECTS"]
  
    def _get_prereqs_from_widget_and_update_item_data(self):
        current_parent_item = self.prereq_parent_list.currentItem()
        if not current_parent_item:
            return
        
        new_children = [self.prereq_assigned_list.item(i).text() for i in range(self.prereq_assigned_list.count())]
        current_parent_item.setData(Qt.UserRole, new_children)

    # --- 保存位置設定タブ ---
    # --- 同時選択不可設定タブ ---
    def create_no_together_tab(self):
        """同時選択不可設定タブを作成（NoTogetherTabクラスを使用）"""
        from config_editor.tabs import NoTogetherTab
        
        # NoTogetherTabクラスをインスタンス化
        self.no_together_tab_widget = NoTogetherTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.no_together_tab_widget)
        self.tab_list.addItem("同時選択不可設定")
        
        # 後方互換性のため古い属性名も維持
        self.no_together_groups_list = self.no_together_tab_widget.groups_list
        self.no_together_assigned_list = self.no_together_tab_widget.assigned_list
        self.no_together_available_list = self.no_together_tab_widget.available_list
        self.no_together_details_pane = self.no_together_tab_widget.details_pane
        self.no_together_group_label = self.no_together_tab_widget.group_label

    def populate_no_together_groups_list(self):
        """同時選択不可グループリストを更新（NoTogetherTabに委譲）"""
        if hasattr(self, 'no_together_tab_widget'):
            self.config_model.update_data(self.config_data)
            self.no_together_tab_widget.populate()
        else:
            # 後方互換性
            self.no_together_groups_list.blockSignals(True)
            self.no_together_groups_list.clear()
            self.no_together_details_pane.setVisible(False)
            
            groups = self.config_data.get("NO_TOGETHER_SUBJECTS", [])
            for i, group in enumerate(groups):
                item_text = ", ".join(sorted(group)) if group else "(空のグループ)"
                item = QListWidgetItem(item_text)
                item.setData(Qt.UserRole, group)
                self.no_together_groups_list.addItem(item)
                
            self.no_together_groups_list.blockSignals(False)
            if self.no_together_groups_list.count() > 0:
                self.no_together_groups_list.setCurrentRow(0)
            else:
                self.on_no_together_group_selected(-1)

    def on_no_together_group_selected(self, index):
        self.current_no_together_group_index = index

        if index < 0:
            self.no_together_details_pane.setVisible(False)
            return

        item = self.no_together_groups_list.item(index)
        if not item:
            return

        self.no_together_details_pane.setVisible(True)
        
        group_subjects = item.data(Qt.UserRole)
        self.no_together_group_label.setText(f"グループ {index + 1} の編集")
        
        self.no_together_assigned_list.clear()
        self.no_together_assigned_list.addItems(sorted(group_subjects))

        master_subjects = self.config_data.get("MASTER_SUBJECTS", [])
        available_subjects = [s for s in master_subjects if s not in group_subjects]
        self.no_together_available_list.clear()
        self.no_together_available_list.addItems(sorted(available_subjects))

    def add_no_together_group(self):
        new_group = []
        item = QListWidgetItem("(空のグループ)")
        item.setData(Qt.UserRole, new_group)
        self.no_together_groups_list.addItem(item)
        self.no_together_groups_list.setCurrentItem(item)
        if self.tutorial_mode:
            self.no_together_group_added.emit()

    def delete_no_together_group(self):
        current_item = self.no_together_groups_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "エラー", "削除するグループを選択してください。" )
            return
        
        reply = QMessageBox.question(self, "確認", "選択したグループを削除しますか？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            row = self.no_together_groups_list.row(current_item)
            self.no_together_groups_list.takeItem(row)

    def move_subject_to_no_together_group(self):
        items_to_move = self.no_together_available_list.selectedItems()
        if not items_to_move: return

        for item in items_to_move:
            self.no_together_assigned_list.addItem(item.text())
        
        # It's safer to get the rows to remove first, then remove them
        rows_to_remove = sorted([self.no_together_available_list.row(item) for item in items_to_move], reverse=True)
        for row in rows_to_remove:
            self.no_together_available_list.takeItem(row)
        
        self._get_no_together_from_widget_and_update_item_data()
        if self.tutorial_mode:
            self.no_together_subject_added.emit()

    def move_subject_from_no_together_group(self):
        items_to_move = self.no_together_assigned_list.selectedItems()
        if not items_to_move: return
        
        texts_to_move = [item.text() for item in items_to_move]
        rows_to_remove = sorted([self.no_together_assigned_list.row(item) for item in items_to_move], reverse=True)

        for text in texts_to_move:
            self.no_together_available_list.addItem(text)
        
        for row in rows_to_remove:
            self.no_together_assigned_list.takeItem(row)

        self._get_no_together_from_widget_and_update_item_data()
        self.no_together_available_list.sortItems()


    
    def _get_no_together_from_ui(self, config_dict):
        all_groups = []
        for i in range(self.no_together_groups_list.count()):
            item = self.no_together_groups_list.item(i)
            if item:
                group_data = item.data(Qt.UserRole)
                if group_data: # Only add non-empty groups
                    all_groups.append(sorted(group_data))
        
        if all_groups:
            config_dict["NO_TOGETHER_SUBJECTS"] = all_groups
        elif "NO_TOGETHER_SUBJECTS" in config_dict:
            del config_dict["NO_TOGETHER_SUBJECTS"]
    
    def _get_no_together_from_widget_and_update_item_data(self):
        current_group_item = self.no_together_groups_list.currentItem()
        if not current_group_item:
            return
        
        new_subjects = [self.no_together_assigned_list.item(i).text() for i in range(self.no_together_assigned_list.count())]
        current_group_item.setData(Qt.UserRole, new_subjects)
        current_group_item.setText(", ".join(sorted(new_subjects)) if new_subjects else "(空のグループ)")
    
    def create_save_position_tab(self):
        """保存位置設定タブを作成（SavePositionTabクラスを使用）"""
        from config_editor.tabs import SavePositionTab
        
        # SavePositionTabクラスをインスタンス化
        self.save_position_tab_widget = SavePositionTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.save_position_tab_widget)
        self.tab_list.addItem("保存位置設定")
        
        # 後方互換性のため古い属性名も維持
        self.save_pos_table = self.save_position_tab_widget.table
        self.save_pos_combo_layout = self.save_position_tab_widget._combo_layout
        self.save_pos_cascading_combos = []

    def add_save_position_row(self):
        if not self.current_save_pos_year: return
        year_suffix = self.current_save_pos_year
        all_slots_key = f"ALL_SLOTS{year_suffix}"
        all_slots = sorted(self.config_data.get(all_slots_key, []))
        
        if not all_slots:
            QMessageBox.warning(self, "スロット未定義", f'「{year_suffix}」のスロットが「スロットレイアウト設定」で定義されていません。')
            return

        used_slots = set()
        for row in range(self.save_pos_table.rowCount()):
            combo = self.save_pos_table.cellWidget(row, 0)
            if combo:
                slot_name = combo.currentText().replace(' (使用中)', '')
                used_slots.add(slot_name)

        row_position = self.save_pos_table.rowCount()
        self.save_pos_table.insertRow(row_position)
        
        combo = QComboBox()
        for s in all_slots:
            display_text = f"{s} (使用中)" if s in used_slots else s
            combo.addItem(display_text)

        self.save_pos_table.setCellWidget(row_position, 0, combo)
        self.save_pos_table.setItem(row_position, 1, QTableWidgetItem(""))

    def delete_save_position_row(self):
        current_row = self.save_pos_table.currentRow()
        if current_row >= 0:
            self.save_pos_table.removeRow(current_row)

    def auto_fill_save_positions(self):
        selected_items = self.save_pos_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "行が選択されていません", "連番記入の開始点となる行を選択してください。" )
            return

        start_row = self.save_pos_table.row(selected_items[0])
        
        start_item = self.save_pos_table.item(start_row, 1)
        if not start_item or not start_item.text():
            QMessageBox.warning(self, "開始セルが空です", "連番記入の開始点となる行の「セル位置」列に値がありません。" )
            return

        start_text = start_item.text()

        for i in range(start_row + 1, self.save_pos_table.rowCount()):
            increment = i - start_row
            
            parts = start_text.split(',')
            new_parts = []
            for part in parts:
                part = part.strip()
                match = re.match(r'^([A-Z]+)(\d+)$', part, re.IGNORECASE)
                if match:
                    prefix = match.group(1)
                    try:
                        number = int(match.group(2))
                        new_number = number + increment
                        new_parts.append(f"{prefix}{new_number}")
                    except ValueError:
                        new_parts.append(part)
                else:
                    new_parts.append(part)

            new_text = ",".join(new_parts)
            
            current_item = self.save_pos_table.item(i, 1)
            if not current_item:
                current_item = QTableWidgetItem(new_text)
                self.save_pos_table.setItem(i, 1, current_item)
            else:
                current_item.setText(new_text)

    def initialize_save_position_selectors(self):
        for combo in self.save_pos_cascading_combos:
            self.save_pos_combo_layout.removeWidget(combo)
            combo.deleteLater()
        self.save_pos_cascading_combos.clear()

        top_level_items = sorted(list(self.config_data.get("YEARS_HIERARCHY", {}).keys()))
        first_combo = QComboBox()
        first_combo.addItem("(未選択)")
        first_combo.addItems(top_level_items)
        first_combo.currentIndexChanged.connect(
            lambda: self.update_save_position_selectors(level=1)
        )
        self.save_pos_cascading_combos.append(first_combo)
        self.save_pos_combo_layout.addWidget(first_combo)
        
        self.update_save_position_selectors(level=1)

    def update_save_position_selectors(self, level):
        while len(self.save_pos_cascading_combos) > level:
            combo = self.save_pos_cascading_combos.pop()
            self.save_pos_combo_layout.removeWidget(combo)
            combo.deleteLater()

        path = []
        for i in range(level):
            combo = self.save_pos_cascading_combos[i]
            if combo.currentIndex() > 0:
                 path.append(combo.currentText())
            else:
                self.on_final_save_position_selection_changed()
                return

        node = self.config_data.get("YEARS_HIERARCHY", {})
        for key in path:
            node = node.get(key, {})
        
        children = node.keys() if node else []

        if children:
            new_combo = QComboBox()
            new_combo.addItem("(未選択)")
            new_combo.addItems(sorted(list(children)))
            new_combo.currentIndexChanged.connect(
                lambda state, l=level + 1: self.update_save_position_selectors(level=l)
            )
            self.save_pos_cascading_combos.append(new_combo)
            self.save_pos_combo_layout.addWidget(new_combo)

        self.on_final_save_position_selection_changed()

    def on_final_save_position_selection_changed(self):
        year_key = None
        path = []
        for combo in self.save_pos_cascading_combos:
            if combo.currentIndex() > 0:
                path.append(combo.currentText())
            else:
                break
        if path:
            year_key = "_".join(path)

        is_leaf = False
        if year_key:
            node = self.config_data.get("YEARS_HIERARCHY", {})
            for key in path:
                node = node.get(key, {})
            if not node:
                is_leaf = True
        
        self.save_pos_table.setEnabled(is_leaf)
        if is_leaf:
            self.save_pos_leaf_selection_changed.emit()

        if self.current_save_pos_year != year_key:
            self._update_current_save_position_to_config_data()
            self.current_save_pos_year = year_key
            self.populate_save_position_table()

    def populate_save_position_table(self):
        self.save_pos_table.setRowCount(0)
        if not self.current_save_pos_year: return

        year_suffix = self.current_save_pos_year
        key = f"SAVE_POSITION{year_suffix}"
        save_pos_data = self.config_data.get(key, {})
        
        all_slots = sorted(self.config_data.get(f"ALL_SLOTS{year_suffix}", []))
        used_slots = set(save_pos_data.keys())

        for slot, pos_list in sorted(save_pos_data.items()):
            row_position = self.save_pos_table.rowCount()
            self.save_pos_table.insertRow(row_position)

            combo = QComboBox()
            for s in all_slots:
                display_text = s
                if s in used_slots and s != slot:
                    display_text = f"{s} (使用中)"
                combo.addItem(display_text)

            combo.setCurrentText(slot)
            self.save_pos_table.setCellWidget(row_position, 0, combo)

            cell_str = ",".join(pos_list[0]) if pos_list and pos_list[0] else ""
            self.save_pos_table.setItem(row_position, 1, QTableWidgetItem(cell_str))

    def _update_current_save_position_to_config_data(self):
        if not self.current_save_pos_year: return

        year_suffix = self.current_save_pos_year
        key = f"SAVE_POSITION{year_suffix}"
        new_data = {}
        for row in range(self.save_pos_table.rowCount()):
            combo = self.save_pos_table.cellWidget(row, 0)
            pos_item = self.save_pos_table.item(row, 1)
            
            if combo and pos_item:
                slot = combo.currentText().replace(" (使用中)", "")
                if slot:
                    pos_list = [p.strip() for p in pos_item.text().split(',') if p.strip()]
                    new_data[slot] = [pos_list]
        
        if new_data:
            self.config_data[key] = new_data
        elif key in self.config_data:
            del self.config_data[key]

    def _get_save_position_from_config_data(self, config_dict):
        if not self.current_save_pos_year: return

        year_suffix = self.current_save_pos_year
        key = f"SAVE_POSITION{year_suffix}"
        new_data = {}
        for row in range(self.save_pos_table.rowCount()):
            combo = self.save_pos_table.cellWidget(row, 0)
            pos_item = self.save_pos_table.item(row, 1)
            
            if combo and pos_item:
                slot = combo.currentText().replace(" (使用中)", "")
                if slot:
                    pos_list = [p.strip() for p in pos_item.text().split(',') if p.strip()]
                    new_data[slot] = [pos_list]
        
        if new_data:
            config_dict[key] = new_data
        elif key in config_dict:
            del config_dict[key]


    def _get_normal_units(self, subject, year):
        if not year: return 0
        year_suffix = year
        
        fixed_slots_key = f"FIXED_SLOTS{year_suffix}"
        fixed_slots = self.config_data.get(fixed_slots_key, {})
        if subject in fixed_slots.values():
            return list(fixed_slots.values()).count(subject)

        slots_key = f"subject_slots_base{year_suffix}"
        
        # Corrected retrieval for slot_groups
        all_subject_slots_data = self.config_data.get(slots_key, [])
        
        # Convert old dict format to new list format for safety, if not already done
        if isinstance(all_subject_slots_data, dict):
            all_subject_slots_data = [{"name": s, "data": d} for s, d in all_subject_slots_data.items()]

        slot_groups = []
        for item in all_subject_slots_data:
            if item.get("name") == subject:
                slot_groups = item.get("data", [])
                break

        if slot_groups and isinstance(slot_groups, list) and len(slot_groups) > 0:
            first_group = slot_groups[0]
            if isinstance(first_group, list):
                return len(first_group)
        
        return 0

    # --- 学年設定タブ ---
    def create_year_settings_tab(self):
        """学年設定タブを作成（YearSettingsTabクラスを使用）"""
        from config_editor.tabs import YearSettingsTab
        
        # YearSettingsTabクラスをインスタンス化
        self.year_settings_tab_widget = YearSettingsTab(
            config_model=self.config_model,
            parent=self,
            icon_creator=lambda path: self._create_icon_from_svg_data(path, self.target_color),
            base_path=self.base_path
        )
        
        # stacked_widgetに追加
        self.stacked_widget.addWidget(self.year_settings_tab_widget)
        self.tab_list.addItem("学年設定")
        
        # 後方互換性のため古い属性名も維持
        self.year_tree_widget = self.year_settings_tab_widget.tree_widget

    def populate_year_list(self):
        # year_settings_tab_widget.populate() を使ってアンカー情報も読み込む
        if hasattr(self, 'year_settings_tab_widget'):
            self.year_settings_tab_widget.populate()
        else:
            # フォールバック（旧方式）
            self.year_tree_widget.clear()
            hierarchy = self.config_data.get("YEARS_HIERARCHY", {})
            self.add_tree_items(self.year_tree_widget, hierarchy)
            self.year_tree_widget.expandAll()

    def get_config_key_for_abnormal_units(self, year_suffix_or_year_name):
        # This function now expects a year SUFFIX (e.g., "2", "3") or a year NAME from which it can derive the suffix.
        # It should not be called with "FIXED".
        year_suffix = year_suffix_or_year_name
        return f"ABNORMAL_SUBJECTS_UNITS{year_suffix}"

    def add_tree_items(self, parent_widget, data):
        for text, children_data in sorted(data.items()):
            item = QTreeWidgetItem(parent_widget, [text])
            item.setFlags(item.flags() | Qt.ItemIsEditable | Qt.ItemIsDragEnabled | Qt.ItemIsDropEnabled)
            self.add_tree_items(item, children_data)

    def on_year_item_double_clicked(self, item, column):
        self.current_editing_item_text = item.text(column)

    def on_year_item_changed(self, item, column):
        if self.current_editing_item_text is None or self.current_editing_item_text == item.text(column):
            return

        old_text = self.current_editing_item_text
        new_text = item.text(column)
        self.current_editing_item_text = None

        parent_item = item.parent()
        data_node = self.find_data_node(parent_item)

        if data_node is None:
            QMessageBox.critical(self, "エラー", "リネーム対象の親データが見つかりませんでした。" )
            item.setText(column, old_text) # Revert
            return

        if new_text in data_node:
            QMessageBox.warning(self, "重複", "同じ階層にその名前の項目が既に存在します。" )
            item.setText(column, old_text) # Revert text
            return

        data_node[new_text] = data_node.pop(old_text)

    def find_data_node(self, item):
        if item is None: # Top-level
            return self.config_data.get("YEARS_HIERARCHY", {})
        
        path = []
        temp_item = item
        while temp_item is not None:
            path.insert(0, temp_item.text(0))
            temp_item = temp_item.parent()

        node = self.config_data.get("YEARS_HIERARCHY", {})
        for key in path:
            if key in node:
                node = node[key]
            else:
                return None
        return node

    def add_year_item(self):
        selected_item = self.year_tree_widget.currentItem()
        if not selected_item:
            self.add_year_top_level_item()
        else:
            self.add_year_child_item()

    def add_year_top_level_item(self):
        text, ok = QInputDialog.getText(self, "親項目の追加", "新しい親項目名を入力してください:")
        if ok and text:
            hierarchy = self.config_data.get("YEARS_HIERARCHY", {})
            if text in hierarchy:
                QMessageBox.warning(self, "重複", "その名前の項目が既に存在します。" )
                return
            hierarchy[text] = {}
            self.populate_year_list()

    def add_year_child_item(self):
        selected_item = self.year_tree_widget.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "選択なし", "親となる項目を選択してください。" )
            return

        text, ok = QInputDialog.getText(self, "子項目の追加", "新しい子項目名を入力してください:")
        if ok and text:
            data_node = self.find_data_node(selected_item)
            if data_node is None:
                QMessageBox.critical(self, "エラー", "子項目を追加する親データが見つかりませんでした。" )
                return
            if text in data_node:
                QMessageBox.warning(self, "重複", "同じ階層にその名前の項目が既に存在します。" )
                return
            data_node[text] = {}
            self.populate_year_list()

    def remove_year_tree_item(self):
        selected_items = self.year_tree_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "選択なし", "削除する項目を1つ以上選択してください。" )
            return

        names_to_delete = [item.text(0) for item in selected_items]
        reply = QMessageBox.question(
            self, 
            "確認", 
            f"""{len(names_to_delete)}個の項目を削除しますか？\n({', '.join(names_to_delete)})

子項目もすべて削除されます。""", QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            for item in selected_items:
                parent_item = item.parent()
                data_node = self.find_data_node(parent_item)

                if data_node is None:
                    continue

                if item.text(0) in data_node:
                    del data_node[item.text(0)]
            
            self.populate_year_list()

    def apply_year_hierarchy_settings(self):
        self._rebuild_hierarchy_from_tree_and_update_config(self.config_data)
        self.update_years_in_config()
        self.reload_all_selectors()
        QMessageBox.information(self, "成功", "設定を適用しました。" )

    def _rebuild_hierarchy_from_tree_and_update_config(self, config_dict):
        new_hierarchy = {}
        
        # アンカーマーカーを除去するヘルパー関数
        def clean_text(text):
            anchor_marker = " ⚓"
            if text.endswith(anchor_marker):
                return text[:-len(anchor_marker)]
            if text.endswith("⚓"):
                return text[:-1].rstrip()
            return text
        
        def recurse(parent_item, parent_dict):
            for i in range(parent_item.childCount()):
                child_item = parent_item.child(i)
                child_text = clean_text(child_item.text(0))  # アンカーマーカーを除去
                parent_dict[child_text] = {}
                recurse(child_item, parent_dict[child_text])

        root = self.year_tree_widget.invisibleRootItem()
        recurse(root, new_hierarchy)
            
        config_dict['YEARS_HIERARCHY'] = new_hierarchy

    def update_years_in_config(self):
        all_paths = self._get_all_hierarchy_paths()
        self.years = sorted([ "_".join(p) for p in all_paths])
    
    def _get_all_hierarchy_paths(self, node=None, path_prefix=[]):
        if node is None:
            node = self.config_data.get("YEARS_HIERARCHY", {})
        
        if not isinstance(node, dict):
            return []

        all_paths = []
        for key, children in node.items():
            new_prefix = path_prefix + [key]
            all_paths.append(new_prefix)
            if children:
                all_paths.extend(self._get_all_hierarchy_paths(children, new_prefix))
        return all_paths

    def update_details_year_selectors(self, level):
        while len(self.details_year_cascading_combos) > level:
            combo = self.details_year_cascading_combos.pop()
            self.details_year_combo_layout.removeWidget(combo)
            combo.deleteLater()
        path = []
        for i in range(level):
             combo = self.details_year_cascading_combos[i]
             if combo.currentIndex() > 0:
                  path.append(combo.currentText())
             else:
                 self.on_final_year_selection_changed()
                 return
        node = self.config_data.get("YEARS_HIERARCHY", {})
        for key in path:
            node = node.get(key, {})
        children = node.keys() if node else []
        if children:
            new_combo = QComboBox()
            new_combo.addItem("(未選択)")
            new_combo.addItems(sorted(list(children)))
            new_combo.currentIndexChanged.connect(lambda state, l=level + 1: self.update_details_year_selectors(level=l))
            self.details_year_cascading_combos.append(new_combo)
            self.details_year_combo_layout.addWidget(new_combo)
        self.on_final_year_selection_changed()

    def on_final_year_selection_changed(self):
        year_key = None
        path = []
        for combo in self.details_year_cascading_combos:
            if combo.currentIndex() > 0:
                path.append(combo.currentText())
            else:
                break
        if path:
            year_key = "_".join(path)
        is_leaf = False
        if year_key:
            node = self.config_data.get("YEARS_HIERARCHY", {})
            for key in path:
                node = node.get(key, {})
            if not node:
                is_leaf = True
        self.set_details_panel_enabled(is_leaf)
        if is_leaf:
            self.leaf_selection_changed.emit()
        if self.current_details_context.get("year") != year_key:
            self.current_details_context["year"] = year_key
            if is_leaf:
                self.update_details_subject_lists()
                self.populate_subject_details()

    def export_to_excel(self):
        if not self.years:
            QMessageBox.warning(self, "エラー", "先に「学部学年選択ダイアログメッセージ設定」で学年を設定してください。")
            return
        dialog = ExcelExportDialog(self.years, self.config_data, self)
        dialog.exec()

    def open_excel_import_dialog(self):
        if not self.years:
            QMessageBox.warning(self, "エラー", "先に「学部学年選択ダイアログメッセージ設定」で学年を設定してください。")
            return
        
        dialog = ExcelImportDialog(self.years, self.config_data, self)
        
        if dialog.exec():

            if dialog.generated_data:

                reply = QMessageBox.question(self, "確認",

                    "Excelから読み込んだデータで現在の設定を上書きしますか？\n"

                    "（この操作は元に戻せません）",

                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

                

                if reply == QMessageBox.Yes:

                    # Update the internal config data

                    for key, value in dialog.generated_data.items():

                        self.config_data[key] = value

                    

                    # Refresh only the relevant UI parts (the subject details tab)

                    # This avoids resetting the entire UI and losing user context.

                    self.update_details_subject_lists()

                    self.populate_subject_details()



                    QMessageBox.information(self, "成功", "Excelからデータをインポートしました。")
        
        
        
def start_main_app():

    """Entry point for the main config editor application."""
    import tempfile

    app = QApplication(sys.argv)

    set_button_styles(app)



    # The main window, MainWindow, is instantiated with default values.

    window = MainWindow(splash=None)

    

    window.show()

    # IPC: Notify launcher that the main window is ready
    ipc_ready_file = Path(tempfile.gettempdir()) / "configeditor_ready.tmp"
    try:
        ipc_ready_file.touch()
        print(f"IPC: Created ready signal file at {ipc_ready_file}")
    except Exception as e:
        print(f"IPC: Failed to create ready signal file: {e}")

    sys.exit(app.exec())
        
        
        
if __name__ == '__main__':

    start_main_app()
        
        