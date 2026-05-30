"""
excel_io.py - Excel I/O操作の堅牢化

設定ファイルのExcelインポート/エクスポート機能を提供します。
エラーハンドリングと詳細なフィードバックを強化しています。
"""
from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from dataclasses import dataclass
from enum import Enum

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelError(Exception):
    """Excel操作の基本例外クラス"""
    pass


class ExcelReadError(ExcelError):
    """Excel読み込みエラー"""
    pass


class ExcelWriteError(ExcelError):
    """Excel書き込みエラー"""
    pass


class ExcelValidationError(ExcelError):
    """Excelデータ検証エラー"""
    pass


@dataclass
class ImportResult:
    """インポート結果を表すデータクラス"""
    success: bool
    data: Dict[str, Any]
    errors: List[str]
    warnings: List[str]
    stats: Dict[str, int]  # 読み込んだ行数等の統計


@dataclass
class ExportResult:
    """エクスポート結果を表すデータクラス"""
    success: bool
    file_path: str
    errors: List[str]
    warnings: List[str]
    stats: Dict[str, int]  # 書き込んだ行数等の統計


class ExcelImporter:
    """
    Excelファイルから設定データをインポートするクラス。
    
    エラーハンドリングと詳細なフィードバックを提供します。
    """
    
    def __init__(self, file_path: str | Path):
        if not OPENPYXL_AVAILABLE:
            raise ExcelError("openpyxlがインストールされていません")
        
        self.file_path = Path(file_path)
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.stats: Dict[str, int] = {}
    
    def import_subjects(self, sheet_name: str = "教科マスタ", column: int = 1) -> ImportResult:
        """
        教科マスタをインポート。
        
        Args:
            sheet_name: シート名
            column: 教科名が記載されている列（1-indexed）
            
        Returns:
            ImportResult
        """
        self.errors = []
        self.warnings = []
        self.stats = {"total_rows": 0, "imported": 0, "skipped": 0}
        
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
        except FileNotFoundError:
            return ImportResult(
                success=False,
                data={},
                errors=[f"ファイルが見つかりません: {self.file_path}"],
                warnings=[],
                stats=self.stats
            )
        except Exception as e:
            return ImportResult(
                success=False,
                data={},
                errors=[f"ファイルを開けません: {e}"],
                warnings=[],
                stats=self.stats
            )
        
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames[:5])
            self.errors.append(f"シート '{sheet_name}' が見つかりません。利用可能: {available}")
            return ImportResult(
                success=False,
                data={},
                errors=self.errors,
                warnings=self.warnings,
                stats=self.stats
            )
        
        sheet = workbook[sheet_name]
        subjects = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, min_col=column, max_col=column), start=2):
            self.stats["total_rows"] += 1
            cell = row[0]
            
            if cell.value is None or str(cell.value).strip() == "":
                self.stats["skipped"] += 1
                continue
            
            subject = str(cell.value).strip()
            
            if subject in subjects:
                self.warnings.append(f"行 {row_num}: 重複する教科 '{subject}' をスキップ")
                self.stats["skipped"] += 1
                continue
            
            subjects.append(subject)
            self.stats["imported"] += 1
        
        workbook.close()
        
        return ImportResult(
            success=len(self.errors) == 0,
            data={"MASTER_SUBJECTS": subjects},
            errors=self.errors,
            warnings=self.warnings,
            stats=self.stats
        )


class ExcelExporter:
    """
    設定データをExcelファイルにエクスポートするクラス。
    
    エラーハンドリングと詳細なフィードバックを提供します。
    """
    
    def __init__(self, file_path: str | Path):
        if not OPENPYXL_AVAILABLE:
            raise ExcelError("openpyxlがインストールされていません")
        
        self.file_path = Path(file_path)
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.stats: Dict[str, int] = {}
    
    def export_subjects(
        self, 
        subjects: List[str], 
        sheet_name: str = "教科マスタ",
        header: str = "教科名"
    ) -> ExportResult:
        """
        教科マスタをエクスポート。
        
        Args:
            subjects: 教科名のリスト
            sheet_name: シート名
            header: ヘッダー行のテキスト
            
        Returns:
            ExportResult
        """
        self.errors = []
        self.warnings = []
        self.stats = {"exported": 0}
        
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            
            # ヘッダー
            header_cell = sheet.cell(row=1, column=1, value=header)
            header_cell.font = Font(bold=True)
            
            # データ
            for i, subject in enumerate(subjects, start=2):
                sheet.cell(row=i, column=1, value=subject)
                self.stats["exported"] += 1
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 30
            
            workbook.save(self.file_path)
            workbook.close()
            
        except PermissionError:
            self.errors.append(f"ファイルに書き込めません（開いている可能性があります）: {self.file_path}")
            return ExportResult(
                success=False,
                file_path=str(self.file_path),
                errors=self.errors,
                warnings=self.warnings,
                stats=self.stats
            )
        except Exception as e:
            self.errors.append(f"エクスポートエラー: {e}")
            return ExportResult(
                success=False,
                file_path=str(self.file_path),
                errors=self.errors,
                warnings=self.warnings,
                stats=self.stats
            )
        
        return ExportResult(
            success=True,
            file_path=str(self.file_path),
            errors=self.errors,
            warnings=self.warnings,
            stats=self.stats
        )


def check_openpyxl_available() -> bool:
    """openpyxlが利用可能かチェック"""
    return OPENPYXL_AVAILABLE
