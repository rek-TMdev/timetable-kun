"""
FileIOManager - ファイル入出力を管理するクラス

Application クラスからファイルI/O関連の責務を分離するために作成。
Excel/専用形式ファイルの読み書きロジックを担当します。

Note:
    ダイアログ表示などのUI操作はApplicationに残し、
    このクラスは純粋なデータI/Oのみを行います。
"""
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl


class FileIOManager:
    """
    ファイル入出力を管理するクラス。
    
    責務:
    - 専用形式ファイル（.tm.json）の読み書き
    - Excelファイルからの時間割データ読み込み
    - Excelファイルへの時間割データ書き込み
    
    Note:
        このクラスはUIに依存しない純粋なI/Oロジックを提供します。
        ファイル選択ダイアログや確認メッセージはApplicationクラスが担当します。
    """
    
    def __init__(self, base_path: Path):
        """
        FileIOManagerを初期化します。
        
        Args:
            base_path: アプリケーションのベースパス
        """
        self.base_path = base_path
    
    # ========================================
    # 専用形式ファイル（.tm.json）
    # ========================================
    
    def save_dedicated_file(
        self, 
        file_path: str, 
        year_data: dict, 
        config: dict,
        art_subject_selections: dict,
        active_profile_name: str
    ) -> None:
        """
        専用形式ファイルに時間割データを保存します。
        
        Args:
            file_path: 保存先のファイルパス
            year_data: 学年ごとの時間割データ
            config: 設定辞書
            art_subject_selections: 芸術科目選択データ
            active_profile_name: 現在のプロファイル名
            
        Raises:
            IOError: ファイル書き込みに失敗した場合
        """
        # シリアライズ可能なデータのみを抽出
        serializable_years = {}
        keys_to_keep = [
            'saved_profiles', 'all_slots', 'profile_lock_settings', 
            'profile_all_slots', 'selected_slots_filter'
        ]
        
        for year, data in year_data.items():
            serializable_years[year] = {}
            for k in keys_to_keep:
                if k in data:
                    val = data[k]
                    serializable_years[year][k] = self._make_serializable(val)
        
        payload = {
            "config": {
                "LAST_UPDATE_DAY": config.get("LAST_UPDATE_DAY"),
                "SCHOOL_NAME": config.get("GENERAL_SETTINGS", {}).get("SCHOOL_NAME")
            },
            "year_data": serializable_years,
            "art_subject_selections": art_subject_selections,
            "active_profile": active_profile_name,
            "saved_at": datetime.now().isoformat()
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
    
    def load_dedicated_file(self, file_path: str) -> dict:
        """
        専用形式ファイルから時間割データを読み込みます。
        
        Args:
            file_path: 読み込むファイルのパス
            
        Returns:
            読み込んだデータのペイロード辞書
            
        Raises:
            FileNotFoundError: ファイルが見つからない場合
            json.JSONDecodeError: JSONの解析に失敗した場合
            ValueError: ファイル形式が不正な場合
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        
        # 専用形式ファイルの検証
        validation_errors = self.validate_dedicated_file_structure(payload)
        if validation_errors:
            raise ValueError(f"ファイル形式が不正です:\n" + "\n".join(validation_errors))
        
        return payload
    
    def validate_dedicated_file_structure(self, payload: dict) -> list[str]:
        """
        専用形式ファイル(.tm.json)の構造を検証します。
        
        Args:
            payload: 読み込んだJSONデータ
            
        Returns:
            エラーメッセージのリスト（空の場合は有効）
        """
        errors = []
        
        if not isinstance(payload, dict):
            errors.append("ファイルのルート要素が辞書ではありません")
            return errors
        
        # 必須キーのチェック
        required_keys = ["config", "year_data"]
        for key in required_keys:
            if key not in payload:
                errors.append(f"必須キー '{key}' が見つかりません")
        
        # configの構造チェック
        if "config" in payload:
            config = payload["config"]
            if not isinstance(config, dict):
                errors.append("'config' は辞書である必要があります")
        
        # year_dataの構造チェック
        if "year_data" in payload:
            year_data = payload["year_data"]
            if not isinstance(year_data, dict):
                errors.append("'year_data' は辞書である必要があります")
            else:
                for year_key, year_val in year_data.items():
                    if not isinstance(year_val, dict):
                        errors.append(f"year_data['{year_key}'] は辞書である必要があります")
        
        return errors
    
    def validate_dedicated_file_school(
        self, 
        payload: dict, 
        current_school_name: Optional[str]
    ) -> tuple[bool, str, str]:
        """
        専用ファイルの学校名が現在の設定と一致するか検証します。
        
        Args:
            payload: 読み込んだファイルのペイロード
            current_school_name: 現在の学校名
            
        Returns:
            (一致するか, ペイロードの学校名, 現在の学校名) のタプル
        """
        payload_config = payload.get("config", {})
        payload_school = payload_config.get("GENERAL_SETTINGS", {}).get("SCHOOL_NAME")
        
        # Noneと空文字を同等として扱う
        normalized_payload = payload_school if payload_school and payload_school.strip() else None
        normalized_current = current_school_name if current_school_name and current_school_name.strip() else None
        
        return (
            normalized_payload == normalized_current,
            normalized_payload or "（未設定）",
            normalized_current or "（未設定）"
        )
    
    # ========================================
    # Excelファイル
    # ========================================
    
    def load_excel_workbook(self, file_path: str) -> openpyxl.Workbook:
        """
        Excelワークブックを読み込みます。
        
        Args:
            file_path: Excelファイルのパス
            
        Returns:
            openpyxl.Workbook オブジェクト
            
        Raises:
            Exception: ファイル読み込みに失敗した場合
        """
        return openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    
    def read_timetable_from_excel(
        self, 
        worksheet, 
        save_position: dict
    ) -> dict:
        """
        Excelワークシートから時間割データを読み取ります。
        
        Args:
            worksheet: openpyxlのワークシートオブジェクト
            save_position: スロット名とセル位置のマッピング
            
        Returns:
            {スロット名: セル値} の辞書
        """
        # セル座標とスロット名のマッピングを作成
        cell_to_slot_map = {}
        for slot_name, v_pos in save_position.items():
            try:
                pos_to_use = v_pos
                if isinstance(pos_to_use, list) and len(pos_to_use) == 1 and isinstance(pos_to_use[0], list):
                    pos_to_use = pos_to_use[0]
                if isinstance(pos_to_use, list) and len(pos_to_use) > 0 and isinstance(pos_to_use[0], str):
                    coords = self._a1_to_coords(pos_to_use[0])
                    if coords is None:
                        raise TypeError()
                    pos_to_use = coords
                if not (isinstance(pos_to_use, list) and len(pos_to_use) == 2):
                    raise TypeError()
                pos_tuple = tuple(pos_to_use)
                cell_to_slot_map[pos_tuple] = slot_name
            except (TypeError, ValueError):
                continue
        
        # セルを読み取り、結果を返す
        result = {}
        for pos, slot_name in cell_to_slot_map.items():
            try:
                cell_value = worksheet.cell(row=pos[0], column=pos[1]).value
                if cell_value is not None and str(cell_value).strip():
                    result[slot_name] = str(cell_value).strip()
            except IndexError:
                continue
        
        return result
    
    def _a1_to_coords(self, a1_string: str) -> Optional[list]:
        """
        A1表記のセル座標を [行, 列] の数値リストに変換します。
        
        Args:
            a1_string: A1表記のセル座標（例: "A1", "B12"）
            
        Returns:
            [行番号, 列番号] のリスト、または無効な場合はNone
        """
        if not isinstance(a1_string, str):
            return None
        
        col_str = ""
        row_str = ""
        for char in a1_string:
            if char.isalpha():
                col_str += char
            elif char.isdigit():
                row_str += char
        
        if not col_str or not row_str:
            return None

        col_idx = 0
        for char in col_str.upper():
            col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
        
        try:
            row_idx = int(row_str)
            return [row_idx, col_idx]
        except ValueError:
            return None
    
    def write_timetable_to_excel(
        self, 
        worksheet, 
        timetable: dict, 
        table_layout: list, 
        fixed_slots: dict
    ) -> None:
        """
        時間割データをExcelワークシートに書き込みます。
        
        Args:
            worksheet: openpyxlのワークシートオブジェクト
            timetable: {スロット名: 教科名} の辞書
            table_layout: 時間割のレイアウト（2次元リスト）
            fixed_slots: 固定スロット（ヘッダーなど）
        """
        for row_idx, row in enumerate(table_layout, start=1):
            for col_idx, slot in enumerate(row, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if slot in timetable:
                    cell.value = timetable[slot]
                elif slot in fixed_slots:
                    cell.value = fixed_slots.get(slot, "")
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )
                else:
                    cell.value = ""
    
    # ========================================
    # ヘルパーメソッド
    # ========================================
    
    def _make_serializable(self, val: Any) -> Any:
        """
        値をJSONシリアライズ可能な形式に変換します。
        
        Args:
            val: 変換する値
            
        Returns:
            シリアライズ可能な値
        """
        if isinstance(val, set):
            return list(val)
        
        if isinstance(val, dict):
            result = {}
            for k, v in val.items():
                result[k] = self._make_serializable(v)
            return result
        
        if isinstance(val, list):
            return [self._make_serializable(item) for item in val]
        
        # JSONシリアライズ可能かテスト
        try:
            json.dumps(val, ensure_ascii=False)
            return val
        except (TypeError, ValueError):
            return str(val)
