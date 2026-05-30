from cProfile import label
import gc
import os
import sys
import math
import json
import re
import darkdetect
import unicodedata
from collections import Counter
from datetime import datetime
from PySide6.QtWidgets import(
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel,
    QPushButton, QCheckBox, QComboBox, QScrollArea, QListWidget, QStackedWidget, QToolTip, QRadioButton,
    QMessageBox, QSlider, QGroupBox, QFileDialog, QDialog, QLineEdit, QSpinBox, QTabWidget, QButtonGroup, QMenu, QListWidgetItem
)
from PySide6.QtCore import Qt, QObject, QEvent, QTimer, Signal, QSortFilterProxyModel
from PySide6.QtGui import QFont, QIcon, QPalette, QColor, QPixmap, QPainter


class ExcludeFileProxyModel(QSortFilterProxyModel):
    """特定のファイル名をファイルダイアログから除外するプロキシモデル"""
    def __init__(self, excluded_filenames, parent=None):
        super().__init__(parent)
        self.excluded_filenames = [f.lower() for f in excluded_filenames]
    
    def filterAcceptsRow(self, source_row, source_parent):
        index = self.sourceModel().index(source_row, 0, source_parent)
        filename = self.sourceModel().data(index)
        if filename and filename.lower() in self.excluded_filenames:
            return False
        return True
from shiboken6 import isValid as widget_is_valid
import openpyxl
from pathlib import Path
from tutorial_manager import TutorialManager, TutorialStep
from config_manager import ConfigManager
from theme_manager import ThemeManager
from file_io_manager import FileIOManager
from profile_manager import ProfileManager
from timetable_worker import TimetableWorker
from timetable_logic import TimetableLogic

# ToolTip 表示用のクラスを追加
class CreateToolTip:
    """ツールチップを作成するヘルパークラス"""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.widget.setToolTip(self.text)
        self.widget.installEventFilter(ToolTipFilter())
        
    def update_text(self, new_text):
        """ツールチップテキストを更新"""
        self.text = new_text
        self.widget.setToolTip(self.text)

def is_windows_light_theme():
    """Windowsでライトテーマが使われているか確認する。"""
    if sys.platform != "win32":
        return True # Windows以外ではライトテーマを既定にする
    try:
        import winreg
        key_path = r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path)
        value, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
        return value == 1
    except Exception:
        # エラー時はライトテーマに戻す。
        return True

# ToolTipFilter クラスはそのまま保持
class ToolTipFilter(QObject):
    def eventFilter(self, obj, event):
        if event.type() == QEvent.ToolTip:
            if hasattr(obj, 'toolTipText'):
                QToolTip.showText(event.globalPos(), obj.toolTipText(), obj)
                return True
        return super().eventFilter(obj, event)


class CustomStepSpinBox(QSpinBox):
    """矢印クリックとスクロールで異なるステップ値を持つスピンボックス"""
    def __init__(self, arrow_step=100, wheel_step=10, parent=None):
        super().__init__(parent)
        self._arrow_step = arrow_step
        self._wheel_step = wheel_step
        self.setSingleStep(arrow_step)
    
    def wheelEvent(self, event):
        # スクロール時は wheel_step を使用
        original_step = self.singleStep()
        self.setSingleStep(self._wheel_step)
        super().wheelEvent(event)
        self.setSingleStep(original_step)



def set_button_styles(app, is_dark=None):
    """テーマに応じたQPushButton用スタイルを適用する。"""
    if is_dark is None:
        try:
            is_dark = darkdetect.theme() == "Dark"
        except Exception:
            is_dark = False

    light_style = """
        QPushButton, QComboBox {
            background-color: #3478f6FF;
            color: black;
            border: 1px solid #747474;
            padding: 4px;
            border-radius: 4px;
        }
        QPushButton:hover, QComboBox:hover {
            background-color: #5F92F1;
            border-color: #285ec6;
        }
        QPushButton:pressed {
            background-color: #1a479b;
            border-color: #1a479b;
        }
        QPushButton:checked {
            background-color: #007AFF;
        }
        QPushButton:disabled, QComboBox:disabled {
            background-color: #E4E4E4FF;
            color: #999999;
            border-color: #999999;
        }
        QComboBox QAbstractItemView {
            background-color: white;
            color: black;
            border: 1px solid #3478f6;
            selection-background-color: #285ec6;
            selection-color: white;
        }
    """

    dark_style = """
        QPushButton, QComboBox {
            background-color: #0A4C8DFF;
            color: white;
            border: 1px solid #EBEBEB;
            padding: 4px;
            border-radius: 4px;
        }
        QPushButton:hover, QComboBox:hover {
            background-color: #0166C5;
            border-color: #0077E8;
        }
        QPushButton:pressed {
            background-color: #0062C2;
            border-color: #0062C2;
        }
        QPushButton:checked {
            background-color: #0A84FF;
        }
        QPushButton:disabled, QComboBox:disabled {
            background-color: #1e1e1e;
            color: #999999;
            border-color: #999999;
        }
        QComboBox QAbstractItemView {
            background-color: #2D2D2D;
            color: white;
            border: 1px solid #0A84FF;
            selection-background-color: #0077E8;
            selection-color: white;
        }
    """

    app.setStyleSheet(dark_style if is_dark else light_style)

class DepartmentSelectionDialog(QDialog):
    def __init__(self, parent=None, hierarchy_keys=None, message="学科を選択してください"):
        super().__init__(parent)
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.setWindowTitle(message)
        self.setModal(True)

        self.selected_department = None

        main_layout = QVBoxLayout(self)

        label = QLabel("選択してください:")
        main_layout.addWidget(label)

        # ラジオボタンをグループ化する
        self.radio_group_box = QGroupBox()
        radio_layout = QVBoxLayout()
        self.radio_button_group = QButtonGroup(self)

        if hierarchy_keys:
            sorted_keys = sorted(hierarchy_keys)
            for i, key in enumerate(sorted_keys):
                radio_button = QRadioButton(key)
                radio_layout.addWidget(radio_button)
                self.radio_button_group.addButton(radio_button, i)
            
            if self.radio_button_group.buttons():
                self.radio_button_group.buttons()[0].setChecked(True)

        self.radio_group_box.setLayout(radio_layout)
        main_layout.addWidget(self.radio_group_box)

        button_box = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("キャンセル")
        cancel_button.clicked.connect(self.reject)

        button_box.addWidget(ok_button)
        button_box.addWidget(cancel_button)
        main_layout.addLayout(button_box)

    def accept(self):
        checked_button = self.radio_button_group.checkedButton()
        if checked_button:
            self.selected_department = checked_button.text()
        super().accept()

    def get_selected_department(self):
        return self.selected_department

class SlotSelectionDialog(QDialog):
    def __init__(self, parent=None, year_label=None, preselected_slots=None):
        super().__init__(parent)
        self.setWindowTitle(f"{year_label}の枠を選択")
        self.setModal(True)
        self.selected_slots = set()

        main_layout = QVBoxLayout(self)

        label = QLabel("検索したい枠を選択してください:")
        main_layout.addWidget(label)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        
        # 親 (Application) の _create_slot_selection_grid メソッドを呼び出す
        grid_widget, self.slot_buttons = self.parent()._create_slot_selection_grid(year_label)
        scroll_area.setWidget(grid_widget)
        main_layout.addWidget(scroll_area)

        # 事前選択されたスロットがあればチェック状態にする
        if preselected_slots:
            for btn in self.slot_buttons:
                if btn.text() in preselected_slots:
                    btn.setChecked(True)
                    self.selected_slots.add(btn.text())
        
        # ボタンのトグルシグナルを接続
        for btn in self.slot_buttons:
            btn.toggled.connect(lambda checked, b=btn: self._on_slot_toggled(checked, b))

        button_box = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        clear_button = QPushButton("クリア")
        clear_button.clicked.connect(self._clear_selection)
        cancel_button = QPushButton("キャンセル")
        cancel_button.clicked.connect(self.reject)

        button_box.addWidget(ok_button)
        button_box.addWidget(clear_button)
        button_box.addWidget(cancel_button)
        main_layout.addLayout(button_box)
    
    def _on_slot_toggled(self, checked, button):
        if checked:
            self.selected_slots.add(button.text())
        else:
            self.selected_slots.discard(button.text())

    def _clear_selection(self):
        for btn in self.slot_buttons:
            btn.setChecked(False)
        self.selected_slots.clear()

    def get_selected_slots(self):
        return list(self.selected_slots)

class SubjectSelectionDialog(QDialog):
    def __init__(self, parent=None, year_label=None, available_subject_groups=None, clicked_slot=None):
        super().__init__(parent)
        self.ini_theme = darkdetect.theme()

        self.setWindowTitle(f"{year_label} - '{clicked_slot}' に教科グループを追加")
        self.setModal(True)
        self.selected_subject_name = None
        self.selected_slot_group = None

        main_layout = QVBoxLayout(self)

        label = QLabel(f"'{clicked_slot}' を含む教科グループを選択してください:")
        main_layout.addWidget(label)

        self.subject_list_widget = QListWidget()
        self.subject_list_widget.setSelectionMode(QListWidget.SingleSelection)
        
        # available_subject_groups は {subject_name: [group1_slots, group2_slots, ...]} の形式を想定
        self.subject_group_map = {} # 表示文字列: (教科名, 時間枠グループ)

        if available_subject_groups:
            # clicked_slot が属するグループを優先的に表示するため、フィルタリング
            filtered_subjects = []
            for subject_name, group_list in available_subject_groups.items():
                for group_slots in group_list:
                    if clicked_slot in group_slots:
                        filtered_subjects.append((subject_name, group_slots))
            
            # フィルタリングされたものがなければ、全て表示
            if not filtered_subjects:
                for subject_name, group_list in available_subject_groups.items():
                    for group_slots in group_list:
                        filtered_subjects.append((subject_name, group_slots))

            for subject_name, group_slots in sorted(filtered_subjects, key=lambda x: x[0]):
                group_text = ", ".join(group_slots)
                displayed_text = f"{subject_name} ({group_text})"
                self.subject_list_widget.addItem(displayed_text)
                self.subject_group_map[displayed_text] = (subject_name, group_slots)

        main_layout.addWidget(self.subject_list_widget)

        button_box = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("キャンセル")
        cancel_button.clicked.connect(self.reject)

        button_box.addWidget(ok_button)
        button_box.addWidget(cancel_button)
        main_layout.addLayout(button_box)

    def accept(self):
        selected_item = self.subject_list_widget.currentItem()
        if selected_item:
            displayed_text = selected_item.text()
            self.selected_subject_name, self.selected_slot_group = self.subject_group_map.get(displayed_text, (None, None))
        super().accept()

    def get_selected_subject_group(self):
        return self.selected_subject_name, self.selected_slot_group

class TimetableSlotButton(QPushButton):
    """左クリックと右クリックを区別して処理するカスタムボタン"""
    def __init__(self, text, slot, year_label, app_instance, is_main_timetable_tab=False):
        super().__init__(text)
        self.slot = slot
        self.year_label = year_label
        self.app = app_instance
        self.is_main_timetable_tab = is_main_timetable_tab

    def mousePressEvent(self, event):
        if self.is_main_timetable_tab:
            # 従来の動作を維持（状態保存・復元は呼ばない）
            self.app._on_timetable_slot_clicked(self.slot, self.year_label, self.text(), event.button())
        else:
            # それ以外の表示では、従来の動作を維持
            if event.button() == Qt.LeftButton:
                self.app.toggle_slot(self.slot, self.year_label)
            elif event.button() == Qt.RightButton:
                self.app.toggle_user_fixation(self.text(), self.year_label)
        
        # clicked()を発火させるため基底クラスのイベントも呼ぶ
        super().mousePressEvent(event)

class SaveCompleteDialog(QDialog):
    """保存完了を通知するノンモーダルダイアlog"""
    def __init__(self, parent=None, message=""):
        super().__init__(parent)
        self.setWindowTitle("保存完了")
        
        # アイコンを設定
        icon_base_name = "時間割くんアイコン.ico"
        icon_path_obj = parent.base_path / icon_base_name
        if icon_path_obj.exists():
            self.setWindowIcon(QIcon(str(icon_path_obj)))

        self.setModal(False)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(message))
        
        # 3秒後に自動で閉じる
        QTimer.singleShot(2000, self.accept)

class ConfigSelectionDialog(QDialog):
    """設定ファイルを選択するための汎用ダイアログ"""
    def __init__(self, parent, title, message, items):
        super().__init__(parent)
        # 【修正】常に最前面かつモーダルに設定
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.setModal(True)
        self.setWindowTitle(title)
        self.selected_item = None

        icon_base_name = "時間割くんアイコン.ico"
        icon_path_obj = parent.base_path / icon_base_name
        icon_path = str(icon_path_obj) # QIcon用に文字列へ変換

        if icon_path_obj.exists(): # pathlibのexists()で存在確認
            icon = QIcon(icon_path) # QIconには文字列パスを渡す
            if not icon.isNull():
                self.setWindowIcon(icon)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(message))

        self.list_widget = QListWidget()
        self.list_widget.addItems(items)
        if items:
            self.list_widget.setCurrentRow(0)
        layout.addWidget(self.list_widget)

        button_box = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("キャンセル")
        cancel_button.clicked.connect(self.reject)
        button_box.addWidget(ok_button)
        button_box.addWidget(cancel_button)
        layout.addLayout(button_box)

    def accept(self):
        if self.list_widget.currentItem():
            self.selected_item = self.list_widget.currentItem().text()
        super().accept()

    def get_selected_item(self):
        return self.selected_item

class Application(QMainWindow):
    tutorial_signal = Signal(dict)
    def __init__(self, splash=None): # splash引数を追加
        super().__init__()
        # 1. まず最初に「終了フラグ」と「スプラッシュの保持」を行う
        self.splash = splash
        self._exit_requested = False
        self.is_tutorial_running = False # チュートリアル実行中フラグを最初に初期化
        self.selected_department_path_key = None
        self.current_profile_name = "Default"

        # 2. パスの確定
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            self.base_path = Path(sys._MEIPASS)
        else:
            try:
                decoded_file_path = Path(__file__).resolve()
                self.base_path = decoded_file_path.parent
            except Exception:
                self.base_path = Path('.')
        
        self.icon_base_name = "時間割くんアイコン.ico"

        # 2.5. ConfigManager の初期化（設定管理を委譲）
        self.config_manager = ConfigManager(self.base_path)

        # 3. ユーザー設定（AppData等のconfig.json）の読み込み
        # ConfigManagerに委譲し、self.user_settings でもアクセス可能に（後方互換性）
        self.user_settings = self.config_manager.load_user_settings()
        self._last_profile_order = self.user_settings.get("profile_order", [])

        # 4. 学校設定ファイル（選んだ学校のjson）のパスを選択
        chosen_config_path = self._select_config_file()

        # キャンセルされた場合は即終了
        if not chosen_config_path or self._exit_requested:
            self._exit_requested = True
            return

        # 5. 選択されたファイルをConfigManager経由で読み込んで「self.config」を作る
        # ★ここより前で _get_setting を呼ぶとエラーになります
        try:
            self.config = self.config_manager.load_config(Path(chosen_config_path))
            self.config_path = chosen_config_path
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"設定ファイルの読み込みに失敗しました:\n{e}")
            self._exit_requested = True
            return

        # 6. マスター設定（共有設定）があるかチェック
        # ここでは self.config がもう存在するので _get_setting が使えます
        master_config_path_str = self._get_setting("MASTER_CONFIG_PATH")
        if master_config_path_str:
            # (以下、既存のマスターチェック処理があればここに記述)
            pass

        # 7. UIの初期設定
        school_name = self._get_setting("SCHOOL_NAME")
        if school_name:
            self.setWindowTitle(f"時間割くん by I.R - [{school_name}]")
        else:
            self.setWindowTitle("時間割くん by I.R")

        self.resize(1000, 800)

        # user_settingsを実行時configの既定値に反映
        try:
            if isinstance(self.user_settings, dict):
                # トップレベルの単純キーはconfigにも写して扱いやすくする
                if 'TIMETABLE_ORDER' in self.user_settings:
                    self.config['TIMETABLE_ORDER'] = self.user_settings['TIMETABLE_ORDER']
                if 'RUN_TUTORIAL_ON_STARTUP' in self.user_settings:
                    self.config['RUN_TUTORIAL_ON_STARTUP'] = self.user_settings['RUN_TUTORIAL_ON_STARTUP']
                if 'APP_FONT_SIZE' in self.user_settings:
                    self.config['APP_FONT_SIZE'] = self.user_settings['APP_FONT_SIZE']
                if 'APP_THEME' in self.user_settings:
                    self.config['APP_THEME'] = self.user_settings['APP_THEME']
                # 年次別単位数は設定ファイルの値を優先し、user_settingsへ反映する
                for k, v in list(self.config.items()):
                    if k.startswith('YEARS_SUBJECTS_UNITS_'):
                        self.user_settings[k] = v  # 設定ファイルの値でuser_settingsを更新
        except Exception as e:
            print(f"Failed applying user settings overrides: {e}")

        # アイコンを読み込む
        icon_path = os.path.join(self.base_path, self.icon_base_name)
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(str(icon_path)))
        else:
            # テーマ別アイコンがない場合は通常アイコンを使用する
            fallback_icon_path = os.path.join(self.base_path, self.icon_base_name)
            if os.path.exists(fallback_icon_path):
                self.setWindowIcon(QIcon(str(fallback_icon_path)))
            else:
                print(f"アイコンファイルが見つかりません: {icon_path} または {fallback_icon_path}")

        # 後方互換性のため_get_setting経由でウィンドウタイトルを設定
        school_name = self._get_setting("SCHOOL_NAME")
        if school_name:
            self.setWindowTitle(f"時間割くん by I.R - [{school_name}]")

        # --- ThemeManager の初期化（テーマ・スタイリング管理を委譲）---
        self.theme_manager = ThemeManager(self.base_path, self.config)
        # 後方互換性のため、self.theme も維持（ThemeManagerから取得）
        self.theme = self.theme_manager.get_all_theme_colors()

        # OSのテーマ設定を反映 (初期化)
        # ThemeManager側で優先順（ユーザー設定 > OS設定）を処理する
        is_dark = self.theme_manager.is_dark_theme()
        set_button_styles(QApplication.instance(), is_dark=is_dark)

        # --- FileIOManager の初期化（ファイル入出力を委譲）---
        self.file_io_manager = FileIOManager(self.base_path)

        # フォントサイズをconfigから読み込んで適用
        font_size = self._get_setting("APP_FONT_SIZE", 12)
        self.change_font_size(font_size)

        # メインウィジェットとレイアウト
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)

        content_layout = QHBoxLayout()


        # 左側のパネル
        left_panel = QWidget()
        left_panel_layout = QVBoxLayout(left_panel)
        left_panel.setFixedWidth(120)

        # 左側のタブリスト
        self.tab_list = QListWidget()
        left_panel_layout.addWidget(self.tab_list)

        # テーマに応じた単色アイコン色を決定する。
        target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"

        # チュートリアル開始ボタン
        icon_path_tutorial = os.path.join(self.base_path, "svgs", "question.svg")
        tutorial_icon = self._create_icon_from_svg_data(icon_path_tutorial, target_color)
        self.tutorial_button = QPushButton(tutorial_icon, "チュートリアル") # コンストラクタでアイコンを直接設定
        # フォントサイズを調整
        font = self.tutorial_button.font()
        font.setPointSize(font.pointSize() - 2) # フォントサイズを2ポイント下げる
        self.tutorial_button.setFont(font)

        self.tutorial_button.clicked.connect(self.on_tutorial_button_clicked)
        left_panel_layout.addWidget(self.tutorial_button)

        # 読み込みボタン
        icon_path_load = os.path.join(self.base_path, "svgs", "load_file.svg")
        load_excel_icon = self._create_icon_from_svg_data(icon_path_load, target_color)
        self.load_button = QPushButton(load_excel_icon, "ファイルを開く")
        load_menu = QMenu(self.load_button)
        load_excel_button = load_menu.addAction("Excelを開く")
        load_excel_button.triggered.connect(self.load_timedate)
        load_dedicated_button = load_menu.addAction("専用ファイルを開く")
        load_dedicated_button.triggered.connect(self.load_dedicated_file)
        self.load_button.setMenu(load_menu)
        # フォントサイズを調整
        font = self.load_button.font()
        font.setPointSize(font.pointSize() - 2)
        self.load_button.setFont(font)
        left_panel_layout.addWidget(self.load_button)

        # 学科変更ボタン
        icon_path_change_dept = os.path.join(self.base_path, "svgs", "logout.svg")
        change_dept_icon = self._create_icon_from_svg_data(icon_path_change_dept, target_color)
        self.change_button = QPushButton(change_dept_icon, "変更")
        change_memu = QMenu(self.change_button)
        change_department = change_memu.addAction("学科を変更")
        change_department.triggered.connect(lambda: self._change_department_clicked())
        change_school_action = change_memu.addAction("設定ファイルを変更")
        change_school_action.triggered.connect(lambda: self._change_school())
        # フォントサイズを調整
        font = self.change_button.font()
        font.setPointSize(font.pointSize() - 2)
        self.change_button.setFont(font)
        self.change_button.setMenu(change_memu)
        left_panel_layout.addWidget(self.change_button)

        content_layout.addWidget(left_panel)

        # 右側のコンテンツエリア
        self.stacked_widget = QStackedWidget()
        content_layout.addWidget(self.stacked_widget)

        self.main_layout.addLayout(content_layout)

        # --- 動的なタブ作成 ---
        self.selected_department_path_key = self.user_settings.get("years_hierarchy")
        if not isinstance(self.selected_department_path_key, str):
            self.selected_department_path_key = None # 保存値が不正な場合に備える
        
        # Noneは未保存扱い、空文字は設定ファイル内の階層を使う指定として扱う
        if self.selected_department_path_key is not None:
            selected_hierarchy_for_years = self._get_hierarchy_dict_from_path_key(self.selected_department_path_key)
        else:
            # 保存済み階層がない場合は選択を行う
            path_key, hierarchy_dict = self._perform_hierarchical_selection()

            if path_key is None: # ルート選択でキャンセル
                QMessageBox.information(self, "終了", "学科が選択されなかったため、アプリケーションを終了します。")
                sys.exit()
            
            if path_key == "SUB_CANCEL":
                selected_hierarchy_for_years = self.config.get("YEARS_HIERARCHY", {})
                self.selected_department_path_key = "" # 個別パス指定なし
            else:  # ユーザーが選択を完了
                self.selected_department_path_key = path_key
                selected_hierarchy_for_years = hierarchy_dict

            self._save_user_settings()

        self.years = self._get_all_leaf_years(selected_hierarchy_for_years)
        self.combination_time_tab = QWidget()
        self.tab_settings = QWidget()
        self.year_tabs = {}

        self.stacked_widget.addWidget(self.combination_time_tab)
        self._add_tab_item("時間割", "timetable")

        for year in self.years:
            tab = QWidget()
            self.year_tabs[year] = tab
            self.stacked_widget.addWidget(tab)
            display_year = year.split('_')[-1]
            self._add_tab_item(display_year, "add")

        self.stacked_widget.addWidget(self.tab_settings)
        self._add_tab_item("設定", "setting")

        # ツールチップを設定
        self.tab_list.item(0).setToolTip("作成した時間割を管理・保存します。")
        for i, year in enumerate(self.years):
            year_label = year.split('_')[-1]
            self.tab_list.item(i + 1).setToolTip(f"{year_label}の時間割を作成します。")
        self.tab_list.item(self.tab_list.count() - 1).setToolTip("アプリケーションの動作設定を行います。")

        self.tab_list.currentRowChanged.connect(self.stacked_widget.setCurrentIndex)
        self.stacked_widget.currentChanged.connect(self.on_tab_changed)
        self.tab_list.setCurrentRow(0)
        
        # 設定値の初期化
        self._initialize_config_values()

        # --- データ構造の初期化 ---
        self.year_data = {}
        for year in self.years:
            # 旧dict形式を新しいlist形式へ変換する補助関数
            def _convert_to_list_format(data_dict):
                if isinstance(data_dict, dict):
                    return [{"name": s, "data": d} for s, d in data_dict.items()]
                return data_dict # すでにlist形式とみなす

            self.year_data[year] = {
                "table_layout": self.config.get(f"table_layout{year}", []),
                "subject_number": _convert_to_list_format(self.config.get(f"subject_number{year}", {})),
                "save_position": self.config.get(f"SAVE_POSITION{year}", {}),
                "subject_slots_base": _convert_to_list_format(self.config.get(f"subject_slots_base{year}", {})),
                "fixed_slots": self.config.get(f"FIXED_SLOTS{year}", {}),
                "all_slots": [slot for row in self.config.get(f"table_layout{year}", []) for slot in row if slot], # all_slotsをtable_layoutから初期化する
                "profile_all_slots": {"Default": [slot for row in self.config.get(f"table_layout{year}", []) for slot in row if slot]},
                "abnormal_units": {self._normalize_subject_name(k): v for k, v in self.config.get(f"ABNORMAL_SUBJECTS_UNITS{year}", {}).items()},
                "required_subjects": self.config.get(f"REQUIRED_SUBJECTS_{year}", {}),
                "saved_profiles": {"Default": {}},
                "profile_ui_states": {"Default": {"checked": set(), "important": set(), "prefixes": {}}},
                "active_profile_name": "Default",
                "complete_combination": {},
                "subject_frames": [],
                "subject_scroll_layout": None,
                "check_vars": {},
                "prefix_vars": {},
                "important_vars": {},
                "duplicate_highlight_status": {"subjects": set(), "widgets": {}},
                "profile_lock_settings": {"Default": {}}
            }

        # 基礎単位数をインスタンス変数として保存
        department_path = self.selected_department_path_key or ""
        config_key = f"YEARS_SUBJECTS_UNITS_{department_path}"
        self.base_units_value = self._get_setting(config_key, None)
        
        # フォールバック: 設定が見つからない場合、YEARS_HIERARCHYのトップレベルキーを使用
        if self.base_units_value is None:
            years_hierarchy = self.config.get("YEARS_HIERARCHY", {})
            if years_hierarchy:
                # トップレベルの最初のキーを使用
                top_level_key = list(years_hierarchy.keys())[0] if years_hierarchy else ""
                fallback_key = f"YEARS_SUBJECTS_UNITS_{top_level_key}"
                self.base_units_value = self._get_setting(fallback_key, 0)
            else:
                self.base_units_value = 0

        # 全学年共通の設定
        self.REQUIRED_SUBJECTS_ALL = self.config.get("REQUIRED_SUBJECTS_ALL", {})
        self.PREREQUISITE_SUBJECTS = self.config.get("PREREQUISITE_SUBJECTS", {})
        self.NO_TOGETHER_SUBJECTS = self.config.get("NO_TOGETHER_SUBJECTS", [])
        self.selected_art_subject = self.config.get("SELECTED_ART_SUBJECT", [])
        self.art_subject_count = self._get_setting("ART_SUBJECT", 1)
        self.SUBJECT_ALIASES = self.config.get("SUBJECT_ALIASES", {})
        self.art_subject_combos = []
        self.art_subject_selections = self.config.get("PROFILE_ART_SELECTIONS", {})
        
        # --- ProfileManager の初期化（プロファイル管理を委譲）---
        self.profile_manager = ProfileManager(self.year_data, self.years)
        self.profile_manager.art_subject_selections = self.art_subject_selections
        
        self.ABNORMAL_SUBJECTS_UNITS_GENERAL = {self._normalize_subject_name(k): v for k, v in self.config.get("ABNORMAL_SUBJECTS_UNITS", {}).items()}

        # 重複ハイライト状態
        for year in self.years:
            if year in self.year_data:
                self.year_data[year]["duplicate_highlight_status"] = {"subjects": set(), "widgets": {}}
        
        # その他の初期化
        self.max_memory_limit = self.MAX_MEMORY_LIMIT
        self.include_fixed = self.INCLUDE_FIXED
        self.display_windows = []
        self.profile_tab_states = {}  # プロファイルタブ状態を保存する辞書
        self.is_rebuilding_profiles = False
        self.is_tutorial_running = False
        
        # 時間割生成ワーカー管理
        self._timetable_worker = None
        self._current_generation_year_label = None

        # メモリリーク対策として、非表示のウィンドウを定期的にクリーンアップするタイマー
        self.cleanup_timer = QTimer(self)
        self.cleanup_timer.setInterval(3000)  # 3秒ごと
        self.cleanup_timer.timeout.connect(self.cleanup_display_windows)
        self.cleanup_timer.start()
        
        # ウィジェットの作成
        self.create_widgets()

        self.combination_save_buttons = []
        self.is_tutorial_running = False
        self._has_active_combinations = False # 新しいフラグを追加
        # TutorialManagerの初期化
        self.tutorial_manager = TutorialManager(self, {
            "tab_list": self.tab_list,
            "get_submit_btn": self._get_submit_btn,
            "get_filter_button": self._get_filter_button,
            "get_subject_cb": self._get_subject_checkbox,
            "get_important_cb": self._get_important_checkbox,
            "get_save_combination_btn": self._get_save_combination_btn,
            "get_current_display_window": self._get_current_display_window,
            "setting_window": lambda: getattr(self, 'setting_window', None),
            "active_filter_subject_cb": lambda: self.active_filter_subject_cb, # 遅延評価
            "include_fixed_cb": lambda: self.include_fixed_cb, # 遅延評価
            "min_subject_spinbox": lambda: self.min_subject_spinbox, # 遅延評価
            "max_subject_spinbox": lambda: self.max_subject_spinbox, # 遅延評価
            "min_units_spinbox": lambda: self.min_units_spinbox, # 遅延評価
            "max_units_spinbox": lambda: self.max_units_spinbox, # 遅延評価
            "get_add_button": lambda: self.add_subject_button,
            "get_eraser_button": lambda: self.eraser_button,
            "get_lock_button": lambda: self.lock_button,
            "get_save_all_button": self._get_save_all_button,
            "get_load_button": lambda: self.load_button,
            "get_change_button": lambda: self.change_button,
            "get_timetable_slot_by_subject": self._get_timetable_slot_by_subject,
            "get_empty_timetable_slot": self._get_empty_timetable_slot,
            "get_disabled_timetable_slot": self._get_disabled_timetable_slot,
            "tutorial_signal": self.tutorial_signal,
        }, self.base_path)
        self.tutorial_manager.tutorial_finished.connect(self.on_tutorial_finished)

        # チュートリアル開始の確認
        if self._get_setting("RUN_TUTORIAL_ON_STARTUP", True):
            # 1. スプラッシュスクリーンが表示されている場合は一時的に隠す
            if hasattr(self, 'splash') and self.splash:
                self.splash.hide()

            msg_box = QMessageBox(self)
            # 2. 最前面フラグを強制的にセットする
            msg_box.setWindowFlags(msg_box.windowFlags() | Qt.WindowStaysOnTopHint)
            
            msg_box.setWindowTitle("チュートリアル")
            msg_box.setText("操作チュートリアルを開始しますか？")
            msg_box.setIcon(QMessageBox.Question)
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg_box.setDefaultButton(QMessageBox.Yes)
            
            cb = QCheckBox("このメッセージを二度と表示しない")
            msg_box.setCheckBox(cb)
            
            # 3. 念押しでフォーカスを要求する
            msg_box.raise_()
            msg_box.activateWindow()
            
            reply = msg_box.exec()

            # 4. ダイアログが閉じたらスプラッシュを再表示する
            # (この後すぐにメイン画面が出るので再表示しなくても良いですが、一応戻すのが安全です)
            if hasattr(self, 'splash') and self.splash:
                self.splash.show()
                self.splash.raise_()

            run_tutorial_on_startup = not cb.isChecked()
            if self._get_setting("RUN_TUTORIAL_ON_STARTUP") != run_tutorial_on_startup:
                self.config["RUN_TUTORIAL_ON_STARTUP"] = run_tutorial_on_startup
                # 設定タブのチェックボックスも同期する
                if hasattr(self, 'run_tutorial_cb'):
                    self.run_tutorial_cb.blockSignals(True)
                    self.run_tutorial_cb.setChecked(run_tutorial_on_startup)
                    self.run_tutorial_cb.blockSignals(False)
                # ユーザー設定ファイルに保存
                self._save_user_settings()

            if reply == QMessageBox.Yes:
                self.start_tutorial()

        # フォントサイズ設定とボトムコントロールを画面下部に配置
        bottom_layout = QHBoxLayout()
        
        # フォントサイズコントロールをグループ化 (左寄せ)
        font_size_controls_layout = QHBoxLayout()
        font_size_controls_layout.addWidget(QLabel("フォントサイズ:"))
        self.main_font_size_spinbox = QSpinBox()
        self.main_font_size_spinbox.setRange(8, 24)
        self.main_font_size_spinbox.setValue(self._get_setting("APP_FONT_SIZE", 14))
        self.main_font_size_spinbox.valueChanged.connect(self.change_font_size)
        self.main_font_size_spinbox.valueChanged.connect(self.save_app_settings)
        font_size_controls_layout.addWidget(self.main_font_size_spinbox)
        
        # --- 時間割タブ用のボトムコントロール ---
        self.timetable_bottom_specific_frame = QFrame()
        timetable_controls_layout = QHBoxLayout(self.timetable_bottom_specific_frame)
        timetable_controls_layout.setContentsMargins(0, 0, 0, 0)
        
        # --- モードボタン ---
        # テーマに応じた単色アイコン色を決定する。
        target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"

        # 共通ベースSVGのパスを定義する。
        icon_path_add = os.path.join(self.base_path, "svgs", "plus.svg")
        icon_path_eraser = os.path.join(self.base_path, "svgs", "eraser.svg")
        icon_path_lock = os.path.join(self.base_path, "svgs", "lock.svg")

        self.add_subject_button = QPushButton()
        add_icon = self._create_icon_from_svg_data(icon_path_add, target_color)
        if add_icon:
            self.add_subject_button.setIcon(add_icon)
        else:
            self.add_subject_button.setText("＋")
        self.add_subject_button.setCheckable(True)
        self.add_subject_button.setToolTip("教科を追加")

        self.eraser_button = QPushButton()
        eraser_icon = self._create_icon_from_svg_data(icon_path_eraser, target_color)
        if eraser_icon:
            self.eraser_button.setIcon(eraser_icon)
        else:
            self.eraser_button.setText("🗑️")
        self.eraser_button.setCheckable(True)
        self.eraser_button.setToolTip("教科を消去")

        self.lock_button = QPushButton()
        lock_icon = self._create_icon_from_svg_data(icon_path_lock, target_color)
        if lock_icon:
            self.lock_button.setIcon(lock_icon)
        else:
            self.lock_button.setText("🔒")
        self.lock_button.setCheckable(True)
        self.lock_button.setToolTip("教科を固定/枠を無効化")
        
        # ボタンとモード文字列を対応づける
        self.mode_button_map = {
            self.add_subject_button: "＋",
            self.eraser_button: "🗑️",
            self.lock_button: "🔒"
        }

        self.timetable_mode_group = QButtonGroup(self)
        self.timetable_mode_group.setExclusive(True)
        self.timetable_mode_group.addButton(self.add_subject_button)
        self.timetable_mode_group.addButton(self.eraser_button)
        self.timetable_mode_group.addButton(self.lock_button)

        self.add_subject_button.setChecked(True)
        self.current_timetable_mode = "＋" # 初期モード

        timetable_controls_layout.addStretch()
        timetable_controls_layout.addWidget(self.add_subject_button)
        timetable_controls_layout.addWidget(self.eraser_button)
        timetable_controls_layout.addWidget(self.lock_button)
        timetable_controls_layout.addStretch()

        self.timetable_mode_group.buttonToggled.connect(self._on_timetable_mode_changed)
        self.timetable_bottom_specific_frame.setVisible(True) # 初期状態から表示

        # ボトムレイアウトにコントロールを配置
        bottom_layout.addLayout(font_size_controls_layout, 0)
        bottom_layout.addStretch(1)
        bottom_layout.addWidget(self.timetable_bottom_specific_frame, 0)
        bottom_layout.addStretch(1)
        
        self.clear_results_button = QPushButton("組み合わせ結果をクリア")
        self.clear_results_button.clicked.connect(self.clear_generated_results)
        self.clear_results_button.setVisible(False)
        # 非表示時もレイアウト上のサイズを維持する
        policy = self.clear_results_button.sizePolicy()
        policy.setRetainSizeWhenHidden(True)
        self.clear_results_button.setSizePolicy(policy)
        bottom_layout.addWidget(self.clear_results_button) # 右寄せのクリアボタン

        bottom_frame = QFrame()
        bottom_frame.setLayout(bottom_layout)
        self.main_layout.addWidget(bottom_frame)
        self._save_user_settings()

        # 単位数を初期表示するために呼び出す
        self.check_subjects_units()

        # イベントフィルターを設定
        self.installEventFilter(self)

    def eventFilter(self, obj, event):
        """イベントフィルターでフォーカスイベントやタブの操作を処理"""
        if event.type() == QEvent.WindowActivate:
            if obj == self:
                self.bring_to_front()
        
        if not hasattr(self, 'profile_tab_widget') or not widget_is_valid(self.profile_tab_widget):
            return super().eventFilter(obj, event)
        
        if obj == self.profile_tab_widget.tabBar():
            if event.type() == QEvent.MouseButtonDblClick:
                index = obj.tabAt(event.pos())
                if index >= 0:
                    self._start_rename_tab(index)
                    return True
            elif event.type() == QEvent.ContextMenu:
                index = obj.tabAt(event.pos())
                if index >= 0:
                    self._show_tab_context_menu(event.globalPos(), index)
                    return True

        return super().eventFilter(obj, event)

    def changeEvent(self, event):
        """システムテーマの変更を検知してスタイルを更新"""
        if event.type() == QEvent.ThemeChange:
            # 1. システム変更を反映するためキャッシュをクリア
            if hasattr(self, 'theme_manager'):
                self.theme_manager.refresh_theme()
                self.theme = self.theme_manager.get_all_theme_colors()
            
            # 2. スタイルを再適用 (現在のテーマ状態を渡す)
            is_dark = self.is_dark_theme()
            set_button_styles(QApplication.instance(), is_dark=is_dark)
            self.update_icons()

            # 3. ハイライト色を即時反映するための更新
            self.update_all_highlights()
            self.check_subjects_units() # 単位数チェックも念のため更新
        super().changeEvent(event)

    def update_icons(self):
        """現在のテーマに基づいてSVGアイコンを更新する"""
        is_dark = self.is_dark_theme()
        target_color = "#FFFFFF" if is_dark else "#1F1F1F"
        
        # ヘルパー: アイコン生成
        def get_icon(name):
            return self._create_icon_from_svg_data(os.path.join(self.base_path, "svgs", f"{name}.svg"), target_color)

        # 1. 固定ボタン
        if hasattr(self, 'tutorial_button'):
            icon = get_icon("question")
            if icon: self.tutorial_button.setIcon(icon)
        
        if hasattr(self, 'load_button'):
            icon = get_icon("load_file")
            if icon: self.load_button.setIcon(icon)

        if hasattr(self, 'change_button'):
            icon = get_icon("logout")
            if icon: self.change_button.setIcon(icon)

        # 2. プロファイル追加ボタン（タブバーのCornerWidget内のボタン）
        if hasattr(self, 'profile_tab_widget'):
            corner = self.profile_tab_widget.cornerWidget(Qt.TopRightCorner)
            if corner:
                buttons = corner.findChildren(QPushButton)
                if buttons:
                    icon = get_icon("plus")
                    if icon: buttons[0].setIcon(icon)
        
        # 3. 左側のタブリト (tab_list)
        if hasattr(self, 'tab_list'):
            for i in range(self.tab_list.count()):
                item = self.tab_list.item(i)
                text = item.text()
                # タブ名に応じてアイコンを決定
                icon_name = "add" # デフォルト（学年タブ用）
                if text == "時間割":
                    icon_name = "timetable"
                elif text == "設定":
                    icon_name = "setting"
                
                icon = get_icon(icon_name)
                if icon: item.setIcon(icon)

        # 4. プロファイルタブ内の保存ボタン
        if hasattr(self, 'profile_tab_widget'):
            save_icon = get_icon("save")
            if save_icon:
                # すべてのページ（プロファイル）を走査
                for i in range(self.profile_tab_widget.count()):
                    page = self.profile_tab_widget.widget(i)
                    if page:
                        # ページ内のすべてのボタンを検索
                        buttons = page.findChildren(QPushButton)
                        for btn in buttons:
                            # テキストで保存ボタンを特定 ("保存" または "Excelに保存")
                            if btn.text() == "保存" or "Excelに保存" in btn.text():
                                btn.setIcon(save_icon)



    def bring_to_front(self):
        """メインウィンドウを前面に持ってくる"""
        if self.is_tutorial_running:
            return
        self.raise_()
        self.activateWindow()
    
    def mousePressEvent(self, event):
        """マウスクリックでメインウィンドウをアクティブにする"""
        self.bring_to_front()
        super().mousePressEvent(event)

    def closeEvent(self, event):
        """
        メインウィンドウが閉じられたときに、未保存データがある場合に警告メッセージを表示し、
        すべての子ウィンドウを閉じてアプリケーションを確実に終了させる。
        """
        self._save_user_settings()

        if self._has_active_combinations:
            reply = QMessageBox.question(self, "未保存のデータ",
                                           "保存されていないデータは削除されます。保存せずに終了しますか？",
                                           QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                           QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                event.ignore()
                return

        # ユーザーが承認した場合、または未保存候補がない場合は終了処理へ進む
        # 開いている表示ウィンドウを閉じる
        if hasattr(self, 'display_windows'):
            for window in self.display_windows:
                # ウィンドウ固有のデータをクリーンアップ
                if hasattr(window, 'window_data'):
                    window_data = window.window_data
                    # 大きなデータを明示的に削除
                    if 'timetables' in window_data:
                        del window_data['timetables']
                    del window.window_data
                window.close()
            self.display_windows.clear()

        # 設定ウィンドウを閉じる
        if hasattr(self, 'setting_window') and self.setting_window.isVisible():
            self.setting_window.close()

        # グローバルな一時データをクリア
        if hasattr(self, 'all_timetables'):
            del self.all_timetables
            
        # ガベージコレクション
        gc.collect()

        # メインウィンドウの終了処理へ進む
        event.accept()

    def clear_generated_results(self):
        """
        生成された時間割の組み合わせ結果をすべてクリアし、メモリを解放する。
        """
        # 1. 表示中の組み合わせウィンドウをすべて閉じる
        if hasattr(self, 'display_windows'):
            for window in list(self.display_windows):
                window.close()
            self.display_windows.clear()

        # 2. すべての学年の組み合わせデータをクリアする
        for year in self.years:
            if year in self.year_data and "complete_combination" in self.year_data[year]:
                self.year_data[year]["complete_combination"].clear()
        
        # 3. メインの時間割タブを再描画して、クリアされたことを反映する
        self.create_combination_time_tab()
        
        # 4. 単位数表示を更新する
        self.check_subjects_units()

        # 5. ユーザーに通知する
        QMessageBox.information(self, "クリア完了", "生成されたすべての組み合わせ結果をメモリから消去しました。")
        self._has_active_combinations = False # フラグをFalseに設定

    def display_timetables(self, timetables, year_label, important_subjects):
        """時間割の組み合わせを表示するウィンドウを作成（Zオーダー修正版）"""
        if not timetables:
            QMessageBox.information(self, "結果", "組み合わせが見つかりませんでした。")
            submit_btn = getattr(self, f"submit_btn_{year_label}", None)
            if submit_btn:
                submit_btn.setEnabled(True)
            return
        
        # ウィンドウ作成時点のプロファイル名をキャプチャ（後で保存ボタンで使用）
        captured_profile_name = self.get_active_profile_name()

        # ジェネレータ対応のため、全件ソートは行わない

        # 新しいウィンドウを作成
        window = QMainWindow()  # 親を指定しない
        window.setWindowTitle(f"{year_label} 時間割候補")
        window.resize(1200, 800)
        
        # アイコンを設定
        icon_base_name = "時間割くんアイコン.ico"
        icon_path_obj = self.base_path / icon_base_name
        if icon_path_obj.exists():
            window.setWindowIcon(QIcon(str(icon_path_obj)))
        else:
            print(f"Icon file not found for timetable window: {icon_path_obj}")
        
        # ウィンドウフラグを通常の独立したウィンドウに設定
        window.setWindowFlags(Qt.Window)
        
        # メインウィンドウの横に表示するように位置調整
        main_geometry = self.frameGeometry()
        window.move(main_geometry.topLeft())
        
        # ウィンドウが閉じられたときのクリーンアップを設定
        window.destroyed.connect(lambda: self.cleanup_display_window_data(year_label))
        
        self.display_windows.append(window)

        central_widget = QWidget()
        window.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        # 左側のパネル（リストとページネーション）
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_panel.setFixedWidth(250)

        # 組み合わせリスト
        total_combinations_label = QLabel(f"組み合わせの総数: {len(timetables)}")
        left_layout.addWidget(total_combinations_label)

        combination_list_widget = QListWidget()
        left_layout.addWidget(combination_list_widget)

        # ページ情報ラベル
        page_label = QLabel("")
        left_layout.addWidget(page_label, alignment=Qt.AlignCenter)

        # ページ移動ボタン
        page_nav_layout = QHBoxLayout()
        prev_button = QPushButton("前へ")
        next_button = QPushButton("次へ")
        page_nav_layout.addWidget(prev_button)
        page_nav_layout.addWidget(next_button)
        left_layout.addLayout(page_nav_layout)
        
        main_layout.addWidget(left_panel)

        # 右側の詳細表示エリア
        details_area = QStackedWidget()
        main_layout.addWidget(details_area)

        # ページネーション関連
        current_page = [0]
        timetables_per_page = self.SUBJECT_COMBINATION_COUNT
        
        # ウィンドウ固有のデータを保持する辞書
        window_data = {
            'timetables': timetables,
            'year_label': year_label,
            'combination_list_widget': combination_list_widget,
            'details_area': details_area,
            'page_label': page_label,
            'current_page': current_page,
            'timetables_per_page': timetables_per_page,
            'combination_save_buttons': []
        }
        
        # ウィンドウオブジェクトにデータを保存
        window.window_data = window_data

        def show_page(page_index):
            # 既存のデータをクリア
            combination_list_widget.clear()
            while details_area.count():
                widget = details_area.widget(0)
                details_area.removeWidget(widget)
                widget.deleteLater()
            
            window_data['combination_save_buttons'].clear()

            start_index = page_index * timetables_per_page
            end_index = start_index + timetables_per_page
            page_timetables = timetables[start_index:end_index]

            for i, timetable in enumerate(page_timetables):
                # 左のリストに項目を追加
                list_item_text = f"組み合わせ {start_index + i + 1}"
                combination_list_widget.addItem(list_item_text)

                # 右の詳細ウィジェットを作成
                details_widget = QWidget()
                details_layout = QVBoxLayout(details_widget)
                
                # 上部情報（番号、単位数）
                info_layout = QHBoxLayout()
                
                num_label = QLabel(f"時間割 {start_index + i + 1}")
                num_label.setFont(QFont("Arial", 10, QFont.Bold))
                info_layout.addWidget(num_label)

                units = self.calculate_timetable_units(timetable, year_label, include_fixed_subjects=True)
                units_label = QLabel(f"単位数: {units}")
                info_layout.addWidget(units_label)

                details_layout.addLayout(info_layout)

                # 時間割テーブル
                table_widget = QWidget()
                table_widget.setLayout(QHBoxLayout())
                self.create_table_slot(table_widget, timetable, self.year_data[year_label]["table_layout"], year_label, self.year_data[year_label]["all_slots"], set())
                details_layout.addWidget(table_widget)

                # 保存ボタン（ウィンドウ作成時点のプロファイル名を使用）
                save_button = QPushButton("この組み合わせを保存")
                save_button.clicked.connect(lambda _, t=timetable, y=year_label, p=captured_profile_name: self.save_timetable(t, y, p))
                details_layout.addWidget(save_button)
                window_data['combination_save_buttons'].append(save_button)

                # 不足している必須科目
                missing_req_frame = QFrame()
                missing_req_layout = QVBoxLayout(missing_req_frame)
                missing_req_label = QLabel("<b>不足している必須科目:</b><br>")
                missing_reqs = self.check_required_subjects(timetable, year_label, include_all=True)
                
                if missing_reqs:
                    missing_req_label.setStyleSheet("font-weight: bold; color: red;")
                    formatted_message = self.format_missing_subjects_message(missing_reqs)
                    missing_req_label.setText(missing_req_label.text() + formatted_message.replace("\n", "<br>"))
                    missing_req_frame.setStyleSheet("background-color: #FFCCCC; border: 1px solid #FF9999; padding: 5px;")
                else:
                    missing_req_label.setText(missing_req_label.text() + "なし")
                
                missing_req_layout.addWidget(missing_req_label)
                details_layout.addWidget(missing_req_frame)

                # 不足している前提科目
                missing_prereq_frame = QFrame()
                missing_prereq_layout = QVBoxLayout(missing_prereq_frame)
                missing_prereq_label = QLabel("<b>不足している前提科目:</b><br>")
                missing_prereqs = self.check_prerequisites(timetable, year_label)
                if missing_prereqs:
                    for subj, prereqs in missing_prereqs.items():
                        missing_prereq_label.setText(missing_prereq_label.text() + f"- {', '.join(prereqs)} (前提: {subj})<br>")
                else:
                    missing_prereq_label.setText(missing_prereq_label.text() + "なし")
                missing_prereq_layout.addWidget(missing_prereq_label)
                details_layout.addWidget(missing_prereq_frame)

                details_layout.addStretch()

                details_area.addWidget(details_widget)

            page_label.setText(f"ページ: {page_index + 1} / {math.ceil(len(timetables) / timetables_per_page)}")
            combination_list_widget.setCurrentRow(0)

        # ページ移動ボタンの接続
        def prev_page():
            if current_page[0] > 0:
                current_page[0] -= 1
                show_page(current_page[0])

        def next_page():
            if (current_page[0] + 1) * timetables_per_page < len(timetables):
                current_page[0] += 1
                show_page(current_page[0])

        prev_button.clicked.connect(prev_page)
        next_button.clicked.connect(next_page)
        
        # リストと詳細の同期
        combination_list_widget.currentRowChanged.connect(details_area.setCurrentIndex)

        show_page(0)
        
        # ウィンドウを表示（アクティブにはしない）
        window.show()
        
        # グローバルな組み合わせデータをクリア
        if hasattr(self, 'all_timetables'):
            del self.all_timetables
        
        # グローバルな組み合わせデータをクリア（不要になったため）
        if hasattr(self, 'all_timetables'):
            del self.all_timetables

    def cleanup_display_window_data(self, year_label):
        """表示ウィンドウのデータをクリーンアップ"""
        # 学年データ内の一時的な組み合わせ情報をクリア
        if year_label in self.year_data:
            # complete_combination は保存された時間割なので保持する
            # 代わりに、一時的な検索結果をクリアするためのフラグなどを追加可能
            pass
        
        # ガベージコレクションを促す
        gc.collect()

    def on_tutorial_button_clicked(self):
        """チュートリアルボタンがクリックされたときの処理"""
        # 実行中の場合は多重起動しない
        if self.is_tutorial_running:
            return
        self.is_tutorial_running = True
        self.tutorial_button.setEnabled(False)
        
        # QTimer.singleShotを使用して、現在のイベント処理が完了した後にチュートリアルを開始する
        QTimer.singleShot(0, self.start_tutorial)

    def start_tutorial(self):
        """チュートリアルを開始する"""
        if not self.years:
            QMessageBox.warning(self, "チュートリアルエラー", "学年が設定されていないため、チュートリアルを開始できません。")
            self.on_tutorial_finished() # 失敗した場合も後処理を呼ぶ
            return
        
        first_year_full = self.years[0]
        first_year_display = first_year_full.split('_')[-1]

        subject_slots = self.year_data[first_year_full]["subject_slots_base"]
        first_subject_item = next(iter(subject_slots), None)
        first_subject_name = first_subject_item["name"] if first_subject_item else "数学Ⅰ"

        tutorial_steps = [
            # --- 基本的な時間割作成 ---
            TutorialStep(message=f"ようこそ！時間割くんの操作チュートリアルへ.\nまず、左の「{first_year_display}」タブをクリックして、時間割の作成を始めましょう.", target_widget_name="tab_list", action="tab_change", expected_value=1),
            TutorialStep(message=f"ここでは{first_year_display}で履修したい科目を選択します.\n試しに「{first_subject_name}」にチェックを入れてみましょう.", target_widget_name="get_subject_cb", action="check", subject_name=first_subject_name, year_label=first_year_full),
            TutorialStep(message="次に、その科目が必ず含まれるように、\n教科名左側の「必含」の欄にチェックを入れます.", target_widget_name="get_important_cb", action="check", subject_name=first_subject_name, year_label=first_year_full),
            TutorialStep(message="素晴らしい！では、右上の「組み合わせを絞り込み」ボタンを押して、\n時間割の条件を設定しましょう.", target_widget_name="get_filter_button", action="click", year_label=first_year_full),
            TutorialStep(message="この画面では、作成する時間割の条件を細かく設定できます.\n今回はデフォルトのまま、下の「閉じる」ボタンで閉じてください.", target_widget_name="setting_window", action="window_close"),
            TutorialStep(message="準備ができました！「実行」ボタンを押して時間割を作成します.", target_widget_name="get_submit_btn", action="click", year_label=first_year_full),
            TutorialStep(message="時間割が作成されました！\n「この組み合わせを保存」ボタンを押して時間割を保存しましょう.", target_widget_name="get_save_combination_btn", action="click"),
            TutorialStep(message="時間割が保存されました！\nウィンドウ右上の「×」ボタンを押してこの画面は閉じてください。\n（「×」ボタンはOSの機能のためハイライトできませんが、このウィンドウを閉じることで次に進みます）", target_widget_name="get_current_display_window", action="window_close"),
            TutorialStep(message="左の「時間割」をクリックして結果を確認しましょう.", target_widget_name="tab_list", action="tab_change", expected_value=0),
            
            # --- 時間割タブの編集機能 (＋ → 🗑️ → 🔒) ---
            TutorialStep(message="これが作成された時間割です。\nここからは、下部のボタンを使って時間割を手動で編集する方法を学びます。", action="info_next_button"),
            
            TutorialStep(
                message="最初は「＋」(追加)モードです。\n空いているコマをクリックして、新しい教科を追加してみましょう。", 
                highlight_target_name="get_empty_timetable_slot",
                action="info_next_button"
            ),

            TutorialStep(message="教科を追加できました。\n次は教科を消してみましょう。下部の「消しゴム」ボタンをクリックしてください。", highlight_target_name="get_eraser_button", action="info_next_button"),

            TutorialStep(
                message=f"消しゴムモードになりました。\n時間割の中から、先ほど保存した「{first_subject_name}」をクリックして消してみましょう。", 
                highlight_target_name="get_timetable_slot_by_subject",
                action="info_next_button",
                subject_name=first_subject_name
            ),

            TutorialStep(message="教科を消せましたね。\n最後に、コマの固定と無効化を試します。下部の「🔒」(錠)ボタンをクリックしてください。", highlight_target_name="get_lock_button", action="info_next_button"),
            TutorialStep(
                message="ロックモードです。時間割上の教科を【右クリック】すると、その教科をその場所に固定（青色表示）できます。試してみましょう。",
                highlight_target_name="get_timetable_slot_by_subject",
                action="info_next_button",
                subject_name=first_subject_name
            ),
            TutorialStep(
                message="教科が固定されました。\nもう一度同じ教科を【右クリック】すると、固定を解除できます。",
                highlight_target_name="get_timetable_slot_by_subject",
                action="info_next_button",
                subject_name=first_subject_name
            ),
            TutorialStep(message="次はコマの無効化です。\n時間割上の【空いているコマ】を【左クリック】すると、そのコマを計算対象から除外（赤色表示）できます。", highlight_target_name="get_empty_timetable_slot", action="info_next_button"),
            TutorialStep(message="コマが無効化されました。\nもう一度同じコマを【左クリック】すると元に戻せます。", highlight_target_name="get_disabled_timetable_slot", action="info_next_button"),
            
            # --- ファイル操作と設定変更 ---
            TutorialStep(message="時間割の編集は以上です。\n次にファイル操作と設定変更機能を試します。", action="info_next_button"),
            TutorialStep(message="作成した時間割はファイルに保存できます。\n右上の「保存」ボタンをクリックし「全体をExcelに保存」をクリックしてください。警告が表示される場合は全て「yes」を押してください。\n保存ダイアログが開きましたら保存したいExcelファイルを選択することで時間割を保存できます。今回は「キャンセル」を押して閉じてください。", highlight_target_name="get_save_all_button", action="info_next_button"),
            TutorialStep(message="次に、専用形式ファイルで保存する機能を試します。\nもう一度右上の「保存」ボタンをクリックし「専用形式ファイルに保存」をクリックしてください。\nこの操作で保存されるファイルは現在作成している時間割くんの進捗を保存することができます。今回は「キャンセル」を押して閉じてください。", highlight_target_name="get_save_all_button", action="info_next_button"),            
            TutorialStep(message="次に、保存したファイルを開く機能を試します。\n画面左の「ファイルを開く」ボタンをクリックし「Excelを開く」または「専用形式ファイルを開く」をクリックしてください。\nファイル選択ウィンドウが開きます。これも「キャンセル」で閉じてください。", highlight_target_name="get_load_button", action="info_next_button"),
            TutorialStep(message="「Excelを開く」と「専用形式ファイルを開く」どちらも作成された時間割を読み込むことができますが、\n「Excelを開く」は保存された時間割だけを読み込むのに対し、「専用形式ファイルを開く」は進捗情報も含めて読み込むことができます。", action="info_next_button"),
            TutorialStep(message="次に、学科を変更する機能を試します。\n画面左の「変更」ボタンをクリックし「学科を変更」をクリックしてください。\n学科を選択する画面が表示されます。今回は「キャンセル」を押して閉じてください。", highlight_target_name="get_change_button", action="info_next_button"),
            TutorialStep(message="最後に、設定ファイルを変更する機能です。\nもう一度「変更」ボタンをクリックし「設定ファイルを変更」をクリックしてください。\n学校データを選択する画面が表示されます。これも「キャンセル」で閉じてください。\nその後選択ダイアログが開きます。これも「キャンセル」を押して閉じてください。", highlight_target_name="get_change_button", action="info_next_button"),
            TutorialStep(message="設定ファイルの変更は、異なる学校などのデータを読み込む際に使用します。\nこの操作により、現在の進捗はリセットされ、新しい学校データが適用されます。", action="info_next_button"),

            # --- 終了 ---
            TutorialStep(message="これでチュートリアルは全て終了です。\n「時間割くん」があなたの学校生活の助けになることを願っています！", target_widget_name=None, action="last_step")
        ]
        self.tutorial_manager.set_steps(tutorial_steps)
        self.tutorial_manager.start_tutorial()

    def on_tutorial_finished(self):
        """チュートリアルが終了したときにフラグとボタンをリセットする"""
        self.is_tutorial_running = False
        if hasattr(self, 'tutorial_button'):
            self.tutorial_button.setEnabled(True)

    def _get_hierarchy_dict_from_path_key(self, path_key):
        if not path_key:
            return self.config.get("YEARS_HIERARCHY", {})

        parts = path_key.split('_')
        current_node = self.config.get("YEARS_HIERARCHY", {})
        
        for part in parts:
            if part in current_node:
                current_node = current_node[part]
            else:
                # 指定パスが見つからない場合は階層全体をフォールバックとして返す
                return self.config.get("YEARS_HIERARCHY", {})

        # 呼び出し側が想定する形に合わせるため、選択パス全体を含む階層dictを再構築する
        # 対話的な選択結果と同じネスト構造にそろえる
        # 選択処理が返す形に合わせる。
        res = current_node
        for key in reversed(parts):
            res = {key: res}
        return res

    def _perform_hierarchical_selection(self):
        """
        Shows a series of dialogs to select a path through the YEARS_HIERARCHY.
        Returns the selected sub-hierarchy, or a special value on cancellation.
        - Success: returns a tuple of (path_key_string, selected_hierarchy_dict).
        - Root Cancel: returns (None, None).
        - Sub-dialog Cancel: returns ("SUB_CANCEL", None).
        
        アンカーが設定されている項目が選択された場合、その時点で選択を確定します。
        """
        if self.is_tutorial_running:
            self.tutorial_signal.emit({'action': 'menu_clicked'})
            dialog = DepartmentSelectionDialog(self, ["（見本）理学部", "（見本）工学部"], "学科を選択してください")
            dialog.finished.connect(lambda result: self.tutorial_signal.emit({'action': 'dialog_closed'}))
            dialog.exec()
            return "SUB_CANCEL", None # チュートリアル中は実際の処理は行わない

        full_hierarchy = self.config.get("YEARS_HIERARCHY", {})
        
        # アンカーリストを取得
        anchors = set(self.config.get("HIERARCHY_ANCHORS", []))

        if not full_hierarchy:
            return "", full_hierarchy

        current_hierarchy = full_hierarchy
        selected_path = []
        max_depth = 10  # 無限ループ防止

        for depth in range(max_depth):
            keys = list(current_hierarchy.keys())
            if not keys:
                break  # 階層の葉ノードに到達

            dialog_title = "選択してください"
            dialog = DepartmentSelectionDialog(self, keys, dialog_title)
            
            # UIイベントを処理し、ダイアログ表示のちらつきを防ぐ
            QApplication.processEvents()
            dialog.raise_()
            dialog.activateWindow()

            if dialog.exec() == QDialog.Accepted:
                choice = dialog.get_selected_department()
                if choice and choice in current_hierarchy:
                    selected_path.append(choice)
                    current_hierarchy = current_hierarchy[choice]
                    
                    # アンカーチェック: 選択されたパスがアンカーに設定されていれば終了
                    path_key = "_".join(selected_path)
                    if path_key in anchors:
                        break  # アンカー項目が選択されたので、これ以上選択ダイアログを表示しない
                else:
                    QMessageBox.warning(self, "選択エラー", "無効な選択です。全ての学科を表示します。")
                    return "", self.config.get("YEARS_HIERARCHY", {})
            else:
                # ユーザーがキャンセル
                if depth == 0:  # 最初のダイアログでキャンセルされた場合
                    return None, None
                else:  # 途中のダイアログでキャンセルされた場合
                    QMessageBox.information(self, "選択キャンセル", "選択がキャンセルされました。全ての学科を表示します。")
                    return "SUB_CANCEL", None

        path_key = "_".join(selected_path)
        if not selected_path:
            return "", full_hierarchy

        # 選択パスから階層を再構築する
        res = current_hierarchy
        # _get_all_leaf_yearsで扱える形へ再構築する
        for key in reversed(selected_path):
            res = {key: res}

        return path_key, res

    def _change_department_clicked(self):
        """学科変更ボタンが押された時の処理"""
        if self.is_tutorial_running:
            self.tutorial_signal.emit({'action': 'menu_clicked'})
            QTimer.singleShot(100, lambda: self.tutorial_signal.emit({'action': 'dialog_opened'}))
            return

        path_key, hierarchy_dict = self._perform_hierarchical_selection()

        if path_key is None:  # ユーザーがキャンセル at the root dialog, do nothing
            QMessageBox.information(self, "選択キャンセル", "学科選択がキャンセルされました。現在の学科設定を維持します。")
            return

        # 新しい選択が完了したか、途中でキャンセルされた。
        if path_key == "SUB_CANCEL":
            # 途中キャンセル時は現在の設定を維持する。
            return
        
        self.setUpdatesEnabled(False)
        try:
            # 選択完了
            self.selected_department_path_key = path_key
            self._save_user_settings()

            # 既存のタブとウィジェットをクリア
            while self.stacked_widget.count() > 0:
                widget = self.stacked_widget.widget(0)
                self.stacked_widget.removeWidget(widget)
                widget.deleteLater()
            self.tab_list.clear()
            self.year_tabs.clear()

            # 新しい選択に基づいて年次を再初期化
            self.years = self._get_all_leaf_years(hierarchy_dict)

            # 削除したタブとウィジェットを再作成
            self.combination_time_tab = QWidget()
            self.tab_settings = QWidget()

            self.stacked_widget.addWidget(self.combination_time_tab)
            self._add_tab_item("時間割", "timetable")

            for year in self.years:
                tab = QWidget()
                self.year_tabs[year] = tab
                self.stacked_widget.addWidget(tab)
                display_year = year.split('_')[-1]
                self._add_tab_item(display_year, "add")

            self.stacked_widget.addWidget(self.tab_settings)
            self._add_tab_item("設定", "setting")

            # 教科データに合わせてyear_dataを再初期化
            self.year_data = {}
            for year in self.years:
                # 旧dict形式を新しいlist形式へ変換する補助関数
                def _convert_to_list_format(data_dict):
                    if isinstance(data_dict, dict):
                        return [{"name": s, "data": d} for s, d in data_dict.items()]
                    return data_dict # すでにlist形式とみなす

                self.year_data[year] = {
                    "table_layout": self.config.get(f"table_layout{year}", []),
                    "subject_number": _convert_to_list_format(self.config.get(f"subject_number{year}", {})),
                    "save_position": self.config.get(f"SAVE_POSITION{year}", {}),
                    "subject_slots_base": _convert_to_list_format(self.config.get(f"subject_slots_base{year}", {})),
                    "fixed_slots": self.config.get(f"FIXED_SLOTS{year}", {}),
                    "all_slots": [slot for row in self.config.get(f"table_layout{year}", []) for slot in row if slot],
                    "profile_all_slots": {"Default": [slot for row in self.config.get(f"table_layout{year}", []) for slot in row if slot]},
                    "abnormal_units": {self._normalize_subject_name(k): v for k, v in self.config.get(f"ABNORMAL_SUBJECTS_UNITS{year}", {}).items()},
                    "required_subjects": self.config.get(f"REQUIRED_SUBJECTS_{year}", {}),
                    "saved_profiles": {"Default": {}},
                    "profile_ui_states": {"Default": {"checked": set(), "important": set(), "prefixes": {}}},
                    "active_profile_name": "Default",
                    "complete_combination": {},
                    "subject_frames": [],
                    "subject_scroll_layout": None,
                    "check_vars": {},
                    "prefix_vars": {},
                    "important_vars": {},
                    "duplicate_highlight_status": {"subjects": set(), "widgets": {}},
                    "profile_lock_settings": {"Default": {}}
                }

            # base_units_value を再計算
            department_path = self.selected_department_path_key or ""
            config_key = f"YEARS_SUBJECTS_UNITS_{department_path}"
            self.base_units_value = self._get_setting(config_key, None)
            
            if self.base_units_value is None:
                years_hierarchy = self.config.get("YEARS_HIERARCHY", {})
                if years_hierarchy:
                    top_level_key = list(years_hierarchy.keys())[0] if years_hierarchy else ""
                    fallback_key = f"YEARS_SUBJECTS_UNITS_{top_level_key}"
                    self.base_units_value = self._get_setting(fallback_key, 0)
                else:
                    self.base_units_value = 0

            # 新しいタブ用のウィジェットを再作成
            # 重複防止のため以前のウィジェットをクリアする
            try:
                # 既存のprofile_tab_widgetなど動的ウィジェットがあれば削除
                if hasattr(self, 'profile_tab_widget'):
                    self.profile_tab_widget.deleteLater()
            except Exception:
                pass
            # タブ内の子ウィジェットを削除して重複を防ぐ
            for i in range(self.stacked_widget.count()):
                w = self.stacked_widget.widget(i)
                for child in w.findChildren(QWidget):
                    # stacked_widget本体など主要コンテナは削除しない
                    if child is not self.stacked_widget and child is not self.combination_time_tab and child is not self.tab_settings:
                        child.deleteLater()
            self.create_widgets()

            # ツールチップを更新（__init__と同内容）
            self.tab_list.item(0).setToolTip("作成した時間割を管理・保存します。")
            for i, year in enumerate(self.years):
                year_label = year.split('_')[-1]
                self.tab_list.item(i + 1).setToolTip(f"{year_label}の時間割を作成します。")
            self.tab_list.item(self.tab_list.count() - 1).setToolTip("アプリケーションの動作設定を行います。")

            self.tab_list.setCurrentRow(0)
            
            # 単位数表示を更新
            self.check_subjects_units()
        finally:
            self.setUpdatesEnabled(True)

    def _change_school(self):
        """設定ファイルを選択し、UIを再読み込みする"""
        if self.is_tutorial_running:
            self.tutorial_signal.emit({'action': 'menu_clicked'})
            QTimer.singleShot(100, lambda: self.tutorial_signal.emit({'action': 'dialog_opened'}))
            return

        new_config_path = self._select_config_file(prefer_last_opened=False)

        # パスが選択されなかったか、現在のファイルと同じ場合は何もしない
        if not new_config_path or (hasattr(self, 'config_path') and Path(new_config_path).resolve() == Path(self.config_path).resolve()):
            return
        
        self.setUpdatesEnabled(False)
        try:
            # 新しい設定ファイルを読み込む
            with open(new_config_path, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            self.config_path = new_config_path
            
            # ConfigManagerのconfigも更新
            self.config_manager.config = self.config

            # ユーザー設定をリセットし、新しい設定ファイルパスを保存
            self._reset_user_settings()
            
            # 新しい設定ファイルのYEARS_SUBJECTS_UNITS_*をuser_settingsに反映
            for key, value in self.config.items():
                if key.startswith('YEARS_SUBJECTS_UNITS_'):
                    self.user_settings[key] = value
            
            # base_units_valueを直接新しい設定ファイルから取得
            years_hierarchy = self.config.get("YEARS_HIERARCHY", {})
            if years_hierarchy:
                top_level_key = list(years_hierarchy.keys())[0]
                config_key = f"YEARS_SUBJECTS_UNITS_{top_level_key}"
                self.base_units_value = self.config.get(config_key, 0)
            
            self._save_user_settings()
            
            # ウィンドウタイトルを更新
            school_name = self.config.get("GENERAL_SETTINGS", {}).get("SCHOOL_NAME")
            if school_name:
                self.setWindowTitle(f"時間割くん by I.R - [{school_name}]")
            else:
                self.setWindowTitle("時間割くん by I.R")

            # UIの再構築をトリガー（学科選択から開始される）
            self._change_department_clicked()
            
            # 単位数表示を更新
            try:
                self.check_subjects_units()
            except Exception:
                pass  # UIが準備できていない場合は無視

        except Exception as e:
            QMessageBox.critical(self, "設定エラー", f"""設定ファイルの読み込みに失敗しました:{new_config_path}
エラー: {e}""")
            return
        finally:
            self.setUpdatesEnabled(True)

    def _get_all_leaf_years(self, hierarchy):
        collected_paths = []
        
        # 階層が空なら空リストを返す
        if not hierarchy:
            return []

        def find_nodes(node, current_path_parts):
            # 現在ノードが空dictなら葉ノードとして扱う
            if not node:
                if current_path_parts: # 初期の空パスではないことを確認
                    collected_paths.append("_".join(current_path_parts))
                return

            # ノードが空でなければ子要素を走査
            for key, value in node.items():
                # 各子要素に対してfind_nodesを再帰呼び出し
                find_nodes(value, current_path_parts + [key])

        find_nodes(hierarchy, [])
        return sorted(list(set(collected_paths)))

    def _get_setting(self, key, default=None):
        """【委譲】設定値の取得をConfigManagerに委譲します。"""
        return self.config_manager.get_setting(key, default)

    def _get_user_settings_path(self):
        """【委譲】ユーザー設定ファイルのパス取得をConfigManagerに委譲します。"""
        return self.config_manager.get_user_settings_path()

    def change_font_size(self, size):
        """アプリケーション全体のフォントサイズを変更し、スピンボックスの値を同期する"""
        font = QApplication.font()
        font.setPointSize(size)
        QApplication.setFont(font)

        # --- 特定ボタンのフォントサイズを手動で更新 ---
        # これらのボタンは基準フォントより2ポイント小さく設定されているため、個別に更新が必要
        special_buttons = set()
        try:
            special_buttons = {self.tutorial_button, self.load_button, self.change_button}
            for button in special_buttons:
                btn_font = button.font()
                btn_font.setPointSize(size - 2)
                button.setFont(btn_font)
                button.update() # ボタン個別に更新を促す
        except AttributeError:
            # ウィジェットがまだ存在しない場合は何もしない
            pass

        # アプリケーション内の全ウィジェットにスタイルを再適用させ、フォント変更を反映させる
        for widget in QApplication.allWidgets():
            if widget not in special_buttons: # 手動で更新したボタンは除外
                widget.style().unpolish(widget)
                widget.style().polish(widget)
        
        self.update() # メインウィンドウ自体の更新も促す

        # ウィジェットが作成される前に呼び出される可能性があるため、存在をチェック
        if hasattr(self, 'main_font_size_spinbox') and hasattr(self, 'settings_font_size_spinbox'):
            # シグナルのループを防ぐために、一時的に接続を解除
            self.main_font_size_spinbox.blockSignals(True)
            self.settings_font_size_spinbox.blockSignals(True)

            self.main_font_size_spinbox.setValue(size)
            self.settings_font_size_spinbox.setValue(size)

            # 接続を再開
            self.main_font_size_spinbox.blockSignals(False)
            self.settings_font_size_spinbox.blockSignals(False)

    def _initialize_config_values(self):
        """config.jsonから読み込んだ値をクラス属性に設定"""
        filter_settings = self.user_settings.get("filter_settings", {})

        # 基本設定値
        self.SUBJECT_COMBINATION_COUNT = self._get_setting("SUBJECT_COMBINATION_COUNT", 100)
        self.MORE_SUBJECT_COMBINATION = self._get_setting("MORE_SUBJECT_COMBINATION", 10)
        self.MIN_SUBJECT_COUNT = filter_settings.get("MIN_SUBJECT_COUNT", 2)
        self.MAX_SUBJECT_COUNT = filter_settings.get("MAX_SUBJECT_COUNT", 10)
        self.MIN_SUBJECT_COUNT_UNITS = filter_settings.get("MIN_SUBJECT_COUNT_UNITS", self._get_setting("MIN_SUBJECT_COUNT_UNITS", 2))
        self.MAX_SUBJECT_COUNT_UNITS = filter_settings.get("MAX_SUBJECT_COUNT_UNITS", self._get_setting("MAX_SUBJECT_COUNT_UNITS", 30))

        # 有効/無効フラグ
        self.ACTIVE_MIN_SUBJECT = filter_settings.get("ACTIVE_MIN_SUBJECT", self._get_setting("ACTIVE_MIN_SUBJECT", False))
        self.ACTIVE_MAX_SUBJECT = filter_settings.get("ACTIVE_MAX_SUBJECT", self._get_setting("ACTIVE_MAX_SUBJECT", False))
        self.ACTIVE_FILTER_SUBJECT = filter_settings.get("ACTIVE_FILTER_SUBJECT", self._get_setting("ACTIVE_FILTER_SUBJECT", False))
        self.ACTIVE_FILTER_SUBJECT_AMOUNT = filter_settings.get("ACTIVE_FILTER_SUBJECT_AMOUNT", self._get_setting("ACTIVE_FILTER_SUBJECT_AMOUNT", False))
        self.ACTIVE_MIN_SUBJECT_UNITS = filter_settings.get("ACTIVE_MIN_SUBJECT_UNITS", self._get_setting("ACTIVE_MIN_SUBJECT_UNITS", False))
        self.ACTIVE_MAX_SUBJECT_UNITS = filter_settings.get("ACTIVE_MAX_SUBJECT_UNITS", self._get_setting("ACTIVE_MAX_SUBJECT_UNITS", False))
        self.ACTIVE_FILTER_SUBJECT_UNITS = filter_settings.get("ACTIVE_FILTER_SUBJECT_UNITS", self._get_setting("ACTIVE_FILTER_SUBJECT_UNITS", False))

        # その他の設定
        self.MAX_MEMORY_LIMIT = self.user_settings.get("MAX_MEMORY_LIMIT", self._get_setting("MAX_MEMORY_LIMIT", 1000))
        self.INCLUDE_FIXED = filter_settings.get("INCLUDE_FIXED", True)

        # 初期値
        self.setting_min_subject_count = self.MIN_SUBJECT_COUNT
        self.max_subject_count = self.MAX_SUBJECT_COUNT
        self.min_subject_units = self.MIN_SUBJECT_COUNT_UNITS
        self.max_subject_units = self.MAX_SUBJECT_COUNT_UNITS

        
    def cleanup_display_windows(self):
        """
        display_windowsリストをクリーンアップして、閉じられた(非表示の)ウィンドウを削除し、
        メモリリークを防ぎます。また、ガベージコレクションを強制実行します。
        """
        if hasattr(self, 'display_windows') and self.display_windows:
            windows_to_remove = []
            for w in self.display_windows:
                if not w.isVisible():
                    if hasattr(w, 'window_data'):
                        window_data = w.window_data
                        if 'timetables' in window_data:
                            del window_data['timetables']
                        del w.window_data
                    windows_to_remove.append(w)
            
            for w in windows_to_remove:
                self.display_windows.remove(w)
                w.deleteLater()
                
            gc.collect()

    def create_widgets(self):
        """すべてのウィジェットを作成"""
        self.create_combination_time_tab()
        for year in self.years:
            tab = self.year_tabs[year]
            data = self.year_data[year]
            self.create_year_tab(tab, data["subject_slots_base"], data["fixed_slots"], data["all_slots"], year)
        self.create_settings_tab()




    def _start_rename_tab(self, index):
        """タブの名前変更を開始する"""
        tab_bar = self.profile_tab_widget.tabBar()
        
        # 既存のQLineEditがあれば、先にそれを処理する
        if hasattr(self, 'rename_edit') and self.rename_edit.isVisible():
            self.rename_edit.editingFinished.emit()

        rect = tab_bar.tabRect(index)
        
        self.rename_edit = QLineEdit(tab_bar)
        self.rename_edit.setText(tab_bar.tabText(index))
        self.rename_edit.setGeometry(rect)
        self.rename_edit.setFocus()
        self.rename_edit.selectAll()
        self.rename_edit.show()

        self.rename_edit.editingFinished.connect(lambda: self._finish_rename_tab(index))

    def _finish_rename_tab(self, index):
        """タブの名前変更を完了する"""
        if not hasattr(self, 'rename_edit') or not self.rename_edit.isVisible():
            return
            
        old_name = self.profile_tab_widget.tabText(index)
        new_name = self.rename_edit.text().strip()
        
        self.rename_edit.deleteLater()
        del self.rename_edit

        if old_name == "Default" and new_name != "Default":
             QMessageBox.warning(self, "エラー", "Defaultプロファイルの名前は変更できません。")
             return

        if not new_name or old_name == new_name:
            return

        all_profiles = list(self.year_data[self.years[0]]["saved_profiles"].keys())
        if new_name in all_profiles:
            QMessageBox.warning(self, "名前の重複", "そのプロファイル名はすでに存在します。")
            return
        
        # データ構造のキーを更新
        for year in self.years:
            if old_name in self.year_data[year]["saved_profiles"]:
                self.year_data[year]["saved_profiles"][new_name] = self.year_data[year]["saved_profiles"].pop(old_name)
            if old_name in self.year_data[year].get("profile_lock_settings", {}):
                self.year_data[year]["profile_lock_settings"][new_name] = self.year_data[year]["profile_lock_settings"].pop(old_name)
            if old_name in self.year_data[year].get("profile_all_slots", {}):
                self.year_data[year]["profile_all_slots"][new_name] = self.year_data[year]["profile_all_slots"].pop(old_name)
            if old_name in self.year_data[year].get("profile_ui_states", {}):
                self.year_data[year]["profile_ui_states"][new_name] = self.year_data[year]["profile_ui_states"].pop(old_name)

        if old_name in self.art_subject_selections:
            self.art_subject_selections[new_name] = self.art_subject_selections.pop(old_name)

        self.profile_tab_widget.setTabText(index, new_name)
        
        # アクティブなプロファイル名も更新
        if self.current_profile_name == old_name:
            self.current_profile_name = new_name

    def _show_tab_context_menu(self, position, index):
        """タブの右クリックメニューを表示する"""
        menu = QMenu()
        
        rename_action = menu.addAction("名前の変更")
        duplicate_action = menu.addAction("複製")
        delete_action = menu.addAction("削除")
        
        if self.profile_tab_widget.tabText(index) == "Default":
            delete_action.setEnabled(False)
            
        action = menu.exec(position)
        
        if action == rename_action:
            self._start_rename_tab(index)
        elif action == duplicate_action:
            self._duplicate_profile_tab(index)
        elif action == delete_action:
            self._delete_profile_tab(index)

    def _delete_profile_tab(self, index):
        """プロファイルを削除する"""
        profile_name = self.profile_tab_widget.tabText(index)

        if profile_name == "Default":
            return
            
        if self.user_settings.get("show_delete_profile_warning", True):
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Question)
            msg_box.setWindowTitle("プロファイルの削除")
            msg_box.setText(f"プロファイル「{profile_name}」を削除しますか？\nこの操作は元に戻せません。")
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg_box.setDefaultButton(QMessageBox.No)
            
            cb = QCheckBox("このメッセージを二度と表示しない")
            msg_box.setCheckBox(cb)
            
            reply = msg_box.exec()

            if cb.isChecked():
                self.user_settings["show_delete_profile_warning"] = False
                self._save_user_settings()
        else:
            reply = QMessageBox.Yes

        if reply == QMessageBox.Yes:
            self.profile_tab_widget.removeTab(index)
            
            for year in self.years:
                if profile_name in self.year_data[year]["saved_profiles"]:
                    del self.year_data[year]["saved_profiles"][profile_name]
                if profile_name in self.year_data[year].get("profile_lock_settings", {}):
                    del self.year_data[year]["profile_lock_settings"][profile_name]
                if profile_name in self.year_data[year].get("profile_all_slots", {}):
                    del self.year_data[year]["profile_all_slots"][profile_name]
                if profile_name in self.year_data[year].get("profile_ui_states", {}):
                    del self.year_data[year]["profile_ui_states"][profile_name]

            if profile_name in self.art_subject_selections:
                del self.art_subject_selections[profile_name]

    def _duplicate_profile_tab(self, index):
        """プロファイルを複製する"""
        original_name = self.profile_tab_widget.tabText(index)
        
        base_name = f"{original_name}のコピー"
        new_name = base_name
        i = 1
        all_profiles = list(self.year_data[self.years[0]]["saved_profiles"].keys())
        while new_name in all_profiles:
            i += 1
            new_name = f"{base_name} {i}"

        for year in self.years:
            original_timetable = self.year_data[year]["saved_profiles"].get(original_name, {})
            self.year_data[year]["saved_profiles"][new_name] = original_timetable.copy()
            
            original_lock_settings = self.year_data[year].get("profile_lock_settings", {}).get(original_name, {})
            self.year_data[year].setdefault("profile_lock_settings", {})[new_name] = original_lock_settings.copy()

            original_all_slots = self.year_data[year].get("profile_all_slots", {}).get(original_name, [])
            self.year_data[year].setdefault("profile_all_slots", {})[new_name] = original_all_slots[:]
            
            original_ui_state = self.year_data[year].get("profile_ui_states", {}).get(original_name, {"checked": [], "important": [], "prefixes": {}})
            self.year_data[year]["profile_ui_states"][new_name] = {
                "checked": list(original_ui_state.get("checked", [])),
                "important": list(original_ui_state.get("important", [])),
                "prefixes": original_ui_state.get("prefixes", {}).copy()
            }
        
        original_art_selections = self.art_subject_selections.get(original_name, [])
        self.art_subject_selections[new_name] = original_art_selections[:]

        self._add_profile_tab(new_name, set_current=True)

    def _initialize_profiles(self):
        """アプリケーション起動時のプロファイル初期化"""
        if not self.year_data[self.years[0]]["saved_profiles"]:
            for year in self.years:
                self.year_data[year]["saved_profiles"]["Default"] = {}
            self.art_subject_selections["Default"] = []

        # 保存された順序を読み込み、存在しないプロファイルを削除し、新しいプロファイルを追加する
        all_profile_names = list(self.year_data[self.years[0]]["saved_profiles"].keys())
        saved_order = self.user_settings.get("profile_order", [])
        
        # 存在しないプロファイルをフィルタリングで削除
        ordered_profiles = [p for p in saved_order if p in all_profile_names]
        
        # 新しいプロファイルをリストの最後に追加
        for p_name in all_profile_names:
            if p_name not in ordered_profiles:
                ordered_profiles.append(p_name)
        
        # 修正された順序を保存
        if self.user_settings.get("profile_order") != ordered_profiles:
            self.user_settings["profile_order"] = ordered_profiles
            self._save_user_settings()

        for name in ordered_profiles:
            self._add_profile_tab(name)

        if self.profile_tab_widget.count() > 0:
            # 破壊的な再構築を避けるため_on_profile_tab_changedは呼ばない。
            # 代わりに初期状態だけを直接設定する。
            active_index = self.profile_tab_widget.currentIndex()
            if active_index == -1:
                active_index = 0
                self.profile_tab_widget.setCurrentIndex(active_index)

            active_profile_name = self.profile_tab_widget.tabText(active_index)
            self.current_profile_name = active_profile_name
            for year in self.years:
                self.year_data[year]["active_profile_name"] = active_profile_name
            
            # 初期選択プロファイルのUI状態を読み込む
            self._load_profile_ui_state(active_profile_name)
            self.check_subjects_units()
            self.update_all_highlights()

    def _add_profile_tab(self, profile_name, set_current=False):
        """指定された名前で新しいタブを作成し、中身を生成する"""
        profile_page = QWidget()
        scroll_layout = QVBoxLayout(profile_page)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        page_content_layout = QVBoxLayout(scroll_content)
        scroll_area.setWidget(scroll_content)
        scroll_layout.addWidget(scroll_area)

        self._populate_tab_content(page_content_layout, profile_name, profile_page)

        index = self.profile_tab_widget.addTab(profile_page, profile_name)

        if set_current:
            self.profile_tab_widget.setCurrentIndex(index)

    def _populate_tab_content(self, layout, profile_name, page_widget):
        """指定されたレイアウトに、特定のプロファイルの時間割コンテンツを生成する"""
        top_ui_frame = QFrame()
        top_ui_layout = QHBoxLayout(top_ui_frame)
        
        saved_selections = self.art_subject_selections.get(profile_name, [])
        
        art_combos = []
        for i in range(self.art_subject_count):
            combo = QComboBox()
            combo.setFixedWidth(120)
            if self.selected_art_subject: 
                combo.addItems(self.selected_art_subject)
            if i < len(saved_selections):
                combo.setCurrentText(saved_selections[i])
            combo.currentTextChanged.connect(lambda text, p=profile_name, idx=i: self._on_art_subject_changed(text, p, idx))
            top_ui_layout.addWidget(combo)
            art_combos.append(combo)
        
        top_ui_layout.addStretch(1)
        
        units_label = QLabel()
        units_label.setFont(QFont("Arial", 12, QFont.Bold))
        top_ui_layout.addWidget(units_label, alignment=Qt.AlignCenter)
        page_widget.units_label = units_label
        top_ui_layout.addStretch(1)

        target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"
        icon_path_save = os.path.join(self.base_path, "svgs", "save.svg")
        save_icon = self._create_icon_from_svg_data(icon_path_save, target_color)
        save_all_button = QPushButton(save_icon, "保存")
        file_menu = QMenu(save_all_button)
        export_excel = file_menu.addAction("全体をExcelに保存")
        export_excel.triggered.connect(lambda: self.save_timedate(year_label="全体"))
        save_action = file_menu.addAction("専用形式ファイルに保存")
        save_action.triggered.connect(self.save_dedicated_file)
        save_all_button.setMenu(file_menu)
        top_ui_layout.addWidget(save_all_button)
        
        layout.addWidget(top_ui_frame)

        common_subjects = set()
        timetables_for_profile = {year: self.year_data[year]["saved_profiles"].get(profile_name, {}) for year in self.years}
        all_subjects_in_profile = {subj for tt in timetables_for_profile.values() for subj in tt.values()}
        if all_subjects_in_profile:
            subject_counts = {subj: list(all_subjects_in_profile).count(subj) for subj in all_subjects_in_profile}
            common_subjects = {subj for subj, count in subject_counts.items() if count > 1}

        for year in self.years:
            data = self.year_data[year]
            time_frame = QFrame()
            time_frame.setFrameShape(QFrame.StyledPanel)
            time_layout = QVBoxLayout(time_frame)

            header_frame = QFrame()
            header_layout = QHBoxLayout(header_frame)
            display_year = year.split('_')[-1]
            label = QLabel(display_year)
            label.setFont(QFont("Arial", 12, QFont.Bold))
            header_layout.addWidget(label, alignment=Qt.AlignLeft)
            save_button = QPushButton(save_icon,f"{display_year}をExcelに保存")
            save_button.clicked.connect(lambda _, y=year: self.save_timedate(year_label=y))
            header_layout.addWidget(save_button)
            time_layout.addWidget(header_frame)

            content_frame = QWidget()
            content_frame.setLayout(QHBoxLayout())
            content_frame.layout().setContentsMargins(20, 10, 20, 10)
            
            timetable = data["saved_profiles"].get(profile_name, {})
            self.create_table_slot(content_frame, timetable, data["table_layout"], year, data["all_slots"], common_subjects, is_main_timetable_tab=True, profile_name=profile_name)
            
            time_layout.addWidget(content_frame)
            layout.addWidget(time_frame)

    def _close_profile_tab(self, index):
        """タブのxボタンが押されたときの処理。プロファイルを削除する。"""
        self._delete_profile_tab(index)

    def _add_new_blank_profile(self):
        """'+'ボタンで空白のプロファイルとタブを作成する"""
        base_name = "無題のプロファイル"
        new_name = base_name
        i = 1
        all_profiles = list(self.year_data[self.years[0]]["saved_profiles"].keys())
        while new_name in all_profiles:
            i += 1
            new_name = f"{base_name} {i}"

        for year in self.years:
            self.year_data[year]["saved_profiles"][new_name] = {}
            self.year_data[year]["profile_lock_settings"][new_name] = {}
            all_slots_list = self.year_data[year].get("all_slots", [])
            self.year_data[year].setdefault("profile_all_slots", {})[new_name] = all_slots_list[:]
            self.year_data[year]["profile_ui_states"][new_name] = {"checked": [], "important": [], "prefixes": {}}
        self.art_subject_selections[new_name] = []

        self._add_profile_tab(new_name, set_current=True)

    def _update_timetable_tab_for_profile(self, profile_name):
        """指定されたプロファイルのデータで、そのプロファイルタブの表示を更新する"""
        if not hasattr(self, 'profile_tab_widget'):
            return

        # 指定されたプロファイル名に対応するタブを見つける
        target_tab = None
        for i in range(self.profile_tab_widget.count()):
            if self.profile_tab_widget.tabText(i) == profile_name:
                target_tab = self.profile_tab_widget.widget(i)
                break
        
        if not target_tab:
            return

        for year in self.years:
            data = self.year_data[year]
            timetable = data["saved_profiles"].get(profile_name, {})
            
            # 固定スロットを取得
            config_fixed_slots = data.get("fixed_slots", {})
            user_fixed_slots = data.get("profile_lock_settings", {}).get(profile_name, {})
            
            # この学年の時間割スロットボタンを、対象タブ内から検索して更新
            for slot in data.get("all_slots", []):
                # 対象タブ内からボタンを検索（他のタブのボタンを誤って更新しない）
                button = target_tab.findChild(TimetableSlotButton, f"slot_{year}_{slot}")
                if button and button.is_main_timetable_tab:
                    # 固定スロットを優先し、次にユーザーロック、最後にtimetableから取得
                    subject_name = config_fixed_slots.get(slot, user_fixed_slots.get(slot, timetable.get(slot, "－")))
                    button.setText(subject_name)
                    # Note: 背景色などの更新は update_all_highlights で別途行われる

    def _on_profile_tab_changed_wrapper(self, index):
        """プロファイルタブが切り替わったときに呼ばれるラッパー。並び替えを検出し、元の処理を呼ぶ。"""
        if not hasattr(self, 'profile_tab_widget'): # ウィジェット未初期化ならスキップ
            self._on_profile_tab_changed(index) # メインハンドラは引き続き呼ぶ
            return

        current_order = [self.profile_tab_widget.tabText(i) for i in range(self.profile_tab_widget.count())]
        
        # 以前の順序と比較し、変更があれば保存
        if hasattr(self, '_last_profile_order') and self._last_profile_order != current_order:
            self._save_profile_order()
        
        # 現在の順序を保存して次回比較できるようにする
        self._last_profile_order = current_order[:] # スライスのコピーを作成
        
        # 元の _on_profile_tab_changed を呼び出す
        self._on_profile_tab_changed(index)

    def _save_profile_order(self):
        """現在のプロファイルタブ順をユーザー設定へ保存する。"""
        if not hasattr(self, 'profile_tab_widget'):
            return

        current_order = [self.profile_tab_widget.tabText(i) for i in range(self.profile_tab_widget.count())]
        
        if isinstance(self.user_settings, dict):
            self.user_settings['profile_order'] = current_order
            self._save_user_settings()
        else:
            print("Warning: Could not save profile order, user_settings is not a dictionary.")

    def _on_profile_tab_changed(self, index):
        """プロファイルタブが切り替わったときの処理（UI再構築なし）"""
        if index == -1 or not hasattr(self, 'profile_tab_widget'):
            return
        
        new_profile_name = self.profile_tab_widget.tabText(index)
        if not new_profile_name or new_profile_name == self.current_profile_name:
            return

        # 1. 古いプロファイルのUI状態を保存
        self._save_profile_ui_state(self.current_profile_name)

        # 2. 現在のプロファイル名を更新
        self.current_profile_name = new_profile_name
        for year in self.years:
            self.year_data[year]["active_profile_name"] = new_profile_name

        # 3. UIの再構築は行わず、既存ウィジェットの内容を更新
        self._update_timetable_tab_for_profile(new_profile_name)
        
        # 4. 新しいプロファイルの状態をUIに反映
        self._load_profile_ui_state(new_profile_name)
        self.check_subjects_units()
        self.update_all_highlights()

    def create_combination_time_tab(self):
        """時間割タブを新しいタブUIで作成"""
        if self.combination_time_tab.layout() is not None:
            while (item := self.combination_time_tab.layout().takeAt(0)) is not None:
                if item.widget(): item.widget().deleteLater()
        else:
            self.combination_time_tab.setLayout(QVBoxLayout())
            self.combination_time_tab.layout().setContentsMargins(0, 0, 0, 0)
        
        self.profile_tab_widget = QTabWidget()
        self.profile_tab_widget.setMovable(True)
        self.profile_tab_widget.setTabsClosable(True)
        
        corner_widget = QWidget()
        corner_layout = QHBoxLayout(corner_widget)
        corner_layout.setContentsMargins(2, 2, 2, 2)
        corner_layout.setSpacing(4)
        
        target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"
        add_icon = self._create_icon_from_svg_data(os.path.join(self.base_path, "svgs", "plus.svg"), target_color)
        add_profile_button = QPushButton(add_icon, "")
        add_profile_button.setToolTip("新しいプロファイルを作成")
        add_profile_button.setFixedSize(24, 24)
        add_profile_button.setIconSize(add_profile_button.size() * 0.8)
        corner_layout.addWidget(add_profile_button)
        self.profile_tab_widget.setCornerWidget(corner_widget, Qt.TopRightCorner)

        self.profile_tab_widget.tabBar().installEventFilter(self)
        self.profile_tab_widget.tabCloseRequested.connect(self._close_profile_tab)
        add_profile_button.clicked.connect(self._add_new_blank_profile)
        self.profile_tab_widget.currentChanged.connect(self._on_profile_tab_changed_wrapper)

        self.combination_time_tab.layout().addWidget(self.profile_tab_widget)
        
        self._initialize_profiles()

    def _on_art_subject_changed(self, text, profile_name=None, idx=None):
        """芸術科目の選択が変更された時の処理"""
        # signal lambdaからprofile_nameが渡された場合はそれを使い、なければ現在のUIから取得
        if profile_name:
            active_profile_name = profile_name
        else:
            active_profile_name = self.year_data[self.years[0]].get("active_profile_name", "Default") if self.years else "Default"
        
        selections = [combo.currentText() for combo in self.art_subject_combos]
        
        self.art_subject_selections[active_profile_name] = selections
        self.save_art_subject_selections()

        self.update_all_highlights()
        self.check_subjects_units()

    def _on_timetable_mode_changed(self, button, checked):
        """時間割操作モードが変更されたときの処理"""
        if checked:
            self.current_timetable_mode = self.mode_button_map.get(button)

    def _save_timetable_ui_state(self):
        """時間割タブの現在のUI状態（プロファイル、スクロール位置）を保存する"""
        state = {
            'profile_index': None,
            'scroll_position': None
        }
        if hasattr(self, 'profile_tab_widget'):
            state['profile_index'] = self.profile_tab_widget.currentIndex()
            current_page = self.profile_tab_widget.currentWidget()
            if current_page:
                scroll_area = current_page.findChild(QScrollArea)
                if scroll_area:
                    state['scroll_position'] = scroll_area.verticalScrollBar().value()
        return state

    def _restore_timetable_ui_state(self, state):
        """保存された時間割タブのUI状態を復元する"""
        if not hasattr(self, 'profile_tab_widget') or not state:
            return

        profile_index = state.get('profile_index')
        scroll_position = state.get('scroll_position')

        if profile_index is not None and hasattr(self, 'profile_tab_widget') and self.profile_tab_widget is not None:
            # タブの数がインデックスより大きいことを確認
            if 0 <= profile_index < self.profile_tab_widget.count():
                self.profile_tab_widget.setCurrentIndex(profile_index)
                
                if scroll_position is not None:
                    def restore_scroll():
                        try:
                            # setCurrentIndex後の正しいウィジェットを再取得
                            current_page = self.profile_tab_widget.widget(profile_index)
                            if current_page:
                                scroll_area = current_page.findChild(QScrollArea)
                                if scroll_area:
                                    scroll_area.verticalScrollBar().setValue(scroll_position)
                        except Exception as e:
                            print(f"スクロール位置の復元に失敗しました: {e}")
                    
                    # UIの更新を待つために短い遅延後に実行
                    QTimer.singleShot(50, restore_scroll)

    def create_settings_tab(self):
        """設定タブを作成"""
        self.tab_settings.setLayout(QVBoxLayout())



        # 時間割順序設定
        order_group = QGroupBox("時間割を単位数が多い順にする:")
        order_layout = QHBoxLayout(order_group)

        self.timetable_order_cb = QCheckBox()
        self.timetable_order_cb.setChecked(self._get_setting("TIMETABLE_ORDER",True))
        self.timetable_order_cb.stateChanged.connect(self.save_app_settings) # 保存処理を接続
        # ユーザー設定ファイルにも即時反映する
        self.timetable_order_cb.stateChanged.connect(self._save_user_settings)
        order_layout.addWidget(self.timetable_order_cb)

        self.tab_settings.layout().addWidget(order_group)

        # フォントサイズ設定
        font_size_group = QGroupBox("フォントサイズ")
        bottom_layout = QHBoxLayout(font_size_group)

        self.settings_font_size_spinbox = QSpinBox(value=self._get_setting("APP_FONT_SIZE",12))
        self.settings_font_size_spinbox.setRange(8, 24)
        self.settings_font_size_spinbox.valueChanged.connect(self.change_font_size)
        self.settings_font_size_spinbox.valueChanged.connect(self.save_app_settings)
        self.settings_font_size_spinbox.valueChanged.connect(self._save_user_settings)
        bottom_layout.addWidget(self.settings_font_size_spinbox)

        self.tab_settings.layout().addWidget(font_size_group)

        # チュートリアル設定
        tutorial_group = QGroupBox("チュートリアル")
        tutorial_layout = QHBoxLayout(tutorial_group)
        self.run_tutorial_cb = QCheckBox("起動時にチュートリアルを確認する")
        self.run_tutorial_cb.setChecked(self._get_setting("RUN_TUTORIAL_ON_STARTUP", False))
        self.run_tutorial_cb.stateChanged.connect(self.save_app_settings)
        self.run_tutorial_cb.stateChanged.connect(self._save_user_settings)
        tutorial_layout.addWidget(self.run_tutorial_cb)
        self.tab_settings.layout().addWidget(tutorial_group)

        # 時間割生成制限設定
        memory_group = QGroupBox("時間割生成制限 (個):")
        memory_layout = QHBoxLayout(memory_group)

        self.memory_limit_spinbox = CustomStepSpinBox(arrow_step=100, wheel_step=10)
        self.memory_limit_spinbox.setRange(100, 7000)
        self.memory_limit_spinbox.setValue(self.max_memory_limit)
        memory_layout.addWidget(self.memory_limit_spinbox)

        self.memory_limit_slider = QSlider(Qt.Horizontal)
        self.memory_limit_slider.setRange(100, 7000)
        self.memory_limit_slider.setValue(self.max_memory_limit)
        self.memory_limit_slider.setSingleStep(100)
        self.memory_limit_slider.setPageStep(100)
        self.memory_limit_slider.setTickPosition(QSlider.TicksBelow)
        self.memory_limit_slider.setTickInterval(1000)
        self.memory_limit_slider.valueChanged.connect(self.save_app_settings) # 保存処理を接続
        self.memory_limit_slider.valueChanged.connect(self._save_user_settings)
        memory_layout.addWidget(self.memory_limit_slider)

        self.memory_limit_slider.valueChanged.connect(self.memory_limit_spinbox.setValue)
        self.memory_limit_spinbox.valueChanged.connect(self.memory_limit_slider.setValue)
        self.memory_limit_spinbox.valueChanged.connect(self.save_app_settings)
        self.memory_limit_spinbox.valueChanged.connect(self._save_user_settings)

        self.tab_settings.layout().addWidget(memory_group)

        # 既存取得単位数設定
        first_year_units_group = QGroupBox("既存の取得単位数:")
        first_year_units_layout = QHBoxLayout(first_year_units_group)

        self.first_year_units_spinbox = QSpinBox()
        self.first_year_units_spinbox.setRange(0, 100)
        # base_units_valueから初期化（YEARS_SUBJECTS_UNITS_*から読み込まれた値）
        self.first_year_units_spinbox.setValue(self.base_units_value)
        # 値変更時にbase_units_valueを更新し、単位数を再計算
        self.first_year_units_spinbox.valueChanged.connect(self._on_base_units_changed)
        self.first_year_units_spinbox.valueChanged.connect(self.save_app_settings)
        self.first_year_units_spinbox.valueChanged.connect(self._save_user_settings)
        first_year_units_layout.addWidget(self.first_year_units_spinbox)

        self.tab_settings.layout().addWidget(first_year_units_group)

        # スペーサー
        self.tab_settings.layout().addStretch()

    def save_app_settings(self):
        """アプリケーション設定をconfig.jsonに保存する"""
        # 実際の値は_save_user_settings内でウィジェットから読み取る
        self.config["TIMETABLE_ORDER"] = self.timetable_order_cb.isChecked()
        self.max_memory_limit = self.memory_limit_slider.value()
        self.config["APP_FONT_SIZE"] = self.main_font_size_spinbox.value()
        
        # YEARS_SUBJECTS_UNITS_*を更新
        years_hierarchy = self.config.get("YEARS_HIERARCHY", {})
        if years_hierarchy:
            top_level_key = list(years_hierarchy.keys())[0]
            config_key = f"YEARS_SUBJECTS_UNITS_{top_level_key}"
            self.config[config_key] = self.first_year_units_spinbox.value()
        
        self._save_user_settings()

    def _on_base_units_changed(self, value):
        """既存の取得単位数スピンボックスの値が変更された時の処理"""
        self.base_units_value = value
        
        # YEARS_SUBJECTS_UNITS_*も更新
        years_hierarchy = self.config.get("YEARS_HIERARCHY", {})
        if years_hierarchy:
            top_level_key = list(years_hierarchy.keys())[0]
            config_key = f"YEARS_SUBJECTS_UNITS_{top_level_key}"
            self.config[config_key] = value
            self.user_settings[config_key] = value
        
        # 単位数を再計算
        self.check_subjects_units()

        # 次回起動時に復元できるよう同じ設定をuser config.jsonへ保存
        try:
            self.user_settings['APP_FONT_SIZE'] = self._get_setting('APP_FONT_SIZE')
            self.user_settings['MAX_MEMORY_LIMIT'] = self.max_memory_limit

            # filter_settingsをマージ
            filter_settings = self.user_settings.get('filter_settings', {}) if isinstance(self.user_settings, dict) else {}
            # 既存のフィルター設定を維持し、INCLUDE_FIXEDも保持する
            filter_settings['INCLUDE_FIXED'] = self.INCLUDE_FIXED
            self.user_settings['filter_settings'] = filter_settings

            # 最後に開いたconfigとyears_hierarchy
            self.user_settings['last_opened_config'] = str(self.config_path) if hasattr(self, 'config_path') else self.user_settings.get('last_opened_config')
            self.user_settings['years_hierarchy'] = self.selected_department_path_key

            # .txtに表示する明示設定も保存（三つのキーを必ず持たせる）
            self.user_settings['TIMETABLE_ORDER'] = bool(self._get_setting('TIMETABLE_ORDER', False))
            self.user_settings['RUN_TUTORIAL_ON_STARTUP'] = bool(self._get_setting('RUN_TUTORIAL_ON_STARTUP', False))
            self.user_settings[config_key] = self._get_setting(config_key)

            # 書き戻す
            path = self._get_user_settings_path()
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(self.user_settings, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Failed to persist settings into config.json: {e}")
    
    def save_art_subject_selections(self):
        """芸術科目選択をconfig.jsonに保存する"""
        self.config["PROFILE_ART_SELECTIONS"] = self.art_subject_selections
        self.save_config_file()
    
    def save_filter_settings(self):
        """フィルター設定をconfig.jsonに保存する"""
        if not hasattr(self, 'setting_window'): # ウィンドウ生成済みか確認
            return
        
        # 保存前に属性値を更新
        self.ACTIVE_FILTER_SUBJECT = self.active_filter_subject_cb.isChecked()
        self.ACTIVE_FILTER_SUBJECT_AMOUNT = self.active_filter_subject_amount_cb.isChecked()
        self.ACTIVE_MIN_SUBJECT = self.active_min_subject_cb.isChecked()
        self.setting_min_subject_count = self.min_subject_spinbox.value()
        self.ACTIVE_MAX_SUBJECT = self.active_max_subject_cb.isChecked()
        self.max_subject_count = self.max_subject_spinbox.value()
        self.ACTIVE_FILTER_SUBJECT_UNITS = self.active_filter_subject_units_cb.isChecked()
        self.ACTIVE_MIN_SUBJECT_UNITS = self.active_min_units_cb.isChecked()
        self.min_subject_units = self.min_units_spinbox.value()
        self.ACTIVE_MAX_SUBJECT_UNITS = self.active_max_units_cb.isChecked()
        self.max_subject_units = self.max_units_spinbox.value()
        self.INCLUDE_FIXED = self.include_fixed_cb.isChecked() if hasattr(self, 'include_fixed_cb') else True

        self._save_user_settings()

    def save_config_file(self):
        """config.jsonへの自動保存は廃止しました（無効化）。設定は専用ファイルで保存してください。"""
        # ユーザー設定によりconfig.jsonへは書き込まない
        return

    def save_dedicated_file(self):
        """時間割くん専用ファイルに、現在の設定と時間割を保存する（UIオブジェクトは含めない）"""
        if self.is_tutorial_running:
            self.tutorial_signal.emit({'action': 'menu_clicked'})
            QTimer.singleShot(100, lambda: self.tutorial_signal.emit({'action': 'dialog_opened'}))
            return

        # 1. UI操作 (ファイルダイアログ)
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "保存（専用形式）", "", "時間割くん専用 (*.tm.json);;JSON (*.json)", options=options)
        if not file_path:
            return

        # 2. ファイル書き込み処理を FileIOManager に委譲
        try:
            self.file_io_manager.save_dedicated_file(
                file_path=file_path,
                year_data=self.year_data,
                config=self.config,
                art_subject_selections=self.art_subject_selections,
                active_profile_name=self.get_active_profile_name()
            )
            # ちらつきを抑えるため、メッセージボックス前後で更新を止めてイベントを処理
            try:
                self.setUpdatesEnabled(False)
            except Exception:
                pass
            QApplication.processEvents()
            QMessageBox.information(self, "保存完了", f"専用ファイルを保存しました:\n{file_path}")
            try:
                self.setUpdatesEnabled(True)
            except Exception:
                pass
            QApplication.processEvents()
        except Exception as e:
            QMessageBox.warning(self, "保存エラー", f"保存に失敗しました:\n{e}")

    def load_dedicated_file(self):
        """専用ファイルを開き、設定と時間割を復元する"""
        if self.is_tutorial_running:
            self.tutorial_signal.emit({'action': 'menu_clicked'})
            QTimer.singleShot(100, lambda: self.tutorial_signal.emit({'action': 'dialog_opened'}))
            return

        # 1. UI操作 (ファイルダイアログ)
        options = QFileDialog.Options()
        file_dialog = QFileDialog(self, "専用形式ファイルを開く", "", "時間割くん専用 (*.tm.json);;JSON (*.json)", options=options)
        file_dialog.setProxyModel(ExcludeFileProxyModel(["config.json"], file_dialog))
        icon_path = os.path.join(self.base_path, "時間割くんアイコン.ico")
        if os.path.exists(icon_path):
            file_dialog.setWindowIcon(QIcon(str(icon_path)))
        
        if file_dialog.exec():
            file_path = file_dialog.selectedFiles()[0] if file_dialog.selectedFiles() else None
        else:
            file_path = None
        if not file_path:
            return

        # 2. ファイル読み込みを FileIOManager に委譲
        try:
            payload = self.file_io_manager.load_dedicated_file(file_path)
        except Exception as e:
            try:
                self.setUpdatesEnabled(False)
                QApplication.processEvents()
                QMessageBox.warning(self, "読み込みエラー", f"ファイルの読み込みに失敗しました:\n{e}")
            finally:
                self.setUpdatesEnabled(True)
                QApplication.processEvents()
            return
        
        # 3. 学校名検証を FileIOManager に委譲
        current_school_name = self.config.get("GENERAL_SETTINGS", {}).get("SCHOOL_NAME")
        is_match, payload_school, current_school = self.file_io_manager.validate_dedicated_file_school(
            payload, current_school_name
        )
        if not is_match:
            try:
                self.setUpdatesEnabled(False)
                QApplication.processEvents()
                QMessageBox.warning(self, "読み込みエラー",
                                    f"現在開いている設定ファイルと学校名が異なります。\n\n"
                                    f"現在の学校: {current_school}\n"
                                    f"ファイル内の学校: {payload_school}")
            finally:
                self.setUpdatesEnabled(True)
                QApplication.processEvents()
            return

        # 基本構造を復元
        self.config = payload.get("config", self.config)
        # 読み込みデータのyear_dataは直列化可能な値だけを持つため慎重に統合する
        loaded_years = payload.get("year_data", {})
        if loaded_years:
            # 読み込んだ直列化可能キーを既存year_dataへマージ
            for year, pdata in loaded_years.items():
                if year not in self.year_data:
                    self.year_data[year] = {}
                # 既知の直列化可能キーをコピー
                for k, v in pdata.items():
                    self.year_data[year][k] = v

        self.art_subject_selections = payload.get("art_subject_selections", self.art_subject_selections)
        # Excel読み込み時と同様に最小限のUI更新でちらつきを抑える
        try:
            # 先に読み込みデータを内部構造へ反映
            active_profile_name = payload.get('active_profile', self.get_active_profile_name())
            # 全体再構築を避け、保存プロファイルなど直列化部分だけ更新
            for year in self.years:
                pdata = self.year_data.get(year, {})
                loaded = loaded_years.get(year, {})
                if 'saved_profiles' in loaded:
                    pdata['saved_profiles'] = loaded.get('saved_profiles', pdata.get('saved_profiles', {}))
                if 'save_position' in loaded:
                    pdata['save_position'] = loaded.get('save_position', pdata.get('save_position', {}))
                if 'profile_lock_settings' in loaded:
                    pdata['profile_lock_settings'] = loaded.get('profile_lock_settings', pdata.get('profile_lock_settings', {}))
                if 'profile_all_slots' in loaded:
                    pdata['profile_all_slots'] = loaded.get('profile_all_slots', pdata.get('profile_all_slots', {}))
                if 'all_slots' in loaded:
                    pdata['all_slots'] = loaded.get('all_slots', pdata.get('all_slots', []))
                self.year_data[year] = pdata

            # load_timedateと同様にチェック状態、時間割タブ、単位計算を更新
            for year in self.years:
                active = self.year_data[year].get('active_profile_name', active_profile_name)
                subjects_in_timetable = set(self.year_data[year].get('saved_profiles', {}).get(active, {}).values())
                for frame in self.year_data[year].get('subject_frames', []):
                    if hasattr(frame, 'subject_cb'):
                        frame.subject_cb.setChecked(frame.subject_name in subjects_in_timetable)

            # 必要箇所のみ更新
            self.create_combination_time_tab()

            # 読み込んだファイルのアクティブタブへ切り替える
            active_profile_name = payload.get('active_profile', self.get_active_profile_name())
            for i in range(self.profile_tab_widget.count()):
                if self.profile_tab_widget.tabText(i) == active_profile_name:
                    self.profile_tab_widget.setCurrentIndex(i)
                    break

            self.check_subjects_units()
            QApplication.processEvents()
            QMessageBox.information(self, "読み込み完了", "専用ファイルを読み込み、環境を復元しました。")
            self._has_active_combinations = True
        except Exception as e:
            QMessageBox.warning(self, "再構築エラー", f"UIの再構築に失敗しました:\n{e}")

    def update_filter_state(self, state, var_name):
        """フィルタリング設定のON/OFF状態を更新する"""
        value = bool(state)
        setattr(self, var_name, value)
        if var_name == "INCLUDE_FIXED":
            self.include_fixed = value
        self.save_filter_settings()

    def create_setting_window(self):
        """設定ウィンドウを作成"""
        self.setting_window = QMainWindow(self)
        self.setting_window.setWindowTitle('絞り込み設定')
        self.setting_window.resize(450, 400)
        
        central_widget = QWidget()
        self.setting_window.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # 絞り込み有効/無効チェックボックス
        top_row_frame = QFrame()
        top_row_layout = QHBoxLayout(top_row_frame)
        top_row_layout.setContentsMargins(0,0,0,0)

        self.active_filter_subject_cb = QCheckBox("絞り込み有効化")
        self.active_filter_subject_cb.setChecked(self.ACTIVE_FILTER_SUBJECT)
        self.active_filter_subject_cb.stateChanged.connect(self.setting_check_system)
        self.active_filter_subject_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_FILTER_SUBJECT"))
        top_row_layout.addWidget(self.active_filter_subject_cb)
        
        top_row_layout.addStretch()
        main_layout.addWidget(top_row_frame)
        
        # 教科数絞り込み設定
        self.active_filter_subject_amount_cb = QCheckBox("教科数絞り込み有効化")
        self.active_filter_subject_amount_cb.setChecked(self.ACTIVE_FILTER_SUBJECT_AMOUNT)
        self.active_filter_subject_amount_cb.stateChanged.connect(self.setting_check_system)
        self.active_filter_subject_amount_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_FILTER_SUBJECT_AMOUNT"))
        main_layout.addWidget(self.active_filter_subject_amount_cb)
        
        # 最小教科数設定
        min_subject_frame = QFrame()
        min_subject_layout = QHBoxLayout(min_subject_frame)
        
        min_subject_label = QLabel("最小教科数:")
        min_subject_layout.addWidget(min_subject_label)
        
        self.min_subject_spinbox = QSpinBox()
        self.min_subject_spinbox.setRange(0, 30)
        self.min_subject_spinbox.setValue(self.setting_min_subject_count)
        min_subject_layout.addWidget(self.min_subject_spinbox)

        self.min_subject_slider = QSlider(Qt.Horizontal)
        self.min_subject_slider.setRange(0, 30)
        self.min_subject_slider.setValue(self.setting_min_subject_count)
        self.min_subject_slider.setTickPosition(QSlider.TicksBelow)
        self.min_subject_slider.setTickInterval(5)
        self.min_subject_slider.valueChanged.connect(self.sync_min_max_subject_count)
        min_subject_layout.addWidget(self.min_subject_slider)

        self.min_subject_slider.valueChanged.connect(self.min_subject_spinbox.setValue)
        self.min_subject_spinbox.valueChanged.connect(self.min_subject_slider.setValue)
        
        self.active_min_subject_cb = QCheckBox("有効化")
        self.active_min_subject_cb.setChecked(self.ACTIVE_MIN_SUBJECT)
        self.active_min_subject_cb.stateChanged.connect(self.setting_check_system)
        self.active_min_subject_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_MIN_SUBJECT"))
        min_subject_layout.addWidget(self.active_min_subject_cb)
        
        main_layout.addWidget(min_subject_frame)
        
        # 最大教科数設定
        max_subject_frame = QFrame()
        max_subject_layout = QHBoxLayout(max_subject_frame)
        
        max_subject_label = QLabel("最大教科数:")
        max_subject_layout.addWidget(max_subject_label)
        
        self.max_subject_spinbox = QSpinBox()
        self.max_subject_spinbox.setRange(0, 30)
        self.max_subject_spinbox.setValue(self.max_subject_count)
        max_subject_layout.addWidget(self.max_subject_spinbox)

        self.max_subject_slider = QSlider(Qt.Horizontal)
        self.max_subject_slider.setRange(0, 30)
        self.max_subject_slider.setValue(self.max_subject_count)
        self.max_subject_slider.setTickPosition(QSlider.TicksBelow)
        self.max_subject_slider.setTickInterval(5)
        self.max_subject_slider.valueChanged.connect(self.sync_min_max_subject_count)
        max_subject_layout.addWidget(self.max_subject_slider)

        self.max_subject_slider.valueChanged.connect(self.max_subject_spinbox.setValue)
        self.max_subject_spinbox.valueChanged.connect(self.max_subject_slider.setValue)
        
        self.active_max_subject_cb = QCheckBox("有効化")
        self.active_max_subject_cb.setChecked(self.ACTIVE_MAX_SUBJECT)
        self.active_max_subject_cb.stateChanged.connect(self.setting_check_system)
        self.active_max_subject_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_MAX_SUBJECT"))
        max_subject_layout.addWidget(self.active_max_subject_cb)
        
        main_layout.addWidget(max_subject_frame)
        
        # 単位数絞り込み設定
        self.active_filter_subject_units_cb = QCheckBox("単位数絞り込み有効化")
        self.active_filter_subject_units_cb.setChecked(self.ACTIVE_FILTER_SUBJECT_UNITS)
        self.active_filter_subject_units_cb.stateChanged.connect(self.setting_check_system)
        self.active_filter_subject_units_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_FILTER_SUBJECT_UNITS"))
        main_layout.addWidget(self.active_filter_subject_units_cb)
        
        # 最小単位数設定
        min_units_frame = QFrame()
        min_units_layout = QHBoxLayout(min_units_frame)
        
        min_units_label = QLabel("最小単位数:")
        min_units_layout.addWidget(min_units_label)
        
        self.min_units_spinbox = QSpinBox()
        self.min_units_spinbox.setRange(0, 30)
        self.min_units_spinbox.setValue(self.min_subject_units)
        min_units_layout.addWidget(self.min_units_spinbox)

        self.min_units_slider = QSlider(Qt.Horizontal)
        self.min_units_slider.setRange(0, 30)
        self.min_units_slider.setValue(self.min_subject_units)
        self.min_units_slider.setTickPosition(QSlider.TicksBelow)
        self.min_units_slider.setTickInterval(5)
        self.min_units_slider.valueChanged.connect(self.sync_min_max_units)
        min_units_layout.addWidget(self.min_units_slider)

        self.min_units_slider.valueChanged.connect(self.min_units_spinbox.setValue)
        self.min_units_spinbox.valueChanged.connect(self.min_units_slider.setValue)
        
        self.active_min_units_cb = QCheckBox("有効化")
        self.active_min_units_cb.setChecked(self.ACTIVE_MIN_SUBJECT_UNITS)
        self.active_min_units_cb.stateChanged.connect(self.setting_check_system)
        self.active_min_units_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_MIN_SUBJECT_UNITS"))
        min_units_layout.addWidget(self.active_min_units_cb)
        
        main_layout.addWidget(min_units_frame)
        
        # 最大単位数設定
        max_units_frame = QFrame()
        max_units_layout = QHBoxLayout(max_units_frame)
        
        max_units_label = QLabel("最大単位数:")
        max_units_layout.addWidget(max_units_label)
        
        self.max_units_spinbox = QSpinBox()
        self.max_units_spinbox.setRange(0, 30)
        self.max_units_spinbox.setValue(self.max_subject_units)
        max_units_layout.addWidget(self.max_units_spinbox)

        self.max_units_slider = QSlider(Qt.Horizontal)
        self.max_units_slider.setRange(0, 30)
        self.max_units_slider.setValue(self.max_subject_units)
        self.max_units_slider.setTickPosition(QSlider.TicksBelow)
        self.max_units_slider.setTickInterval(5)
        self.max_units_slider.valueChanged.connect(self.sync_min_max_units)
        max_units_layout.addWidget(self.max_units_slider)

        self.max_units_slider.valueChanged.connect(self.max_units_spinbox.setValue)
        self.max_units_spinbox.valueChanged.connect(self.max_units_slider.setValue)
        
        self.active_max_units_cb = QCheckBox("有効化")
        self.active_max_units_cb.setChecked(self.ACTIVE_MAX_SUBJECT_UNITS)
        self.active_max_units_cb.stateChanged.connect(lambda state: self.update_filter_state(state, "ACTIVE_MAX_SUBJECT_UNITS"))
        max_units_layout.addWidget(self.active_max_units_cb)
        
        main_layout.addWidget(max_units_frame)
        
        # 閉じるボタン
        close_button = QPushButton("閉じる")
        close_button.clicked.connect(self.setting_window.close)
        main_layout.addWidget(close_button)
        
        # 初期状態を設定
        self.setting_check_system()
        self.setting_window.show()
    
    def _on_subject_check(self, checked):
        """科目のチェックボックスが変更された時の処理"""
        cb = self.sender()
        # parent_frameはcreate_year_tabで設定されるカスタム属性
        frame = cb.parent_frame if hasattr(cb, 'parent_frame') else None

        if not frame or not hasattr(frame, 'subject_name'):
            return

        subject = frame.subject_name
        year_label = frame.year_label

        if not year_label or year_label not in self.year_data:
            return

        data = self.year_data[year_label]
        prefix_vars = data["prefix_vars"]

        if subject not in prefix_vars:
            return

        self._block_prefix_sync = True
        try:
            for prefix_cb in prefix_vars[subject].values():
                prefix_cb.blockSignals(True)
                prefix_cb.setChecked(checked)
                prefix_cb.blockSignals(False)

            data["check_vars"][subject] = checked
            frame.important_cb.setEnabled(checked)
            if not checked:
                frame.important_cb.setChecked(False)
        finally:
            self._block_prefix_sync = False

    def _on_prefix_check(self, checked):
        """特定の枠のチェックボックスが変更された時の処理"""
        if hasattr(self, '_block_prefix_sync') and self._block_prefix_sync:
            return

        prefix_cb = self.sender()
        frame = prefix_cb.parent_frame

        if not frame or not hasattr(frame, 'subject_name'):
            return

        subject = frame.subject_name
        year_label = frame.year_label

        if not year_label or year_label not in self.year_data:
            return

        data = self.year_data[year_label]
        prefix_vars = data["prefix_vars"]

        if subject not in prefix_vars:
            return

        all_checked = all(cb.isChecked() for cb in prefix_vars[subject].values())
        any_checked = any(cb.isChecked() for cb in prefix_vars[subject].values())

        frame.subject_cb.blockSignals(True)
        frame.subject_cb.setChecked(all_checked)
        frame.subject_cb.blockSignals(False)

        frame.important_cb.setEnabled(any_checked)
        if not any_checked:
            frame.important_cb.setChecked(False)

        data["check_vars"][subject] = all_checked

    def _on_important_check(self, state, subject, year_label):
        """必含チェックボックスが変更された時の処理"""
        if year_label not in self.year_data:
            return
        data = self.year_data[year_label]
        data["important_vars"][subject] = bool(state)

        # 同じ教科の選択チェックボックスも自動的にチェックする
        if state and not data["check_vars"].get(subject):
            data["check_vars"].update({subject: True})
            # UI更新のためにチェックボックスの状態を変更
            tab = self.year_tabs.get(year_label)
            if tab:
                scroll_area = tab.findChild(QScrollArea)
                if scroll_area:
                    scroll_content = scroll_area.widget()
                    if scroll_content:
                        for frame in scroll_content.findChildren(QFrame):
                            if hasattr(frame, 'subject_name') and frame.subject_name == subject:
                                if hasattr(frame, 'subject_cb'):
                                    frame.subject_cb.setChecked(True)
                                break

    def sync_min_max_subject_count(self, value):
        """最小教科数と最大教科数の値を同期"""
        sender = self.sender()
        
        if sender == self.min_subject_slider:
            self.setting_min_subject_count = value
            if self.active_max_subject_cb.isChecked() and value > self.max_subject_slider.value():
                self.max_subject_slider.setValue(value)
        elif sender == self.max_subject_slider:
            self.max_subject_count = value
            if self.active_min_subject_cb.isChecked() and value < self.min_subject_slider.value():
                self.min_subject_slider.setValue(value)
    
    def sync_min_max_units(self, value):
        """最小単位数と最大単位数の値を同期"""
        sender = self.sender()
        
        if sender == self.min_units_slider:
            self.min_subject_units = value
            if self.active_max_units_cb.isChecked() and value > self.max_units_slider.value():
                self.max_units_slider.setValue(value)
        elif sender == self.max_units_slider:
            self.max_subject_units = value
            if self.active_min_units_cb.isChecked() and value < self.min_units_slider.value():
                self.min_units_slider.setValue(value)
    
    def setting_check_system(self):
        """設定チェックシステム - 絞り込み条件の有効/無効を管理"""
        # 絞り込み全体の有効/無効状態
        is_filter_active = self.active_filter_subject_cb.isChecked()
        
        # 教科数絞り込み関連ウィジェットの状態更新
        is_amount_filter_active = is_filter_active and self.active_filter_subject_amount_cb.isChecked()
        self.active_filter_subject_amount_cb.setEnabled(is_filter_active)
        
        # 最小教科数設定
        self.active_min_subject_cb.setEnabled(is_amount_filter_active)
        is_min_subject_enabled = is_amount_filter_active and self.active_min_subject_cb.isChecked()
        self.min_subject_slider.setEnabled(is_min_subject_enabled)
        if hasattr(self, 'min_subject_spinbox'):
            self.min_subject_spinbox.setEnabled(is_min_subject_enabled)
        
        # 最大教科数設定
        self.active_max_subject_cb.setEnabled(is_amount_filter_active)
        is_max_subject_enabled = is_amount_filter_active and self.active_max_subject_cb.isChecked()
        self.max_subject_slider.setEnabled(is_max_subject_enabled)
        if hasattr(self, 'max_subject_spinbox'):
            self.max_subject_spinbox.setEnabled(is_max_subject_enabled)
        
        # 単位数絞り込み関連ウィジェットの状態更新
        is_units_filter_active = is_filter_active and self.active_filter_subject_units_cb.isChecked()
        self.active_filter_subject_units_cb.setEnabled(is_filter_active)
        
        # 最小単位数設定
        self.active_min_units_cb.setEnabled(is_units_filter_active)
        is_min_units_enabled = is_units_filter_active and self.active_min_units_cb.isChecked()
        self.min_units_slider.setEnabled(is_min_units_enabled)
        if hasattr(self, 'min_units_spinbox'):
            self.min_units_spinbox.setEnabled(is_min_units_enabled)
        
        # 最大単位数設定
        self.active_max_units_cb.setEnabled(is_units_filter_active)
        is_max_units_enabled = is_units_filter_active and self.active_max_units_cb.isChecked()
        self.max_units_slider.setEnabled(is_max_units_enabled)
        if hasattr(self, 'max_units_spinbox'):
            self.max_units_spinbox.setEnabled(is_max_units_enabled)
    
    def _get_requirement_info(self, subject, year_label):
        """必須科目情報を取得（データ構造の多様性に対応）"""
        default_color = self._get_setting("REQUIRED_SUBJECTS_COLOR", "#FF0000")

        def find_subject_in_req(req_data):
            """要件データ（リストまたは辞書）内で教科を検索する"""
            if isinstance(req_data, list):
                if subject in req_data:
                    return {"mark": "★", "color": default_color}
            elif isinstance(req_data, dict):
                for req_info in req_data.values():
                    color = req_info.get("color", default_color)
                    if "subjects" in req_info and subject in req_info["subjects"]:
                        return {"mark": "★", "color": color}
                    if "conditions" in req_info:
                        for condition in req_info["conditions"]:
                            if "subjects" in condition and subject in condition["subjects"]:
                                return {"mark": "★", "color": color}
            return None

        final_req_info = {}

        # 1. 全学年共通の必須科目を確認 (REQUIRED_SUBJECTS_ALL)
        req_data_all = self.config.get("REQUIRED_SUBJECTS_ALL", {})
        req_info_all = find_subject_in_req(req_data_all)
        if req_info_all:
            final_req_info.update(req_info_all)
        
        # 2. 階層的な必須科目を確認 (例: REQUIRED_SUBJECTS_全日制)
        parts = year_label.split('_')
        department_label = []
        for y in parts:
            if y: # 学科部分があることを確認
                department_label.append(y)
                label = '_'.join(department_label)
                department_req_key = f"REQUIRED_SUBJECTS_{label}"
                department_req = self.config.get(department_req_key, {})
                department_req_info = find_subject_in_req(department_req)
                if department_req_info:
                    final_req_info.update(department_req_info)

        return final_req_info if final_req_info else None    

    def create_table_slot(self, parent, timetable, table_layout, year_label, all_slots, common_subjects=None, is_main_timetable_tab=False, profile_name=None):
        """時間割表を作成"""
        if not parent.layout():
            parent.setLayout(QVBoxLayout())
        
        table_frame = QFrame(parent)
        table_layout_main = QVBoxLayout(table_frame)
        
        for row in table_layout:
            table_row = QFrame()
            row_layout = QHBoxLayout(table_row)
            row_layout.setSpacing(1)
            
            for slot in row:
                cell_frame = QFrame()
                cell_layout = QVBoxLayout(cell_frame)
                cell_layout.setSpacing(1)
                
                if slot:
                    slot_label = QLabel(slot)
                    slot_label.setFont(QFont("Arial", 6))
                    slot_label.setStyleSheet("color: gray;")
                    cell_layout.addWidget(slot_label, alignment=Qt.AlignTop | Qt.AlignLeft)
                
                # 固定枠の処理
                config_fixed_slots = self.year_data[year_label]["fixed_slots"]
                
                if is_main_timetable_tab and profile_name:
                    user_fixed_slots = self.year_data[year_label].get("profile_lock_settings", {}).get(profile_name, {})
                    current_profile_all_slots = self.year_data[year_label].get("profile_all_slots", {}).get(profile_name, [])
                else:
                    user_fixed_slots = self.get_current_profile_lock_settings(year_label)
                    current_profile_all_slots = self.get_current_profile_all_slots(year_label)

                is_config_fixed = slot in config_fixed_slots
                is_user_locked = slot in user_fixed_slots
                is_excluded = slot not in current_profile_all_slots
                
                subject_name = config_fixed_slots.get(slot, user_fixed_slots.get(slot, timetable.get(slot, "－")))
                
                # セルボタン作成
                cell_button = TimetableSlotButton(subject_name, slot, year_label, self, is_main_timetable_tab)
                cell_button.setFixedSize(120, 30)
                # オブジェクト名を設定して、後からページ内検索できるようにする
                try:
                    cell_button.setObjectName(f"slot_{year_label}_{slot}")
                except Exception:
                    pass
                
                # 背景色の決定とスタイルシートの設定
                bg_color_str = ""
                if is_config_fixed:
                    # ボタンは後で無効化されるため、見た目はテーマに任せる。
                    pass
                elif is_user_locked:
                    # ユーザーロック枠は青系で強調
                    highlight_color = QApplication.palette().color(QPalette.Highlight)
                    highlight_color.setAlpha(80)
                    bg_color_str = f"background-color: {highlight_color.name(QColor.HexArgb)};"
                elif is_excluded:
                    # 除外枠は赤系で強調
                    bg_color_str = "background-color: rgba(255, 0, 0, 70);"
                elif common_subjects and subject_name in common_subjects and subject_name not in config_fixed_slots.values() and subject_name not in user_fixed_slots.values():
                    # 共通教科は黄系で強調
                    tooltip_color = QApplication.palette().color(QPalette.ToolTipBase)
                    tooltip_text_color = QApplication.palette().color(QPalette.ToolTipText)
                    bg_color_str = f"background-color: {tooltip_color.name()}; color: {tooltip_text_color.name()};"
                
                if bg_color_str:
                    cell_button.setStyleSheet(bg_color_str)

                # クリック動作はカスタムボタン内で処理される

                # configで固定された枠は操作不可にする
                if is_config_fixed:
                    cell_button.setEnabled(False)
                
                cell_layout.addWidget(cell_button, alignment=Qt.AlignBottom)
                setattr(self, f"slot_{year_label}_{slot}", cell_button)
                
                row_layout.addWidget(cell_frame)
            
            table_layout_main.addWidget(table_row)
        
        parent.layout().addWidget(table_frame)
    
    def toggle_slot(self, slot, year_label):
        """（左クリック）枠を計算対象から除外、または復帰させる"""
        try:
            data = self.year_data[year_label]
            if slot in data["fixed_slots"]:
                return

            # 青色（位置固定）だった場合は、まずそれを解除する
            user_fixed_slots = data.get("user_fixed_slots", {})
            if slot in user_fixed_slots:
                subject_to_unfix = user_fixed_slots[slot]
                slots_to_unfix = [s for s, sub in user_fixed_slots.items() if sub == subject_to_unfix]
                for s in slots_to_unfix:
                    del user_fixed_slots[s]

            # 除外状態をトグルする
            if slot in data["all_slots"]:
                data["all_slots"].remove(slot)
            else:
                data["all_slots"].append(slot)
            
            self.create_combination_time_tab() # UIを再描画して状態を完全に反映
            self.update_year_tab_visuals(year_label)

        except Exception as e:
            print(f"枠切り替えエラー: {e}")

    def toggle_user_fixation(self, subject_to_toggle, year_label):
        """（右クリック）教科を一時的に位置固定、または解除する"""
        try:
            if subject_to_toggle == "－": return
            data = self.year_data.get(year_label)
            if not data: return

            current_timetable = data.get("complete_combination", {}).get(1, {})
            if not current_timetable: return

            slots_for_subject = [s for s, subj in current_timetable.items() if subj == subject_to_toggle]
            if not slots_for_subject: return

            user_fixed_slots = data.setdefault("user_fixed_slots", {})
            is_any_slot_fixed = any(s in user_fixed_slots for s in slots_for_subject)

            if is_any_slot_fixed:
                for s in slots_for_subject:
                    if s in user_fixed_slots:
                        del user_fixed_slots[s]
            else:
                # 固定する前に、まず赤色（除外）状態を解除する
                for s in slots_for_subject:
                    if s not in data["all_slots"]:
                        data["all_slots"].append(s)
                # その後、青色（位置固定）にする
                for s in slots_for_subject:
                    user_fixed_slots[s] = subject_to_toggle
            
            self.create_combination_time_tab()
            self.update_year_tab_visuals(year_label)
        except Exception as e:
            print(f"ユーザー固定エラー: {e}")

    def _on_timetable_slot_clicked(self, slot, year_label, subject_in_slot, button):
        """時間割スロットがクリックされたときのモード別ディスパッチャ"""
        if self.is_tutorial_running:
            current_step = self.tutorial_manager.get_current_step()
            # subject_addedステップでは事前シグナルは不要。
            if current_step and current_step.expected_value and current_step.expected_value.get('action') == 'subject_added':
                pass
            else:
                # このシグナルは他のチュートリアル手順で右クリック判定などに使う。
                self.tutorial_signal.emit({'button': button, 'subject': subject_in_slot, 'slot': slot})

        if self.current_timetable_mode in ["＋", "🗑️"]:
            if button != Qt.LeftButton:
                return
            if self.current_timetable_mode == "＋":
                if subject_in_slot in ["－", "-"]:
                    self._handle_add_subject_mode(slot, year_label)
            elif self.current_timetable_mode == "🗑️":
                if subject_in_slot not in ["－", "-"]:
                    self._handle_eraser_mode(slot, year_label, subject_in_slot)
        
        elif self.current_timetable_mode == "🔒":
            self._handle_lock_mode(slot, year_label, subject_in_slot, button)

    def _handle_add_subject_mode(self, clicked_slot, year_label):
            data = self.year_data[year_label]
            active_profile_name = self.get_active_profile_name()
            
            # 固定枠（configとユーザーロック）かチェック
            fixed_slots = data.get("fixed_slots", {})
            user_fixed_slots = data.get("profile_lock_settings", {}).get(active_profile_name, {})
            if clicked_slot in fixed_slots or clicked_slot in user_fixed_slots:
                QMessageBox.warning(self, "追加不可", f"'{clicked_slot}' は固定された枠のため、手動で教科を追加することはできません。")
                return

            saved_state = self._save_timetable_ui_state()
            current_timetable = data["saved_profiles"].get(active_profile_name, {})
            all_valid_slots = set(self.get_current_profile_all_slots(year_label))
            available_subject_groups_dict = {item["name"]: item["data"] for item in data["subject_slots_base"]}

            # 現在のプロファイルに存在する教科をセットとして取得
            current_subjects = {v for v in current_timetable.values() if v and v.strip() not in ["－", "-"]}

            displayable_subject_groups = {}
            for subject_name, group_list in available_subject_groups_dict.items():
                # 既にプロファイルに存在する教科はスキップ
                if subject_name in current_subjects:
                    continue

                for group_slots in group_list:
                    if clicked_slot in group_slots:
                        is_conflict = False
                        for slot_in_group in group_slots:
                            if (slot_in_group in current_timetable and current_timetable[slot_in_group] != "－" and current_timetable[slot_in_group] != subject_name):
                                is_conflict = True
                                break
                            if slot_in_group not in all_valid_slots:
                                is_conflict = True
                                break
                        if not is_conflict:
                            if subject_name not in displayable_subject_groups:
                                displayable_subject_groups[subject_name] = []
                            displayable_subject_groups[subject_name].append(group_slots)
            
            if not displayable_subject_groups:
                QMessageBox.information(self, "追加不可", f"'{clicked_slot}' を含む追加可能な教科グループが見つかりませんでした。")
                return

            dialog = SubjectSelectionDialog(self, year_label, displayable_subject_groups, clicked_slot)
            
            if dialog.exec() == QDialog.Accepted:
                selected_subject, selected_slot_group = dialog.get_selected_subject_group()
                if selected_subject and selected_slot_group:
                    conflict_slots = []
                    for slot_to_add in selected_slot_group:
                        if (slot_to_add in current_timetable and current_timetable[slot_to_add] != "－" and current_timetable[slot_to_add] != selected_subject):
                            conflict_slots.append(slot_to_add)
                        if slot_to_add not in all_valid_slots:
                            conflict_slots.append(slot_to_add)
                    
                    if conflict_slots:
                        QMessageBox.warning(self, "追加エラー", f"選択された教科グループは、以下のスロットで競合または無効な枠と衝突します:\n{', '.join(conflict_slots)}")
                    else:
                        for slot_to_add in selected_slot_group:
                            data["saved_profiles"][active_profile_name][slot_to_add] = selected_subject
                        
                        # === 修正箇所: チュートリアル分岐の強化 ===
                        if not self.is_tutorial_running:
                            # 通常モード
                            if self.user_settings.get("show_add_subject_success_message", True):
                                msg_box = QMessageBox(self)
                                msg_box.setIcon(QMessageBox.Information)
                                msg_box.setWindowTitle("追加完了")
                                msg_box.setText(f"「{selected_subject}」を時間割に追加しました。")
                                
                                cb = QCheckBox("このメッセージを二度と表示しない")
                                msg_box.setCheckBox(cb)
                                
                                msg_box.exec()

                                if cb.isChecked():
                                    self.user_settings["show_add_subject_success_message"] = False
                                    self._save_user_settings()
                            
                            self._update_timetable_tab_for_profile(active_profile_name)
                            self.check_subjects_units()
                            self.update_all_highlights()
                        else:
                            # チュートリアルモード
                            # 1. 完了メッセージボックスは邪魔になるので出さない
                            
                            # 2. UIの手動更新（create_combination_time_tabを呼ぶと状態がリセットされやすいため手動更新）
                            for slot_to_update in selected_slot_group:
                                button_to_update = self.findChild(TimetableSlotButton, f"slot_{year_label}_{slot_to_update}")
                                if button_to_update:
                                    button_to_update.setText(selected_subject)
                            self.check_subjects_units()
                            self.update_all_highlights()

                            # 3. 強制進行処理
                            # シグナル待ち等は無視し、マネージャーの内部状態を強制リセットして次へ進める
                            try:
                                # 進行中フラグを強制的に解除（これでnext_stepの早期リターンを回避）
                                self.tutorial_manager.is_advancing = False
                                
                                # 現在のイベント処理が完了した直後に実行するようにタイマーをセット
                                QTimer.singleShot(10, self.tutorial_manager.next_step)
                            except Exception as e:
                                print(f"Tutorial force advance failed: {e}")

            self._restore_timetable_ui_state(saved_state)

    def _handle_eraser_mode(self, slot, year_label, subject_in_slot):
        saved_state = self._save_timetable_ui_state()
        if subject_in_slot == "－" or not subject_in_slot:
            return

        data = self.year_data[year_label]
        active_profile_name = self.get_active_profile_name()
        current_timetable = data["saved_profiles"][active_profile_name]

        target_group = None
        subject_slots_base = data.get("subject_slots_base", [])
        for subject_info in subject_slots_base:
            if subject_info.get("name") == subject_in_slot:
                for group in subject_info.get("data", []):
                    if slot in group:
                        target_group = group
                        break
            if target_group:
                break
        
        if target_group is None:
            slots_for_subject = {s for s, sub in current_timetable.items() if sub == subject_in_slot}
            if slots_for_subject:
                target_group = sorted(list(slots_for_subject))
            else:
                target_group = [slot]

        do_delete = False
        if self._get_setting("CONFIRM_ON_ERASE", True):
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("教科グループを消去")
            group_text = ", ".join(target_group)
            msg_box.setText(f"'{subject_in_slot}' のグループ({group_text})を時間割から消去しますか？")
            msg_box.setIcon(QMessageBox.Question)
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg_box.setDefaultButton(QMessageBox.No)
            
            cb = QCheckBox("今後はこのメッセージを表示しない")
            msg_box.setCheckBox(cb)
            
            reply = msg_box.exec()

            if cb.isChecked():
                self.config["CONFIRM_ON_ERASE"] = False
                self.save_config_file()

            if reply == QMessageBox.Yes:
                do_delete = True
        else:
            do_delete = True

        if do_delete:
            for s in target_group:
                if s in current_timetable:
                    current_timetable[s] = "－"

            user_fixed_slots = self.get_current_profile_lock_settings(year_label)
            for s in target_group:
                if s in user_fixed_slots:
                    del user_fixed_slots[s]
            
            if not self.is_tutorial_running:
                self._update_timetable_tab_for_profile(active_profile_name)
                self.check_subjects_units()
                self.update_all_highlights()
            else:
                for slot_to_update in target_group:
                    button_to_update = self.findChild(TimetableSlotButton, f"slot_{year_label}_{slot_to_update}")
                    if button_to_update:
                        button_to_update.setText("－")
                self.check_subjects_units()
                self.update_all_highlights()
            
            if self.is_tutorial_running:
                self.tutorial_signal.emit({'action': 'subject_erased'})
                
        self._restore_timetable_ui_state(saved_state)

    def _handle_lock_mode(self, slot, year_label, subject_in_slot, button):
        saved_state = self._save_timetable_ui_state()
        data = self.year_data[year_label]
        user_fixed_slots = self.get_current_profile_lock_settings(year_label)

        if button == Qt.RightButton:  # 右クリック：教科固定
            if subject_in_slot != "－" and subject_in_slot:
                
                # クリックされた枠が属する教科グループを探す。
                # まず定義済みグループを確認する。
                subject_slots_base = data.get("subject_slots_base", [])
                target_group = None
                for subject_info in subject_slots_base:
                    if subject_info.get("name") == subject_in_slot:
                        for group in subject_info.get("data", []):
                            if slot in group:
                                target_group = group
                                break
                    if target_group:
                        break
                
                # 定義済みグループがなければ現在の時間割からグループを探す。
                if target_group is None:
                    active_profile_name = self.year_data[year_label].get("active_profile_name", "Default")
                    current_timetable = data["saved_profiles"].get(active_profile_name, {})
                    slots_for_subject = {s for s, sub in current_timetable.items() if sub == subject_in_slot}
                    if slots_for_subject:
                        target_group = sorted(list(slots_for_subject))
                    else:
                        target_group = [slot] # 単一枠として扱う


                # 対象グループ内に同じ教科で固定済みの枠があるか確認する。
                is_already_fixed = any(s in user_fixed_slots and user_fixed_slots[s] == subject_in_slot for s in target_group)

                if is_already_fixed:
                    # グループ内の固定をすべて解除する。
                    for s in target_group:
                        if s in user_fixed_slots:
                            del user_fixed_slots[s]
                else:
                    # グループ内の枠をすべて固定する。
                    # 固定前に他のユーザー固定教科との競合を確認する。
                    conflict_found = False
                    for s in target_group:
                        if s in user_fixed_slots:
                            conflict_found = True
                            break
                    if not conflict_found:
                        for s in target_group:
                            user_fixed_slots[s] = subject_in_slot
            else:
                # 空枠の右クリックでは何もしない。
                return

        elif button == Qt.LeftButton:  # 左クリック：枠無効化
            profile_name = self.get_active_profile_name()
            profile_slots = self.year_data[year_label].setdefault("profile_all_slots", {}).setdefault(profile_name, [])
            # 枠の有効/無効を切り替える。
            if slot in profile_slots:
                profile_slots.remove(slot)
            else:
                profile_slots.append(slot)
        else:
            return  # その他のクリックは無視する。

        # UIを更新する。
        if not self.is_tutorial_running:
            self.create_combination_time_tab()
            self.check_subjects_units()
            self.update_all_highlights()
            self.update_year_tab_visuals(year_label)
            self._restore_timetable_ui_state(saved_state)
        else:
            # チュートリアル中は、UIの完全な再構築を避けて手動で更新
            # 現在のプロファイルタブ内のボタンを探してスタイルを更新
            page = self.profile_tab_widget.currentWidget()
            if page:
                buttons = page.findChildren(TimetableSlotButton)
                for btn in buttons:
                    # このボタンの状態を再評価
                    is_user_locked_style = btn.slot in user_fixed_slots
                    profile_name = self.get_active_profile_name()
                    profile_slots = self.year_data[year_label].setdefault("profile_all_slots", {}).setdefault(profile_name, [])
                    is_excluded_style = btn.slot not in profile_slots

                    bg_color_str = ""
                    if is_user_locked_style:
                        highlight_color = QApplication.palette().color(QPalette.Highlight)
                        highlight_color.setAlpha(80)
                        bg_color_str = f"background-color: {highlight_color.name(QColor.HexArgb)};"
                    elif is_excluded_style:
                        bg_color_str = "background-color: rgba(255, 0, 0, 70);"
                    
                    # 条件に合わない場合はスタイルを戻し、合う場合は新しいスタイルを適用
                    btn.setStyleSheet(bg_color_str)

            self.check_subjects_units()
            self.update_all_highlights()
            self.update_year_tab_visuals(year_label)

    def on_tab_changed(self, index):
        """タブが変更されたときにタブの表示を更新する"""
        # メインウィンドウをアクティブに保つ
        self.raise_()
        self.activateWindow()
        
        # QListWidgetのインデックスとQStackedWidgetのインデックスは同じ
        # 0: 時間割タブ, 1-N: 学年タブ, N+1: 設定タブ
        if index == 0: # 時間割タブ
            # ちらつき抑制のため、現在表示中でなければcombination_time_tabへ切り替える
            try:
                current = self.stacked_widget.currentWidget()
            except Exception:
                current = None
            if current is not self.combination_time_tab:
                # すでに表示中なら再構築しない
                self.create_combination_time_tab()
            self.clear_results_button.setVisible(False)
            if hasattr(self, 'timetable_bottom_specific_frame'):
                self.timetable_bottom_specific_frame.setVisible(True)
        elif 0 < index <= len(self.years): # 学年タブ
            year_label = self.years[index - 1]
            # 削除済みウィジェットへの更新を防ぐ
            try:
                self.update_year_tab_visuals(year_label)
            except RuntimeError:
                # ウィジェットが削除済みなら記録して無視する
                print(f"Warning: update_year_tab_visuals skipped for {year_label} due to deleted widgets")
            self.clear_results_button.setVisible(True)
            if hasattr(self, 'timetable_bottom_specific_frame'):
                self.timetable_bottom_specific_frame.setVisible(False)
        else: # 設定タブなど
            self.clear_results_button.setVisible(False)
            if hasattr(self, 'timetable_bottom_specific_frame'):
                self.timetable_bottom_specific_frame.setVisible(False)



    def _add_tab_item(self, text, icon_name):
        """テーマに応じたアイコン付きのQListWidgetItemを作成して追加する。"""
        # アイコンの基準パスを決定
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(os.path.dirname(__file__))
        
        # テーマに応じた単色アイコン色を決定する。
        target_color = "#FFFFFF" if self.is_dark_theme() else "#1F1F1F"

        # 共通ベースSVGファイルのパスを組み立てる
        icon_path = os.path.join(base_path, "svgs", f"{icon_name}.svg")

        # 色差し替えローダーでアイコンを生成
        icon = self._create_icon_from_svg_data(icon_path, target_color)

        # リスト項目を生成して追加
        item = QListWidgetItem(text)
        if icon:
            item.setIcon(icon)
        self.tab_list.addItem(item)

    def update_year_tab_visuals(self, year_label):
        """指定された学年タブの全教科の見た目と順序を、ユーザー固定の状態に基づいて更新する"""
        data = self.year_data.get(year_label)
        if not data or not data.get("subject_frames") or not data.get("subject_scroll_layout"):
            return

        # --- 状態の取得 ---
        user_fixed_slots = self.get_current_profile_lock_settings(year_label)
        locked_subjects = set(user_fixed_slots.values())
        locked_slots = set(user_fixed_slots.keys())
        subject_frames = data["subject_frames"]
        layout_box = data["subject_scroll_layout"]
        subject_slots_map = {item["name"]: item["data"] for item in data["subject_slots_base"]}

        # --- 各教科の状態を事前に判断 ---
        subject_states = {}
        for frame in subject_frames:
            subject_name = frame.subject_name
            tooltip_text = ""
            state = "normal"

            if subject_name in locked_subjects:
                state = "locked"
                tooltip_text = "この教科は時間割タブで手動で固定されているため、選択できません。"
            else:
                possible_groups = subject_slots_map.get(subject_name, [])
                is_placeable = False
                conflicting_subject_info = ""

                if not possible_groups: # 時間枠未定義なら常に配置可能
                    is_placeable = True
                else:
                    can_be_placed = False
                    # 利用可能な時間枠グループが一つ以上あるか確認
                    for group in possible_groups:
                        intersecting_slots = set(group) & locked_slots
                        if not intersecting_slots:
                            can_be_placed = True
                            break  # この時間枠グループは利用可能なので配置可能
                        else:
                            # このグループは競合しているため、原因のロック済み教科を探す。
                            slot = next(iter(intersecting_slots))
                            conflicting_subject = user_fixed_slots.get(slot, "不明")
                            conflicting_subject_info = f"手動で固定された「{conflicting_subject}」と時間（{slot}）が競合します。"
                    
                    if can_be_placed:
                        is_placeable = True

                if not is_placeable:
                    state = "conflicted"
                    tooltip_text = f"この教科は選択できません。理由: {conflicting_subject_info}"

            subject_states[subject_name] = {"state": state, "tooltip": tooltip_text}

        # --- 並び替え ---
        initial_order_map = {self._normalize_subject_name(item["name"]): i for i, item in enumerate(data["subject_number"])}
        subject_frames.sort(key=lambda frame: (
            subject_states.get(frame.subject_name, {}).get("state", "normal") != "normal",
            initial_order_map.get(frame.subject_name, 999)
        ))

        # --- レイアウトの再構築 ---
        # C++側オブジェクトが削除済みのフレームを無視し、安全に再配置する
        for frame in list(subject_frames):
            # ウィジェットが有効か確認してからレイアウト操作
            if widget_is_valid(frame):
                layout_box.removeWidget(frame)
                layout_box.addWidget(frame)

        # --- 各行の見た目を更新 ---
        for frame in subject_frames:
            subject_name = frame.subject_name
            state_info = subject_states.get(subject_name, {"state": "normal", "tooltip": ""})
            state = state_info['state']
            tooltip = state_info['tooltip']

            try:
                frame.subject_cb.toggled.disconnect()
            except (TypeError, RuntimeError):
                pass
            
            frame.setToolTip(tooltip)
            frame.subject_cb.setToolTip(tooltip)

            if state == "locked":
                frame.subject_cb.setChecked(False)
                frame.subject_cb.setEnabled(False)
                for prefix_cb in frame.prefix_cbs.values():
                    prefix_cb.setChecked(False)
                frame.setStyleSheet(f"background-color: {self.theme['user_locked_subject']};")

            elif state == "conflicted":
                frame.subject_cb.setChecked(False)
                frame.subject_cb.setEnabled(False)
                for prefix_cb in frame.prefix_cbs.values():
                    prefix_cb.setChecked(False)
                frame.setStyleSheet(f"background-color: {self.theme['user_locked_conflict']};")

            else: # stateがnormalの場合
                frame.setEnabled(True)
                frame.setStyleSheet("")
                frame.setToolTip("")
                frame.subject_cb.setToolTip("")
                frame.subject_cb.setChecked(data["check_vars"].get(subject_name, False))
                frame.subject_cb.toggled.connect(self._on_subject_check)

    def _create_slot_selection_grid(self, year_label):
        """枠選択ダイアログ内に表示する、クリック可能なスロット選択グリッドを作成する"""
        grid_widget = QWidget()
        grid_layout = QVBoxLayout(grid_widget)
        grid_layout.setSpacing(1)

        data = self.year_data.get(year_label, {})
        table_layout_config = data.get("table_layout", [])
        
        slot_buttons = []

        # table_layout_configは[[],[],[]]の形式なので、各行をループ
        for row_data in table_layout_config:
            row_layout = QHBoxLayout()
            row_layout.setSpacing(1)
            for slot_name in row_data: # 行内の各スロット名をループ
                if not slot_name: # スロット名が空の場合はスキップ
                    row_layout.addWidget(QFrame()) # 空のフレームを追加してスペースを埋める
                    continue

                button = QPushButton(slot_name)
                button.setCheckable(True) # チェック可能にする
                button.setFixedSize(60, 30) # サイズを固定
                button.setStyleSheet("QPushButton:checked { background-color: #ADD8E6; border: 1px solid #0078D7; }") # チェック時のスタイル

                slot_buttons.append(button)
                row_layout.addWidget(button)
            grid_layout.addLayout(row_layout)
        
        return grid_widget, slot_buttons # グリッドウィジェットとボタンリストを返す

    def _open_slot_selection_dialog(self, year_label):
        """枠選択ダイアログを開き、選択された枠で教科リストをフィルタリングする"""
        current_selected_slots = self.year_data[year_label].get("selected_slots_filter", [])
        dialog = SlotSelectionDialog(self, year_label, current_selected_slots)
        if dialog.exec() == QDialog.Accepted:
            selected_slots = dialog.get_selected_slots()
            self.year_data[year_label]["selected_slots_filter"] = selected_slots
            
            # 選択されたスロットと現在の検索バーのテキストに基づいてフィルタを適用
            search_text_bar = self.year_data[year_label]["search_bar"]
            if search_text_bar:
                self._filter_subjects(search_text_bar.text(), year_label, selected_slots=selected_slots)
        else:
            # ダイアログがキャンセルされた場合は、現在のフィルタを再適用（スロット選択は変更しない）
            search_text_bar = self.year_data[year_label]["search_bar"]
            if search_text_bar:
                self._filter_subjects(search_text_bar.text(), year_label, selected_slots=current_selected_slots)

    def check_required_subjects(self, timetable, year_label, include_all=False):
        """必須科目チェック（動的階層対応）
        【委譲】TimetableLogic.check_required_subjectsに処理を委譲します。
        """
        current_subjects = [s for s in timetable.values() if s and s.strip() not in ["－", "-", ""]]
        
        # チェック対象の科目を準備
        subjects_to_check = set(current_subjects)
        if include_all:
            for other_year, other_data in self.year_data.items():
                if other_year == year_label:
                    continue
                
                other_timetable = other_data.get("complete_combination", {}).get(1, {})
                if other_timetable:
                    other_subjects = {s for s in other_timetable.values() if s and s.strip() not in ["－", "-", ""]}
                    subjects_to_check.update(other_subjects)

        # --- チェックする要件セットを収集 ---
        requirement_sources = []
        
        # include_allがFalseの場合、現在の学年の科目のみをチェック対象とする
        subjects_for_local_reqs = list(subjects_to_check) if include_all else current_subjects

        # 1. 学年固有の要件
        year_req = self.year_data[year_label].get("required_subjects", {})
        if year_req:
            requirement_sources.append({
                "source_label": year_label,
                "requirements": year_req,
                "subjects_to_check": subjects_for_local_reqs
            })

        # 2. 階層的な学科要件
        parts = year_label.split('_')
        if len(parts) > 1:
            department_label = []
            for part in parts[:-1]:
                department_label.append(part)
                label = '_'.join(department_label)
                department_req_key = f"REQUIRED_SUBJECTS_{label}"
                department_req = self.config.get(department_req_key, {})
                if department_req:
                    requirement_sources.append({
                        "source_label": label,
                        "requirements": department_req,
                        "subjects_to_check": subjects_for_local_reqs
                    })

        # 3. 全学年共通の要件 (include_allがTrueの場合のみ)
        if include_all:
            global_req = self.config.get("REQUIRED_SUBJECTS_ALL", {})
            if global_req:
                requirement_sources.append({
                    "source_label": "全学年",
                    "requirements": global_req,
                    "subjects_to_check": list(subjects_to_check)
                })

        # --- TimetableLogicに処理を委譲 ---
        return TimetableLogic.check_required_subjects(
            subjects_to_check=list(subjects_to_check),
            requirement_sources=requirement_sources
        )

    def format_missing_subjects_message(self, missing_info):
        """不足科目情報のリストをユーザーフレンドリーな文字列にフォーマットする
        【委譲】TimetableLogic.format_missing_subjects_messageに処理を委譲します。
        """
        return TimetableLogic.format_missing_subjects_message(missing_info)

    def get_prefix(self, slots):
        """スロットのリストからプレフィックスを生成
        【委譲】TimetableLogic.get_prefixに処理を委譲します。
        """
        return TimetableLogic.get_prefix(slots)

    def update_duplicate_highlight_status(self):
        """重複選択と前提教科のハイライト状態を更新"""
        self.update_all_highlights()

    def check_prerequisites(self, timetable, year_label):
        """前提科目の不足をチェック
        【委譲】TimetableLogic.check_prerequisitesに処理を委譲します。
        """
        # 芸術科目を収集（UIから取得）
        art_subjects = set()
        for combo in self.art_subject_combos:
            selected_art = combo.currentText()
            if selected_art:
                art_subjects.add(selected_art)
        
        return TimetableLogic.check_prerequisites(
            timetable=timetable,
            art_subjects=art_subjects,
            prerequisite_config=self.PREREQUISITE_SUBJECTS
        )

    def _a1_to_coords(self, a1_string):
        """【委譲】A1表記のセル座標変換をFileIOManagerに委譲します。"""
        return self.file_io_manager._a1_to_coords(a1_string)

    def find_subject_for_slot_and_number(self, year: str, slot_name: str, num_str: str) -> str | None:
        """【委譲】スロットと番号による教科名検索をTimetableLogicに委譲します。"""
        year_data = self.year_data.get(year)
        if not year_data:
            return None
        subject_number_list = year_data.get("subject_number", [])
        
        return TimetableLogic.find_subject_for_slot_and_number(
            subject_number_list, slot_name, num_str
        )

    def load_timedate(self):
        """Excelファイルから時間割を復元し、時間割タブに表示する"""
        if self.is_tutorial_running:
            self.tutorial_signal.emit({'action': 'menu_clicked'})
            QTimer.singleShot(100, lambda: self.tutorial_signal.emit({'action': 'dialog_opened'}))
            return

        file_dialog = QFileDialog(self, "時間割ファイルを開く", "", "Excel Files (*.xlsx *.xlsm)")
        icon_path = os.path.join(self.base_path, "時間割くんアイコン.ico")
        if os.path.exists(icon_path):
            file_dialog.setWindowIcon(QIcon(str(icon_path)))
        
        file_path, _ = file_dialog.getOpenFileName()
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "ファイルエラー", f"Excelファイルの読み込みに失敗しました:\n{e}")
            return

        # アクティブなプロファイルを取得
        active_profile_name = self.get_active_profile_name()
        
        # 読み込み前にアクティブプロファイルのデータをクリア
        for year in self.years:
            self.year_data[year]["saved_profiles"][active_profile_name] = {}
            self.year_data[year]["complete_combination"] = {}
            if active_profile_name in self.year_data[year]["profile_lock_settings"]:
                self.year_data[year]["profile_lock_settings"][active_profile_name] = {}

        for year in self.years:
            data = self.year_data[year]
            save_pos = data.get("save_position", {})
            if not save_pos:
                continue

            # 1. 読み取るべきセルの座標とスロット名のマッピングを作成
            cell_to_slot_map = {}
            for slot_name, v_pos in save_pos.items():
                try:
                    pos_to_use = v_pos
                    if isinstance(pos_to_use, list) and len(pos_to_use) == 1 and isinstance(pos_to_use[0], list):
                        pos_to_use = pos_to_use[0]
                    if isinstance(pos_to_use, list) and len(pos_to_use) > 0 and isinstance(pos_to_use[0], str):
                        coords = self._a1_to_coords(pos_to_use[0])
                        if coords is None: raise TypeError()
                        pos_to_use = coords
                    if not (isinstance(pos_to_use, list) and len(pos_to_use) == 2):
                        raise TypeError()
                    pos_tuple = tuple(pos_to_use)
                    cell_to_slot_map[pos_tuple] = slot_name
                except (TypeError, ValueError):
                    continue

            # 2. セルを読み、スロットと教科名を特定して時間割を構築
            new_timetable = {}
            for pos, slot_name in cell_to_slot_map.items():
                try:
                    cell_value = ws.cell(row=pos[0], column=pos[1]).value
                    if cell_value is not None and str(cell_value).strip():
                        number_str = str(cell_value).strip()
                        subject_name = self.find_subject_for_slot_and_number(year, slot_name, number_str)
                        if subject_name:
                            new_timetable[slot_name] = subject_name
                except IndexError:
                    continue

            # 3. 構築した時間割をアクティブプロファイルと古いデータ構造に格納
            if new_timetable:
                self.year_data[year]["saved_profiles"][active_profile_name] = new_timetable
                self.year_data[year]["complete_combination"] = {1: new_timetable}

        # 4. チェックボックスの状態を更新
        for year in self.years:
            subjects_in_timetable = set(self.year_data[year].get("saved_profiles", {}).get(active_profile_name, {}).values())
            for frame in self.year_data[year].get("subject_frames", []):
                if hasattr(frame, 'subject_cb'):
                    is_checked = frame.subject_name in subjects_in_timetable
                    frame.subject_cb.setChecked(is_checked)

        # 5. 時間割タブの表示と単位数を更新
        self.create_combination_time_tab()
        self.check_subjects_units()
        QApplication.processEvents()
        QMessageBox.information(self, "読み込み完了", f"Excelファイルから「{active_profile_name}」プロファイルに時間割を復元しました。")
        self._has_active_combinations = True # フラグをTrueに設定

    def _filter_subjects(self, text, year_label, selected_slots=None):
        """教科リストをテキストと枠名（スロット）でフィルタリングする"""
        data = self.year_data.get(year_label)
        if not data:
            return

        # selected_slotsがNoneの場合、year_dataから取得
        if selected_slots is None:
            selected_slots = data.get("selected_slots_filter", [])

        # 検索テキストを正規化し、キーワードに分割
        search_text = unicodedata.normalize("NFKC", text).lower()
        text_keywords = search_text.split()
        
        # 教科ごとのスロット情報をマップに変換して効率化
        subject_slots_map = {item["name"]: item["data"] for item in data.get("subject_slots_base", [])}

        # 各教科の一致スコア（一致キーワード数 + 一致枠数）を計算する
        initial_order_map = {item["name"]: i for i, item in enumerate(data.get("subject_slots_base", []))}
        visible_with_score = []

        for frame in data.get("subject_frames", []):
            subject_name = frame.subject_name

            # 検索対象テキストを準備
            subject_name_normalized = unicodedata.normalize("NFKC", subject_name).lower()
            aliases = [unicodedata.normalize("NFKC", alias).lower() for alias in self.SUBJECT_ALIASES.get(subject_name, [])]
            search_targets = [subject_name_normalized] + aliases

            # テキスト一致数（キーワードはOR条件で一致数を数える）
            text_match_count = 0
            if text_keywords:
                for kw in text_keywords:
                    if any(kw in target for target in search_targets):
                        text_match_count += 1

            # 枠一致数（選択枠のうち教科で利用可能な枠数を数える）
            slot_match_count = 0
            possible_slots_for_subject = {s.lower() for grp in subject_slots_map.get(subject_name, []) for s in grp}
            if selected_slots:
                for slot in selected_slots:
                    if slot.lower() in possible_slots_for_subject:
                        slot_match_count += 1

            total_matches = text_match_count + slot_match_count

            # フィルタリングロジック
            is_visible = False
            if not text_keywords and not selected_slots:
                # フィルタが指定されていない場合は全て表示
                is_visible = True
            else:
                # テキストフィルタの一致判定 (AND条件)
                text_matches_all_keywords = True
                if text_keywords:
                    for kw in text_keywords:
                        if not any(kw in target for target in search_targets):
                            text_matches_all_keywords = False
                            break
                else: # テキストキーワードがない場合はテキスト条件は満たされたとみなす
                    text_matches_all_keywords = True

                # スロットフィルタの一致判定 (AND条件)
                slot_matches_all_slots = True
                if selected_slots:
                    possible_slots_for_subject = {s.lower() for grp in subject_slots_map.get(subject_name, []) for s in grp}
                    for slot in selected_slots:
                        if slot.lower() not in possible_slots_for_subject:
                            slot_matches_all_slots = False
                            break
                else: # 選択スロットがない場合はスロット条件は満たされたとみなす
                    slot_matches_all_slots = True
                
                # 最終的な表示判定 (テキスト条件とスロット条件のAND)
                is_visible = text_matches_all_keywords and slot_matches_all_slots

            if is_visible:
                frame.show()
                visible_with_score.append((frame, total_matches))
            else:
                frame.hide()

        # 表示中フレームを一致数の降順、次に初期順で並べ替える
        if visible_with_score:
            visible_with_score.sort(key=lambda x: (-x[1], initial_order_map.get(x[0].subject_name, 999)))
            # 対象フレームを保持するレイアウトを取得する
            layout_box = None
            for frame, _ in visible_with_score:
                parent = frame.parent()
                if parent is not None and parent.layout() is not None:
                    layout_box = parent.layout()
                    break
            if layout_box is not None:
                for frame, _ in visible_with_score:
                    # ウィジェットが有効か確認してからレイアウト操作
                    if widget_is_valid(frame):
                        layout_box.removeWidget(frame)
                        layout_box.addWidget(frame)

    def _show_legend(self):
        """UIの凡例をメッセージボックスで表示する"""
        legend_text = f"""
        <b>UI要素の凡例</b>
        <hr>
        <b>必含:</b> チェックすると、この科目が<u>必ず含まれる</u>時間割が作成されます。<br><br>
        <b>★ (色付き):</b> 必須選択科目<br>
        <br>
        <b>背景色:</b>
        <ul>
        <li style="background-color:{self.theme['duplicate_subject']};"> 重複教科: 他の学年でも選択されている科目</li>
        <li style="background-color:{self.theme['no_together_conflict']};"> 同時履修不可: 共に履修できない組み合わせの科目</li>
        <li style="background-color:{self.theme['prerequisite_missing']};"> 前提科目不足: 未選択な前提科目</li>
        <li style="background-color:{self.theme['user_locked_subject']};"> 手動固定: 時間割タブで手動で固定した科目（行全体）</li>
        <li style="background-color:{self.theme['user_locked_conflict']};"> 競合: 手動固定により選択できなくなった科目（行全体）</li>
        </ul>
        """
        QMessageBox.information(self, "凡例", legend_text)



    def create_year_tab_header(self, widths):
        """学年タブのヘッダーフレームを作成する"""
        header_frame = QFrame()
        header_frame.setFrameShape(QFrame.StyledPanel)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setSpacing(0)
        header_layout.setContentsMargins(0, 0, 0, 0)

        headers = ["必含", "", "教科名", "選択", "特定の枠を選択"]

        for i, (header, width) in enumerate(zip(headers, widths)):
            column_frame = QFrame()
            if width > 0:
                column_frame.setFixedWidth(width)
            column_layout = QHBoxLayout(column_frame)
            column_layout.setContentsMargins(5, 5, 5, 5)
            column_layout.setSpacing(0)
            column_layout.setAlignment(Qt.AlignCenter)

            label = QLabel(header)
            label.setAlignment(Qt.AlignCenter)

            if i == 1:
                label.setText("")

            column_layout.addWidget(label)
            header_layout.addWidget(column_frame, stretch=1 if width == -1 else 0)
        return header_frame

    def create_year_tab(self, tab, subject_slots, fixed_slots, all_slots, year_label):
        """学年タブ作成（必含選択チェックボックス対応版）"""
        # ヘッダー項目と幅の設定（必含列を追加）
        widths = [40, 25, 200, 60, -1]  # -1は可変幅

        tab.setLayout(QVBoxLayout())

        # 上部コントロールフレーム
        top_frame = QFrame()
        top_layout = QHBoxLayout(top_frame)

        # 検索バー
        search_bar = QLineEdit()
        search_bar.setPlaceholderText("教科名を検索...")
        self.year_data[year_label]["search_bar"] = search_bar # 検索バーへの参照を保存
        search_bar.textChanged.connect(lambda text: self._filter_subjects(text, year_label))
        top_layout.addWidget(search_bar)

        # 枠で検索ボタン
        select_slots_button = QPushButton("枠で検索")
        select_slots_button.clicked.connect(lambda: self._open_slot_selection_dialog(year_label))
        top_layout.addWidget(select_slots_button)
        
        # 絞り込みボタン
        filter_button = QPushButton('組み合わせを絞り込み')
        filter_button.clicked.connect(self.create_setting_window)
        top_layout.addWidget(filter_button)

        # 実行ボタン
        submit_btn = QPushButton('実行')
        submit_btn.clicked.connect(lambda: self.execution(
            subject_slots,
            {**self.year_data[year_label]["fixed_slots"], **self.get_current_profile_lock_settings(year_label)},
            self.get_current_profile_all_slots(year_label),
            year_label
        ))
        top_layout.addWidget(submit_btn)

        # 凡例ボタン
        legend_button = QPushButton("?")
        legend_button.setFixedSize(30, 30) # 必要なサイズに調整
        legend_button.setStyleSheet("background-color: #2196F3; color: white; border-radius: 15px;") # 青背景、白文字、円形にする
        legend_button.setToolTip("このボタンは凡例を表示します。") # ツールチップを追加
        legend_button.clicked.connect(self._show_legend)
        top_layout.addWidget(legend_button)

        # ボタンをインスタンス変数として保存
        setattr(self, f"submit_btn_{year_label}", submit_btn)
        setattr(self, f"filter_button_{year_label}", filter_button)

        tab.layout().addWidget(top_frame)

        # ステータスメッセージ用ラベル
        status_label = QLabel()
        status_label.setWordWrap(True)
        status_label.setAlignment(Qt.AlignCenter)
        tab.layout().addWidget(status_label)
        self.year_data[year_label]["status_label"] = status_label

        # スクロールエリアの作成
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        # ヘッダーフレームの作成
        header_frame = self.create_year_tab_header(widths)
        scroll_layout.addWidget(header_frame)

        # チェックボックス変数の初期化
        data = self.year_data[year_label]
        data["check_vars"] = {self._normalize_subject_name(item["name"]): False for item in data["subject_number"]}
        check_vars = data["check_vars"]
        data["prefix_vars"] = {self._normalize_subject_name(item["name"]): {} for item in data["subject_number"]}
        prefix_vars = data["prefix_vars"]
        data["important_vars"] = {self._normalize_subject_name(item["name"]): False for item in data["subject_number"]}
        important_vars = data["important_vars"]

        # 他方の学年で使用されている科目を取得
        other_years_subjects = set().union(*(d.get("complete_combination", {}).get(1, {}).values() for y, d in self.year_data.items() if y != year_label))
        used_in_other_year = {s for s in other_years_subjects if s and s not in (set().union(*(d["fixed_slots"].values() for d in self.year_data.values())))}

        # 科目ごとの行を作成
        subject_slots_map = {self._normalize_subject_name(i["name"]): i["data"] for i in data["subject_slots_base"]}
        for item in data["subject_number"]:
            subject_raw = item["name"]
            subject_data = subject_slots_map.get(self._normalize_subject_name(subject_raw), [])
            subject = self._normalize_subject_name(subject_raw)
            if not subject: # 無効な教科名はスキップ
                continue

            row_frame = QFrame()
            row_frame.setFrameShape(QFrame.StyledPanel)
            row_layout = QHBoxLayout(row_frame)
            row_layout.setSpacing(0)
            row_layout.setContentsMargins(0, 0, 0, 0)

            # 1. 必含チェックボックス列
            important_frame = QFrame()
            important_frame.setFixedWidth(widths[0])
            important_layout = QHBoxLayout(important_frame)
            important_layout.setContentsMargins(5, 5, 5, 5)
            important_layout.setAlignment(Qt.AlignCenter)

            important_cb = QCheckBox()
            important_cb.setChecked(important_vars[subject])
            important_cb.setEnabled(False)  # 初期状態は無効
            important_cb.stateChanged.connect(lambda state, s=subject: self._on_important_check(state, s, year_label))
            important_layout.addWidget(important_cb)
            row_layout.addWidget(important_frame)

            # 2. 空白列 (マーク用)
            blank_frame = QFrame()
            blank_frame.setFixedWidth(widths[1])
            blank_layout = QHBoxLayout(blank_frame)
            blank_layout.setContentsMargins(5, 5, 5, 5)

            # 必須科目マークを追加
            req_info = self._get_requirement_info(subject_raw, year_label)
            if req_info:
                mark_label = QLabel(req_info["mark"])
                mark_label.setStyleSheet(f"color: {req_info['color']}; font-weight: bold;")
                blank_layout.addWidget(mark_label)
            else:
                mark_label = QLabel("")
                blank_layout.addWidget(mark_label)
            row_frame.required_mark_label = mark_label

            row_layout.addWidget(blank_frame)

            # 3. 教科名列
            subject_frame = QFrame()
            subject_frame.setFixedWidth(widths[2])
            subject_layout = QHBoxLayout(subject_frame)
            subject_layout.setContentsMargins(5, 5, 5, 5)
            subject_layout.setAlignment(Qt.AlignCenter)

            subject_label = QLabel(subject_raw)
            subject_label.setAlignment(Qt.AlignCenter)
            subject_label.setObjectName('subject_label')
            subject_layout.addWidget(subject_label)
            row_layout.addWidget(subject_frame)

            # 4. 選択チェックボックス列
            select_frame = QFrame()
            select_frame.setFixedWidth(widths[3])
            select_layout = QHBoxLayout(select_frame)
            select_layout.setContentsMargins(5, 5, 5, 5)
            select_layout.setAlignment(Qt.AlignCenter)

            bg_color = self.theme['duplicate_subject'] if subject in used_in_other_year else None
            cb = QCheckBox()
            cb.setChecked(check_vars[subject])
            cb.setObjectName(f"cb_{subject}")
            cb.parent_frame = row_frame # 親フレームへの参照を追加
            row_frame.subject_cb = cb
            cb.toggled.connect(self._on_subject_check)
            if bg_color:
                cb.setStyleSheet(f"background-color: {bg_color};")
            select_layout.addWidget(cb)
            row_layout.addWidget(select_frame)

            # 行フレームに参照を保存
            row_frame.subject_name = subject
            row_frame.year_label = year_label
            row_frame.subject_cb = cb
            row_frame.important_cb = important_cb

            # 5. 特定の枠選択列 (可変幅)
            prefix_frame = QFrame()
            prefix_layout = QHBoxLayout(prefix_frame)
            prefix_layout.setContentsMargins(5, 5, 5, 5)
            prefix_layout.setAlignment(Qt.AlignLeft)

            unique_prefixes = set()
            for slots in subject_data: # subject_dataを使用
                prefix = "".join(sorted(set(slot[0] for slot in slots)))
                unique_prefixes.add(prefix)

            prefix_cbs = {}
            for prefix in sorted(unique_prefixes):
                prefix_cb = QCheckBox(prefix)
                prefix_cb.setChecked(False)

                prefix_cb.subject = subject
                prefix_cb.parent_frame = row_frame

                prefix_cb.setObjectName(f"prefix_cb_{subject}_{prefix}")
                prefix_cb.toggled.connect(self._on_prefix_check)

                if bg_color:
                    prefix_cb.setStyleSheet(f"background-color: {bg_color};")

                prefix_layout.addWidget(prefix_cb)
                prefix_cbs[prefix] = prefix_cb

            row_frame.prefix_cbs = prefix_cbs
            prefix_vars[subject] = prefix_cbs

            row_layout.addWidget(prefix_frame, stretch=1)

            # 検索用にフレームを保存
            data["subject_frames"].append(row_frame)

            # 重複ハイライト用ウィジェット登録
            if subject not in data["duplicate_highlight_status"]["widgets"]:
                data["duplicate_highlight_status"]["widgets"][subject] = []
            data["duplicate_highlight_status"]["widgets"][subject].append(cb)
            data["duplicate_highlight_status"]["widgets"][subject].append(subject_label)

            scroll_layout.addWidget(row_frame)

        # 初期重複状態を適用

        scroll_area.setWidget(scroll_content)
        tab.layout().addWidget(scroll_area)

        # 検索用にレイアウトを保存
        data["subject_scroll_layout"] = scroll_layout


    def _get_submit_btn(self, year_label):
        return getattr(self, f"submit_btn_{year_label}", None)

    def _get_filter_button(self, year_label):
        return getattr(self, f"filter_button_{year_label}", None)

    def _get_subject_checkbox(self, subject_name, year_label):
        """指定された教科名と学年ラベルのsubject_cbを取得するヘルパー"""
        normalized_subject_name = self._normalize_subject_name(subject_name)
        tab = self.year_tabs.get(year_label)
        if not tab: return None
        scroll_area = tab.findChild(QScrollArea)
        if scroll_area:
            scroll_content = scroll_area.widget()
            if scroll_content:
                for frame in scroll_content.findChildren(QFrame):
                    # --- ↓ 比較対象を修正 ↓ ---
                    if hasattr(frame, 'subject_name') and frame.subject_name == normalized_subject_name:
                    # -------------------------
                        return frame.subject_cb
        return None

    def _get_important_checkbox(self, subject_name, year_label):
        """指定された教科名と学年ラベルのimportant_cbを取得するヘルパー"""
        # --- ↓ この1行を追加 ↓ ---
        normalized_subject_name = self._normalize_subject_name(subject_name)
        # -------------------------

        tab = self.year_tabs.get(year_label)
        if not tab: return None
        scroll_area = tab.findChild(QScrollArea)
        if scroll_area:
            scroll_content = scroll_area.widget()
            if scroll_content:
                for frame in scroll_content.findChildren(QFrame):
                    # --- ↓ 比較対象を修正 ↓ ---
                    if hasattr(frame, 'subject_name') and frame.subject_name == normalized_subject_name:
                    # -------------------------
                        return frame.important_cb
        return None

    def _get_save_combination_btn(self):
        """表示されている「この組み合わせを保存」ボタンを取得するヘルパー"""
        window = self._get_current_display_window()
        if window and hasattr(window, 'window_data'):
            buttons = window.window_data.get('combination_save_buttons', [])
            list_widget = window.window_data.get('combination_list_widget')
            if buttons and list_widget:
                current_row = list_widget.currentRow()
                if current_row == -1: # 未選択なら先頭を既定にする
                    current_row = 0
                if 0 <= current_row < len(buttons):
                    return buttons[current_row]
        return None

    def _get_current_display_window(self):
        if hasattr(self, 'display_windows') and self.display_windows:
            return self.display_windows[-1]
        return None

    def highlight_subjects(self, subjects, color, force=False):
        for year in self.years:
            if year not in self.year_data:
                continue
            status = self.year_data[year]["duplicate_highlight_status"]
            widgets_dict = status.get("widgets", {})
            is_duplicate_set = status.get("subjects", set())

            for subject in subjects:
                widgets = widgets_dict.get(subject, [])
                for widget in widgets:
                    current_color = widget.palette().color(widget.backgroundRole()).name()
                    # 前提教科ハイライト ("#FF9999") は強制上書き以外は残す
                    if not force and current_color.lower() == "#ff9999":
                        continue
                    # 重複状態の優先色 ("#FFD699")
                    if subject in is_duplicate_set and color != "#FF9999":
                        widget.setStyleSheet("background-color: #FFD699;")
                    else:
                        widget.setStyleSheet(f"background-color: {color};")

    def update_all_highlights(self):
        """すべてのハイライト処理を統括管理"""
        # 全学年の選択科目を取得（固定科目は除外）
        year_subjects = {}
        active_profile_name = self.get_active_profile_name()

        for year in self.years:
            if year in self.year_data:
                data = self.year_data[year]
                
                # 現在プロファイルと同期しているcheck_varsから教科を取得
                checked_subjects = {subject for subject, is_checked in data.get("check_vars", {}).items() if is_checked}
                
                # メイン時間割の強調表示のため、保存済み時間割の教科も含める
                saved_timetable = data.get("saved_profiles", {}).get(active_profile_name, {})
                subjects_from_timetable = {v for v in saved_timetable.values() if v and v.strip() not in ["－", "-", ""]}

                combined_subjects = checked_subjects.union(subjects_from_timetable)

                fixed_slots_for_year = data.get("fixed_slots", {})
                
                year_subjects[year] = {
                    s for s in combined_subjects if s not in fixed_slots_for_year.values()
                }

        # 芸術科目を追加 (現在アクティブなプロファイルから)
        art_selections = self.art_subject_selections.get(active_profile_name, [])
        for art_subject in art_selections:
            if art_subject:
                 # 学年間の前提教科チェック用に全学年へ追加
                for year in self.years:
                    if year in year_subjects:
                        year_subjects[year].add(art_subject)

        all_subjects_combined = set().union(*year_subjects.values())

        # 前提教科の不足をチェック
        prerequisite_targets = set()
        # configの "本体科目": ["前提科目1", "前提科目2"] という構造をループ
        for main_subject, prerequisites in self.PREREQUISITE_SUBJECTS.items():
            # 教科名を正規化
            norm_main_subject = self._normalize_subject_name(main_subject)
            # もし「本体科目」(例: 数学Ⅲ)が選択されていたら...
            if norm_main_subject in all_subjects_combined:
                # その前提科目を一つずつチェック
                for prereq in prerequisites:
                    # 前提科目名も正規化
                    norm_prereq = self._normalize_subject_name(prereq)
                    # もし「前提科目」(例: 数学Ⅱ)が選択されていなければ...
                    if norm_prereq not in all_subjects_combined:
                        # 不足している「前提科目」(数学Ⅱ)をハイライト対象に追加する
                        prerequisite_targets.add(norm_prereq)

        # 同時選択不可科目
        notogether_highlight = set()
        for subject_group in self.NO_TOGETHER_SUBJECTS:
            selected_subjects_in_group = [s for s in subject_group if s in all_subjects_combined]
            if len(selected_subjects_in_group) > 1:
                for subject in selected_subjects_in_group:
                    notogether_highlight.add(subject)

        # ハイライト適用
        for year in self.years:
            if year not in self.year_data:
                continue
            data = self.year_data[year] # 現在年次のデータを取得
            
            # --- 必須教科マークラベルを更新 ---
            for frame in data.get("subject_frames", []): # 教科フレームを走査
                subject = frame.subject_name
                if hasattr(frame, 'required_mark_label'):
                    req_info = self._get_requirement_info(subject, year)
                    if req_info:
                        frame.required_mark_label.setText(req_info["mark"])
                        frame.required_mark_label.setStyleSheet(f"color: {req_info['color']}; font-weight: bold;")
                    else:
                        frame.required_mark_label.setText("")
                        frame.required_mark_label.setStyleSheet("") # スタイルシートをリセット

            # --- その他の強調表示（重複、前提不足、同時選択不可）を適用 ---
            widgets_dict = data.get("duplicate_highlight_status", {}).get("widgets", {})
            other_years_subjects = set().union(*(s for y, s in year_subjects.items() if y != year))

            for subject, widgets in widgets_dict.items():
                for widget in list(widgets):
                    # ウィジェットが有効か確認
                    if not widget_is_valid(widget):
                        continue
                    
                    if isinstance(widget, QLabel) and widget.objectName() == 'subject_label':
                        if subject in prerequisite_targets:
                            widget.setStyleSheet(f"background-color: {self.theme['prerequisite_missing']};")
                        elif subject in other_years_subjects:
                            widget.setStyleSheet(f"background-color: {self.theme['duplicate_subject']};")
                        elif subject in notogether_highlight:
                            widget.setStyleSheet(f"background-color: {self.theme['no_together_conflict']};")
                        else:
                            widget.setStyleSheet("")  # リセット

    def save_timetable(self, timetable, year_label, profile_name=None):
        """時間割保存処理（UI再構築なし）"""
        year_name = year_label.split('_')[-1]
        if year_label not in self.year_data:
            QMessageBox.warning(self, "保存エラー", f"{year_name}のデータが見つかりません。")
            return

        # profile_nameが指定されていなければ現在のアクティブプロファイルを使用
        active_profile_name = profile_name if profile_name else self.get_active_profile_name()
        # 1. データモデルを更新 (saved_profiles が唯一の信頼できる情報源)
        self.year_data[year_label]["saved_profiles"][active_profile_name] = timetable
        
        # 2. 既存のUIの内容を更新
        self._update_timetable_tab_for_profile(active_profile_name)
        self.check_subjects_units()
        self.update_all_highlights()

        # 3. 確認メッセージを表示
        if not self.is_tutorial_running:
            dialog = SaveCompleteDialog(self, f"「{active_profile_name}」プロファイルに{year_name}の時間割を保存しました。")
            dialog.show()
        self._has_active_combinations = True

        # 4. この組み合わせを保存したウィンドウを閉じる
        sender_button = self.sender()
        if sender_button and isinstance(sender_button, QPushButton):
            # ボタンが属する最上位ウィンドウを取得
            parent_window = sender_button.window()
            if parent_window is not self: # メインウィンドウは閉じない
                parent_window.close()

    def _force_update_year_tab(self, year_label):
        """学年タブのUIを強制更新"""
        tab = self.year_tabs.get(year_label)
        if tab:
            scroll_area = tab.findChild(QScrollArea)
            if scroll_area:
                scroll_content = scroll_area.widget()
                if scroll_content:
                    scroll_content.update()

    def _calculate_duplicate_subjects(self, year_label):
        """指定年の重複科目を検出しセットで返す"""
        current_subjects = set()
        other_subjects = set()

        if year_label in self.year_data:
            data = self.year_data[year_label]
            timetable = data["complete_combination"].get(1, {})
            fixed_slots = data["fixed_slots"]
            current_subjects.update(
                v for v in timetable.values()
                if v and v.strip() not in ["－", "-", ""] and v not in fixed_slots.values()
            )

        for other_year, other_data in self.year_data.items():
            if other_year == year_label: continue
            timetable = other_data["complete_combination"].get(1, {})
            fixed_slots = other_data["fixed_slots"]
            other_subjects.update(
                v for v in timetable.values()
                if v and v.strip() not in ["－", "-", ""] and v not in fixed_slots.values()
            )

        # 芸術科目を考慮
        selected_art = self.selected_art_subject_combo.currentText()
        if selected_art:
            other_subjects.add(selected_art)

        return current_subjects & other_subjects

    def _get_prerequisite_highlight_targets_per_year(self, all_subjects):
        """
        前提教科ハイライト対象となる子教科セットを返す（全学年の科目を考慮）
        親教科がどちらかの学年の all_subjects に含まれていて、
        かつ対応する子教科がどの学年でも未選択の場合のみハイライト対象とする
        """
        highlight_subjects = set()

        # 全学年の選択科目を結合
        all_subjects_combined = set()
        for year in self.years:
            if year in self.year_data:
                data = self.year_data[year]
                timetable = data["complete_combination"].get(1, {})
                fixed_slots = data["fixed_slots"]
                all_subjects_combined.update(
                    v for v in timetable.values()
                    if v and v.strip() not in ["－", "-", ""] and v not in fixed_slots.values()
                )

        # 芸術科目を追加
        selected_art = self.selected_art_subject_combo.currentText() if hasattr(self, 'selected_art_subject_combo') else None
        if selected_art:
            all_subjects_combined.add(selected_art)

        for parent, children in self.PREREQUISITE_SUBJECTS.items():
            # 親教科がどちらかの学年で選択されているか
            if parent in all_subjects_combined:
                for child in children:
                    # 子教科が現在の学年で選択されていない場合のみハイライト
                    if child not in all_subjects:
                        highlight_subjects.add(child)

        return highlight_subjects

    def _on_art_subject_changed(self, text, profile_name=None, idx=None):
        """芸術科目の選択が変更された時の処理"""
        if profile_name:
            active_profile_name = profile_name
        else:
            active_profile_name = self.year_data[self.years[0]].get("active_profile_name", "Default") if self.years else "Default"
        
        selections = [combo.currentText() for combo in self.art_subject_combos]
        
        self.art_subject_selections[active_profile_name] = selections
        self.save_art_subject_selections()

        self.update_all_highlights()
        self.check_subjects_units()

    def check_no_together_subjects(self, all_subjects):
        """同一選択不可科目のチェック（教科名ラベルのみを対象にハイライト）"""
        selected_art = self.selected_art_subject_combo.currentText()
        if selected_art:
            all_subjects.add(selected_art)

        # 対象教科を記録して後でラベルだけを処理
        to_reset = set()
        to_highlight = set()

        for subject_group in self.NO_TOGETHER_SUBJECTS:
            for subject in subject_group:
                to_reset.add(subject)

            selected_count = sum(1 for subject in subject_group if subject in all_subjects)
            if selected_count > 1:
                for subject in subject_group:
                    if subject not in all_subjects:
                        to_highlight.add(subject)

        # ラベルに適用
        for year in self.years:
            if year not in self.year_data:
                continue
            widgets_dict = self.year_data[year].get("duplicate_highlight_status", {}).get("widgets", {})
            for subject, widgets in widgets_dict.items():
                for widget in widgets:
                    if hasattr(widget, 'objectName') and widget.objectName() == 'subject_label':
                        if subject in to_highlight:
                            widget.setStyleSheet("background-color: #7CAAFA;")
                        elif subject in to_reset:
                            widget.setStyleSheet("background-color: #FFFFFF;")

    def _update_duplicate_status(self, target_year, new_duplicates):
        """重複状態を更新（教科名ラベルのみをハイライト）"""
        if target_year not in self.year_data:
            return
        status = self.year_data[target_year]["duplicate_highlight_status"]
        status["subjects"] = new_duplicates

        # タブを取得
        tab = self.year_tabs.get(target_year)
        if not tab:
            return

        # スクロールエリア内のウィジェットを検索
        scroll_area = tab.findChild(QScrollArea)
        if not scroll_area:
            return

        scroll_content = scroll_area.widget()
        if not scroll_content:
            return

        # フレーム内の教科名ラベルのみを更新
        for frame in scroll_content.findChildren(QFrame):
            if hasattr(frame, 'subject_name'):
                subject_name = frame.subject_name

                for label in frame.findChildren(QLabel):
                    if label.objectName() == 'subject_label':
                        if subject_name in new_duplicates:
                            label.setStyleSheet("background-color: #FFD699;")
                        else:
                            label.setStyleSheet("background-color: #FFFFFF;")

    def execution(self, subject_slots, fixed_slots, all_slots, year_label):
        """時間割の組み合わせを生成（QThread非同期版）"""
        if year_label not in self.year_data:
            QMessageBox.critical(self, "内部エラー", f"学年データが見つかりません: {year_label}")
            return
        
        # 既に生成中の場合は警告
        if self._timetable_worker is not None and self._timetable_worker.isRunning():
            QMessageBox.warning(self, "生成中", "時間割の生成が既に実行中です。完了をお待ちください。")
            return
            
        data = self.year_data[year_label]
        check_vars = data["check_vars"]
        prefix_vars = data["prefix_vars"]
        norm_to_raw_map = {self._normalize_subject_name(item["name"]): item["name"] for item in data["subject_number"]}
        subject_slots_dict = {self._normalize_subject_name(item["name"]): item["data"] for item in data["subject_slots_base"]}

        selected_subjects = {}
        for subject_raw, var in check_vars.items():
            subject = self._normalize_subject_name(subject_raw)
            if not subject:
                continue

            selected_prefixes = [p for p, p_var in prefix_vars.get(subject, {}).items() if p_var.isChecked()]
            valid_slots = [
                slots for slots in subject_slots_dict.get(subject, [])
                if not selected_prefixes or self.get_prefix(slots) in selected_prefixes
            ]
            if valid_slots and (var or selected_prefixes):
                selected_subjects[subject] = valid_slots

        if not selected_subjects:
            QMessageBox.information(self, "入力エラー", "教科を1つ以上選択してください。")
            return

        status_label = self.year_data[year_label].get("status_label")
        if status_label:
            status_label.setText("時間割を生成中...")

        submit_btn = getattr(self, f"submit_btn_{year_label}", None)
        if submit_btn:
            submit_btn.setEnabled(False)
        
        # 必含選択科目を取得
        tab = self.year_tabs.get(year_label)
        if not tab:
            QMessageBox.critical(self, "内部エラー", f"タブが見つかりません: {year_label}")
            if submit_btn:
                submit_btn.setEnabled(True)
            return

        scroll_area = tab.findChild(QScrollArea)
        scroll_content = scroll_area.widget() if scroll_area else None

        important_subjects = []
        if scroll_content:
            for frame in scroll_content.findChildren(QFrame):
                if hasattr(frame, 'important_cb') and frame.important_cb.isChecked():
                    important_subjects.append(frame.subject_name)
        
        # 現在の生成対象の学年を記録
        self._current_generation_year_label = year_label
        
        # フィルタリング設定を辞書にまとめる
        filter_settings = {
            'active_filter_subject': self.ACTIVE_FILTER_SUBJECT,
            'active_filter_subject_amount': self.ACTIVE_FILTER_SUBJECT_AMOUNT,
            'active_min_subject': self.ACTIVE_MIN_SUBJECT,
            'active_max_subject': self.ACTIVE_MAX_SUBJECT,
            'active_filter_subject_units': self.ACTIVE_FILTER_SUBJECT_UNITS,
            'active_min_subject_units': self.ACTIVE_MIN_SUBJECT_UNITS,
            'active_max_subject_units': self.ACTIVE_MAX_SUBJECT_UNITS,
            'min_subject_count': self.setting_min_subject_count,
            'max_subject_count': self.max_subject_count,
            'min_subject_units': self.min_subject_units,
            'max_subject_units': self.max_subject_units,
        }
        
        # unit_dataを作成（TimetableLogic用）
        abnormal_units_general = {
            self._normalize_subject_name(k): v 
            for k, v in self.config.get("ABNORMAL_SUBJECTS_UNITS", {}).items()
        }
        all_year_abnormal_units = []
        for year in self.years:
            year_key = f"ABNORMAL_SUBJECTS_UNITS{year}"
            year_abnormal = {
                self._normalize_subject_name(k): v 
                for k, v in self.config.get(year_key, {}).items()
            }
            all_year_abnormal_units.append(year_abnormal)
        
        unit_data = {
            'abnormal_units_general': abnormal_units_general,
            'all_year_abnormal_units': all_year_abnormal_units,
        }
        
        # TimetableWorkerを作成
        self._timetable_worker = TimetableWorker(
            selected_subjects=selected_subjects,
            fixed_slots=fixed_slots,
            all_slots=all_slots,
            year_label=year_label,
            important_subjects=important_subjects,
            norm_to_raw_map=norm_to_raw_map,
            max_limit=self.max_memory_limit,
            filter_settings=filter_settings,
            unit_data=unit_data,
            parent=self
        )
        
        # シグナルを接続
        self._timetable_worker.progress.connect(self._on_worker_progress)
        self._timetable_worker.finished_with_results.connect(self._on_worker_finished)
        self._timetable_worker.error.connect(self._on_worker_error)
        
        # ワーカーを開始
        self._timetable_worker.start()
    
    def _on_worker_progress(self, count):
        """時間割生成の進捗を更新"""
        year_label = self._current_generation_year_label
        if year_label and year_label in self.year_data:
            status_label = self.year_data[year_label].get("status_label")
            if status_label:
                status_label.setText(f"時間割を生成中... {count}件発見")
    
    def _on_worker_finished(self, timetables, year_label, important_subjects):
        """時間割生成完了時の処理"""
        status_label = self.year_data[year_label].get("status_label") if year_label in self.year_data else None
        submit_btn = getattr(self, f"submit_btn_{year_label}", None)
        
        try:
            if timetables:
                if status_label:
                    text = f"{len(timetables)}通りの組み合わせが見つかりました。"
                    if len(timetables) >= self.max_memory_limit:
                        text += f" (上限: {self.max_memory_limit}件)"
                    status_label.setText(text)
                self.display_timetables(timetables, year_label, important_subjects)
            else:
                error_message = "組み合わせが見つかりませんでした。\n"
                error_message += "以下の条件を確認してください:\n"
                active_filters = []
                if self.ACTIVE_FILTER_SUBJECT and self.ACTIVE_FILTER_SUBJECT_AMOUNT:
                    if self.ACTIVE_MIN_SUBJECT:
                        active_filters.append(f"・最小教科数: {self.setting_min_subject_count}")
                    if self.ACTIVE_MAX_SUBJECT:
                        active_filters.append(f"・最大教科数: {self.max_subject_count}")
                
                if self.ACTIVE_FILTER_SUBJECT and self.ACTIVE_FILTER_SUBJECT_UNITS:
                    if self.ACTIVE_MIN_SUBJECT_UNITS:
                        active_filters.append(f"・最小単位数: {self.min_subject_units}")
                    if self.ACTIVE_MAX_SUBJECT_UNITS:
                        active_filters.append(f"・最大単位数: {self.max_subject_units}")

                if important_subjects:
                    active_filters.append(f"・必含選択: {', '.join(important_subjects)}")

                if not active_filters:
                    error_message = "組み合わせが見つかりませんでした。\n選択した教科のコマが重複している可能性があります。"
                else:
                    error_message += "\n".join(active_filters)

                if status_label:
                    status_label.setText(error_message)
                else:
                    QMessageBox.information(self, "結果", error_message)
        finally:
            if submit_btn:
                submit_btn.setEnabled(True)
            self._timetable_worker = None
            self._current_generation_year_label = None
    
    def _on_worker_error(self, error_msg):
        """時間割生成エラー時の処理"""
        year_label = self._current_generation_year_label
        status_label = self.year_data[year_label].get("status_label") if year_label and year_label in self.year_data else None
        submit_btn = getattr(self, f"submit_btn_{year_label}", None) if year_label else None
        
        if status_label:
            status_label.setText(f"エラーが発生しました: {error_msg}")
        QMessageBox.critical(self, "エラー", f"時間割の生成中にエラーが発生しました: {error_msg}")
        
        if submit_btn:
            submit_btn.setEnabled(True)
        self._timetable_worker = None
        self._current_generation_year_label = None

    def display_combination(self, timetables, year_label):
        if not timetables:
            QMessageBox.information(self, "結果", "条件に合う組み合わせは見つかりませんでした。")
            return
        
        # ウィンドウ作成時点のプロファイル名をキャプチャ
        captured_profile_name = self.get_active_profile_name()

        display_window = QMainWindow(self)
        display_window.setWindowTitle(f"{year_label} 時間割の組み合わせ")
        display_window.resize(800, 600)
        self.display_windows.append(display_window)

        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)
        display_window.setCentralWidget(main_widget)

        tab_widget = QTabWidget()
        main_layout.addWidget(tab_widget)

        page_count = math.ceil(len(timetables) / self.SUBJECT_COMBINATION_COUNT)
        current_page = 0

        page_nav_frame = QFrame()
        main_layout.addWidget(page_nav_frame)
        page_nav_layout = QHBoxLayout(page_nav_frame)

        more_button_frame = QFrame()
        main_layout.addWidget(more_button_frame)
        more_button_layout = QHBoxLayout(more_button_frame)

        def show_page(page_index):
            nonlocal current_page
            current_page = page_index
            
            while tab_widget.count() > 0:
                tab_widget.removeTab(0)

            start_index = page_index * self.SUBJECT_COMBINATION_COUNT
            end_index = start_index + self.SUBJECT_COMBINATION_COUNT
            
            self.combination_save_buttons = []

            for i, timetable in enumerate(timetables[start_index:end_index], start=start_index):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                
                scroll_area = QScrollArea()
                scroll_area.setWidgetResizable(True)
                scroll_content = QWidget()
                scroll_layout = QVBoxLayout(scroll_content)

                label.setFont(QFont("Arial", 12, QFont.Bold))
                scroll_layout.addWidget(label)

                time_frame = QFrame()
                time_frame.setFrameShape(QFrame.StyledPanel)
                
                data = self.year_data[year_label]
                table_layout_data = data["table_layout"]
                all_slots_data = data["all_slots"]
                
                common_subjects = self._get_common_subjects()
                self.create_table_slot(time_frame, timetable, table_layout_data, year_label, all_slots_data, common_subjects)
                scroll_layout.addWidget(time_frame)

                button_frame = QFrame()
                button_layout = QHBoxLayout(button_frame)
                
                save_button = QPushButton("この組み合わせを保存")
                save_button.clicked.connect(lambda _, t=timetable, y=year_label, p=captured_profile_name: self.save_timetable(t, y, p))
                button_layout.addWidget(save_button)
                self.combination_save_buttons.append(save_button)
                
                scroll_layout.addWidget(button_frame)

                # 不足している必須科目
                missing_req_frame = QFrame()
                missing_req_layout = QVBoxLayout(missing_req_frame)
                missing_req_label = QLabel("<b>不足している必須科目:</b><br>")
                missing_reqs = self.check_required_subjects(timetable, year_label, include_all=True)
                
                if missing_reqs:
                    missing_req_label.setStyleSheet("font-weight: bold; color: red;")
                    formatted_message = self.format_missing_subjects_message(missing_reqs)
                    missing_req_label.setText(missing_req_label.text() + formatted_message.replace("\n", "<br>"))
                    missing_req_frame.setStyleSheet("background-color: #FFCCCC; border: 1px solid #FF9999; padding: 5px;")
                else:
                    missing_req_label.setText(missing_req_label.text() + "なし")
                
                missing_req_layout.addWidget(missing_req_label)
                scroll_layout.addWidget(missing_req_frame)

                missing_prereq_frame = QFrame()
                missing_prereq_layout = QVBoxLayout(missing_prereq_frame)
                missing_prereq_label = QLabel("不足している前提科目:\n")
                missing_prereqs = self.check_prerequisites(timetable, year_label)
                if missing_prereqs:
                    for subj, prereqs in missing_prereqs.items():
                        missing_prereq_label.setText(missing_prereq_label.text() + f"- {subj} (前提: {', '.join(prereqs)})\n")
                else:
                    missing_prereq_label.setText(missing_prereq_label.text() + "なし")
                missing_prereq_layout.addWidget(missing_prereq_label)
                scroll_layout.addWidget(missing_prereq_frame)

                scroll_content.setLayout(scroll_layout)
                scroll_area.setWidget(scroll_content)
                layout.addWidget(scroll_area)
                tab_widget.addTab(tab, f"{i+1}")

            # ナビゲーションをクリアして更新
            for i in reversed(range(page_nav_layout.count())):
                page_nav_layout.itemAt(i).widget().setParent(None)
            for i in reversed(range(more_button_layout.count())):
                more_button_layout.itemAt(i).widget().setParent(None)

            if page_count > 1:
                prev_page_button = QPushButton("前のページ")
                prev_page_button.setEnabled(current_page > 0)
                prev_page_button.clicked.connect(lambda: show_page(current_page - 1))
                page_nav_layout.addWidget(prev_page_button)

                page_label = QLabel(f"ページ {current_page + 1} / {page_count}")
                page_nav_layout.addWidget(page_label, alignment=Qt.AlignCenter)

                next_page_button = QPushButton("次のページ")
                next_page_button.setEnabled(current_page < page_count - 1)
                next_page_button.clicked.connect(lambda: show_page(current_page + 1))
                page_nav_layout.addWidget(next_page_button)

        show_page(0)
        display_window.show()

    def show_more_combinations(self):
        """さらに組み合わせを表示する"""
        if not hasattr(self, 'combination_list') or not hasattr(self, 'details_stack'):
            return
        
        # ウィンドウ作成時点のプロファイル名をキャプチャ
        captured_profile_name = self.get_active_profile_name()

        self.combination_list.clear()
        while self.details_stack.count() > 0:
            widget = self.details_stack.widget(0)
            self.details_stack.removeWidget(widget)
            widget.deleteLater()

        start_index = self.current_combination_index
        end_index = start_index + self.MORE_SUBJECT_COMBINATION
        
        self.combination_save_buttons = []

        for i, timetable in enumerate(self.timetables_to_display[start_index:end_index], start=start_index):
            self.combination_list.addItem(f"組み合わせ {i + 1}")

            details_widget = QWidget()
            details_layout = QVBoxLayout(details_widget)

            time_frame = QFrame()
            time_frame.setFrameShape(QFrame.StyledPanel)
            time_frame.setLayout(QVBoxLayout())
            
            data = self.year_data[self.year_label_to_display]
            table_layout_data = data["table_layout"]
            all_slots_data = data["all_slots"]
            
            common_subjects = self._get_common_subjects()
            self.create_table_slot(time_frame, timetable, table_layout_data, self.year_label_to_display, all_slots_data, common_subjects)
            details_layout.addWidget(time_frame)

            button_frame = QFrame()
            button_layout = QHBoxLayout(button_frame)
            
            save_button = QPushButton("この組み合わせを保存")
            save_button.clicked.connect(lambda _, t=timetable, y=self.year_label_to_display, p=captured_profile_name: self.save_timetable(t, y, p))
            button_layout.addWidget(save_button)
            self.combination_save_buttons.append(save_button)
            
            details_layout.addWidget(button_frame)

            missing_req_frame = QFrame()
            missing_req_layout = QVBoxLayout(missing_req_frame)
            
            missing_req_label = QLabel("不足している必須科目:\n")
            missing_reqs = self.check_required_subjects(timetable, self.year_label_to_display, include_all=True)
            
            # 不足要件をcondition_numとis_or_groupでグループ化
            grouped_missing_info = {}
            for info in missing_reqs:
                key = (info['condition_num'], info['is_or_group'])
                if key not in grouped_missing_info:
                    grouped_missing_info[key] = []
                grouped_missing_info[key].append(info)

            if missing_reqs: # 不足要件があるか確認
                missing_req_label.setStyleSheet("font-weight: bold; color: red;")
                
                for key, infos_list in grouped_missing_info.items():
                    condition_num, is_or_group = key
                    
                    if is_or_group:
                        # OR条件では不足している選択肢を列挙
                        missing_req_label.setText(missing_req_label.text() + f"--- 条件 {condition_num} ---")
                        for info in infos_list:
                            subjects_str = ', '.join(info['subjects'])
                            missing_req_label.setText(missing_req_label.text() + f"  - 教科: {subjects_str} (現在: {info['matched']} / 必要: {info['required']})")
                        missing_req_label.setText(missing_req_label.text() + "\n") # 区切り用の空行を追加
                    else:
                        info = infos_list[0] # 非OR条件ではキーごとにinfoは一つだけ
                        subjects_str = ', '.join(info['subjects'])
                        missing_req_label.setText(missing_req_label.text() + f"- {subjects_str} (必要: {info['required']}, 一致: {info['matched']})")
                
                missing_req_frame.setStyleSheet("background-color: #FFCCCC; border: 1px solid #FF9999; padding: 5px;") # 薄赤の背景、枠線、余白を設定
            else:
                missing_req_label.setText(missing_req_label.text() + "なし")
            
            missing_req_layout.addWidget(missing_req_label)
            details_layout.addWidget(missing_req_frame)

            missing_prereq_frame = QFrame()
            missing_prereq_layout = QVBoxLayout(missing_prereq_frame)
            missing_prereq_label = QLabel("不足している前提科目:\n")
            missing_prereqs = self.check_prerequisites(timetable, self.year_label_to_display)
            if missing_prereqs:
                for subj, prereqs in missing_prereqs.items():
                    missing_prereq_label.setText(missing_prereq_label.text() + f"- {subj} (前提: {', '.join(prereqs)})\n")
            else:
                missing_prereq_label.setText(missing_prereq_label.text() + "なし")
            missing_prereq_layout.addWidget(missing_prereq_label)
            details_layout.addWidget(missing_prereq_frame)

            self.details_stack.addWidget(details_widget)

        if hasattr(self, 'page_label'):
            total_pages = math.ceil(len(self.timetables_to_display) / self.MORE_SUBJECT_COMBINATION)
            current_page = (self.current_combination_index // self.MORE_SUBJECT_COMBINATION) + 1
            self.page_label.setText(f"{current_page} / {total_pages}")

    def show_prev_page(self):
        """前のページを表示"""
        if self.current_combination_index > 0:
            self.current_combination_index -= self.MORE_SUBJECT_COMBINATION
            self.show_more_combinations()

    def show_next_page(self):
        """次のページを表示"""
        if self.current_combination_index + self.MORE_SUBJECT_COMBINATION < len(self.timetables_to_display):
            self.current_combination_index += self.MORE_SUBJECT_COMBINATION
            self.show_more_combinations()

    def _on_important_check(self, state, subject, year_label):
        """必含チェックボックスが変更された時の処理"""
        if year_label not in self.year_data:
            return
        data = self.year_data[year_label]
        data["important_vars"].update({subject: bool(state)})

        # 同じ教科の選択チェックボックスも自動的にチェックする
        if state and not data["check_vars"].get(subject):
            data["check_vars"].update({subject: True})
            # UI更新のためにチェックボックスの状態を変更
            tab = self.year_tabs.get(year_label)
            if tab:
                scroll_area = tab.findChild(QScrollArea)
                if scroll_area:
                    scroll_content = scroll_area.widget()
                    if scroll_content:
                        for frame in scroll_content.findChildren(QFrame):
                            if hasattr(frame, 'subject_name') and frame.subject_name == subject:
                                if hasattr(frame, 'subject_cb'):
                                    frame.subject_cb.setChecked(True)
                                break

    def _show_more(self):
        """表示する組み合わせ数を増やす"""
        if hasattr(self, 'current_display_window'):
            window = self.current_display_window
            if hasattr(window, 'displayed_count'):
                window.displayed_count += self.MORE_SUBJECT_COMBINATION
                window.show_timetables()

    def _change_page(self, delta):
        """ページを変更する（MORE_SUBJECT_COMBINATIONの数だけ表示し、スクロールバーを一番上に戻す）"""
        if hasattr(self, 'current_display_window'):
            window = self.current_display_window
            if hasattr(window, 'current_page') and hasattr(window, 'total_pages'):
                new_page = window.current_page + delta
                if 0 <= new_page < window.total_pages:
                    window.current_page = new_page
                    window.displayed_count = self.MORE_SUBJECT_COMBINATION  # 表示数をリセット
                    window.show_timetables()
                    window.page_label.setText(f"ページ {window.current_page + 1}/{window.total_pages}")

                    # スクロールバーを一番上に戻す
                    scroll_area = window.findChild(QScrollArea)
                    if scroll_area:
                        scroll_area.verticalScrollBar().setValue(0)

    def collect_duplicate_info(self, year_label, other_combo):
        """重複選択情報を収集"""
        duplicate_info = {}
        if not other_combo or 1 not in other_combo:
            return duplicate_info
        
        fixed_subjects = set(self.FIXED_SLOTS2.values()) | set(self.FIXED_SLOTS3.values())
        other_timetable = other_combo[1]
        
        # 他方の学年の科目と枠を収集
        for slot, subject in other_timetable.items():
            if subject and subject not in fixed_subjects:
                if subject not in duplicate_info:
                    duplicate_info[subject] = []
                duplicate_info[subject].append(slot)
        
        return duplicate_info

    def save_combination(self, year_label, timetable):
        """組み合わせを保存し、ハイライトを更新"""
        if year_label == "2年":
            self.complete_combination_2 = {1: timetable}  # ← 常に1に固定
        elif year_label == "3年":
            self.complete_combination_3 = {1: timetable}

        # ハイライトの更新
        self.update_all_highlights()

    def save_timedate(self, year_label):
        """時間割をExcelファイルに保存（新しいデータ形式対応）"""
        active_profile_name = self.get_active_profile_name()
        selected_art = self.art_subject_selections.get(active_profile_name, [])
        
        years_to_process = self.years if year_label == "全体" else [year_label]

        # --- 1. 選択されている時間割があるかチェック ---
        has_content = any(self.year_data[y]["saved_profiles"].get(active_profile_name) for y in years_to_process)
        if not has_content:
            reply = QMessageBox.question(
                self, "確認",
                "保存対象のプロファイルに時間割がありません。空の状態で保存しますか？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # --- 2. 共通科目チェック ---
        if len(self.years) > 1 and year_label == "全体":
            all_fixed_subjects = set().union(*(d["fixed_slots"].values() for d in self.year_data.values()))
            
            subject_sets = []
            for y in self.years:
                timetable = self.year_data[y].get("complete_combination", {}).get(1, {})
                if timetable:
                    subject_sets.append({
                        subj for subj in timetable.values() 
                        if subj and subj.strip() not in ["－", "-", ""] and subj not in all_fixed_subjects
                    })
            
            if len(subject_sets) > 1:
                common_subjects = set.intersection(*subject_sets)
                if common_subjects:
                    subject_list = sorted(common_subjects)
                    msg = f"以下の教科が両学年で共通しています:\n{', '.join(subject_list)}\n\n保存しますか？"
                    reply = QMessageBox.question(
                        self, "共通教科の確認", msg,
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    if reply == QMessageBox.StandardButton.No:
                        return

        # --- 3. 全学年の科目を収集 ---
        all_subjects = set()
        if selected_art:
            all_subjects.update(selected_art)

        for y in self.years:
            timetable = self.year_data[y]["saved_profiles"].get(active_profile_name, {})
            all_subjects.update(
                subj for subj in timetable.values() if subj and subj.strip() not in ["－", "-", ""]
            )

        # --- 4. 同時選択不可教科チェック ---
        conflicting_groups_found = []
        normalized_all_subjects = {self._normalize_subject_name(s) for s in all_subjects}

        for subject_group in self.NO_TOGETHER_SUBJECTS:
            normalized_group = [self._normalize_subject_name(s) for s in subject_group]
            selected_in_group = [s for s in normalized_group if s in normalized_all_subjects]
            
            if len(selected_in_group) > 1:
                # 表示用に元の教科名を取得
                original_names_in_conflict = [s_raw for s_raw in subject_group if self._normalize_subject_name(s_raw) in selected_in_group]
                conflicting_groups_found.append(original_names_in_conflict)

        if conflicting_groups_found:
            message_parts = ["以下の教科は同時に履修できません:\n"]
            for group in conflicting_groups_found:
                message_parts.append(f"- [{', '.join(group)}]")
            
            msg = "\n".join(message_parts)
            msg += "\n\nこのまま保存しますか？"
            
            reply = QMessageBox.warning(
                self,
                "同時履修不可の教科",
                msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # --- 5. 前提科目チェック ---
        missing_prerequisites = []
        for main_subject, prerequisites in self.PREREQUISITE_SUBJECTS.items():
            if self._normalize_subject_name(main_subject) in normalized_all_subjects:
                for prereq in prerequisites:
                    if self._normalize_subject_name(prereq) not in normalized_all_subjects:
                        missing_prerequisites.append(f"「{main_subject}」には前提科目「{prereq}」が必要です。")
        
        if missing_prerequisites:
            msg = "以下の前提科目が不足しています（全学年チェック）:\n\n" + "\n".join(missing_prerequisites)
            msg += "\n\nこのまま保存しますか？"
            reply = QMessageBox.warning(
                self,
                "前提科目不足（全学年）",
                msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # --- 6. 必須科目チェック ---
        all_missing_info = []
        for y in years_to_process:
            timetable = self.year_data[y]["saved_profiles"].get(active_profile_name, {})
            all_missing_info.extend(self.check_required_subjects(timetable, y, include_all=True))

        

        unique_missing_info = []
        seen = set()
        for d in all_missing_info:
            hashable_d = tuple(sorted((k, tuple(v) if isinstance(v, list) else v) for k, v in d.items() if k not in ['color', 'source_type']))
            if hashable_d not in seen:
                unique_missing_info.append(d)
                seen.add(hashable_d)

        if unique_missing_info:
            grouped_missing_info = {}
            for info in unique_missing_info:
                # 非ORグループでは一意性確保のため教科もキーに含める
                if info['is_or_group']:
                    key = (info['condition_num'], True)
                else:
                    key = (info['condition_num'], False, frozenset(info['subjects']))

                if key not in grouped_missing_info:
                    grouped_missing_info[key] = []
                grouped_missing_info[key].append(info)

            if unique_missing_info:
                formatted_message = self.format_missing_subjects_message(unique_missing_info)
                msg = "【警告】以下の必須科目が不足しています！\n\n" + formatted_message + "\n\nこのまま保存しますか？"

                reply = QMessageBox.warning(
                    self,
                    "必須科目不足",
                    msg,
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return

        # --- 7. ファイル保存 ---
        file_path, _ = QFileDialog.getOpenFileName(
            self, "保存するExcelファイルを選択", str(Path.cwd()), "Excelファイル (*.xlsx)")

        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.worksheets[0]

            for y in years_to_process:
                data = self.year_data[y]
                timetable = data["saved_profiles"].get(active_profile_name, {})
                subject_number = data["subject_number"]
                save_position = data["save_position"]

                for slot, positions in save_position.items():
                    subject = timetable.get(slot)
                    number = None
                    if subject and subject.strip() not in ["－", "-", ""]:
                        for entry in subject_number:
                            if entry["name"] == subject:
                                number = entry["data"].get(slot, "")
                                break
                    
                    for pos_info in positions:
                        cell_addr = pos_info[0] 
                        # 結合を気にせず左上のセルに直接書き込む
                        ws[cell_addr] = number if number else None

            wb.save(file_path)
            try:
                self.setUpdatesEnabled(False)
                QApplication.processEvents()
                QMessageBox.information(self, "保存完了", f"時間割を保存しました: {file_path}")
            finally:
                self.setUpdatesEnabled(True)
                QApplication.processEvents()
            self._has_active_combinations = False # 保存成功時にフラグをFalseに設定

        except Exception as e:
            try:
                self.setUpdatesEnabled(False)
                QApplication.processEvents()
                QMessageBox.critical(self, "保存エラー", f"ファイルの保存中にエラーが発生しました:\n{e}")
            finally:
                self.setUpdatesEnabled(True)
                QApplication.processEvents()



    def _write_timetable_to_sheet(self, worksheet, timetable, table_layout, fixed_slots):
        """時間割をExcelシートに書き込む（空白でも明示的に消去）"""
        import openpyxl.styles

        for row_idx, row in enumerate(table_layout, 1):
            for col_idx, slot in enumerate(row, 1):
                if not slot:
                    continue  # 空の枠はスキップ

                # 科目取得
                if slot in fixed_slots:
                    subject = fixed_slots[slot]
                else:
                    subject = timetable.get(slot, "")

                # 「None」や「－」や空白も含めて常に空白で上書き
                if subject in ["－", "-", " ", None]:
                    subject = ""

                # 書き込み実行
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = subject  # ← 空白でも上書き！

                # デバッグログ（確認用）
                print(f"[DEBUG WRITE] row={row_idx}, col={col_idx}, slot={slot}, value='{subject}'")

                # 固定科目の背景色（視覚的区別）
                if slot in fixed_slots:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )


    def is_dark_theme(self):
        """Determines if a dark theme is active.
        【委譲】ThemeManagerに処理を委譲します。
        """
        # ThemeManagerが初期化前に呼ばれる場合はフォールバック
        if hasattr(self, 'theme_manager'):
            return self.theme_manager.is_dark_theme()
        # フォールバック：darkdetectを直接使用
        ini_theme = darkdetect.theme()
        if ini_theme == "Dark":
            return True
        if ini_theme == "Light":
            return False
        return not is_windows_light_theme()

    def _normalize_subject_name(self, name):
        """【委譲】教科名の正規化をTimetableLogicに委譲します。"""
        return TimetableLogic.normalize_subject_name(name)

    def _create_icon_from_svg_data(self, svg_path, color_hex):
        """指定された色のSVGアイコンを作成する
        【委譲】ThemeManagerに処理を委譲します。
        """
        if hasattr(self, 'theme_manager'):
            return self.theme_manager.create_icon_from_svg(svg_path, color_hex)
        # フォールバック（ThemeManager初期化前）
        return None

    def get_active_profile_name(self):
        """現在アクティブなプロファイル名を取得する"""
        # self.current_profile_name は _on_profile_tab_changed で確実に更新される
        if hasattr(self, 'current_profile_name') and self.current_profile_name:
            return self.current_profile_name
        # フォールバック：タブウィジェットから取得（初期化途中の場合）
        if hasattr(self, 'profile_tab_widget'):
            current_index = self.profile_tab_widget.currentIndex()
            if current_index != -1:
                return self.profile_tab_widget.tabText(current_index)
        return "Default"

    def get_current_profile_lock_settings(self, year_label):
        """現在アクティブなプロファイルのロック設定を取得する"""
        profile_name = self.get_active_profile_name()
        # ロック設定辞書にプロファイルが存在することを保証
        if profile_name not in self.year_data[year_label]["profile_lock_settings"]:
            self.year_data[year_label]["profile_lock_settings"][profile_name] = {}
        return self.year_data[year_label]["profile_lock_settings"][profile_name]

    def get_current_profile_all_slots(self, year_label):
        """現在アクティブなプロファイルの有効なスロットリストを取得する"""
        profile_name = self.get_active_profile_name()
        if profile_name not in self.year_data[year_label].get("profile_all_slots", {}):
            # プロファイルがなければ基本のall_slotsから初期化
            base_slots = self.year_data[year_label].get("all_slots", [])
            self.year_data[year_label].setdefault("profile_all_slots", {})[profile_name] = base_slots[:]
        return self.year_data[year_label]["profile_all_slots"][profile_name]



    def check_subjects_units(self):
        """履修科目全体の単位数を計算して、現在のアクティブタブのラベルを更新
        【委譲】TimetableLogic.calculate_total_unitsに処理を委譲します。
        """
        if not hasattr(self, 'profile_tab_widget'):
            return 0

        current_page_widget = self.profile_tab_widget.currentWidget()
        if not current_page_widget or not hasattr(current_page_widget, 'units_label'):
            return 0

        units_label = current_page_widget.units_label
        active_profile_name = self.get_active_profile_name()
        
        if not active_profile_name:
            units_label.setText("単位数: 0")
            return 0

        # TimetableLogicに計算を委譲
        total_units = TimetableLogic.calculate_total_units(
            base_units=self.base_units_value,
            years=self.years,
            year_data=self.year_data,
            active_profile=active_profile_name,
            art_selections=self.art_subject_selections,
            abnormal_units_general=self.ABNORMAL_SUBJECTS_UNITS_GENERAL
        )
        
        units_label.setText(f"単位数: {total_units}")
        return total_units

    def calculate_timetable_units(self, timetable, year_label, include_fixed_subjects=False):
        """時間割の単位数を計算し、オプションで固定単位数を加算する
        【委譲】TimetableLogic.calculate_timetable_unitsに処理を委譲します。
        """
        year_fixed_slots = self.year_data[year_label].get("fixed_slots", {})
        all_year_abnormal_units = [self.year_data[y].get("abnormal_units", {}) for y in self.years]
        
        return TimetableLogic.calculate_timetable_units(
            timetable=timetable,
            year_fixed_slots=year_fixed_slots,
            include_fixed_subjects=include_fixed_subjects,
            abnormal_units_general=self.ABNORMAL_SUBJECTS_UNITS_GENERAL,
            all_year_abnormal_units=all_year_abnormal_units
        )

    def is_exit_requested(self):
        """ランチャーがこの値をチェックして終了するか決めます"""
        return self._exit_requested

    def _select_config_file(self, prefer_last_opened=True):
            """複数のconfig.json候補から使用するファイルを選択する"""
            if self.is_tutorial_running:
                self.tutorial_signal.emit({'action': 'menu_clicked'})
                dialog = ConfigSelectionDialog(self, "学校の選択", "使用する学校を選択してください:", ["（見本）A高校", "（見本）B高校"])
                dialog.finished.connect(lambda result: self.tutorial_signal.emit({'action': 'dialog_closed'}))
                if dialog.exec():
                    return None 
                else:
                    self._exit_requested = True # キャンセル
                    return None

            # --- 前回開いたファイルをチェック ---
            if prefer_last_opened and self.user_settings and "last_opened_config" in self.user_settings:
                last_config_path_str = self.user_settings["last_opened_config"]
                if last_config_path_str:
                    last_config_path = Path(last_config_path_str)
                    if last_config_path.is_file():
                        try:
                            with open(last_config_path, 'r', encoding='utf-8') as f:
                                json.load(f)
                            return str(last_config_path)
                        except (json.JSONDecodeError, IOError):
                            pass

            # 1. 候補パスを収集
            base_path = self.base_path
            search_dirs = {str(Path(base_path).resolve()), str(Path(os.getcwd()).resolve())}
            possible_paths = set()
            for directory in search_dirs:
                try:
                    for filename in os.listdir(directory):
                        if filename.endswith('.json') and not filename.endswith('.tm.json') and not filename.endswith('config.json'):
                            possible_paths.add(os.path.join(directory, filename))
                except OSError:
                    continue
            
            valid_configs = []
            seen_paths = set()
            for path_str in possible_paths:
                path = Path(path_str).resolve()
                if path in seen_paths or not path.is_file():
                    continue
                seen_paths.add(path)
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        valid_configs.append({
                            "path": str(path),
                            "school": data.get("SCHOOL_NAME", "（学校名未設定）"),
                            "date": data.get("LAST_UPDATE_DAY", "（更新日時なし）")
                        })
                except (json.JSONDecodeError, IOError):
                    continue

            # 2. 候補がゼロの場合
            if not valid_configs:
                # スプラッシュを一時的に隠す
                if self.splash: self.splash.hide()
                
                msg_box = QMessageBox(self)
                msg_box.setWindowFlags(msg_box.windowFlags() | Qt.WindowStaysOnTopHint) # 最前面
                msg_box.setIcon(QMessageBox.Information)
                msg_box.setWindowTitle("設定ファイルが見つかりません")
                msg_box.setText("有効な設定ファイルが見つかりません。手動で選択してください。")
                msg_box.exec()

                file_dialog = QFileDialog(None, "設定ファイルを開く", str(Path.cwd()), "JSON Files (*.json)")
                file_dialog.setProxyModel(ExcludeFileProxyModel(["config.json"], file_dialog))
                file_dialog.setWindowFlags(file_dialog.windowFlags() | Qt.WindowStaysOnTopHint)
                icon_path = os.path.join(self.base_path, "時間割くんアイコン.ico")
                if os.path.exists(icon_path):
                    file_dialog.setWindowIcon(QIcon(str(icon_path)))

                file_path = None
                if file_dialog.exec():
                    selected_files = file_dialog.selectedFiles()
                    file_path = selected_files[0] if selected_files else None
                
                if self.splash: self.splash.show() # スプラッシュ復帰
                
                if not file_path:
                    self._exit_requested = True
                return file_path

            # 2.5 1つだけ見つかった場合（既存機能：サイレント自動選択を避ける）
            if len(valid_configs) == 1:
                schools = {}
                for config in valid_configs:
                    school_name = config["school"]
                    if school_name not in schools:
                        schools[school_name] = []
                    schools[school_name].append(config)
                school_names = sorted(schools.keys())
                school_dialog = ConfigSelectionDialog(self, "学校の確認", "使用する学校を確認してください:", school_names)
                if not school_dialog.exec():
                    # キャンセル時は手動選択へ
                    if self.splash: self.splash.hide()
                    file_dialog = QFileDialog(None, "設定ファイルを開く", "", "JSON Files (*.json)")
                    file_dialog.setProxyModel(ExcludeFileProxyModel(["config.json"], file_dialog))
                    file_dialog.setWindowFlags(file_dialog.windowFlags() | Qt.WindowStaysOnTopHint)
                    res = None
                    if file_dialog.exec(): res = file_dialog.selectedFiles()[0]
                    if self.splash: self.splash.show()
                    if not res: self._exit_requested = True
                    return res
                
                selected_school = school_dialog.get_selected_item()
                configs_for_school = schools[selected_school]
                if len(configs_for_school) == 1:
                    return configs_for_school[0]["path"]
                # 複数バージョンある場合は後のロジックへ流れる

            # 3. 学校名でグループ化
            schools = {}
            for config in valid_configs:
                school_name = config["school"]
                if school_name not in schools:
                    schools[school_name] = []
                schools[school_name].append(config)

            # 4. 学校名選択ダイアログ
            school_names = sorted(schools.keys())
            school_dialog = ConfigSelectionDialog(self, "学校の選択", "複数の設定ファイルが見つかりました。\n使用する学校を選択してください:", school_names)
            if not school_dialog.exec():
                # ここから元の複雑なアイコン・手動選択ロジックを維持
                if self.splash: self.splash.hide()
                file_dialog = QFileDialog(None, "設定ファイルを開く", "", "JSON Files (*.json)")
                file_dialog.setProxyModel(ExcludeFileProxyModel(["config.json"], file_dialog))
                file_dialog.setWindowFlags(file_dialog.windowFlags() | Qt.WindowStaysOnTopHint)
                
                icon_path = os.path.join(self.base_path, "時間割くんアイコン.ico")
                if os.path.exists(icon_path):
                    pixmap = QPixmap(icon_path)
                    if not pixmap.isNull():
                        file_dialog.setWindowIcon(QIcon(pixmap))

                file_path = None
                if file_dialog.exec():
                    selected_files = file_dialog.selectedFiles()
                    file_path = selected_files[0] if selected_files else None
                
                if self.splash: self.splash.show()
                if not file_path: self._exit_requested = True
                return file_path

            selected_school = school_dialog.get_selected_item()
            if not selected_school:
                self._exit_requested = True
                return None

            # 5. 日付選択ダイアログ
            configs_for_school = schools[selected_school]
            if len(configs_for_school) == 1:
                return configs_for_school[0]["path"]
            else:
                configs_for_school.sort(key=lambda x: x['date'], reverse=True)
                date_items = [f"{c['date']} ({c['path']})" for c in configs_for_school]
                date_dialog = ConfigSelectionDialog(self, "バージョンの選択", f"「{selected_school}」の複数のバージョンが見つかりました。\n使用する更新日時を選択してください:", date_items)

                if not date_dialog.exec():
                    # 既存機能：バージョン選択キャンセル時のメッセージと手動選択
                    if self.splash: self.splash.hide()
                    msg_box = QMessageBox(self)
                    msg_box.setWindowFlags(msg_box.windowFlags() | Qt.WindowStaysOnTopHint)
                    msg_box.setText("バージョン選択がキャンセルされました。手動で選択してください。")
                    msg_box.exec()
                    
                    file_dialog = QFileDialog(None, "設定ファイルを開く", "", "JSON Files (*.json)")
                    file_dialog.setProxyModel(ExcludeFileProxyModel(["config.json"], file_dialog))
                    file_dialog.setWindowFlags(file_dialog.windowFlags() | Qt.WindowStaysOnTopHint)
                    res = None
                    if file_dialog.exec(): res = file_dialog.selectedFiles()[0]
                    if self.splash: self.splash.show()
                    if not res: self._exit_requested = True
                    return res
                
                selected_date_item = date_dialog.get_selected_item()
                for config in configs_for_school:
                    if f"{config['date']} ({config['path']})" == selected_date_item:
                        return config['path']
            
            self._exit_requested = True
            return None

    def _save_user_settings(self):
        """現在の設定をユーザー設定ファイル(config.json)に保存する"""
        
        filter_settings = {}
        if hasattr(self, 'active_filter_subject_cb'):
            filter_settings = {
                "ACTIVE_MIN_SUBJECT": self.active_min_subject_cb.isChecked(),
                "ACTIVE_MAX_SUBJECT": self.active_max_subject_cb.isChecked(),
                "ACTIVE_FILTER_SUBJECT": self.active_filter_subject_cb.isChecked(),
                "ACTIVE_FILTER_SUBJECT_AMOUNT": self.active_filter_subject_amount_cb.isChecked(),
                "ACTIVE_MIN_SUBJECT_UNITS": self.active_min_units_cb.isChecked(),
                "ACTIVE_MAX_SUBJECT_UNITS": self.active_max_units_cb.isChecked(),
                "ACTIVE_FILTER_SUBJECT_UNITS": self.active_filter_subject_units_cb.isChecked(),
                "MIN_SUBJECT_COUNT": self.min_subject_spinbox.value(),
                "MAX_SUBJECT_COUNT": self.max_subject_spinbox.value(),
                "MIN_SUBJECT_COUNT_UNITS": self.min_units_spinbox.value(),
                "MAX_SUBJECT_COUNT_UNITS": self.max_units_spinbox.value(),
                "INCLUDE_FIXED": self.include_fixed_cb.isChecked() if hasattr(self, 'include_fixed_cb') else True
            }

        settings = {
            "APP_FONT_SIZE": self.main_font_size_spinbox.value() if hasattr(self, 'main_font_size_spinbox') else self._get_setting('APP_FONT_SIZE', 12),
            "MAX_MEMORY_LIMIT": self.memory_limit_slider.value() if hasattr(self, 'memory_limit_slider') else 1000,
            "last_opened_config": str(self.config_path) if hasattr(self, 'config_path') else None,
            "years_hierarchy": self.selected_department_path_key,
            "filter_settings": filter_settings,
        }
        # 古いself.configを避けるため、明示キーは現在のウィジェットから読み取る
        try:
            settings['TIMETABLE_ORDER'] = self.timetable_order_cb.isChecked()
        except Exception:
            if 'TIMETABLE_ORDER' in self.config:
                settings['TIMETABLE_ORDER'] = bool(self.config['TIMETABLE_ORDER'])

        try:
            settings['RUN_TUTORIAL_ON_STARTUP'] = self.run_tutorial_cb.isChecked()
        except Exception:
            if 'RUN_TUTORIAL_ON_STARTUP' in self.config:
                settings['RUN_TUTORIAL_ON_STARTUP'] = bool(self.config['RUN_TUTORIAL_ON_STARTUP'])

        department_path = self.selected_department_path_key or ""
        config_key = f"YEARS_SUBJECTS_UNITS_{department_path}"
        try:
            settings[config_key] = self.first_year_units_spinbox.value()
        except Exception:
            if config_key in self.config:
                settings[config_key] = self.config[config_key]

        # ConfigManagerに委譲して保存
        self.config_manager.save_user_settings(settings)
        self.user_settings = settings  # ローカルキャッシュも更新

    def _reset_user_settings(self):
        """ユーザー設定ファイル(config.json)をリセットする
        【委譲】ConfigManagerに処理を委譲します。
        """
        self.config_manager.reset_user_settings()
        self.user_settings = {}
        # 現在のUI状態で新しい設定ファイルを作成
        self._save_user_settings()

    def _save_profile_ui_state(self, profile_name):
        return self._save_profile_tab_state(profile_name)

    def _save_profile_tab_state(self, profile_name=None):
        """あとで復元できるよう現在のプロファイルタブUI状態を保存する。"""
        if profile_name is None:
            profile_name = self.get_active_profile_name()
        if not profile_name:
            return
        
        # プロファイルタブの状態を保存（プロファイルタブがある場合）
        if hasattr(self, 'profile_tab_widget'):
            current_profile_index = self.profile_tab_widget.currentIndex()
        else:
            current_profile_index = -1
        
        # 設定ウィンドウのスライドバー・スピンボックス状態を保存
        filter_settings = {}
        if hasattr(self, 'setting_window') and self.setting_window:
            filter_settings = {
                'min_subject': self.min_subject_spinbox.value() if hasattr(self, 'min_subject_spinbox') else None,
                'max_subject': self.max_subject_spinbox.value() if hasattr(self, 'max_subject_spinbox') else None,
                'min_units': self.min_units_spinbox.value() if hasattr(self, 'min_units_spinbox') else None,
                'max_units': self.max_units_spinbox.value() if hasattr(self, 'max_units_spinbox') else None,
                'active_filter_subject': self.active_filter_subject_cb.isChecked() if hasattr(self, 'active_filter_subject_cb') else None,
                'active_min_subject': self.active_min_subject_cb.isChecked() if hasattr(self, 'active_min_subject_cb') else None,
                'active_max_subject': self.active_max_subject_cb.isChecked() if hasattr(self, 'active_max_subject_cb') else None,
                'active_min_units': self.active_min_units_cb.isChecked() if hasattr(self, 'active_min_units_cb') else None,
                'active_max_units': self.active_max_units_cb.isChecked() if hasattr(self, 'active_max_units_cb') else None,
            }
        
        # profile_nameに対応する状態を保存
        if profile_name not in self.profile_tab_states:
            self.profile_tab_states[profile_name] = {}
        self.profile_tab_states[profile_name]['profile_index'] = current_profile_index
        self.profile_tab_states[profile_name]['filter_settings'] = filter_settings
        
        for year in self.years:
            if year in self.year_data:
                year_data = self.year_data[year]
                checked_subjects = set()
                important_subjects = set()
                prefixes = {}

                for frame in year_data.get("subject_frames", []):
                    # ウィジェットが有効か確認
                    if not widget_is_valid(frame):
                        continue
                    
                    subject_name = frame.subject_name
                    if frame.subject_cb.isChecked():
                        checked_subjects.add(subject_name)
                    if frame.important_cb.isChecked():
                        important_subjects.add(subject_name)
                    subject_prefixes = []
                    for prefix, cb in frame.prefix_cbs.items():
                        if cb.isChecked():
                            subject_prefixes.append(prefix)
                    if subject_prefixes:
                        prefixes[subject_name] = subject_prefixes

                # 素早く復元できるよう一時領域へ保存
                year_data.setdefault("_saved_ui_snapshot", {})[profile_name] = {
                    "checked": list(checked_subjects),
                    "important": list(important_subjects),
                    "prefixes": prefixes,
                }

    def _restore_profile_tab_state(self, profile_name=None):
        """保存済みのプロファイルタブUI状態を復元する。"""
        if profile_name is None:
            profile_name = self.get_active_profile_name()
        if not profile_name:
            return
        
        # プロファイルタブの状態を復元
        if hasattr(self, 'profile_tab_widget') and profile_name in self.profile_tab_states:
            saved_index = self.profile_tab_states[profile_name].get('profile_index', -1)
            if saved_index >= 0 and saved_index < self.profile_tab_widget.count():
                self.profile_tab_widget.setCurrentIndex(saved_index)
        
        # 注記: スライドバー・スピンボックスの状態復元はUIのスクロール位置変更につながるため、スキップします
        # 設定値は処理中に必要に応じてデータとして保存・使用されます
        
        for year in self.years:
            if year in self.year_data:
                year_data = self.year_data[year]
                snap = year_data.get("_saved_ui_snapshot", {}).get(profile_name)
                if not snap:
                    continue

                checked_subjects = set(snap.get("checked", []))
                important_subjects = set(snap.get("important", []))
                prefixes = snap.get("prefixes", {})

                for frame in year_data.get("subject_frames", []):
                    # ウィジェットが有効か確認
                    if not widget_is_valid(frame):
                        continue
                    
                    subject_name = frame.subject_name
                    frame.subject_cb.blockSignals(True)
                    frame.important_cb.blockSignals(True)
                    is_checked = subject_name in checked_subjects
                    frame.subject_cb.setChecked(is_checked)
                    is_important = subject_name in important_subjects
                    frame.important_cb.setChecked(is_important)
                    frame.important_cb.setEnabled(is_checked)
                    frame.subject_cb.blockSignals(False)
                    frame.important_cb.blockSignals(False)

                    subject_prefixes = prefixes.get(subject_name, [])
                    for prefix, cb in frame.prefix_cbs.items():
                        cb.blockSignals(True)
                        cb.setChecked(prefix in subject_prefixes)
                        cb.blockSignals(False)

                # 内部変数を同期
                if "check_vars" in year_data:
                    for subject in year_data["check_vars"]:
                        year_data["check_vars"][subject] = subject in checked_subjects
                if "important_vars" in year_data:
                    for subject in year_data["important_vars"]:
                        year_data["important_vars"][subject] = subject in important_subjects

                self.update_year_tab_visuals(year)

    def _save_profile_ui_state(self, profile_name):
        """指定されたプロファイルの現在の学年タブのUI状態を保存する"""
        if not profile_name:
            return
        for year in self.years:
            if year in self.year_data:
                year_data = self.year_data[year]
                
                checked_subjects = set()
                important_subjects = set()
                prefixes = {}

                for frame in year_data.get("subject_frames", []):
                    # ウィジェットが有効かチェック
                    if not widget_is_valid(frame):
                        continue
                    
                    subject_name = frame.subject_name
                    if frame.subject_cb.isChecked():
                        checked_subjects.add(subject_name)
                    if frame.important_cb.isChecked():
                        important_subjects.add(subject_name)
                    
                    subject_prefixes = []
                    for prefix, cb in frame.prefix_cbs.items():
                        if cb.isChecked():
                            subject_prefixes.append(prefix)
                    if subject_prefixes:
                        prefixes[subject_name] = subject_prefixes

                # 状態を保存
                year_data["profile_ui_states"][profile_name] = {
                    "checked": list(checked_subjects),
                    "important": list(important_subjects),
                    "prefixes": prefixes,
                }

    def _load_profile_ui_state(self, profile_name):
        """指定されたプロファイルのUI状態を学年タブに読み込む"""
        if not profile_name:
            return
        for year in self.years:
            if year in self.year_data:
                year_data = self.year_data[year]

                # プロファイルの状態がなければ初期化
                if profile_name not in year_data["profile_ui_states"]:
                    year_data["profile_ui_states"][profile_name] = {"checked": [], "important": [], "prefixes": {}}
                
                state = year_data["profile_ui_states"][profile_name]
                checked_subjects = set(state.get("checked", []))
                important_subjects = set(state.get("important", []))
                prefixes = state.get("prefixes", {})

                # UIウィジェットの状態を更新
                for frame in year_data.get("subject_frames", []):
                    # ウィジェットが有効かチェック
                    if not widget_is_valid(frame):
                        continue
                    
                    subject_name = frame.subject_name
                    
                    # 意図しない副作用を防ぐためシグナルを一時停止
                    frame.subject_cb.blockSignals(True)
                    frame.important_cb.blockSignals(True)
                    
                    is_checked = subject_name in checked_subjects
                    frame.subject_cb.setChecked(is_checked)
                    
                    is_important = subject_name in important_subjects
                    frame.important_cb.setChecked(is_important)
                    frame.important_cb.setEnabled(is_checked) # 教科が選択されている場合だけimportantを有効化

                    frame.subject_cb.blockSignals(False)
                    frame.important_cb.blockSignals(False)

                    subject_prefixes = prefixes.get(subject_name, [])
                    for prefix, cb in frame.prefix_cbs.items():
                        cb.blockSignals(True)
                        cb.setChecked(prefix in subject_prefixes)
                        cb.blockSignals(False)
                
                # 内部データをUIの状態と同期させる
                if "check_vars" in year_data:
                    for subject in year_data["check_vars"]:
                        year_data["check_vars"][subject] = subject in checked_subjects
                if "important_vars" in year_data:
                    for subject in year_data["important_vars"]:
                        year_data["important_vars"][subject] = subject in important_subjects

                # UIの視覚的な更新
                self.update_year_tab_visuals(year)

    def _get_save_all_button(self):
        """チュートリアル用に右上の「保存」ボタンを取得する"""
        if hasattr(self, 'profile_tab_widget'):
            current_page = self.profile_tab_widget.currentWidget()
            if current_page:
                # findChildrenで見つける方が安定的
                buttons = current_page.findChildren(QPushButton)
                for btn in buttons:
                    if btn.text() == "保存":
                        return btn
        return None

    def _get_timetable_slot_by_subject(self, subject_name=None):
        """チュートリアル用に、指定された教科名が表示されている最初の時間割スロットボタンを取得する"""
        page = self.profile_tab_widget.currentWidget()
        if not page: return None
        buttons = page.findChildren(TimetableSlotButton)

        if subject_name:
            for button in buttons:
                if button.text() == subject_name:
                    return button
        else: # subject_nameがない場合は、何かしら教科が入っているボタンを探す
            for button in buttons:
                if button.text() not in ["－", "-", ""]:
                    return button
        return None

    def _get_empty_timetable_slot(self):
        """チュートリアル用に、空いている最初の時間割スロットボタンを取得する"""
        page = self.profile_tab_widget.currentWidget()
        if not page: return None
        buttons = page.findChildren(TimetableSlotButton)
        for button in buttons:
            if button.text() in ["－", "-", ""]:
                return button
        return None

    def _get_disabled_timetable_slot(self):
        """チュートリアル用に、無効化されている（赤色表示）最初の時間割スロットボタンを取得する"""
        page = self.profile_tab_widget.currentWidget()
        if not page: return None
        buttons = page.findChildren(TimetableSlotButton)
        for button in buttons:
            if "rgba(255, 0, 0, 70)" in button.styleSheet():
                return button
        return None

def start_main_app():
    """本体アプリ実行ファイルのエントリーポイント。"""
    import tempfile
    
    # スクリプトから直接起動された場合の処理
    app = QApplication(sys.argv)
    set_button_styles(app) # ファイルスコープの関数で初期スタイルを適用

    # スプラッシュ画面は外部ランチャーが扱うためApplicationにはNoneを渡す
    # スプラッシュ画面は外部ランチャー側で表示する。
    window = Application(splash=None)

    # 初期化中にユーザーがダイアログをキャンセルした場合は正常終了する。
    if hasattr(window, 'is_exit_requested') and window.is_exit_requested():
        sys.exit(0)
    
    window.show()
    
    # IPC: メインウィンドウの準備完了をランチャーへ通知
    ipc_ready_file = Path(tempfile.gettempdir()) / "timemanager_ready.tmp"
    try:
        ipc_ready_file.touch()
        print(f"IPC: Created ready signal file at {ipc_ready_file}")
    except Exception as e:
        print(f"IPC: Failed to create ready signal file: {e}")
    
    sys.exit(app.exec())

if __name__ == '__main__':
    try:
        start_main_app()
    except Exception as e:
        import traceback
        traceback.print_exc()
        input("Press Enter to exit...")

